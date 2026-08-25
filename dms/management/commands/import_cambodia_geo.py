import os
import sys
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from dms.models import CambodiaProvince, CambodiaDistrict, CambodiaCommune, CambodiaVillage

class Command(BaseCommand):
    help = "Imports Cambodia administrative geography (Provinces, Districts, Communes, Villages) from Cambodia All List2025.xlsx"

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='Cambodia All List2025.xlsx',
            help='Path to Cambodia All List2025.xlsx'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        if not os.path.isabs(file_path):
            file_path = os.path.join(os.getcwd(), file_path)

        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f"File not found at: {file_path}"))
            return

        self.stdout.write(self.style.SUCCESS(f"Loading workbook: {file_path}..."))
        wb = openpyxl.load_workbook(file_path, data_only=True)

        with transaction.atomic():
            # 1. Sheet: CambodiaProvinceList2025
            self.stdout.write("1/4 Importing Provinces...")
            ws_prov = wb['CambodiaProvinceList2025']
            prov_rows = list(ws_prov.iter_rows(values_only=True))
            header_prov = [str(c).strip().lower() if c else '' for c in prov_rows[0]]
            
            p_code_idx = header_prov.index('province_code')
            p_kh_idx = header_prov.index('province_kh')
            p_en_idx = header_prov.index('province_en') if 'province_en' in header_prov else -1

            provinces_dict = {}
            for r in prov_rows[1:]:
                if not r[p_code_idx]:
                    continue
                code = str(r[p_code_idx]).strip()
                name_kh = str(r[p_kh_idx] or '').strip()
                name_en = str(r[p_en_idx] or '').strip() if p_en_idx >= 0 and len(r) > p_en_idx else ''
                provinces_dict[code] = CambodiaProvince(code=code, name_kh=name_kh, name_en=name_en)

            # 2. Sheet: CambodiaDistrictList2025
            self.stdout.write("2/4 Importing Districts...")
            ws_dist = wb['CambodiaDistrictList2025']
            dist_rows = list(ws_dist.iter_rows(values_only=True))
            header_dist = [str(c).strip().lower() if c else '' for c in dist_rows[0]]
            
            d_pcode_idx = header_dist.index('province_code')
            d_code_idx = header_dist.index('district_code')
            d_kh_idx = header_dist.index('district_kh')
            d_en_idx = header_dist.index('district_en') if 'district_en' in header_dist else -1

            districts_dict = {}
            for r in dist_rows[1:]:
                if not r[d_code_idx]:
                    continue
                p_code = str(r[d_pcode_idx]).strip()
                code = str(r[d_code_idx]).strip()
                name_kh = str(r[d_kh_idx] or '').strip()
                name_en = str(r[d_en_idx] or '').strip() if d_en_idx >= 0 and len(r) > d_en_idx else ''
                
                # Ensure province exists
                if p_code not in provinces_dict:
                    provinces_dict[p_code] = CambodiaProvince(code=p_code, name_kh=f"ខេត្ត/រាជធានី ({p_code})", name_en='')

                districts_dict[code] = CambodiaDistrict(code=code, province_id=p_code, name_kh=name_kh, name_en=name_en)

            # 3. Sheet: CambodiaCommuneList2025
            self.stdout.write("3/4 Importing Communes...")
            ws_com = wb['CambodiaCommuneList2025']
            com_rows = list(ws_com.iter_rows(values_only=True))
            header_com = [str(c).strip().lower() if c else '' for c in com_rows[0]]

            c_pcode_idx = header_com.index('province_code')
            c_dcode_idx = header_com.index('district_code')
            c_code_idx = header_com.index('commune_code')
            c_kh_idx = header_com.index('commune_kh')
            c_en_idx = header_com.index('commune_en') if 'commune_en' in header_com else -1

            communes_dict = {}
            for r in com_rows[1:]:
                if not r[c_code_idx]:
                    continue
                p_code = str(r[c_pcode_idx]).strip()
                d_code = str(r[c_dcode_idx]).strip()
                code = str(r[c_code_idx]).strip()
                name_kh = str(r[c_kh_idx] or '').strip()
                name_en = str(r[c_en_idx] or '').strip() if c_en_idx >= 0 and len(r) > c_en_idx else ''
                
                if p_code not in provinces_dict:
                    provinces_dict[p_code] = CambodiaProvince(code=p_code, name_kh=f"ខេត្ត/រាជធានី ({p_code})", name_en='')
                if d_code not in districts_dict:
                    districts_dict[d_code] = CambodiaDistrict(code=d_code, province_id=p_code, name_kh=f"ស្រុក/ខណ្ឌ ({d_code})", name_en='')

                communes_dict[code] = CambodiaCommune(code=code, district_id=d_code, province_id=p_code, name_kh=name_kh, name_en=name_en)

            # 4. Sheet: CambodiaVillagesList2025
            self.stdout.write("4/4 Importing Villages...")
            ws_vil = wb['CambodiaVillagesList2025']
            vil_rows = list(ws_vil.iter_rows(values_only=True))
            header_vil = [str(c).strip().lower() if c else '' for c in vil_rows[0]]

            v_pcode_idx = header_vil.index('province_code')
            v_dcode_idx = header_vil.index('district_code')
            v_ccode_idx = header_vil.index('commune_code')
            v_code_idx = header_vil.index('village_code')
            v_kh_idx = header_vil.index('village_kh')
            v_en_idx = header_vil.index('village_en') if 'village_en' in header_vil else -1

            villages_dict = {}
            for r in vil_rows[1:]:
                if not r[v_code_idx]:
                    continue
                p_code = str(r[v_pcode_idx]).strip()
                d_code = str(r[v_dcode_idx]).strip()
                c_code = str(r[v_ccode_idx]).strip()
                code = str(r[v_code_idx]).strip()
                name_kh = str(r[v_kh_idx] or '').strip()
                name_en = str(r[v_en_idx] or '').strip() if v_en_idx >= 0 and len(r) > v_en_idx else ''

                if p_code not in provinces_dict:
                    provinces_dict[p_code] = CambodiaProvince(code=p_code, name_kh=f"ខេត្ត/រាជធានី ({p_code})", name_en='')
                if d_code not in districts_dict:
                    districts_dict[d_code] = CambodiaDistrict(code=d_code, province_id=p_code, name_kh=f"ស្រុក/ខណ្ឌ ({d_code})", name_en='')
                if c_code not in communes_dict:
                    communes_dict[c_code] = CambodiaCommune(code=c_code, district_id=d_code, province_id=p_code, name_kh=f"ឃុំ/សង្កាត់ ({c_code})", name_en='')

                villages_dict[code] = CambodiaVillage(code=code, commune_id=c_code, district_id=d_code, province_id=p_code, name_kh=name_kh, name_en=name_en)

            # Database write
            self.stdout.write("Writing to database...")
            CambodiaVillage.objects.all().delete()
            CambodiaCommune.objects.all().delete()
            CambodiaDistrict.objects.all().delete()
            CambodiaProvince.objects.all().delete()

            CambodiaProvince.objects.bulk_create(provinces_dict.values())
            self.stdout.write(self.style.SUCCESS(f"  -> Saved {len(provinces_dict)} Provinces"))

            CambodiaDistrict.objects.bulk_create(districts_dict.values())
            self.stdout.write(self.style.SUCCESS(f"  -> Saved {len(districts_dict)} Districts"))

            CambodiaCommune.objects.bulk_create(communes_dict.values(), batch_size=2000)
            self.stdout.write(self.style.SUCCESS(f"  -> Saved {len(communes_dict)} Communes"))

            CambodiaVillage.objects.bulk_create(villages_dict.values(), batch_size=2000)
            self.stdout.write(self.style.SUCCESS(f"  -> Saved {len(villages_dict)} Villages"))

        self.stdout.write(self.style.SUCCESS("[OK] Successfully imported all Cambodia 2025 geographic data!"))
