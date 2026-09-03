import os
import mimetypes
import calendar
from datetime import datetime, date, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse, FileResponse, HttpResponseForbidden
from django.core.paginator import Paginator
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from django.contrib.auth.models import User
from django.conf import settings
from django.core.files.base import ContentFile
from .models import (
    Document, DocumentVersion, LeadershipAnnotation, DocumentRouting,
    Notification, Department, UserProfile, OTPVerification, ChatMessage,
    CivilServantProfile, OfficerAttachment, OfficerAuditLog, OfficerEditWindowSetting,
    OfficerDepartmentTransferHistory,
    CambodiaProvince, CambodiaDistrict, CambodiaCommune, CambodiaVillage,
    OfficerPromotionRequest, OfficerMedalRequest,
    ContractOfficer, ContractOfficerAttachment, ContractOfficerRenewalHistory,
    ContractOfficerTransferHistory,
    Vehicle, VehicleRequest, VehicleRequestAttachment,
    AttendanceRecord, get_attendance_position_sort_weight, get_rank_step_sort_weight,
    normalize_khmer_role,
    to_arabic_digits
)
from .forms import (
    DocumentForm, DocumentVersionForm, LeadershipAnnotationForm, DocumentRoutingForm,
    FlexibleLoginForm, UserRegistrationForm, OTPVerificationForm, DepartmentForm,
    UserManagementEditForm
)
from .otp_service import send_otp_notification
from django.views.decorators.csrf import csrf_exempt


@csrf_exempt
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        form = FlexibleLoginForm(request.POST)
        if form.is_valid():
            identifier = form.cleaned_data['identifier'].strip()
            password = form.cleaned_data['password']

            # Find matching user by username, email, or phone
            user_obj = None
            # 1. Try username
            user_obj = User.objects.filter(username__iexact=identifier).first()
            # 2. Try email
            if not user_obj and '@' in identifier:
                user_obj = User.objects.filter(email__iexact=identifier).first()
            # 3. Try phone in UserProfile
            if not user_obj:
                phone_clean = identifier.replace(' ', '')
                profile_obj = UserProfile.objects.filter(phone=phone_clean).select_related('user').first()
                if profile_obj:
                    user_obj = profile_obj.user

            user = None
            if user_obj:
                user = authenticate(request, username=user_obj.username, password=password)

            if user is not None:
                is_admin_account = user.is_superuser or (user.username.upper() == 'ADMIN')
                user_profile = getattr(user, 'profile', None)

                # Check if account is approved by ADMIN
                if not is_admin_account:
                    if not user_profile or not user_profile.is_approved:
                        messages.warning(
                            request,
                            "⚠️ គណនីរបស់អ្នកត្រូវបានបង្កើត និងផ្ទៀងផ្ទាត់ OTP រួចហើយ ប៉ុន្តែកំពុងស្ថិតក្នុងដំណាក់កាល «រង់ចាំការអនុម័តពី ADMIN»។ សូមរង់ចាំ ឬទាក់ទង ADMIN ដើម្បីបើកសិទ្ធិប្រើប្រាស់!"
                        )
                        return render(request, 'dms/login.html', {'form': form})

                login(request, user)
                full_name = user.get_full_name() or user.username
                role_label = user_profile.get_role_display() if user_profile else 'អ្នកគ្រប់គ្រងប្រព័ន្ធ'
                messages.success(request, f"ស្វាគមន៍មកកាន់ប្រព័ន្ធ! ចូលប្រព័ន្ធជោគជ័យជា {full_name} ({role_label})")
                return redirect('dashboard')
            else:
                messages.error(request, "ឈ្មោះគណនី/អ៊ីមែល/លេខទូរស័ព្ទ ឬ ពាក្យសម្ងាត់មិនត្រឹមត្រូវឡើយ!")
    else:
        form = FlexibleLoginForm()

    return render(request, 'dms/login.html', {
        'form': form,
    })


@csrf_exempt
def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            
            full_name = data['full_name'].strip()
            parts = full_name.split(' ', 1)
            # In Khmer: First part is family name, second is given name
            first_name = parts[0] if len(parts) > 1 else full_name
            last_name = parts[1] if len(parts) > 1 else ''

            pos = data.get('position_title', '')
            admin_pers_dept = Department.objects.filter(Q(code='ADMIN_PERS') | Q(name_kh='ការិយាល័យរដ្ឋបាល បុគ្គលិក')).first()
            lead_dept = Department.objects.filter(Q(code='LEAD') | Q(name_kh='ថ្នាក់ដឹកនាំមន្ទីរ')).first()
            if pos in ['ប្រធានមន្ទីរ', 'អនុប្រធានមន្ទីរ']:
                user_role = 'LEADERSHIP'
                target_dept = lead_dept or (admin_pers_dept if admin_pers_dept else data['department'])
            elif data['department'].code in ['ADMIN', 'ADMIN_PERS'] or (admin_pers_dept and data['department'].id == admin_pers_dept.id):
                user_role = 'ADMIN'
                target_dept = data['department']
            else:
                user_role = 'SPECIALIZED'
                target_dept = data['department']

            # Direct User Creation without OTP - Waiting for Admin Approval
            try:
                user = User.objects.create_user(
                    username=data['username'],
                    email=data.get('email', ''),
                    password=data['password'],
                    first_name=first_name,
                    last_name=last_name
                )

                UserProfile.objects.create(
                    user=user,
                    department=target_dept,
                    role=user_role,
                    position_title=pos,
                    phone=data.get('phone', ''),
                    raw_password_display=data['password'],
                    is_approved=False
                )

                messages.success(
                    request,
                    f"🎉 ការចុះឈ្មោះគណនី «{user.username}» ទទួលបានជោគជ័យ! ដោយសារប្រព័ន្ធតម្រូវឱ្យមានការត្រួតពិនិត្យសុវត្ថិភាព គណនីរបស់អ្នកត្រូវ «រង់ចាំការអនុម័តពី ADMIN» ជាមុនសិន ទើបអាចចូលប្រើប្រាស់បាន។"
                )
                return redirect('login')
            except Exception as e:
                messages.error(request, f"មានបញ្ហាក្នុងការបង្កើតគណនី៖ {e}")
    else:
        form = UserRegistrationForm()

    return render(request, 'dms/register.html', {
        'form': form
    })


@csrf_exempt
def otp_verify_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    otp_id = request.session.get('pending_otp_id')
    identifier = request.session.get('pending_otp_identifier')
    phone = request.session.get('pending_otp_phone', '')

    if not otp_id:
        messages.warning(request, "មិនមានប្រតិបត្តិការផ្ទៀងផ្ទាត់ដែលកំពុងរង់ចាំឡើយ។ សូមចុះឈ្មោះម្តងទៀត។")
        return redirect('register')

    otp_record = OTPVerification.objects.filter(id=otp_id).first()
    if not otp_record or otp_record.is_used:
        messages.warning(request, "លេខកូដផ្ទៀងផ្ទាត់នេះត្រូវបានប្រើប្រាស់រួចហើយ ឬមិនមានសុពលភាព។")
        return redirect('register')

    if request.method == 'POST':
        form = OTPVerificationForm(request.POST)
        if form.is_valid():
            submitted_code = form.cleaned_data['otp_code'].strip()

            if not otp_record.is_valid():
                messages.error(request, "លេខកូដ OTP នេះបានផុតកំណត់សុពលភាពហើយ (១០ នាទី)។ សូមចុច «ផ្ញើកូដឡើងវិញ»។")
            elif otp_record.otp_code != submitted_code:
                messages.error(request, "លេខកូដ ៦ ខ្ទង់មិនត្រឹមត្រូវឡើយ! សូមពិនិត្យ និងវាយម្តងទៀត។")
            else:
                # Valid OTP! Create User and Profile with is_approved = False
                payload = otp_record.payload_data
                try:
                    user = User.objects.create_user(
                        username=payload['username'],
                        email=payload['email'],
                        password=payload['password'],
                        first_name=payload['first_name'],
                        last_name=payload['last_name']
                    )

                    dept = Department.objects.filter(id=payload['department_id']).first()
                    UserProfile.objects.create(
                        user=user,
                        department=dept,
                        role=payload['role'],
                        position_title=payload.get('position_title', ''),
                        phone=payload.get('phone', ''),
                        raw_password_display=payload['password'],
                        is_approved=False
                    )

                    otp_record.is_used = True
                    otp_record.save()

                    # Clean session
                    request.session.pop('pending_otp_id', None)
                    request.session.pop('pending_otp_identifier', None)
                    request.session.pop('pending_otp_phone', None)

                    messages.success(
                        request,
                        f"🎉 ការផ្ទៀងផ្ទាត់លេខកូដជោគជ័យ! គណនី «{user.username}» ត្រូវបានបង្កើតរួចរាល់។ ដោយសារប្រព័ន្ធតម្រូវឱ្យមានការត្រួតពិនិត្យសុវត្ថិភាព គណនីរបស់អ្នកនឹងត្រូវ «រង់ចាំការអនុម័តពី ADMIN» ជាមុនសិន ទើបអាចចូលប្រើប្រាស់បាន។"
                    )
                    return redirect('login')
                except Exception as e:
                    messages.error(request, f"មានបញ្ហាក្នុងការបង្កើតគណនី៖ {e}")
    else:
        form = OTPVerificationForm()

    return render(request, 'dms/otp_verify.html', {
        'form': form,
        'identifier': identifier,
        'phone': phone,
        'dev_otp': otp_record.otp_code if settings.DEBUG else None,
    })


@csrf_exempt
def otp_resend_view(request):
    otp_id = request.session.get('pending_otp_id')
    if not otp_id:
        messages.error(request, "មិនមានសំណើផ្ទៀងផ្ទាត់ឡើយ!")
        return redirect('register')

    old_otp = OTPVerification.objects.filter(id=otp_id).first()
    if not old_otp:
        messages.error(request, "សំណើមិនត្រឹមត្រូវ!")
        return redirect('register')

    # Generate new OTP
    new_otp = OTPVerification.generate_otp(
        identifier=old_otp.identifier,
        purpose=old_otp.purpose,
        payload_data=old_otp.payload_data,
        expiry_minutes=10
    )

    request.session['pending_otp_id'] = new_otp.id
    
    full_name = ''
    if new_otp.payload_data:
        full_name = f"{new_otp.payload_data.get('last_name', '')} {new_otp.payload_data.get('first_name', '')}".strip()

    send_otp_notification(
        identifier=new_otp.identifier,
        otp_code=new_otp.otp_code,
        purpose=new_otp.purpose,
        full_name=full_name
    )

    messages.success(request, f"លេខកូដ OTP ថ្មី ៦ ខ្ទង់ត្រូវបានផ្ញើទៅកាន់ {new_otp.identifier} រួចរាល់ហើយ!")
    return redirect('otp_verify')


@login_required
def user_approvals_view(request):
    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        messages.error(request, "លោកអ្នកគ្មានសិទ្ធិចូលទំព័រនេះឡើយ!")
        return redirect('dashboard')

    status_filter = request.GET.get('status', 'pending')
    users_qs = UserProfile.objects.exclude(user__is_superuser=True).exclude(user__username__iexact='ADMIN').select_related('user', 'department').order_by('-created_at')

    if status_filter == 'pending':
        users_qs = users_qs.filter(is_approved=False)
    elif status_filter == 'approved':
        users_qs = users_qs.filter(is_approved=True)

    pending_count = UserProfile.objects.filter(is_approved=False).exclude(user__is_superuser=True).exclude(user__username__iexact='ADMIN').count()
    approved_count = UserProfile.objects.filter(is_approved=True).exclude(user__is_superuser=True).exclude(user__username__iexact='ADMIN').count()

    return render(request, 'dms/user_approvals.html', {
        'profiles': users_qs,
        'status_filter': status_filter,
        'pending_count': pending_count,
        'approved_count': approved_count,
    })


@login_required
def user_approve_action(request, user_id):
    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        messages.error(request, "លោកអ្នកគ្មានសិទ្ធិអនុវត្តសកម្មភាពនេះឡើយ!")
        return redirect('dashboard')

    target_user = get_object_or_404(User, id=user_id)
    profile = getattr(target_user, 'profile', None)
    if profile:
        profile.is_approved = True
        profile.save()
        messages.success(request, f"បានអនុម័ត (Approve) គណនី {target_user.username} ({target_user.get_full_name()}) រួចរាល់ដោយជោគជ័យ!")
        
        Notification.objects.create(
            recipient=target_user,
            title="គណនីរបស់អ្នកត្រូវបានអនុម័ត",
            message=f"គណនីរបស់អ្នកត្រូវបានអនុម័តដោយ ADMIN រួចហើយ! ឥឡូវនេះអ្នកអាចចូលប្រើប្រាស់មុខងារទាំងអស់ក្នុងប្រព័ន្ធបាន។"
        )
    return redirect('user_approvals')


@login_required
def user_revoke_action(request, user_id):
    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        messages.error(request, "លោកអ្នកគ្មានសិទ្ធិអនុវត្តសកម្មភាពនេះឡើយ!")
        return redirect('dashboard')

    target_user = get_object_or_404(User, id=user_id)
    profile = getattr(target_user, 'profile', None)
    if profile:
        profile.is_approved = False
        profile.save()
        messages.warning(request, f"បានផ្អាក/ដកសិទ្ធិ (Revoke) គណនី {target_user.username} រួចរាល់!")
    return redirect('user_approvals')


@login_required
def user_delete_action(request, user_id):
    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        messages.error(request, "លោកអ្នកគ្មានសិទ្ធិអនុវត្តសកម្មភាពនេះឡើយ!")
        return redirect('dashboard')

    target_user = get_object_or_404(User, id=user_id)
    if target_user.is_superuser or target_user.username.upper() == 'ADMIN':
        messages.error(request, "មិនអាចលុបគណនី ADMIN បានឡើយ!")
        return redirect('user_approvals')

    uname = target_user.username
    target_user.delete()
    messages.success(request, f"បានលុបគណនី {uname} ចេញពីប្រព័ន្ធរួចរាល់!")
    return redirect('user_approvals')


@login_required
def dev_switch_user(request, user_id):
    """
    DEV PHASE: Allows ADMIN to instantly switch (impersonate) into any user account for testing,
    and retain the ability to switch back to ADMIN seamlessly.
    """
    is_admin = request.user.is_superuser or (request.user.username.upper() == 'ADMIN') or ('original_admin_id' in request.session)
    if not is_admin:
        messages.error(request, "⚠️ លោកអ្នកគ្មានសិទ្ធិប្រើប្រាស់មុខងារ Dev Switch User ឡើយ!")
        return redirect('dashboard')

    target_user = get_object_or_404(User, id=user_id)

    # Track original admin user ID across session switch
    orig_id = request.session.get('original_admin_id', request.user.id)

    # Perform switch
    login(request, target_user, backend='django.contrib.auth.backends.ModelBackend')

    if target_user.id != orig_id:
        request.session['original_admin_id'] = orig_id
    else:
        request.session.pop('original_admin_id', None)

    profile = getattr(target_user, 'profile', None)
    role_name = profile.get_role_display() if profile else 'អ្នកប្រើប្រាស់'
    dept_name = profile.department.name_kh if (profile and profile.department) else 'គ្មានការិយាល័យ'
    full_name = target_user.get_full_name() or target_user.username

    messages.success(
        request,
        f"⚡ [DEV MODE] បាន Switch ចូលប្រើប្រាស់ជា៖ «{full_name}» (@{target_user.username} - {role_name} - {dept_name}) រួចរាល់!"
    )
    return redirect('dashboard')


@login_required
def dev_switch_back_admin(request):
    """
    DEV PHASE: Instantly switch back to the original ADMIN account.
    """
    orig_id = request.session.pop('original_admin_id', None)
    admin_user = None
    if orig_id:
        admin_user = User.objects.filter(id=orig_id).first()

    if not admin_user:
        admin_user = User.objects.filter(username__iexact='ADMIN').first() or User.objects.filter(is_superuser=True).first()

    if admin_user:
        login(request, admin_user, backend='django.contrib.auth.backends.ModelBackend')
        messages.success(request, f"🛡️ [DEV MODE] បានត្រឡប់មកកាន់គណនី ADMIN ({admin_user.username}) វិញដោយជោគជ័យ!")
        return redirect('user_approvals')
    else:
        messages.warning(request, "មិនអាចស្វែងរកគណនី Admin ដើមបានឡើយ!")
        return redirect('login')


@login_required
def user_edit_view(request, user_id):
    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        messages.error(request, "លោកអ្នកគ្មានសិទ្ធិអនុវត្តសកម្មភាពនេះឡើយ!")
        return redirect('dashboard')

    target_user = get_object_or_404(User, id=user_id)
    profile, _ = UserProfile.objects.get_or_create(user=target_user)

    if request.method == 'POST':
        form = UserManagementEditForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            
            full_name = data['full_name'].strip()
            parts = full_name.split(' ', 1)
            target_user.first_name = parts[0] if len(parts) > 1 else full_name
            target_user.last_name = parts[1] if len(parts) > 1 else ''
            
            new_username = data['username'].strip()
            if User.objects.filter(username__iexact=new_username).exclude(id=target_user.id).exists():
                messages.error(request, f"ឈ្មោះគណនី «{new_username}» មានអ្នកប្រើប្រាស់រួចហើយ!")
                return render(request, 'dms/user_edit.html', {'form': form, 'target_user': target_user, 'profile': profile})
            
            target_user.username = new_username
            target_user.email = data.get('email', '').strip()

            new_pass = data.get('new_password', '').strip()
            if new_pass:
                target_user.set_password(new_pass)
                profile.raw_password_display = new_pass

            target_user.save()

            profile.position_title = data.get('position_title', 'មន្ត្រី')
            if profile.position_title in ['ប្រធានមន្ទីរ', 'អនុប្រធានមន្ទីរ']:
                lead_dept = Department.objects.filter(Q(code='LEAD') | Q(name_kh='ថ្នាក់ដឹកនាំមន្ទីរ')).first()
                if lead_dept:
                    profile.department = lead_dept
                profile.role = 'LEADERSHIP'
                profile.can_annotate = True
            else:
                profile.department = data.get('department')
                profile.role = data.get('role', 'SPECIALIZED')
            profile.phone = data.get('phone', '').strip()
            profile.is_approved = data.get('is_approved', False)

            # Save Feature Permissions
            profile.can_create_document = data.get('can_create_document', False)
            profile.can_edit_document = data.get('can_edit_document', False)
            profile.can_route_document = data.get('can_route_document', False)
            profile.can_annotate = data.get('can_annotate', False)
            profile.can_complete_document = data.get('can_complete_document', False)
            profile.can_delete_document = data.get('can_delete_document', False)
            profile.can_view_reports = data.get('can_view_reports', False)
            profile.can_print = data.get('can_print', False)
            profile.can_export_contract_excel = data.get('can_export_contract_excel', False)
            profile.can_export_civil_servant_excel = data.get('can_export_civil_servant_excel', False)
            profile.can_view_e2_report = data.get('can_view_e2_report', False)
            profile.save()

            messages.success(request, f"បានកែប្រែព័ត៌មាន សិទ្ធិប្រើប្រាស់ និងពាក្យសម្ងាត់របស់គណនី «{target_user.username}» ដោយជោគជ័យ!")
            return redirect('user_approvals')
    else:
        full_name = f"{target_user.first_name} {target_user.last_name}".strip() or target_user.username
        initial_data = {
            'full_name': full_name,
            'username': target_user.username,
            'email': target_user.email,
            'phone': profile.phone,
            'department': profile.department,
            'position_title': profile.position_title,
            'role': profile.role,
            'is_approved': profile.is_approved,
            'can_create_document': profile.can_create_document,
            'can_edit_document': profile.can_edit_document,
            'can_route_document': profile.can_route_document,
            'can_annotate': profile.can_annotate,
            'can_complete_document': profile.can_complete_document,
            'can_delete_document': profile.can_delete_document,
            'can_view_reports': profile.can_view_reports,
            'can_print': profile.can_print,
            'can_export_contract_excel': getattr(profile, 'can_export_contract_excel', False),
            'can_export_civil_servant_excel': getattr(profile, 'can_export_civil_servant_excel', False),
            'can_view_e2_report': getattr(profile, 'can_view_e2_report', False),
        }
        form = UserManagementEditForm(initial=initial_data)

    return render(request, 'dms/user_edit.html', {
        'form': form,
        'target_user': target_user,
        'profile': profile,
    })


# ==========================================
# DEPARTMENT & ORGANIZATION MANAGEMENT (ADMIN)
# ==========================================
@login_required
def department_list_view(request):
    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        messages.error(request, "លោកអ្នកគ្មានសិទ្ធិចូលកាន់ទំព័រនេះឡើយ!")
        return redirect('dashboard')

    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            dept = form.save()
            messages.success(request, f"បានបង្កើតការិយាល័យ/អង្គភាព «{dept.name_kh}» ({dept.code}) ដោយជោគជ័យ!")
            return redirect('department_list_manage')
        else:
            messages.error(request, "សូមពិនិត្យទិន្នន័យដែលបានបំពេញឡើងវិញ!")
    else:
        form = DepartmentForm()

    tab = request.GET.get('tab', 'all')
    queryset = Department.objects.annotate(
        user_count=Count('users', distinct=True),
        officer_count=Count('officers', distinct=True),
        doc_count=Count('current_documents', distinct=True)
    ).order_by('order_index', 'id')

    total_count = queryset.count()
    old_count = queryset.filter(structure_type='OLD').count()
    new_count = queryset.filter(structure_type='NEW').count()
    active_count = queryset.filter(is_active=True).count()
    inactive_count = queryset.filter(is_active=False).count()

    if tab == 'old':
        departments = queryset.filter(structure_type='OLD')
    elif tab == 'new':
        departments = queryset.filter(structure_type='NEW')
    elif tab == 'active':
        departments = queryset.filter(is_active=True)
    elif tab == 'inactive':
        departments = queryset.filter(is_active=False)
    else:
        departments = queryset

    return render(request, 'dms/department_list.html', {
        'departments': departments,
        'form': form,
        'tab': tab,
        'total_count': total_count,
        'old_count': old_count,
        'new_count': new_count,
        'active_count': active_count,
        'inactive_count': inactive_count,
    })


@login_required
def department_toggle_active_view(request, pk):
    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        messages.error(request, "លោកអ្នកគ្មានសិទ្ធិអនុវត្តសកម្មភាពនេះឡើយ!")
        return redirect('dashboard')

    dept = get_object_or_404(Department, pk=pk)
    if dept.code == 'LEAD':
        messages.warning(request, "⚠️ «ថ្នាក់ដឹកនាំមន្ទីរ» ជារចនាសម្ព័ន្ធស្នូលរបស់ប្រព័ន្ធ មិនអាចបិទដំណើរការបានឡើយ!")
        return redirect('department_list_manage')

    dept.is_active = not dept.is_active
    dept.save()
    if dept.is_active:
        messages.success(request, f"✅ បានបើកដំណើរការការិយាល័យ «{dept.name_kh}» ({dept.code}) ឡើងវិញដោយជោគជ័យ!")
    else:
        messages.info(request, f"⏸️ បានបិទដំណើរការការិយាល័យ «{dept.name_kh}» ({dept.code}) ជាបណ្ដោះអាសន្ន!")
    
    return redirect(request.META.get('HTTP_REFERER') or 'department_list_manage')


@login_required
def department_switch_structure_view(request):
    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        messages.error(request, "លោកអ្នកគ្មានសិទ្ធិអនុវត្តសកម្មភាពនេះឡើយ!")
        return redirect('dashboard')

    if request.method == 'POST':
        target = request.POST.get('target_structure', 'OLD')
        if target == 'NEW':
            # Activate New Structure, Pause Old Structure
            Department.objects.filter(structure_type='NEW').update(is_active=True)
            Department.objects.filter(structure_type='OLD').update(is_active=False)
            Department.objects.filter(structure_type='CORE').update(is_active=True)
            messages.success(request, "🚀 បានប្តូរទៅប្រើ «រចនាសម្ព័ន្ធថ្មី» ដោយជោគជ័យ! ការិយាល័យរចនាសម្ព័ន្ធចាស់ត្រូវបានបិទបណ្ដោះអាសន្ន។")
        else:
            # Activate Old Structure, Pause New Structure
            Department.objects.filter(structure_type='OLD').update(is_active=True)
            Department.objects.filter(structure_type='NEW').update(is_active=False)
            Department.objects.filter(structure_type='CORE').update(is_active=True)
            messages.success(request, "🏛️ បានប្តូរទៅប្រើ «រចនាសម្ព័ន្ធចាស់» (ការិយាល័យទាំង១០ + ខណ្ឌទាំង២) ដោយជោគជ័យ! ការិយាល័យរចនាសម្ព័ន្ធថ្មីត្រូវបានបិទបណ្ដោះអាសន្ន។")

    return redirect('department_list_manage')


@login_required
def department_edit_view(request, pk):
    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        messages.error(request, "លោកអ្នកគ្មានសិទ្ធិអនុវត្តសកម្មភាពនេះឡើយ!")
        return redirect('dashboard')

    dept = get_object_or_404(Department, pk=pk)

    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=dept)
        if form.is_valid():
            form.save()
            messages.success(request, f"បានកែប្រែព័ត៌មានការិយាល័យ «{dept.name_kh}» រួចរាល់ដោយជោគជ័យ!")
            return redirect('department_list_manage')
    else:
        form = DepartmentForm(instance=dept)

    return render(request, 'dms/department_edit.html', {
        'form': form,
        'department': dept,
    })


@login_required
def department_delete_view(request, pk):
    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        messages.error(request, "លោកអ្នកគ្មានសិទ្ធិអនុវត្តសកម្មភាពនេះឡើយ!")
        return redirect('dashboard')

    dept = get_object_or_404(Department, pk=pk)

    if dept.code in ['LEAD', 'ADMIN', 'ADMIN_PERS'] or dept.name_kh == 'ថ្នាក់ដឹកនាំមន្ទីរ':
        messages.error(request, "⚠️ មិនអាចលុប «ថ្នាក់ដឹកនាំមន្ទីរ» ឬ «ការិយាល័យរដ្ឋបាល បុគ្គលិក» បានឡើយ ព្រោះជារចនាសម្ព័ន្ធស្នូលរបស់ប្រព័ន្ធ!")
        return redirect('department_list_manage')

    admin_dept = Department.objects.filter(code='ADMIN_PERS').first() or Department.objects.filter(code='ADMIN').first()

    # Reassign users and documents safely before deleting
    if admin_dept:
        dept.users.update(department=admin_dept)
        dept.officers.update(department=admin_dept)
        Document.objects.filter(origin_department=dept).update(origin_department=admin_dept)
        Document.objects.filter(current_department=dept).update(current_department=admin_dept)
        DocumentRouting.objects.filter(from_dept=dept).update(from_dept=admin_dept)
        DocumentRouting.objects.filter(to_dept=dept).update(to_dept=admin_dept)
        ChatMessage.objects.filter(department=dept).update(department=admin_dept)
        ChatMessage.objects.filter(target_department=dept).update(target_department=admin_dept)

    name = dept.name_kh
    dept.delete()
    messages.success(request, f"បានលុបការិយាល័យ «{name}» ចេញពីប្រព័ន្ធដោយជោគជ័យ!")
    return redirect('department_list_manage')


@login_required
def document_delete(request, pk):
    document = get_object_or_404(Document, pk=pk)
    profile = getattr(request.user, 'profile', None)

    is_admin = request.user.is_superuser or request.user.username.upper() == 'ADMIN'
    is_sender_or_creator = (document.created_by == request.user) or (profile and profile.department and document.origin_department == profile.department)
    has_perm = bool(profile and profile.can_delete_document)

    if not (is_admin or is_sender_or_creator or has_perm):
        messages.error(request, "⚠️ លោកអ្នកគ្មានសិទ្ធិលុបឯកសារនេះឡើយ!")
        return redirect('document_detail', pk=pk)

    # Business Rule: If receiving office has already accepted/completed, sender CANNOT delete!
    if document.status == 'COMPLETED' and not is_admin:
        messages.error(request, "⚠️ ឯកសារនេះត្រូវបានការិយាល័យទទួល ចុះយល់ព្រមទទួលយករួចរាល់ហើយ ដូច្នេះដាច់ខាតមិនអាចលុបចេញពីប្រព័ន្ធបានឡើយ!")
        return redirect('document_detail', pk=pk)

    reg_num = document.registry_number
    document.delete()
    messages.success(request, f"បានលុបឯកសារលេខ {reg_num} ចេញពីប្រព័ន្ធដោយជោគជ័យ!")

    next_url = request.META.get('HTTP_REFERER')
    if next_url and 'outbound' in next_url:
        return redirect('outbound_documents')
    elif next_url and 'inbound' in next_url:
        return redirect('inbound_documents')
    return redirect('document_list')



def logout_view(request):
    request.session.pop('original_admin_id', None)
    if request.user.is_authenticated:
        logout(request)
        messages.info(request, "បានចាកចេញពីប្រព័ន្ធដោយជោគជ័យ!")
    return redirect('login')


@login_required
def dashboard(request):
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None

    is_super_or_lead = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (profile and profile.is_leadership)

    if is_super_or_lead:
        # Admin / Leadership see global statistics across entire ministry
        base_docs = Document.objects.all()
        pending_my_action = base_docs.filter(status='PENDING_LEADERSHIP') if (profile and profile.is_leadership) else base_docs.filter(status__in=['PENDING_ADMIN', 'PENDING_LEADERSHIP'])
        recent_documents = base_docs.order_by('-updated_at')[:8]
        recent_routings = DocumentRouting.objects.select_related('document', 'from_user', 'to_dept').order_by('-routed_at')[:6]
        inbound_count = base_docs.filter(doc_type='INBOUND').count()
        # Outbound count strictly for this department's own outbound letters
        outbound_count = Document.objects.filter(origin_department=dept).count() if dept else base_docs.filter(doc_type__in=['OUTBOUND', 'INTERNAL']).count()
    else:
        # Specialized Office: STRICTLY scoped to their own department ONLY
        if dept:
            base_docs = Document.objects.filter(
                Q(origin_department=dept) | Q(current_department=dept) | Q(routings__to_dept=dept)
            ).distinct()
            pending_my_action = base_docs.filter(current_department=dept).exclude(status='COMPLETED')
            recent_documents = base_docs.order_by('-updated_at')[:8]
            recent_routings = DocumentRouting.objects.filter(
                Q(from_dept=dept) | Q(to_dept=dept)
            ).select_related('document', 'from_user', 'to_dept').order_by('-routed_at')[:6]
            
            # Inbound: External Inbound OR routed to dept from other offices
            inbound_count = base_docs.filter(
                Q(doc_type='INBOUND', current_department=dept) |
                Q(doc_type='INBOUND', routings__to_dept=dept) |
                Q(routings__to_dept=dept)
            ).exclude(origin_department=dept).distinct().count()
            
            # Outbound: Originating from this department
            outbound_count = base_docs.filter(origin_department=dept).count()
        else:
            base_docs = Document.objects.none()
            pending_my_action = base_docs
            recent_documents = base_docs
            recent_routings = DocumentRouting.objects.none()
            inbound_count = 0
            outbound_count = 0

    total_docs = base_docs.count()
    completed_count = base_docs.filter(status='COMPLETED').count()
    pending_count = pending_my_action.count()

    context = {
        'total_docs': total_docs,
        'inbound_count': inbound_count,
        'outbound_count': outbound_count,
        'completed_count': completed_count,
        'pending_count': pending_count,
        'recent_documents': recent_documents,
        'recent_routings': recent_routings,
        'profile': profile,
        'dept': dept,
    }
    return render(request, 'dms/dashboard.html', context)


@login_required
def document_list(request):
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    is_super_or_lead = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (profile and profile.is_leadership)

    tab = request.GET.get('tab', 'all')
    search_q = request.GET.get('q', '').strip()
    urgency_filter = request.GET.get('urgency', '')
    doc_type_filter = request.GET.get('doc_type', '')
    dept_filter = request.GET.get('department', '')

    if is_super_or_lead:
        queryset = Document.objects.all()
    else:
        if dept:
            queryset = Document.objects.filter(
                Q(origin_department=dept) | Q(current_department=dept) | Q(routings__to_dept=dept)
            ).distinct()
        else:
            queryset = Document.objects.none()

    # Tab filter
    if tab == 'pending_me':
        if profile and profile.is_leadership:
            queryset = queryset.filter(status='PENDING_LEADERSHIP')
        elif dept:
            queryset = queryset.filter(current_department=dept).exclude(status='COMPLETED')
    elif tab == 'inbound':
        if is_super_or_lead:
            queryset = queryset.filter(doc_type='INBOUND')
        else:
            queryset = queryset.filter(
                Q(doc_type='INBOUND') | Q(routings__to_dept=dept)
            ).exclude(origin_department=dept).distinct()
    elif tab == 'outbound':
        if is_super_or_lead:
            queryset = queryset.filter(doc_type__in=['OUTBOUND', 'INTERNAL'])
        else:
            queryset = queryset.filter(origin_department=dept)
    elif tab == 'completed':
        queryset = queryset.filter(status='COMPLETED')
    elif tab == 'my_office' and dept:
        queryset = queryset.filter(Q(current_department=dept) | Q(origin_department=dept))

    # Apply search filters
    if search_q:
        queryset = queryset.filter(
            Q(registry_number__icontains=search_q) |
            Q(title__icontains=search_q) |
            Q(origin_org__icontains=search_q) |
            Q(destination_org__icontains=search_q) |
            Q(summary__icontains=search_q)
        )

    if urgency_filter:
        queryset = queryset.filter(urgency=urgency_filter)

    if doc_type_filter:
        queryset = queryset.filter(doc_type=doc_type_filter)

    if dept_filter and is_super_or_lead:
        queryset = queryset.filter(current_department_id=dept_filter)

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')

    context = {
        'documents': queryset.order_by('-received_date', '-id'),
        'tab': tab,
        'search_q': search_q,
        'urgency_filter': urgency_filter,
        'doc_type_filter': doc_type_filter,
        'dept_filter': dept_filter,
        'departments': departments,
        'dept': dept,
    }
    return render(request, 'dms/document_list.html', context)


@login_required
def document_detail(request, pk):
    document = get_object_or_404(Document, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    is_super_or_lead = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (profile and profile.is_leadership)

    # Strict department scope check for specialized officers
    if not is_super_or_lead:
        if dept:
            has_access = (
                document.origin_department == dept or
                document.current_department == dept or
                document.annotations.filter(target_departments=dept).exists() or
                document.routings.filter(to_dept=dept, is_cancelled=False).exists()
            )
            if not has_access:
                messages.error(request, "⚠️ ឯកសារនេះមិនស្ថិតក្នុងដែនសមត្ថកិច្ចនៃការិយាល័យរបស់អ្នកឡើយ!")
                return redirect('dashboard')
        else:
            messages.error(request, "⚠️ គណនីរបស់អ្នកមិនទាន់បានភ្ជាប់ជាមួយការិយាល័យណាមួយឡើយ!")
            return redirect('dashboard')

    annotation_form = LeadershipAnnotationForm()
    routing_form = DocumentRoutingForm(user=request.user, profile=profile)
    version_form = DocumentVersionForm()

    versions = document.versions.order_by('-uploaded_at')
    routings = document.routings.select_related('from_user', 'from_dept', 'to_dept').order_by('routed_at')
    annotations = document.annotations.select_related('leader').prefetch_related('target_departments').order_by('-signed_at')

    # Business Rule:
    # 1. Admin (General Affairs) CAN complete/acknowledge if:
    #    - Document was sent FROM a Specialized Office (origin != admin_dept)
    #    - OR Leadership Annotation specifically targeted General Affairs Office (admin_dept in annotations target_departments)
    #    - OR Document was routed to General Affairs Office (admin_dept in routings to_dept)
    # 2. Specialized Offices CAN complete/acknowledge if:
    #    - Document is targeted/routed to their department (current_dept, annotations target_departments, or routings to_dept)
    admin_dept = Department.objects.filter(code='ADMIN').first()
    is_admin_user = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (dept and dept.code == 'ADMIN')

    can_complete_action = False
    if document.status != 'COMPLETED':
        if is_admin_user:
            is_annotated_to_admin = bool(admin_dept and document.annotations.filter(target_departments=admin_dept).exists())
            is_routed_to_admin = bool(admin_dept and document.routings.filter(to_dept=admin_dept, is_cancelled=False).exists())
            from_specialized = bool(document.origin_department and document.origin_department != admin_dept)

            if from_specialized or is_annotated_to_admin or is_routed_to_admin:
                can_complete_action = True
        else:
            is_annotated_to_dept = bool(dept and document.annotations.filter(target_departments=dept).exists())
            is_routed_to_dept = bool(dept and document.routings.filter(to_dept=dept, is_cancelled=False).exists())
            is_curr_dept = bool(dept and document.current_department == dept)

            if is_annotated_to_dept or is_routed_to_dept or is_curr_dept:
                if document.origin_department != dept or document.status in ['ROUTED', 'ANNOTATED', 'IN_PROGRESS']:
                    can_complete_action = True

    context = {
        'document': document,
        'versions': versions,
        'routings': routings,
        'annotations': annotations,
        'annotation_form': annotation_form,
        'routing_form': routing_form,
        'version_form': version_form,
        'profile': profile,
        'dept': dept,
        'can_complete_action': can_complete_action,
    }
    return render(request, 'dms/document_detail.html', context)


@login_required
def document_create(request):
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None

    # Check permission
    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        if not profile or not profile.can_create_document:
            messages.error(request, "⚠️ គណនីរបស់អ្នកមិនត្រូវបាន ADMIN អនុញ្ញាតឱ្យចុះបញ្ជីលិខិតថ្មីឡើយ!")
            return redirect('document_list')

    req_doc_type = request.GET.get('doc_type', '').upper()

    if request.method == 'POST':
        mode = 'INBOUND' if request.POST.get('doc_type') == 'INBOUND' or req_doc_type == 'INBOUND' else 'OUTBOUND'
        form = DocumentForm(request.POST, request.FILES, user=request.user, profile=profile, doc_type_mode=mode)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.created_by = request.user
            doc.origin_department = dept or Department.objects.first()
            doc.current_department = dept or Department.objects.first()

            # Auto-fill origin and destination if left empty
            if not doc.origin_org.strip():
                if doc.doc_type == 'INBOUND':
                    doc.origin_org = "ប្រភពខាងក្រៅ/ក្រសួងស្ថាប័ន"
                else:
                    doc.origin_org = dept.name_kh if dept else "មន្ទីររដ្ឋបាល"
            if not doc.destination_org.strip():
                if doc.doc_type == 'INBOUND':
                    doc.destination_org = dept.name_kh if dept else "មន្ទីររដ្ឋបាល"
                else:
                    doc.destination_org = "ក្រសួង/អង្គភាពពាក់ព័ន្ធ"

            # Auto-generate registry code if empty
            if not doc.registry_number.strip():
                year_str = datetime.now().strftime('%y')
                dept_code = dept.code if dept else 'ADMIN'
                count = Document.objects.filter(origin_department=dept).count() + 1
                doc.registry_number = f"{count:03d}/{year_str}/{dept_code}"

            # Check optional immediate auto-routing
            initial_dept = form.cleaned_data.get('initial_route_dept')
            initial_broadcast = form.cleaned_data.get('initial_send_to_all')

            if initial_broadcast:
                doc.status = 'ROUTED'
                doc.destination_org = 'គ្រប់ការិយាល័យ (All)'
            elif initial_dept:
                doc.current_department = initial_dept
                doc.destination_org = initial_dept.name_kh
                doc.status = 'ROUTED'
            else:
                # Specialized office: direct in-progress; Admin/Leadership: pending leadership
                if profile and profile.is_admin:
                    doc.status = 'PENDING_LEADERSHIP'
                elif profile and profile.is_specialized:
                    doc.status = 'IN_PROGRESS'
                else:
                    doc.status = 'PENDING_LEADERSHIP'

            doc.save()

            # Handle initial file attachment version
            file = request.FILES.get('file')
            if file:
                DocumentVersion.objects.create(
                    document=doc,
                    version_number='1.0',
                    file=file,
                    file_name=file.name,
                    uploaded_by=request.user,
                    change_summary='ឯកសារដើម Upload លើកដំបូង'
                )

            # Log routing history
            if initial_broadcast:
                for d in Department.objects.all():
                    DocumentRouting.objects.create(
                        document=doc,
                        from_user=request.user,
                        from_dept=dept,
                        to_dept=d,
                        action_taken="បង្កើត និងបញ្ជូនផ្សព្វផ្សាយទៅគ្រប់ការិយាល័យ (Broadcast)",
                        notes="បញ្ជូនស្វ័យប្រវត្តិតាមការចុះបញ្ជីថ្មី",
                        is_broadcast=True
                    )
            elif initial_dept:
                DocumentRouting.objects.create(
                    document=doc,
                    from_user=request.user,
                    from_dept=dept,
                    to_dept=initial_dept,
                    action_taken=f"បង្កើត និងបញ្ជូនទៅកាន់ {initial_dept.name_kh}",
                    notes="បញ្ជូនស្វ័យប្រវត្តិតាមការចុះបញ្ជីថ្មី"
                )
            else:
                DocumentRouting.objects.create(
                    document=doc,
                    from_user=request.user,
                    from_dept=dept,
                    to_dept=doc.current_department,
                    action_taken="បង្កើត និងចុះបញ្ជីឯកសារថ្មី",
                    notes=f"ប្រភេទលិខិត៖ {doc.get_doc_type_display()}"
                )

            # Notify Admin or Leadership
            if doc.status == 'PENDING_ADMIN':
                admin_dept = Department.objects.filter(code='ADMIN').first()
                if admin_dept:
                    admin_users = [p.user for p in UserProfile.objects.filter(department=admin_dept)]
                    for u in admin_users:
                        Notification.objects.create(
                            recipient=u,
                            document=doc,
                            title="លិខិតថ្មីរង់ចាំការពិនិត្យ",
                            message=f"ឯកសារលេខ {doc.registry_number} ត្រូវ​បាន​បញ្ជូន​មកកាន់​រដ្ឋបាល"
                        )
            elif doc.status == 'PENDING_LEADERSHIP':
                lead_users = [p.user for p in UserProfile.objects.filter(role='LEADERSHIP')]
                for u in lead_users:
                    Notification.objects.create(
                        recipient=u,
                        document=doc,
                        title="លិខិតរង់ចាំចំណារពីថ្នាក់ដឹកនាំ",
                        message=f"ឯកសារលេខ {doc.registry_number} កំពុងរង់ចាំចំណារ"
                    )

            messages.success(request, f"បានចុះលេខ និងរក្សាទុកឯកសារ {doc.registry_number} ដោយជោគជ័យ!")
            return redirect('document_detail', pk=doc.pk)
    else:
        # Pre-fill form initial values
        year_str = datetime.now().strftime('%y')
        dept_code = dept.code if dept else 'ADMIN'
        count = Document.objects.filter(origin_department=dept).count() + 1
        default_reg = f"{count:03d}/{year_str}/{dept_code}"
        
        mode = 'INBOUND' if req_doc_type == 'INBOUND' else 'OUTBOUND'
        initial_doc_type = 'INBOUND' if mode == 'INBOUND' else 'OUTBOUND'

        initial_dict = {
            'doc_type': initial_doc_type,
            'registry_number': default_reg,
            'origin_org': dept.name_kh if dept else 'មន្ទីររដ្ឋបាល',
            'destination_org': 'ថ្នាក់ដឹកនាំមន្ទីរ' if (profile and profile.is_specialized) else 'ក្រសួង/អង្គភាពពាក់ព័ន្ធ'
        }
        if initial_doc_type == 'INBOUND':
            initial_dict['origin_org'] = ''
            initial_dict['destination_org'] = dept.name_kh if dept else 'មន្ទីររដ្ឋបាល'

        form = DocumentForm(initial=initial_dict, user=request.user, profile=profile, doc_type_mode=mode)

    if (request.method == 'GET' and req_doc_type == 'INBOUND') or (request.method == 'POST' and request.POST.get('doc_type') == 'INBOUND'):
        page_title = 'ចុះបញ្ជីលិខិតចូល (New Inbound Document)'
        current_type = 'INBOUND'
    else:
        page_title = 'ចុះបញ្ជីលិខិតចេញ (New Outbound Document)'
        current_type = 'OUTBOUND'

    return render(request, 'dms/document_form.html', {'form': form, 'title': page_title, 'doc_type': current_type})


@login_required
def annotate_document(request, pk):
    document = get_object_or_404(Document, pk=pk)
    profile = getattr(request.user, 'profile', None)

    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        if not profile or not profile.can_annotate:
            messages.error(request, "⚠️ គណនីរបស់អ្នកមិនត្រូវបាន ADMIN អនុញ្ញាតឱ្យធ្វើចំណារថ្នាក់ដឹកនាំឡើយ!")
            return redirect('document_detail', pk=pk)

    if request.method == 'POST':
        form = LeadershipAnnotationForm(request.POST)
        if form.is_valid():
            annotation = form.save(commit=False)
            annotation.document = document
            annotation.leader = request.user
            annotation.save()

            target_depts = form.cleaned_data['target_departments']
            annotation.target_departments.set(target_depts)

            # Update document status and current department
            first_target = target_depts.first()
            document.status = 'ROUTED'
            if first_target:
                document.current_department = first_target
            document.save()

            # Record Routing history for EACH target department
            target_names = ", ".join([d.name_kh for d in target_depts])
            for d in target_depts:
                DocumentRouting.objects.create(
                    document=document,
                    from_user=request.user,
                    from_dept=profile.department if profile else None,
                    to_dept=d,
                    action_taken=f"ធ្វើចំណារ: {annotation.get_decision_display()}",
                    notes=f"ចាត់ចែងជូន៖ {target_names}\nខ្លឹមសារ៖ {annotation.annotation_text}"
                )

            # Notify target departments
            for d in target_depts:
                dept_users = [p.user for p in UserProfile.objects.filter(department=d)]
                for u in dept_users:
                    Notification.objects.create(
                        recipient=u,
                        document=document,
                        title="ឯកសារទទួលបានចំណារពីថ្នាក់ដឹកនាំ",
                        message=f"ឯកសារលេខ {document.registry_number} ត្រូវ​បាន​ថ្នាក់ដឹកនាំ​ធ្វើចំណារចាត់ចែងមកកាន់ការិយាល័យលោកអ្នក"
                    )

            messages.success(request, f"បានរក្សាទុកចំណារ និងបញ្ជូនឯកសារ {document.registry_number} រួចរាល់!")
    return redirect('document_detail', pk=pk)


@login_required
def route_document(request, pk):
    document = get_object_or_404(Document, pk=pk)
    profile = getattr(request.user, 'profile', None)

    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        if not profile or not profile.can_route_document:
            messages.error(request, "⚠️ គណនីរបស់អ្នកមិនត្រូវបាន ADMIN អនុញ្ញាតឱ្យបញ្ជូន ឬចាត់ចែងលិខិតឡើយ!")
            return redirect('document_detail', pk=pk)

    if request.method == 'POST':
        form = DocumentRoutingForm(request.POST, user=request.user, profile=profile)
        if form.is_valid():
            to_dept = form.cleaned_data.get('to_dept')
            send_to_all = form.cleaned_data.get('send_to_all')
            notes = form.cleaned_data.get('notes', '')

            old_dept = document.current_department

            if send_to_all:
                # Broadcast to all departments
                document.status = 'ROUTED'
                document.save()

                all_depts = Department.objects.all()
                for d in all_depts:
                    DocumentRouting.objects.create(
                        document=document,
                        from_user=request.user,
                        from_dept=profile.department if profile else None,
                        to_dept=d,
                        action_taken="បញ្ជូនផ្សព្វផ្សាយទៅគ្រប់ការិយាល័យ (Broadcast to All)",
                        notes=notes,
                        is_broadcast=True
                    )
                    for u in [p.user for p in UserProfile.objects.filter(department=d)]:
                        if u != request.user:
                            Notification.objects.create(
                                recipient=u,
                                document=document,
                                title="ឯកសារផ្សព្វផ្សាយទូទាំងមន្ទីរ (Broadcast)",
                                message=f"ឯកសារលេខ {document.registry_number} «{document.title}» ត្រូវបានបញ្ជូនផ្សព្វផ្សាយមកកាន់គ្រប់ការិយាល័យទាំងអស់។"
                            )

                messages.success(request, f"បានបញ្ជូនផ្សព្វផ្សាយឯកសារ {document.registry_number} ទៅកាន់ «គ្រប់ការិយាល័យទាំងអស់» ដោយជោគជ័យ!")
            elif to_dept:
                document.current_department = to_dept

                if to_dept.code == 'LEAD':
                    document.status = 'PENDING_LEADERSHIP'
                elif to_dept.code == 'ADMIN':
                    document.status = 'PENDING_ADMIN'
                else:
                    document.status = 'ROUTED'

                document.save()

                # Log routing
                DocumentRouting.objects.create(
                    document=document,
                    from_user=request.user,
                    from_dept=profile.department if profile else None,
                    to_dept=to_dept,
                    action_taken=f"បញ្ជូនតទៅកាន់ {to_dept.name_kh}",
                    notes=notes
                )

                # Notify target dept
                dept_users = [p.user for p in UserProfile.objects.filter(department=to_dept)]
                for u in dept_users:
                    if u != request.user:
                        Notification.objects.create(
                            recipient=u,
                            document=document,
                            title="ឯកសារត្រូវបានបញ្ជូនមកកាន់ការិយាល័យ",
                            message=f"ឯកសារលេខ {document.registry_number} ត្រូវបានបញ្ជូនមកពី {old_dept}"
                        )

                messages.success(request, f"បានបញ្ជូនឯកសារទៅកាន់ {to_dept.name_kh} ដោយជោគជ័យ!")
            else:
                messages.error(request, "សូមជ្រើសរើសការិយាល័យទទួល ឬធីកលើ «បញ្ជូនទៅគ្រប់ការិយាល័យទាំងអស់»!")
    return redirect('document_detail', pk=pk)


@login_required
def cancel_route_document(request, pk, routing_id):
    document = get_object_or_404(Document, pk=pk)
    routing = get_object_or_404(DocumentRouting, pk=routing_id, document=document)
    profile = getattr(request.user, 'profile', None)

    is_admin = request.user.is_superuser or request.user.username.upper() == 'ADMIN'
    is_leader = profile and profile.is_leadership
    is_sender = routing.from_user == request.user
    is_same_dept = profile and profile.department and routing.from_dept and (profile.department == routing.from_dept)

    if not (is_admin or is_leader or is_sender or is_same_dept):
        messages.error(request, "⚠️ លោកអ្នកគ្មានសិទ្ធិ Cancel ការបញ្ជូនលិខិតនេះឡើយ!")
        return redirect('document_detail', pk=pk)

    target_name = routing.to_dept.name_kh if routing.to_dept else "គ្រប់ការិយាល័យ"

    # Cancel this routing (and any other broadcast routings created at the same time if broadcast)
    if routing.is_broadcast:
        same_broadcasts = DocumentRouting.objects.filter(
            document=document,
            from_user=routing.from_user,
            is_broadcast=True,
            is_cancelled=False,
            routed_at__date=routing.routed_at.date()
        )
        same_broadcasts.update(
            is_cancelled=True,
            cancelled_at=timezone.now(),
            cancelled_by=request.user,
            cancellation_reason="បានលុបចោលការបញ្ជូនផ្សព្វផ្សាយទូទាំងមន្ទីរ"
        )
        Notification.objects.filter(document=document).exclude(recipient=request.user).delete()
    else:
        routing.is_cancelled = True
        routing.cancelled_at = timezone.now()
        routing.cancelled_by = request.user
        routing.cancellation_reason = "បានលុបចោលការបញ្ជូនទៅកាន់ការិយាល័យ"
        routing.save()
        if routing.to_dept:
            Notification.objects.filter(document=document, recipient__profile__department=routing.to_dept).delete()

    # Revert document status and current_department
    document.current_department = document.origin_department
    if document.status == 'ROUTED':
        document.status = 'PENDING_ADMIN'
    document.save()

    # Log cancellation in timeline
    DocumentRouting.objects.create(
        document=document,
        from_user=request.user,
        from_dept=profile.department if profile else None,
        to_dept=document.origin_department,
        action_taken="🚫 បានលុបចោល (Cancel) ការបញ្ជូនលិខិត",
        notes=f"បានដកការបញ្ជូនពី «{target_name}» ត្រឡប់មកវិញ",
        is_cancelled=False
    )

    messages.success(request, f"✅ បានលុបចោល (Cancel) ការបញ្ជូនលិខិត {document.registry_number} ទៅកាន់ {target_name} ដោយជោគជ័យ!")
    return redirect('document_detail', pk=pk)


@login_required
def complete_document(request, pk):
    document = get_object_or_404(Document, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None

    if document.status == 'COMPLETED':
        messages.info(request, "ឯកសារនេះត្រូវបានចុះយល់ព្រមទទួលយករួចរាល់ហើយ!")
        return redirect('document_detail', pk=pk)

    is_admin = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (dept and dept.code == 'ADMIN')
    admin_dept = Department.objects.filter(code='ADMIN').first()

    can_complete = False
    if is_admin:
        is_annotated_to_admin = bool(admin_dept and document.annotations.filter(target_departments=admin_dept).exists())
        is_routed_to_admin = bool(admin_dept and document.routings.filter(to_dept=admin_dept, is_cancelled=False).exists())
        from_specialized = bool(document.origin_department and document.origin_department != admin_dept)

        if from_specialized or is_annotated_to_admin or is_routed_to_admin:
            can_complete = True
        else:
            messages.error(request, "⚠️ ឯកសារនេះជាលិខិតរបស់រដ្ឋបាលផ្ញើទៅការិយាល័យជំនាញ ដោយពុំមានការចាត់ចែងជូនការិយាល័យកិច្ចការទូទៅឡើយ!")
            return redirect('document_detail', pk=pk)
    else:
        # For Specialized Office:
        is_annotated_to_dept = bool(dept and document.annotations.filter(target_departments=dept).exists())
        is_routed_to_dept = bool(dept and document.routings.filter(to_dept=dept, is_cancelled=False).exists())
        is_curr_dept = bool(dept and document.current_department == dept)

        if is_annotated_to_dept or is_routed_to_dept or is_curr_dept:
            can_complete = True

    if not can_complete:
        messages.error(request, "⚠️ លោកអ្នកគ្មានសិទ្ធិ ឬមិនមែនជាការិយាល័យទទួលឯកសារនេះឡើយ!")
        return redirect('document_detail', pk=pk)

    document.status = 'COMPLETED'
    document.save()

    dept_name = profile.department.name_kh if profile and profile.department else "អ្នកទទួល"
    user_name = request.user.get_full_name() or request.user.username

    DocumentRouting.objects.create(
        document=document,
        from_user=request.user,
        from_dept=profile.department if profile else None,
        to_dept=document.current_department,
        action_taken=f"✅ {dept_name} បានចុះយល់ព្រម និងទទួលយកឯកសារ (Completed)",
        notes=f"បានចុះយល់ព្រមទទួលយកដោយ៖ {user_name} ({dept_name})"
    )

    # Notify creator if different user
    if document.created_by and document.created_by != request.user:
        Notification.objects.create(
            recipient=document.created_by,
            document=document,
            title="ការិយាល័យទទួលបានចុះយល់ព្រមឯកសារ",
            message=f"ឯកសារលេខ {document.registry_number} ត្រូវ​បាន «{dept_name}» ចុះយល់ព្រមទទួលយក និងបិទបញ្ចប់រួចរាល់។"
        )

    messages.success(request, f"✅ ការិយាល័យ «{dept_name}» បានចុះយល់ព្រម និងទទួលយកឯកសារលេខ {document.registry_number} រួចរាល់ជាស្ថាពរ!")
    return redirect('document_detail', pk=pk)


@login_required
def document_version_add(request, pk):
    document = get_object_or_404(Document, pk=pk)
    profile = getattr(request.user, 'profile', None)

    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        if not profile or not profile.can_edit_document:
            messages.error(request, "⚠️ គណនីរបស់អ្នកមិនត្រូវបាន ADMIN អនុញ្ញាតឱ្យកែប្រែ ឬបន្ថែម Version ឯកសារឡើយ!")
            return redirect('document_detail', pk=pk)

    if request.method == 'POST':
        form = DocumentVersionForm(request.POST, request.FILES)
        if form.is_valid():
            version = form.save(commit=False)
            version.document = document
            version.uploaded_by = request.user
            if request.FILES.get('file'):
                version.file_name = request.FILES['file'].name
            version.save()

            # If uploaded by Admin / General Affairs for a specialized office document:
            is_admin_or_lead = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (profile and profile.department and profile.department.code in ['ADMIN', 'LEAD'])
            if is_admin_or_lead and document.origin_department and document.origin_department != getattr(profile, 'department', None):
                document.status = 'COMPLETED'
                document.current_department = document.origin_department
                document.save()

                DocumentRouting.objects.create(
                    document=document,
                    from_user=request.user,
                    from_dept=profile.department if profile else None,
                    to_dept=document.origin_department,
                    action_taken=f"📤 រដ្ឋបាលបាន Upload ឯកសារសម្រេច V{version.version_number} និងបញ្ជូនត្រឡប់មកការិយាល័យដើមវិញ (Completed)",
                    notes=f"ឯកសារសម្រេចត្រូវបាន Upload ដោយរដ្ឋបាល៖ {version.change_summary or 'ឯកសារសម្រេចពីថ្នាក់ដឹកនាំ'}"
                )

                # Send notification to origin department users
                origin_users = [p.user for p in UserProfile.objects.filter(department=document.origin_department)]
                for u in origin_users:
                    Notification.objects.create(
                        recipient=u,
                        document=document,
                        title="ឯកសារសម្រេចត្រូវបាន Upload រួចរាល់",
                        message=f"ឯកសារលេខ {document.registry_number} ត្រូវបានរដ្ឋបាល Upload ឯកសារសម្រេចពីថ្នាក់ដឹកនាំ និងបញ្ជូនត្រឡប់មកការិយាល័យលោកអ្នកវិញរួចរាល់។"
                    )

                messages.success(request, f"✅ បាន Upload ឯកសារសម្រេច V{version.version_number} និងបញ្ជូនត្រឡប់ទៅកាន់ «{document.origin_department.name_kh}» រួចរាល់ជាស្ថាពរ!")
            else:
                messages.success(request, f"បាន Upload ជំនាន់ថ្មី V{version.version_number} សម្រាប់ឯកសារ {document.registry_number}!")
    return redirect('document_detail', pk=pk)


@login_required
def document_version_download(request, pk, version_id):
    document = get_object_or_404(Document, pk=pk)
    version = get_object_or_404(DocumentVersion, pk=version_id, document=document)

    if not version.file:
        messages.error(request, "⚠️ ឯកសារនេះមិនមាន File សម្រាប់ទាញយកឡើយ!")
        return redirect('document_detail', pk=pk)

    try:
        file_path = version.file.path
        if not os.path.exists(file_path):
            messages.error(request, "⚠️ រកមិនឃើញ File ឯកសារក្នុង Server ឡើយ!")
            return redirect('document_detail', pk=pk)

        filename = version.file_name or os.path.basename(file_path)
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename)
    except Exception as e:
        messages.error(request, f"⚠️ បរាជ័យក្នុងការទាញយក File: {str(e)}")
        return redirect('document_detail', pk=pk)


from django.views.decorators.clickjacking import xframe_options_exempt, xframe_options_sameorigin

@login_required
@xframe_options_exempt
def document_version_view(request, pk, version_id):
    """
    Allows users (both specialized offices and admin/leadership) to view/preview document files inline in browser/iframe.
    """
    import mimetypes
    document = get_object_or_404(Document, pk=pk)
    version = get_object_or_404(DocumentVersion, pk=version_id, document=document)

    if not version.file:
        messages.error(request, "⚠️ ឯកសារនេះមិនមាន File សម្រាប់បើកមើលឡើយ!")
        return redirect('document_detail', pk=pk)

    try:
        file_path = version.file.path
        if not os.path.exists(file_path):
            messages.error(request, "⚠️ រកមិនឃើញ File ឯកសារក្នុង Server ឡើយ!")
            return redirect('document_detail', pk=pk)

        filename = version.file_name or os.path.basename(file_path)
        content_type, _ = mimetypes.guess_type(file_path)
        if not content_type:
            if filename.lower().endswith('.pdf'):
                content_type = 'application/pdf'
            elif filename.lower().endswith(('.jpg', '.jpeg')):
                content_type = 'image/jpeg'
            elif filename.lower().endswith('.png'):
                content_type = 'image/png'
            else:
                content_type = 'application/octet-stream'

        response = FileResponse(open(file_path, 'rb'), content_type=content_type)
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        response['X-Frame-Options'] = 'ALLOWALL'
        return response
    except Exception as e:
        messages.error(request, f"⚠️ បរាជ័យក្នុងការបើកមើល File: {str(e)}")
        return redirect('document_detail', pk=pk)


@login_required
def document_print(request, pk):
    document = get_object_or_404(Document, pk=pk)
    profile = getattr(request.user, 'profile', None)

    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        if not profile or not profile.can_print:
            messages.error(request, "⚠️ គណនីរបស់អ្នកមិនត្រូវបាន ADMIN អនុញ្ញាតឱ្យបោះពុម្ពប័ណ្ណតាមដានឡើយ!")
            return redirect('document_detail', pk=pk)

    latest_annotation = document.latest_annotation
    routings = document.routings.all().order_by('routed_at')

    context = {
        'document': document,
        'annotation': latest_annotation,
        'routings': routings,
        'today': timezone.now().date(),
    }
    return render(request, 'dms/document_print.html', context)


def get_report_date_range(params):
    today = timezone.now().date()
    period = params.get('period', 'monthly')
    
    try:
        year = int(params.get('year', today.year)) if params.get('year') else today.year
    except ValueError:
        year = today.year

    try:
        month = int(params.get('month', today.month)) if params.get('month') else today.month
    except ValueError:
        month = today.month

    try:
        quarter = int(params.get('quarter', (today.month - 1) // 3 + 1)) if params.get('quarter') else (today.month - 1) // 3 + 1
    except ValueError:
        quarter = 1

    try:
        semester = int(params.get('semester', 1 if today.month <= 6 else 2)) if params.get('semester') else (1 if today.month <= 6 else 2)
    except ValueError:
        semester = 1

    custom_day = params.get('custom_day', '')
    start_date_str = params.get('start_date', '')
    end_date_str = params.get('end_date', '')

    month_names_kh = ['', 'មករា', 'កុម្ភៈ', 'មីនា', 'មេសា', 'ឧសភា', 'មិថុនា', 'កក្កដា', 'សីហា', 'កញ្ញា', 'តុលា', 'វិច្ឆិកា', 'ធ្នូ']

    if period == 'daily':
        if custom_day:
            try:
                target_date = datetime.strptime(custom_day, '%Y-%m-%d').date()
            except ValueError:
                target_date = today
        else:
            target_date = today
        return target_date, target_date, f"ប្រចាំថ្ងៃ {target_date.strftime('%d/%m/%Y')}", period

    elif period == 'weekly':
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start, end, f"ប្រចាំសប្តាហ៍ ({start.strftime('%d/%m/%Y')} ដល់ {end.strftime('%d/%m/%Y')})", period

    elif period == 'monthly':
        last_day = calendar.monthrange(year, month)[1]
        start = date(year, month, 1)
        end = date(year, month, last_day)
        m_name = month_names_kh[month] if (1 <= month <= 12) else f"ខែ {month}"
        return start, end, f"ប្រចាំខែ {m_name} ឆ្នាំ {year}", period

    elif period == 'quarterly':
        start_month = (quarter - 1) * 3 + 1
        end_month = start_month + 2
        last_day = calendar.monthrange(year, end_month)[1]
        start = date(year, start_month, 1)
        end = date(year, end_month, last_day)
        return start, end, f"ប្រចាំត្រីមាសទី {quarter} ឆ្នាំ {year} (ខែ {start_month:02d} - {end_month:02d})", period

    elif period == 'semester':
        if semester == 1:
            start = date(year, 1, 1)
            end = date(year, 6, 30)
            label = f"ប្រចាំឆមាសទី ១ ឆ្នាំ {year} (មករា - មិថុនា)"
        else:
            start = date(year, 7, 1)
            end = date(year, 12, 31)
            label = f"ប្រចាំឆមាសទី ២ ឆ្នាំ {year} (កក្កដា - ធ្នូ)"
        return start, end, label, period

    elif period == 'nine_month':
        start = date(year, 1, 1)
        end = date(year, 9, 30)
        return start, end, f"ប្រចាំនព្វមាស ឆ្នាំ {year} (៩ខែដើមឆ្នាំ៖ មករា - កញ្ញា)", period

    elif period == 'yearly':
        start = date(year, 1, 1)
        end = date(year, 12, 31)
        return start, end, f"ប្រចាំឆ្នាំ {year}", period

    elif period == 'custom' and start_date_str and end_date_str:
        try:
            start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            return start, end, f"ចន្លោះកាលបរិច្ឆេទ {start.strftime('%d/%m/%Y')} ដល់ {end.strftime('%d/%m/%Y')}", period
        except ValueError:
            pass

    # Default fallback to current month
    last_day = calendar.monthrange(year, month)[1]
    m_name = month_names_kh[month] if (1 <= month <= 12) else f"ខែ {month}"
    return date(year, month, 1), date(year, month, last_day), f"ប្រចាំខែ {m_name} ឆ្នាំ {year}", 'monthly'


@login_required
def reports_view(request):
    profile = getattr(request.user, 'profile', None)

    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        if not profile or not profile.can_view_reports:
            messages.error(request, "⚠️ គណនីរបស់អ្នកមិនត្រូវបាន ADMIN អនុញ្ញាតឱ្យមើលរបាយការណ៍ឡើយ!")
            return redirect('dashboard')

    is_super_or_lead = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (profile and profile.is_leadership)

    # Department scoping
    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')
    selected_dept = None
    dept_id = request.GET.get('department_id', '')

    if is_super_or_lead:
        if dept_id and dept_id != 'all':
            try:
                selected_dept = Department.objects.get(id=int(dept_id))
            except (Department.DoesNotExist, ValueError):
                selected_dept = None
    else:
        # Regular specialized officer: STRICTLY locked to own department
        selected_dept = profile.department if profile else None

    # Calculate date range from period parameters
    start_date, end_date, period_label, period_type = get_report_date_range(request.GET)

    # Base queryset filtered by date range
    docs = Document.objects.filter(received_date__gte=start_date, received_date__lte=end_date)

    # Apply department filter
    if selected_dept:
        docs = docs.filter(
            Q(origin_department=selected_dept) | Q(current_department=selected_dept) | Q(routings__to_dept=selected_dept)
        ).distinct()

    # Breakdown statistics
    total_count = docs.count()
    inbound_count = docs.filter(doc_type='INBOUND').count()
    outbound_count = docs.filter(doc_type='OUTBOUND').count()
    completed_count = docs.filter(status='COMPLETED').count()
    pending_count = docs.exclude(status='COMPLETED').count()

    # New registrations within period originating from dept
    if selected_dept:
        new_registered_count = docs.filter(origin_department=selected_dept).count()
    else:
        new_registered_count = docs.count()

    urgency_stats = docs.values('urgency').annotate(count=Count('id'))

    # Available years list (e.g. 2024 to current_year + 1)
    current_year = timezone.now().year
    years_list = list(range(current_year - 3, current_year + 2))

    context = {
        'docs': docs.order_by('-received_date', '-id'),
        'total_count': total_count,
        'inbound_count': inbound_count,
        'outbound_count': outbound_count,
        'new_registered_count': new_registered_count,
        'completed_count': completed_count,
        'pending_count': pending_count,
        'urgency_stats': urgency_stats,
        'period_label': period_label,
        'period_type': period_type,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'is_super_or_lead': is_super_or_lead,
        'departments': departments,
        'selected_dept': selected_dept,
        'years_list': years_list,
        'current_year': int(request.GET.get('year', current_year)) if request.GET.get('year') else current_year,
        'current_month': int(request.GET.get('month', timezone.now().month)) if request.GET.get('month') else timezone.now().month,
        'current_quarter': int(request.GET.get('quarter', (timezone.now().month - 1) // 3 + 1)) if request.GET.get('quarter') else (timezone.now().month - 1) // 3 + 1,
        'current_semester': int(request.GET.get('semester', 1 if timezone.now().month <= 6 else 2)) if request.GET.get('semester') else (1 if timezone.now().month <= 6 else 2),
        'custom_day': request.GET.get('custom_day', ''),
    }
    return render(request, 'dms/reports.html', context)


@login_required
def export_excel_report(request):
    profile = getattr(request.user, 'profile', None)

    if not request.user.is_superuser and request.user.username.upper() != 'ADMIN':
        if not profile or not profile.can_view_reports:
            messages.error(request, "⚠️ គណនីរបស់អ្នកមិនត្រូវបាន ADMIN អនុញ្ញាតឱ្យទាញយករបាយការណ៍ Excel ឡើយ!")
            return redirect('dashboard')

    is_super_or_lead = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (profile and profile.is_leadership)

    selected_dept = None
    dept_id = request.GET.get('department_id', '')
    if is_super_or_lead:
        if dept_id and dept_id != 'all':
            try:
                selected_dept = Department.objects.get(id=int(dept_id))
            except (Department.DoesNotExist, ValueError):
                selected_dept = None
    else:
        selected_dept = profile.department if profile else None

    start_date, end_date, period_label, _ = get_report_date_range(request.GET)

    docs = Document.objects.filter(received_date__gte=start_date, received_date__lte=end_date)
    if selected_dept:
        docs = docs.filter(
            Q(origin_department=selected_dept) | Q(current_department=selected_dept) | Q(routings__to_dept=selected_dept)
        ).distinct()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "របាយការណ៍ឯកសារ"

    # Header styling
    header_font = Font(name='Khmer OS Battambang', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')

    # Title Banner
    ws.merge_cells('A1:J1')
    dept_title = f" - {selected_dept.name_kh}" if selected_dept else " - ទូទាំងមន្ទីរ"
    ws['A1'] = f"របាយការណ៍ស្ថិតិឯកសាររដ្ឋបាល ({period_label}){dept_title}"
    ws['A1'].font = Font(name='Khmer OS Muol Light', size=13, bold=True, color='1F4E78')
    ws['A1'].alignment = align_center

    ws.merge_cells('A2:J2')
    ws['A2'] = f"ចន្លោះកាលបរិច្ឆេទ៖ {start_date.strftime('%d/%m/%Y')} ដល់ {end_date.strftime('%d/%m/%Y')} | សរុប៖ {docs.count()} ឯកសារ (លិខិតចូល: {docs.filter(doc_type='INBOUND').count()} / លិខិតចេញ: {docs.filter(doc_type='OUTBOUND').count()})"
    ws['A2'].font = Font(name='Khmer OS Battambang', size=10, italic=True, color='555555')
    ws['A2'].alignment = align_center

    headers = [
        'ល.រ', 'លេខចុះលិខិត', 'ប្រភេទ', 'កម្មវត្ថុ / ចំណងជើង',
        'ប្រភពលិខិត', 'អ្នកទទួល', 'កម្រិតប្រញាប់', 'ទីតាំងបច្ចុប្បន្ន', 'ស្ថានភាព', 'ថ្ងៃចុះបញ្ជី'
    ]

    ws.append([]) # row 3 blank
    ws.append(headers) # row 4

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    # Add Data
    for idx, d in enumerate(docs.order_by('-received_date', '-id'), start=1):
        row = [
            idx,
            d.registry_number,
            d.get_doc_type_display(),
            d.title,
            d.origin_org,
            d.destination_org,
            d.get_urgency_display(),
            d.current_department.name_kh if d.current_department else 'មន្ទីរ',
            d.get_status_display(),
            d.received_date.strftime('%Y-%m-%d')
        ]
        ws.append(row)

    # Column width auto adjustment
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"dms_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def mark_notification_read(request, pk):
    notif = get_object_or_404(Notification, pk=pk, recipient=request.user)
    notif.is_read = True
    notif.save()
    if notif.document:
        return redirect('document_detail', pk=notif.document.pk)
    return redirect('dashboard')


@login_required
def mark_all_notifications_read(request):
    request.user.notifications.filter(is_read=False).update(is_read=True)
    messages.success(request, "បានផ្លាស់ប្តូរសារជូនដំណឹងទាំងអស់ជា 'បានអាន' រួចរាល់!")
    return redirect('dashboard')


# ==========================================
# 📥 DEDICATED INBOUND DOCUMENTS WORKSPACE
# ==========================================
@login_required
def inbound_documents_view(request):
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    is_super_or_lead = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (profile and profile.is_leadership)

    inbound_source = request.GET.get('source', 'all') # 'all', 'external', 'internal'
    status_filter = request.GET.get('status', '')
    search_q = request.GET.get('q', '').strip()
    urgency_filter = request.GET.get('urgency', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    if is_super_or_lead:
        # Admin / Leadership global view
        if inbound_source == 'external':
            base_queryset = Document.objects.filter(doc_type='INBOUND')
        elif inbound_source == 'internal':
            base_queryset = Document.objects.filter(routings__is_cancelled=False).distinct()
        else:
            base_queryset = Document.objects.filter(
                Q(doc_type='INBOUND') | Q(routings__is_cancelled=False)
            ).distinct()
    else:
        # Specialized Office:
        # - External Inbound: Registered as INBOUND assigned/routed to dept
        # - Internal Inbound: Routed from another office/leadership to this dept (no registration needed!)
        if dept:
            if inbound_source == 'external':
                base_queryset = Document.objects.filter(
                    doc_type='INBOUND'
                ).filter(
                    Q(current_department=dept) | Q(routings__to_dept=dept, routings__is_cancelled=False)
                ).distinct()
            elif inbound_source == 'internal':
                base_queryset = Document.objects.filter(
                    routings__to_dept=dept, routings__is_cancelled=False
                ).exclude(origin_department=dept).distinct()
            else: # 'all'
                base_queryset = Document.objects.filter(
                    Q(doc_type='INBOUND', current_department=dept) |
                    Q(doc_type='INBOUND', routings__to_dept=dept, routings__is_cancelled=False) |
                    Q(routings__to_dept=dept, routings__is_cancelled=False)
                ).exclude(origin_department=dept).distinct()
        else:
            base_queryset = Document.objects.none()

    queryset = base_queryset

    if status_filter:
        queryset = queryset.filter(status=status_filter)
    if urgency_filter:
        queryset = queryset.filter(urgency=urgency_filter)
    if start_date:
        queryset = queryset.filter(received_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(received_date__lte=end_date)
    if search_q:
        queryset = queryset.filter(
            Q(registry_number__icontains=search_q) |
            Q(title__icontains=search_q) |
            Q(origin_org__icontains=search_q) |
            Q(destination_org__icontains=search_q) |
            Q(summary__icontains=search_q)
        )

    # Scoped stats calculation
    if is_super_or_lead:
        ext_count = Document.objects.filter(doc_type='INBOUND').count()
        int_count = Document.objects.filter(routings__is_cancelled=False).distinct().count()
    else:
        ext_count = Document.objects.filter(doc_type='INBOUND').filter(Q(current_department=dept) | Q(routings__to_dept=dept, routings__is_cancelled=False)).distinct().count() if dept else 0
        int_count = Document.objects.filter(routings__to_dept=dept, routings__is_cancelled=False).exclude(origin_department=dept).distinct().count() if dept else 0

    stats = {
        'total': base_queryset.count(),
        'external': ext_count,
        'internal': int_count,
        'pending': base_queryset.exclude(status='COMPLETED').count(),
        'completed': base_queryset.filter(status='COMPLETED').count(),
        'urgent': base_queryset.filter(urgency__in=['URGENT', 'VERY_URGENT']).count(),
    }

    return render(request, 'dms/inbound_list.html', {
        'documents': queryset.order_by('-received_date', '-id'),
        'stats': stats,
        'inbound_source': inbound_source,
        'status_filter': status_filter,
        'search_q': search_q,
        'urgency_filter': urgency_filter,
        'start_date': start_date,
        'end_date': end_date,
        'dept': dept,
    })


# ==========================================
# 📤 DEDICATED OUTBOUND DOCUMENTS WORKSPACE
# ==========================================
@login_required
def outbound_documents_view(request):
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    is_super_or_lead = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (profile and profile.is_leadership)

    status_filter = request.GET.get('status', '')
    search_q = request.GET.get('q', '').strip()
    urgency_filter = request.GET.get('urgency', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    if not dept and (request.user.is_superuser or request.user.username.upper() == 'ADMIN'):
        dept = Department.objects.filter(code='ADMIN').first()

    # Outbound documents must STRICTLY belong to the user's origin department
    if dept:
        base_queryset = Document.objects.filter(origin_department=dept)
    else:
        base_queryset = Document.objects.filter(created_by=request.user)

    queryset = base_queryset

    if status_filter:
        queryset = queryset.filter(status=status_filter)
    if urgency_filter:
        queryset = queryset.filter(urgency=urgency_filter)
    if start_date:
        queryset = queryset.filter(received_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(received_date__lte=end_date)
    if search_q:
        queryset = queryset.filter(
            Q(registry_number__icontains=search_q) |
            Q(title__icontains=search_q) |
            Q(origin_org__icontains=search_q) |
            Q(destination_org__icontains=search_q) |
            Q(summary__icontains=search_q)
        )

    stats = {
        'total': base_queryset.count(),
        'pending_lead': base_queryset.filter(status='PENDING_LEADERSHIP').count(),
        'routed': base_queryset.filter(status='ROUTED').count(),
        'completed': base_queryset.filter(status='COMPLETED').count(),
        'urgent': base_queryset.filter(urgency__in=['URGENT', 'VERY_URGENT']).count(),
    }

    return render(request, 'dms/outbound_list.html', {
        'documents': queryset.order_by('-received_date', '-id'),
        'stats': stats,
        'status_filter': status_filter,
        'search_q': search_q,
        'urgency_filter': urgency_filter,
        'start_date': start_date,
        'end_date': end_date,
        'dept': dept,
    })


# ==========================================
# 🏢 MY DEPARTMENT ARCHIVE & FLOW WORKSPACE
# ==========================================
@login_required
def my_department_documents_view(request):
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None

    if not dept:
        messages.warning(request, "គណនីរបស់អ្នកមិនទាន់បានភ្ជាប់ជាមួយការិយាល័យណាជាក់លាក់ឡើយ!")
        return redirect('dashboard')

    doc_type_filter = request.GET.get('doc_type', '')
    status_filter = request.GET.get('status', '')
    search_q = request.GET.get('q', '').strip()

    queryset = Document.objects.filter(
        Q(origin_department=dept) | Q(current_department=dept) | Q(routings__to_dept=dept, routings__is_cancelled=False)
    ).distinct()

    if doc_type_filter:
        queryset = queryset.filter(doc_type=doc_type_filter)
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    if search_q:
        queryset = queryset.filter(
            Q(registry_number__icontains=search_q) |
            Q(title__icontains=search_q) |
            Q(origin_org__icontains=search_q) |
            Q(destination_org__icontains=search_q)
        )

    dept_docs = queryset.order_by('-updated_at')

    stats = {
        'total': queryset.count(),
        'inbound': queryset.filter(doc_type='INBOUND').count(),
        'outbound': queryset.filter(doc_type='OUTBOUND').count(),
        'completed': queryset.filter(status='COMPLETED').count(),
        'pending_action': queryset.filter(current_department=dept).exclude(status='COMPLETED').count(),
    }

    return render(request, 'dms/my_department_docs.html', {
        'department': dept,
        'documents': dept_docs,
        'stats': stats,
        'doc_type_filter': doc_type_filter,
        'status_filter': status_filter,
        'search_q': search_q,
    })


# ==========================================
# 💬 INTERNAL DEPARTMENT LIVE CHAT SYSTEM (STRICT CONFIDENTIALITY)
# ==========================================
@login_required
def live_chat_view(request):
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    is_super_or_lead = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (profile and profile.is_leadership)

    room_code = request.GET.get('room', 'all')
    target_dept = None

    if room_code != 'all':
        try:
            target_dept = Department.objects.get(id=int(room_code))
        except (Department.DoesNotExist, ValueError):
            room_code = 'all'

    if room_code == 'all':
        # Broadcast / Announcements channel
        messages_qs = ChatMessage.objects.filter(target_department__isnull=True)
    else:
        if is_super_or_lead:
            # Super admin or leadership can view channel
            messages_qs = ChatMessage.objects.filter(
                Q(target_department=target_dept) | Q(department=target_dept)
            )
        else:
            if user_dept:
                if target_dept == user_dept:
                    # Internal room within user's own department
                    messages_qs = ChatMessage.objects.filter(
                        department=user_dept, target_department=user_dept
                    )
                else:
                    # STRICT CONFIDENTIALITY: 1-to-1 conversation between user_dept and target_dept ONLY
                    messages_qs = ChatMessage.objects.filter(
                        (Q(department=user_dept) & Q(target_department=target_dept)) |
                        (Q(department=target_dept) & Q(target_department=user_dept))
                    )
            else:
                messages_qs = ChatMessage.objects.none()

    messages_qs = messages_qs.select_related('sender', 'department', 'target_department', 'related_document')
    recent_messages = list(messages_qs.order_by('-created_at')[:80])
    recent_messages.reverse()

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'id')

    # Calculate bell notification count for each department
    dept_bell_counts = {}
    for d in departments:
        if user_dept:
            if d.id == user_dept.id:
                # Internal room count
                c = ChatMessage.objects.filter(department=user_dept, target_department=user_dept).count()
            else:
                # Direct messages from this department to user_dept
                c = ChatMessage.objects.filter(department=d, target_department=user_dept).count()
        else:
            c = ChatMessage.objects.filter(department=d).count()
        dept_bell_counts[d.id] = c

    # Broadcast count
    broadcast_bell_count = ChatMessage.objects.filter(target_department__isnull=True).count()

    return render(request, 'dms/chat.html', {
        'departments': departments,
        'room_code': room_code,
        'target_dept': target_dept,
        'user_dept': user_dept,
        'is_super_or_lead': is_super_or_lead,
        'chat_messages': recent_messages,
        'dept_bell_counts': dept_bell_counts,
        'broadcast_bell_count': broadcast_bell_count,
    })


@login_required
def send_chat_api(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)

    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None

    text = request.POST.get('message', '').strip()
    target_dept_id = request.POST.get('target_department_id', '')

    if not text:
        return JsonResponse({'status': 'error', 'message': 'សូមវាយបញ្ចូលសារ! (ការបញ្ជូនឯកសារ/Files ផ្លូវការ ត្រូវធ្វើតាមមុខងារបញ្ជូនលិខិតផ្លូវការ)'}, status=400)

    target_dept = None
    if target_dept_id and target_dept_id != 'all':
        try:
            target_dept = Department.objects.get(id=int(target_dept_id))
        except (Department.DoesNotExist, ValueError):
            target_dept = None

    chat_msg = ChatMessage.objects.create(
        sender=request.user,
        department=user_dept,
        target_department=target_dept,
        message=text,
        attachment=None
    )

    return JsonResponse({
        'status': 'success',
        'message_id': chat_msg.id,
        'sender_name': request.user.get_full_name() or request.user.username,
        'sender_username': request.user.username,
        'sender_dept': user_dept.name_kh if user_dept else 'មន្ទីរ',
        'sender_position': profile.position_title if profile else '',
        'text': chat_msg.message,
        'is_edited': chat_msg.is_edited,
        'attachment_url': '',
        'attachment_name': '',
        'created_at': chat_msg.created_at.strftime('%H:%M'),
        'created_date': chat_msg.created_at.strftime('%d/%m/%Y'),
        'is_me': True,
        'can_modify': True,
    })


@login_required
def edit_chat_message_api(request, pk):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)

    chat_msg = get_object_or_404(ChatMessage, pk=pk)
    is_admin = request.user.is_superuser or request.user.username.upper() == 'ADMIN'

    if chat_msg.sender != request.user and not is_admin:
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    new_text = request.POST.get('message', '').strip()
    if not new_text:
        return JsonResponse({'status': 'error', 'message': 'ខ្លឹមសារសារមិនអាចទទេបានឡើយ!'}, status=400)

    chat_msg.message = new_text
    chat_msg.is_edited = True
    chat_msg.save()

    return JsonResponse({
        'status': 'success',
        'message_id': chat_msg.id,
        'text': chat_msg.message,
        'is_edited': True,
    })


@login_required
def delete_chat_message_api(request, pk):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)

    chat_msg = get_object_or_404(ChatMessage, pk=pk)
    is_admin = request.user.is_superuser or request.user.username.upper() == 'ADMIN'

    if chat_msg.sender != request.user and not is_admin:
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    msg_id = chat_msg.id
    chat_msg.delete()

    return JsonResponse({
        'status': 'success',
        'message_id': msg_id,
    })


@login_required
def get_chat_messages_api(request):
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    is_super_or_lead = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (profile and profile.is_leadership)

    room_code = request.GET.get('room', 'all')
    after_id = int(request.GET.get('after_id', 0))

    if room_code == 'all':
        messages_qs = ChatMessage.objects.filter(target_department__isnull=True, id__gt=after_id)
    else:
        try:
            dept_id = int(room_code)
            target_dept = Department.objects.get(id=dept_id)

            if is_super_or_lead:
                messages_qs = ChatMessage.objects.filter(
                    Q(target_department=target_dept) | Q(department=target_dept),
                    id__gt=after_id
                )
            else:
                if user_dept:
                    if target_dept == user_dept:
                        messages_qs = ChatMessage.objects.filter(
                            department=user_dept, target_department=user_dept, id__gt=after_id
                        )
                    else:
                        # STRICT CONFIDENTIALITY: Only messages between user_dept and target_dept
                        messages_qs = ChatMessage.objects.filter(
                            (Q(department=user_dept) & Q(target_department=target_dept)) |
                            (Q(department=target_dept) & Q(target_department=user_dept)),
                            id__gt=after_id
                        )
                else:
                    messages_qs = ChatMessage.objects.none()
        except (Department.DoesNotExist, ValueError):
            messages_qs = ChatMessage.objects.none()

    messages_qs = messages_qs.select_related('sender', 'department', 'target_department', 'related_document').order_by('created_at')[:50]

    data = []
    for m in messages_qs:
        sender_profile = getattr(m.sender, 'profile', None)
        can_mod = (m.sender == request.user) or is_super_or_lead
        data.append({
            'id': m.id,
            'sender_id': m.sender.id,
            'sender_name': m.sender.get_full_name() or m.sender.username,
            'sender_username': m.sender.username,
            'sender_dept': m.department.name_kh if m.department else 'មន្ទីរ',
            'sender_position': sender_profile.position_title if sender_profile else '',
            'text': m.message,
            'is_edited': m.is_edited,
            'attachment_url': m.attachment.url if m.attachment else '',
            'attachment_name': m.attachment.name.split('/')[-1] if m.attachment else '',
            'related_doc_id': m.related_document.id if m.related_document else None,
            'related_doc_title': m.related_document.title if m.related_document else '',
            'related_doc_reg': m.related_document.registry_number if m.related_document else '',
            'created_at': m.created_at.strftime('%H:%M'),
            'created_date': m.created_at.strftime('%d/%m/%Y'),
            'is_me': m.sender == request.user,
            'can_modify': can_mod,
        })

    return JsonResponse({'status': 'success', 'messages': data})


@login_required
def check_new_chat_notifications_api(request):
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    is_super_or_lead = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (profile and profile.is_leadership)

    try:
        after_id = int(request.GET.get('after_id', 0))
    except ValueError:
        after_id = 0

    if is_super_or_lead:
        base_qs = ChatMessage.objects.exclude(sender=request.user)
    else:
        if user_dept:
            base_qs = ChatMessage.objects.filter(
                Q(target_department=user_dept) | Q(target_department__isnull=True)
            ).exclude(sender=request.user)
        else:
            base_qs = ChatMessage.objects.none()

    if after_id == 0:
        latest_msg = base_qs.order_by('-id').first()
        latest_id = latest_msg.id if latest_msg else 0
        return JsonResponse({
            'status': 'success',
            'latest_id': latest_id,
            'new_messages': [],
            'count': 0,
        })

    new_messages = base_qs.filter(id__gt=after_id).select_related('sender', 'department').order_by('created_at')[:10]

    data = []
    latest_id = after_id
    for m in new_messages:
        sender_profile = getattr(m.sender, 'profile', None)
        sender_dept_name = m.department.name_kh if m.department else 'មន្ទីរ'
        sender_pos = sender_profile.position_title if sender_profile else ''
        
        # Room URL to reply
        if m.target_department is None:
            room_url = "/chat/?room=all"
        else:
            room_url = f"/chat/?room={m.department.id if m.department else 'all'}"

        data.append({
            'id': m.id,
            'sender_name': m.sender.get_full_name() or m.sender.username,
            'sender_dept': sender_dept_name,
            'sender_position': sender_pos,
            'text': m.message,
            'time': m.created_at.strftime('%H:%M'),
            'room_url': room_url,
        })
        if m.id > latest_id:
            latest_id = m.id

    return JsonResponse({
        'status': 'success',
        'latest_id': latest_id,
        'new_messages': data,
        'count': len(data),
    })


# =========================================================================
# 👥 CIVIL SERVANT PROFILE & HR MANAGEMENT (គ្រប់គ្រងជីវប្រវត្តិមន្ត្រីរាជការ)
# រៀបចំតាមស្តង់ដារផ្លូវការ ក ដល់ ឆ / ជ
# =========================================================================

def _parse_tabular_json(post_data, prefix, keys):
    """
    Parses dynamic rows sent from HTML table inputs into a list of dicts.
    """
    import json
    raw = post_data.get(f"{prefix}_json_raw")
    if raw:
        try:
            return json.loads(raw)
        except Exception:
            pass

    first_key = keys[0]
    first_list = post_data.getlist(f"{prefix}_{first_key}")
    if not first_list:
        return []

    num_rows = len(first_list)
    key_lists = {k: post_data.getlist(f"{prefix}_{k}") for k in keys}
    rows = []
    for i in range(num_rows):
        row = {}
        has_value = False
        for k in keys:
            val_list = key_lists[k]
            val = val_list[i].strip() if i < len(val_list) else ''
            row[k] = val
            if val:
                has_value = True
        if has_value:
            rows.append(row)
    return rows


def check_is_system_admin(user, request=None):
    """
    Checks if a user has full administrative privileges over civil servant profiles.
    Strictly checks user profile role and admin credentials.
    """
    if not user or not user.is_authenticated:
        return False
    
    # If the user has a specialized office or leadership role, they are NOT system admin!
    profile = getattr(user, 'profile', None)
    if profile:
        if profile.role in ['SPECIALIZED', 'LEADERSHIP']:
            return False
        if profile.role == 'ADMIN':
            return True
            
    if (user.username or '').upper() == 'ADMIN':
        return True
    if user.is_superuser:
        return True
    return False


@login_required
def officer_list(request):
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    # Submission & Edit Window Setting
    edit_window = OfficerEditWindowSetting.get_setting()
    is_window_open = edit_window.is_open_for_department(dept)
    can_edit_officers = is_system_admin or is_window_open

    search_q = request.GET.get('q', '').strip()
    dept_filter = request.GET.get('department', '').strip()
    gender_filter = request.GET.get('gender', '').strip()
    marital_filter = request.GET.get('marital_status', '').strip()

    queryset = CivilServantProfile.objects.select_related('department', 'user').all()

    # Specialized Technical Offices can ONLY view officers in their own department
    if not is_admin_or_lead:
        if dept:
            queryset = queryset.filter(department=dept)
        else:
            queryset = queryset.none()
    elif dept_filter:
        queryset = queryset.filter(department_id=dept_filter)

    if search_q:
        q_arabic = to_arabic_digits(search_q)
        queryset = queryset.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_last_name__icontains=search_q) |
            Q(latin_first_name__icontains=search_q) |
            Q(officer_id_number__icontains=search_q) |
            Q(officer_id_number__icontains=q_arabic) |
            Q(national_id_number__icontains=search_q) |
            Q(national_id_number__icontains=q_arabic) |
            Q(phone__icontains=search_q) |
            Q(phone__icontains=q_arabic) |
            Q(email__icontains=search_q) |
            Q(current_rank_and_step__icontains=search_q) |
            Q(current_position_title__icontains=search_q)
        )

    if gender_filter:
        queryset = queryset.filter(gender=gender_filter)

    if marital_filter:
        queryset = queryset.filter(marital_status=marital_filter)

    total_officers = queryset.count()
    male_count = queryset.filter(gender='MALE').count()
    female_count = queryset.filter(gender='FEMALE').count()
    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh') if is_admin_or_lead else (Department.objects.filter(id=dept.id) if dept else [])

    from .models import officer_sort_key
    officers_list = list(queryset)
    officers_list.sort(key=officer_sort_key)

    context = {
        'officers': officers_list,
        'total_officers': total_officers,
        'male_count': male_count,
        'female_count': female_count,
        'departments': departments,
        'search_q': search_q,
        'dept_filter': dept_filter,
        'gender_filter': gender_filter,
        'marital_filter': marital_filter,
        'dept': dept,
        'is_admin_or_lead': is_admin_or_lead,
        'has_global_access': has_global_access,
        'is_system_admin': is_system_admin,
        'edit_window': edit_window,
        'is_window_open': is_window_open,
        'can_edit_officers': can_edit_officers,
        'can_export_officer_excel': can_export_civil_servants_to_excel(request.user, profile),
        'can_view_e2_report': can_access_officer_e2_report(request.user, profile),
    }
    return render(request, 'dms/officer_list.html', context)


def log_officer_action(request, officer, action_type, attachment_title="", category="", details=""):
    """
    Records an audit log entry for any action performed on an officer profile or attachment.
    """
    try:
        ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', ''))
        if ip and ',' in ip:
            ip = ip.split(',')[0].strip()
        OfficerAuditLog.objects.create(
            officer=officer,
            officer_name_cache=officer.full_name_kh if officer else '',
            user=request.user if request.user.is_authenticated else None,
            action_type=action_type,
            attachment_title=attachment_title,
            category=category,
            details=details,
            ip_address=ip or '127.0.0.1'
        )
    except Exception:
        pass


def _handle_form_attachments(request, officer):
    """
    Saves uploaded reference documents submitted from each individual section (ក, ខ, ឃ, ច, ឆ, ជ, ផ្សេងៗ).
    Enforces maximum file size of 2MB per attachment.
    """
    MAX_OFFICER_ATTACHMENT_SIZE = 2 * 1024 * 1024  # 2MB
    categories = ['PERSONAL', 'FAMILY', 'EDUCATION', 'APPOINTMENT', 'PROMOTION', 'AWARD', 'OTHER']
    for cat in categories:
        files = request.FILES.getlist(f'ref_file_{cat}')
        titles = request.POST.getlist(f'ref_title_{cat}')
        doc_nos = request.POST.getlist(f'ref_doc_no_{cat}')
        dates = request.POST.getlist(f'ref_date_{cat}')
        descs = request.POST.getlist(f'ref_desc_{cat}')

        for idx, f in enumerate(files):
            if f.size > MAX_OFFICER_ATTACHMENT_SIZE:
                file_size_mb = f.size / (1024 * 1024)
                messages.warning(request, f"⚠️ File «{f.name}» មានទំហំ {file_size_mb:.2f}MB (លើសពី 2MB) មិនត្រូវបានរក្សាទុកឡើយ!")
                continue

            title = titles[idx].strip() if idx < len(titles) and titles[idx].strip() else f.name
            doc_no = doc_nos[idx].strip() if idx < len(doc_nos) else ''
            issued_date = dates[idx].strip() if idx < len(dates) else ''
            desc = descs[idx].strip() if idx < len(descs) else ''

            att = OfficerAttachment.objects.create(
                officer=officer,
                category=cat,
                title=title,
                doc_number=doc_no,
                issued_date=issued_date,
                description=desc,
                file=f,
                file_size=f.size,
                uploaded_by=request.user
            )

            log_officer_action(
                request,
                officer=officer,
                action_type='UPLOAD',
                attachment_title=att.title,
                category=att.get_category_display(),
                details=f"បានបញ្ចូលឯកសារយោងក្នុងផ្នែក «{att.get_category_display()}»៖ «{att.title}» (ទំហំ៖ {att.formatted_file_size})"
            )


def _resolve_geo_address(province_code, district_code, commune_code, village_code,
                         province_name, district_name, commune_name, village_name):
    """
    Ensures bidirectional resolution between Cambodian Geographic Codes and Khmer Names.
    """
    province_code = (province_code or '').strip()
    district_code = (district_code or '').strip()
    commune_code = (commune_code or '').strip()
    village_code = (village_code or '').strip()
    province_name = (province_name or '').strip()
    district_name = (district_name or '').strip()
    commune_name = (commune_name or '').strip()
    village_name = (village_name or '').strip()

    # If code is given but name is empty, lookup name
    if province_code and not province_name:
        p = CambodiaProvince.objects.filter(code=province_code).first()
        if p:
            province_name = p.name_kh
    if district_code and not district_name:
        d = CambodiaDistrict.objects.filter(code=district_code).first()
        if d:
            district_name = d.name_kh
    if commune_code and not commune_name:
        c = CambodiaCommune.objects.filter(code=commune_code).first()
        if c:
            commune_name = c.name_kh
    if village_code and not village_name:
        v = CambodiaVillage.objects.filter(code=village_code).first()
        if v:
            village_name = v.name_kh

    # If name is given but code is empty, lookup code
    if province_name and not province_code:
        p = CambodiaProvince.objects.filter(name_kh=province_name).first()
        if p:
            province_code = p.code
    if district_name and not district_code:
        d_qs = CambodiaDistrict.objects.filter(name_kh=district_name)
        if province_code:
            d_qs = d_qs.filter(province_id=province_code)
        d = d_qs.first()
        if d:
            district_code = d.code
    if commune_name and not commune_code:
        c_qs = CambodiaCommune.objects.filter(name_kh=commune_name)
        if district_code:
            c_qs = c_qs.filter(district_id=district_code)
        c = c_qs.first()
        if c:
            commune_code = c.code
    if village_name and not village_code:
        v_qs = CambodiaVillage.objects.filter(name_kh=village_name)
        if commune_code:
            v_qs = v_qs.filter(commune_id=commune_code)
        v = v_qs.first()
        if v:
            village_code = v.code

    return (province_code or None, district_code or None, commune_code or None, village_code or None,
            province_name, district_name, commune_name, village_name)


@login_required
def officer_create(request):
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    # Edit window check for non-admin users
    edit_window = OfficerEditWindowSetting.get_setting()
    if not is_system_admin and not edit_window.is_open_for_department(dept):
        messages.warning(request, f"🔒 ប្រព័ន្ធត្រូវបានបិទមិនឱ្យបញ្ចូលជីវប្រវត្តិមន្ត្រីថ្មីឡើយ! ({edit_window.status_label_kh})")
        return redirect('officer_list')

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh') if is_admin_or_lead else (Department.objects.filter(id=dept.id) if dept else [])
    users = User.objects.filter(is_active=True).order_by('username')
    provinces = CambodiaProvince.objects.all().order_by('code')

    if request.method == 'POST':
        try:
            pos_title = request.POST.get('current_position_title', '').strip()
            if pos_title in ['ប្រធានមន្ទីរ', 'អនុប្រធានមន្ទីរ']:
                lead_dept = Department.objects.filter(Q(code='LEAD') | Q(name_kh='ថ្នាក់ដឹកនាំមន្ទីរ')).first()
                target_dept = lead_dept or dept
            elif not is_admin_or_lead:
                target_dept = dept
            else:
                target_dept_id = request.POST.get('department')
                target_dept = Department.objects.filter(id=target_dept_id).first() if target_dept_id else None

            # Resolve POB Address
            pob_pcode, pob_dcode, pob_ccode, pob_vcode, pob_pname, pob_dname, pob_cname, pob_vname = _resolve_geo_address(
                request.POST.get('pob_province_code'), request.POST.get('pob_district_code'),
                request.POST.get('pob_commune_code'), request.POST.get('pob_village_code'),
                request.POST.get('pob_province'), request.POST.get('pob_district'),
                request.POST.get('pob_commune'), request.POST.get('pob_village')
            )

            # Resolve Current Address
            cur_pcode, cur_dcode, cur_ccode, cur_vcode, cur_pname, cur_dname, cur_cname, cur_vname = _resolve_geo_address(
                request.POST.get('current_province_code'), request.POST.get('current_district_code'),
                request.POST.get('current_commune_code'), request.POST.get('current_village_code'),
                request.POST.get('current_province'), request.POST.get('current_district'),
                request.POST.get('current_commune'), request.POST.get('current_village')
            )

            # Resolve Permanent Address
            perm_same = (request.POST.get('perm_same_as_current') == 'on' or request.POST.get('perm_same_as_current') == 'true')
            if perm_same:
                perm_pcode, perm_dcode, perm_ccode, perm_vcode = cur_pcode, cur_dcode, cur_ccode, cur_vcode
                perm_pname, perm_dname, perm_cname, perm_vname = cur_pname, cur_dname, cur_cname, cur_vname
                perm_hno, perm_st = request.POST.get('current_house_no', '').strip(), request.POST.get('current_street', '').strip()
            else:
                perm_pcode, perm_dcode, perm_ccode, perm_vcode, perm_pname, perm_dname, perm_cname, perm_vname = _resolve_geo_address(
                    request.POST.get('perm_province_code'), request.POST.get('perm_district_code'),
                    request.POST.get('perm_commune_code'), request.POST.get('perm_village_code'),
                    request.POST.get('perm_province'), request.POST.get('perm_district'),
                    request.POST.get('perm_commune'), request.POST.get('perm_village')
                )
                perm_hno = request.POST.get('perm_house_no', '').strip()
                perm_st = request.POST.get('perm_street', '').strip()

            officer = CivilServantProfile(
                created_by=request.user,
                department=target_dept,
                user_id=request.POST.get('user') or None,

                # ក- ព័ត៌មានផ្ទាល់ខ្លួន
                khmer_last_name=request.POST.get('khmer_last_name', '').strip(),
                khmer_first_name=request.POST.get('khmer_first_name', '').strip(),
                latin_last_name=request.POST.get('latin_last_name', '').strip().upper(),
                latin_first_name=request.POST.get('latin_first_name', '').strip().upper(),
                gender=request.POST.get('gender', 'MALE'),
                dob=request.POST.get('dob', '').strip(),
                ethnicity=request.POST.get('ethnicity', 'ខ្មែរ').strip(),
                nationality=request.POST.get('nationality', 'ខ្មែរ').strip(),

                pob_province_code=pob_pcode,
                pob_district_code=pob_dcode,
                pob_commune_code=pob_ccode,
                pob_village_code=pob_vcode,
                pob_village=pob_vname,
                pob_commune=pob_cname,
                pob_district=pob_dname,
                pob_province=pob_pname,

                current_province_code=cur_pcode,
                current_district_code=cur_dcode,
                current_commune_code=cur_ccode,
                current_village_code=cur_vcode,
                current_house_no=request.POST.get('current_house_no', '').strip(),
                current_street=request.POST.get('current_street', '').strip(),
                current_village=cur_vname,
                current_commune=cur_cname,
                current_district=cur_dname,
                current_province=cur_pname,

                perm_same_as_current=perm_same,
                perm_province_code=perm_pcode,
                perm_district_code=perm_dcode,
                perm_commune_code=perm_ccode,
                perm_village_code=perm_vcode,
                perm_house_no=perm_hno,
                perm_street=perm_st,
                perm_village=perm_vname,
                perm_commune=perm_cname,
                perm_district=perm_dname,
                perm_province=perm_pname,

                phone=request.POST.get('phone', '').strip(),
                email=request.POST.get('email', '').strip(),
                officer_id_number=request.POST.get('officer_id_number', '').strip(),
                national_id_number=request.POST.get('national_id_number', '').strip(),
                national_id_valid_from=request.POST.get('national_id_valid_from', '').strip(),
                national_id_valid_to=request.POST.get('national_id_valid_to', '').strip(),
                physical_condition=request.POST.get('physical_condition', 'គ្រប់គ្រាន់').strip(),
                disability_detail=request.POST.get('disability_detail', '').strip(),

                # ខ- ព័ត៌មានគ្រួសារ
                marital_status=request.POST.get('marital_status', 'SINGLE'),
                spouse_marriage_cert_no=request.POST.get('spouse_marriage_cert_no', '').strip(),
                spouse_name_kh=request.POST.get('spouse_name_kh', '').strip(),
                spouse_is_alive=(request.POST.get('spouse_is_alive', 'true') == 'true'),
                spouse_name_latin=request.POST.get('spouse_name_latin', '').strip().upper(),
                spouse_dob=request.POST.get('spouse_dob', '').strip(),
                spouse_national_id=request.POST.get('spouse_national_id', '').strip(),
                spouse_pob=request.POST.get('spouse_pob', '').strip(),
                spouse_occupation=request.POST.get('spouse_occupation', '').strip(),
                spouse_current_address=request.POST.get('spouse_current_address', '').strip(),
                spouse_organization=request.POST.get('spouse_organization', '').strip(),

                father_name=request.POST.get('father_name', '').strip(),
                father_is_alive=(request.POST.get('father_is_alive', 'true') == 'true'),
                father_pob=request.POST.get('father_pob', '').strip(),
                father_occupation=request.POST.get('father_occupation', '').strip(),

                mother_name=request.POST.get('mother_name', '').strip(),
                mother_is_alive=(request.POST.get('mother_is_alive', 'true') == 'true'),
                mother_pob=request.POST.get('mother_pob', '').strip(),
                mother_occupation=request.POST.get('mother_occupation', '').strip(),

                # គ- ទំនាក់ទំនងអាសន្ន
                emergency_last_name=request.POST.get('emergency_last_name', '').strip(),
                emergency_first_name=request.POST.get('emergency_first_name', '').strip(),
                emergency_gender=request.POST.get('emergency_gender', 'FEMALE'),
                emergency_relationship=request.POST.get('emergency_relationship', '').strip(),
                emergency_occupation=request.POST.get('emergency_occupation', '').strip(),
                emergency_address=request.POST.get('emergency_address', '').strip(),
                emergency_phone=request.POST.get('emergency_phone', '').strip(),
                emergency_email=request.POST.get('emergency_email', '').strip(),

                # ច- ក្របខ័ណ្ឌការងារ
                civil_service_start_date=request.POST.get('civil_service_start_date', '').strip(),
                civil_service_permanent_date=request.POST.get('civil_service_permanent_date', '').strip(),
                framework_name=request.POST.get('framework_name', '').strip(),
                framework_category=request.POST.get('framework_category', '').strip() or None,
                current_rank_and_step=request.POST.get('current_rank_and_step', '').strip(),
                current_position_title=request.POST.get('current_position_title', '').strip(),
                position_code=request.POST.get('position_code', '').strip() or CivilServantProfile.POSITION_CODE_MAP.get(request.POST.get('current_position_title', '').strip(), ''),
                highest_degree=request.POST.get('highest_degree', '').strip() or None,
                officer_status=request.POST.get('officer_status', 'ACTIVE').strip() or 'ACTIVE',

                # JSON Tables
                children_data=_parse_tabular_json(request.POST, 'child', ['name', 'gender', 'dob', 'occupation']),
                education_data=_parse_tabular_json(request.POST, 'edu', ['category', 'level_label', 'school', 'location', 'degree', 'skill', 'start_date', 'end_date']),
                languages_data=_parse_tabular_json(request.POST, 'lang', ['language', 'reading', 'speaking', 'writing']),
                history_public_sector=_parse_tabular_json(request.POST, 'pub', ['start_date', 'end_date', 'ministry', 'department', 'position', 'skill']),
                history_private_sector=_parse_tabular_json(request.POST, 'priv', ['start_date', 'end_date', 'org', 'role', 'skill']),
                promotions_by_seniority=_parse_tabular_json(request.POST, 'promo_sen', ['effective_date', 'ministry', 'department', 'office', 'old_rank_step', 'new_rank_step', 'promo_type']),
                promotions_by_degree=_parse_tabular_json(request.POST, 'promo_deg', ['effective_date', 'school', 'location', 'degree', 'old_rank_step', 'new_rank_step']),
                outside_framework_status=_parse_tabular_json(request.POST, 'out_fw', ['start_date', 'end_date', 'ministry', 'position']),
                unpaid_leave_status=_parse_tabular_json(request.POST, 'unpaid', ['start_date', 'end_date', 'ministry', 'duration']),
                awards_data=_parse_tabular_json(request.POST, 'award', ['doc_number', 'date', 'ministry', 'description', 'type']),
                sanctions_data=_parse_tabular_json(request.POST, 'sanct', ['doc_number', 'date', 'ministry', 'description', 'type']),
            )

            if 'photo' in request.FILES:
                officer.photo = request.FILES['photo']

            officer.save()

            # Process attached reference files from sections
            _handle_form_attachments(request, officer)

            # Record Audit Log
            log_officer_action(
                request,
                officer=officer,
                action_type='PROFILE_CREATE',
                details=f"បានបង្កើតជីវប្រវត្តិមន្ត្រីថ្មី «{officer.full_name_kh}» (អត្តលេខ: {officer.officer_id_number or '-'})"
            )

            active_tab = request.POST.get('active_tab', 'tab-a')
            save_action = request.POST.get('save_action', 'save_and_continue')

            if save_action == 'save_and_continue':
                messages.success(request, f"✅ បានរក្សាទុកទិន្នន័យ «{officer.full_name_kh}» ដោយជោគជ័យ! លោកអ្នកអាចបន្តបញ្ចូលព័ត៌មានបន្ថែមបាន។")
                return redirect(f"/officers/{officer.pk}/edit/?tab={active_tab}")
            else:
                messages.success(request, f"✅ បានបញ្ចូល និងរក្សាទុកជីវប្រវត្តិមន្ត្រី «{officer.full_name_kh}» ដោយជោគជ័យ!")
                return redirect('officer_detail', pk=officer.pk)
        except Exception as e:
            messages.error(request, f"⚠️ មានបញ្ហាក្នុងការរក្សាទុកទិន្នន័យ៖ {str(e)}")

    context = {
        'departments': departments,
        'users': users,
        'provinces': provinces,
        'is_edit': False,
        'dept': dept,
        'is_admin_or_lead': is_admin_or_lead,
        'has_global_access': has_global_access,
        'attachments': [],
        'attachments_by_category': {},
        'attachment_categories': OfficerAttachment.CATEGORY_CHOICES,
    }
    return render(request, 'dms/officer_form.html', context)


@login_required
def officer_import_docx(request):
    """
    Imports officer data automatically from uploaded Cambodian Civil Servant Word (.docx) file(s).
    Supports both traditional single file submission and AJAX batch per-file progress uploading.
    """
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.POST.get('format') == 'json' or 'application/json' in request.headers.get('Accept', '')
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = request.user.is_superuser or request.user.username.upper() == 'ADMIN' or (profile and profile.is_leadership) or has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    # Edit window check for non-admin users
    edit_window = OfficerEditWindowSetting.get_setting()
    if not is_system_admin and not edit_window.is_open_for_department(dept):
        msg = f"🔒 ប្រព័ន្ធត្រូវបានបិទមិនឱ្យនាំចូលជីវប្រវត្តិមន្ត្រីឡើយ! ({edit_window.status_label_kh})"
        if is_ajax:
            return JsonResponse({'success': False, 'error': msg}, status=403)
        messages.warning(request, msg)
        return redirect('officer_list')

    if request.method == 'POST':
        docx_file = request.FILES.get('docx_file')
        if not docx_file:
            if is_ajax:
                return JsonResponse({'success': False, 'error': 'សូមជ្រើសរើសឯកសារ Word (.docx) ដើម្បីនាំចូល!'})
            messages.error(request, "⚠️ សូមជ្រើសរើសឯកសារ Word (.docx) ដើម្បីនាំចូល!")
            return redirect('officer_list')

        if not docx_file.name.lower().endswith('.docx'):
            if is_ajax:
                return JsonResponse({'success': False, 'error': 'ប្រព័ន្ធគាំទ្រតែឯកសារប្រភេទ Word (.docx) ប៉ុណ្ណោះ!'})
            messages.error(request, "⚠️ ប្រព័ន្ធគាំទ្រតែឯកសារប្រភេទ Word (.docx) ប៉ុណ្ណោះ!")
            return redirect('officer_list')

        try:
            from .docx_parser import parse_docx_officer
            data = parse_docx_officer(docx_file, original_filename=docx_file.name)

            if not data.get('khmer_last_name') and not data.get('khmer_first_name'):
                if is_ajax:
                    return JsonResponse({'success': False, 'error': 'មិនអាចស្វែងរកឈ្មោះមន្ត្រីក្នុងឯកសារ Word នេះបានឡើយ។'})
                messages.error(request, "⚠️ មិនអាចស្វែងរកឈ្មោះមន្ត្រីក្នុងឯកសារ Word នេះបានឡើយ។ សូមពិនិត្យមើលទម្រង់ឯកសារម្តងទៀត។")
                return redirect('officer_list')

            # Auto-detect framework_category if not set
            rk = data.get('current_rank_and_step', '') or ''
            if rk.startswith('ក'):
                data['framework_category'] = 'A'
            elif rk.startswith('ខ'):
                data['framework_category'] = 'B'
            elif rk.startswith('គ'):
                data['framework_category'] = 'C'
            elif rk.startswith('ឃ'):
                data['framework_category'] = 'D'

            # Auto-detect highest degree
            edu_list = data.get('education_data', [])
            for edu in edu_list:
                deg = edu.get('degree', '') + ' ' + edu.get('level_label', '')
                if any(kw in deg for kw in ['បណ្ឌិត', 'PhD', 'Doctor', 'បណ្ឌិតវិទ្យាសាស្ត្រ']):
                    data['highest_degree'] = 'DOCTORATE'
                    break
                elif any(kw in deg for kw in ['អនុបណ្ឌិត', 'Master', 'MA', 'MS']):
                    if data.get('highest_degree') != 'DOCTORATE':
                        data['highest_degree'] = 'MASTER'
                elif any(kw in deg for kw in ['បរិញ្ញាបត្រ', 'Bachelor', 'BA', 'BS', 'វិស្វករ', 'វេជ្ជបណ្ឌិត']):
                    if data.get('highest_degree') not in ['DOCTORATE', 'MASTER']:
                        data['highest_degree'] = 'BACHELOR'
                elif any(kw in deg for kw in ['បរិញ្ញាបត្ររង', 'Associate', 'ជាន់ខ្ពស់']):
                    if data.get('highest_degree') not in ['DOCTORATE', 'MASTER', 'BACHELOR']:
                        data['highest_degree'] = 'ASSOCIATE'
                elif any(kw in deg for kw in ['មធ្យមសិក្សាទុតិយភូមិ', 'បាក់ឌុប', 'ទុតិយភូមិ', 'High School']):
                    if not data.get('highest_degree'):
                        data['highest_degree'] = 'HIGHSCHOOL'

            target_dept_id = request.POST.get('department')
            target_dept = None
            if target_dept_id:
                target_dept = Department.objects.filter(id=target_dept_id).first()
            elif not has_global_access and dept:
                target_dept = dept
            else:
                full_history_text = " ".join([h.get('department', '') + " " + h.get('ministry', '') for h in data.get('history_public_sector', [])])
                pos = data.get('current_position_title', '')
                if pos in ['ប្រធានមន្ទីរ', 'អនុប្រធានមន្ទីរ'] or 'ប្រធានមន្ទីរ' in full_history_text or 'ថ្នាក់ដឹកនាំ' in full_history_text:
                    target_dept = Department.objects.filter(Q(code='LEAD') | Q(name_kh='ថ្នាក់ដឹកនាំមន្ទីរ')).first()
                elif 'ព្រៃឈើ' in full_history_text:
                    target_dept = Department.objects.filter(Q(code='CANTON_FOREST') | Q(code='FOREST') | Q(code='FOREST_NEW')).first()
                elif 'ជលផល' in full_history_text:
                    target_dept = Department.objects.filter(Q(code='CANTON_FISH') | Q(code='FISH') | Q(code='FISH_NEW')).first()
                elif 'ក្សេត្រសាស្ត្រ' in full_history_text:
                    target_dept = Department.objects.filter(Q(code='AGRO_PROD') | Q(code='AGRO') | Q(code='AGRO_NEW')).first()
                elif 'សុខភាពសត្វ' in full_history_text or 'បសុព្យាបាល' in full_history_text or 'ផលិតកម្ម' in full_history_text:
                    target_dept = Department.objects.filter(Q(code='PROD_VET') | Q(code='AHVP') | Q(code='AHVP_NEW')).first()
                elif 'សហគមន៍' in full_history_text:
                    target_dept = Department.objects.filter(Q(code='AGRI_COMM') | Q(code='ACD')).first()
                elif 'គ្រឿងយន្ត' in full_history_text:
                    target_dept = Department.objects.filter(Q(code='AGRI_MACH') | Q(code='ENG') | Q(code='ENG_NEW')).first()
                elif 'កសិឧស្សាហកម្ម' in full_history_text:
                    target_dept = Department.objects.filter(code='AGRO_IND').first()
                elif 'កៅស៊ូ' in full_history_text:
                    target_dept = Department.objects.filter(code='RUBBER').first()
                elif 'ផ្សព្វផ្សាយ' in full_history_text:
                    target_dept = Department.objects.filter(code='AGRI_EXT').first()
                elif 'នីតិកម្ម' in full_history_text:
                    target_dept = Department.objects.filter(code='AGRI_LEG').first()
                elif 'ផែនការ' in full_history_text or 'គណនេយ្យ' in full_history_text:
                    target_dept = Department.objects.filter(code='PLAN_ACC').first()
                else:
                    target_dept = Department.objects.filter(code='ADMIN_PERS').first() or Department.objects.filter(code='ADMIN').first() or Department.objects.filter(is_active=True).first()

            duplicate_mode = request.POST.get('duplicate_mode', 'review') # review, auto_update, skip, force_update

            from .docx_parser import to_arabic_digits, compare_officer_data

            officer_id_num = to_arabic_digits(data.get('officer_id_number', ''))
            officer = None
            if officer_id_num:
                officer = CivilServantProfile.objects.filter(
                    Q(officer_id_number=officer_id_num) |
                    Q(officer_id_number=data.get('officer_id_number', ''))
                ).first()

            if not officer:
                officer = CivilServantProfile.objects.filter(
                    khmer_last_name=data['khmer_last_name'],
                    khmer_first_name=data['khmer_first_name']
                ).first()

            is_new = officer is None

            if not is_new and not has_global_access and officer.department and officer.department != dept:
                dept_name = officer.department.name_kh
                err_msg = f"មន្ត្រី «{officer.full_name_kh}» (អត្តលេខ: {officer.officer_id_number or '-'}) ស្ថិតនៅក្នុង «{dept_name}» រួចហើយ! លោកអ្នកគ្មានសិទ្ធិកែប្រែ ឬបន្ថែមទិន្នន័យមន្ត្រីក្រៅការិយាល័យឡើយ។"
                if is_ajax:
                    return JsonResponse({'success': False, 'error': err_msg})
                messages.error(request, f"⚠️ {err_msg}")
                return redirect('officer_list')

            # Auto-resolve Geographic administrative IDs
            p_pcode, p_dcode, p_ccode, p_vcode, p_pname, p_dname, p_cname, p_vname = _resolve_geo_address(
                data.get('pob_province_code'), data.get('pob_district_code'), data.get('pob_commune_code'), data.get('pob_village_code'),
                data.get('pob_province'), data.get('pob_district'), data.get('pob_commune'), data.get('pob_village')
            )
            data['pob_province_code'] = p_pcode
            data['pob_district_code'] = p_dcode
            data['pob_commune_code'] = p_ccode
            data['pob_village_code'] = p_vcode
            data['pob_province'] = p_pname
            data['pob_district'] = p_dname
            data['pob_commune'] = p_cname
            data['pob_village'] = p_vname

            c_pcode, c_dcode, c_ccode, c_vcode, c_pname, c_dname, c_cname, c_vname = _resolve_geo_address(
                data.get('current_province_code'), data.get('current_district_code'), data.get('current_commune_code'), data.get('current_village_code'),
                data.get('current_province'), data.get('current_district'), data.get('current_commune'), data.get('current_village')
            )
            data['current_province_code'] = c_pcode
            data['current_district_code'] = c_dcode
            data['current_commune_code'] = c_ccode
            data['current_village_code'] = c_vcode
            data['current_province'] = c_pname
            data['current_district'] = c_dname
            data['current_commune'] = c_cname
            data['current_village'] = c_vname

            if data.get('perm_same_as_current', True):
                data['perm_province_code'] = c_pcode
                data['perm_district_code'] = c_dcode
                data['perm_commune_code'] = c_ccode
                data['perm_village_code'] = c_vcode
                data['perm_province'] = c_pname
                data['perm_district'] = c_dname
                data['perm_commune'] = c_cname
                data['perm_village'] = c_vname

            if is_new:
                officer = CivilServantProfile(
                    department=target_dept,
                    created_by=request.user,
                    **data
                )
                officer.save()
                
                # Save uploaded original .docx as attachment
                try:
                    docx_file.seek(0)
                    OfficerAttachment.objects.create(
                        officer=officer,
                        category='OTHER',
                        title=f"ជីវប្រវត្តិដើមពី Word ({docx_file.name})",
                        file=docx_file,
                        file_size=docx_file.size,
                        description="បាននាំចូលដោយស្វ័យប្រវត្តិតាមរយៈប្រព័ន្ធ Import Word CV",
                        uploaded_by=request.user
                    )
                except Exception:
                    pass

                log_officer_action(
                    request,
                    officer=officer,
                    action_type='PROFILE_CREATE',
                    details=f"បាននាំចូលជីវប្រវត្តិមន្ត្រីថ្មីពីឯកសារ Word «{docx_file.name}» សម្រាប់មន្ត្រី «{officer.full_name_kh}»"
                )

                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'status': 'SUCCESS_NEW',
                        'is_new': True,
                        'officer_id': officer.pk,
                        'officer_name_kh': officer.full_name_kh,
                        'officer_name_latin': officer.full_name_latin,
                        'officer_id_number': officer.officer_id_number or '-',
                        'department_name': officer.department.name_kh if officer.department else '-',
                        'position': officer.current_position_title or '-',
                        'message': f"បានបញ្ចូលជីវប្រវត្តិថ្មី «{officer.full_name_kh}» ({officer.officer_id_number or '-'}) ដោយជោគជ័យ!"
                    })

            else:
                # Existing officer found
                diffs = compare_officer_data(officer, data)

                if duplicate_mode == 'skip':
                    if is_ajax:
                        return JsonResponse({
                            'success': True,
                            'status': 'EXISTING_SKIPPED',
                            'is_new': False,
                            'officer_id': officer.pk,
                            'officer_name_kh': officer.full_name_kh,
                            'officer_name_latin': officer.full_name_latin,
                            'officer_id_number': officer.officer_id_number or '-',
                            'department_name': officer.department.name_kh if officer.department else '-',
                            'position': officer.current_position_title or '-',
                            'message': f"បានរំលង៖ មន្ត្រី «{officer.full_name_kh}» មានក្នុងប្រព័ន្ធរួចហើយ"
                        })

                elif duplicate_mode == 'review':
                    if len(diffs) == 0:
                        if is_ajax:
                            return JsonResponse({
                                'success': True,
                                'status': 'EXISTING_IDENTICAL',
                                'is_new': False,
                                'officer_id': officer.pk,
                                'officer_name_kh': officer.full_name_kh,
                                'officer_name_latin': officer.full_name_latin,
                                'officer_id_number': officer.officer_id_number or '-',
                                'department_name': officer.department.name_kh if officer.department else '-',
                                'position': officer.current_position_title or '-',
                                'message': f"ទិន្នន័យដូចគ្នាទាំងស្រុង (មានក្នុងប្រព័ន្ធរួចហើយ មិនបាច់កែប្រែ)"
                            })
                    else:
                        # Differences detected, report diffs for user decision
                        if is_ajax:
                            return JsonResponse({
                                'success': True,
                                'status': 'DIFF_DETECTED',
                                'is_new': False,
                                'has_diff': True,
                                'differences': diffs,
                                'officer_id': officer.pk,
                                'officer_name_kh': officer.full_name_kh,
                                'officer_name_latin': officer.full_name_latin,
                                'officer_id_number': officer.officer_id_number or '-',
                                'department_name': officer.department.name_kh if officer.department else '-',
                                'position': officer.current_position_title or '-',
                                'message': f"រកឃើញទិន្នន័យខុសគ្នា {len(diffs)} ចំណុច"
                            })

                # auto_update or force_update (or review with 0 diffs in traditional submit)
                if len(diffs) > 0 or duplicate_mode == 'force_update':
                    for k, v in data.items():
                        setattr(officer, k, v)
                    if target_dept:
                        officer.department = target_dept
                    officer.save()

                    try:
                        docx_file.seek(0)
                        OfficerAttachment.objects.create(
                            officer=officer,
                            category='OTHER',
                            title=f"ជីវប្រវត្តិដើមពី Word ({docx_file.name})",
                            file=docx_file,
                            file_size=docx_file.size,
                            description="បាន Update ស្វ័យប្រវត្តិតាមរយៈប្រព័ន្ធ Import Word CV",
                            uploaded_by=request.user
                        )
                    except Exception:
                        pass

                    diff_summary = ", ".join([f"{d['label']}: {d['old_val']} ➔ {d['new_val']}" for d in diffs[:5]])
                    log_officer_action(
                        request,
                        officer=officer,
                        action_type='PROFILE_EDIT',
                        details=f"បាន Update ទិន្នន័យពីឯកសារ Word «{docx_file.name}» សម្រាប់មន្ត្រី «{officer.full_name_kh}» ({len(diffs)} ចំណុច៖ {diff_summary})"
                    )

                    if is_ajax:
                        return JsonResponse({
                            'success': True,
                            'status': 'SUCCESS_UPDATED',
                            'is_new': False,
                            'differences': diffs,
                            'officer_id': officer.pk,
                            'officer_name_kh': officer.full_name_kh,
                            'officer_name_latin': officer.full_name_latin,
                            'officer_id_number': officer.officer_id_number or '-',
                            'department_name': officer.department.name_kh if officer.department else '-',
                            'position': officer.current_position_title or '-',
                            'message': f"បានធ្វើបច្ចុប្បន្នភាពទិន្នន័យថ្មី ({len(diffs)} ចំណុច) សម្រាប់ «{officer.full_name_kh}»!"
                        })
                else:
                    if is_ajax:
                        return JsonResponse({
                            'success': True,
                            'status': 'EXISTING_IDENTICAL',
                            'is_new': False,
                            'officer_id': officer.pk,
                            'officer_name_kh': officer.full_name_kh,
                            'officer_name_latin': officer.full_name_latin,
                            'officer_id_number': officer.officer_id_number or '-',
                            'department_name': officer.department.name_kh if officer.department else '-',
                            'position': officer.current_position_title or '-',
                            'message': f"ទិន្នន័យដូចគ្នាទាំងស្រុង (មានក្នុងប្រព័ន្ធរួចហើយ)"
                        })

            action_label = "បញ្ចូលថ្មី" if is_new else "ធ្វើបច្ចុប្បន្នភាព"
            messages.success(request, f"✅ បាន{action_label}ជីវប្រវត្តិមន្ត្រី «{officer.full_name_kh}» (អត្តលេខ: {officer.officer_id_number or '-'}) ដោយជោគជ័យពី Word!")
            return redirect('officer_detail', pk=officer.pk)
        except Exception as e:
            if is_ajax:
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(request, f"⚠️ មានបញ្ហាក្នុងដំណើរការឯកសារ Word៖ {str(e)}")
            return redirect('officer_list')

    if is_ajax:
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    return redirect('officer_list')


@login_required
def api_officers_for_photo_match(request):
    """
    Returns lightweight JSON officer profiles for instant client-side matching
    when dragging and dropping batch officer photos.
    """
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    if is_admin_or_lead:
        officers_qs = CivilServantProfile.objects.all().select_related('department')
    elif dept:
        officers_qs = CivilServantProfile.objects.filter(department=dept).select_related('department')
    else:
        officers_qs = CivilServantProfile.objects.none()

    from .models import officer_sort_key
    officers_list = list(officers_qs)
    officers_list.sort(key=officer_sort_key)

    data = []
    for o in officers_list:
        data.append({
            'id': o.pk,
            'name_kh': o.full_name_kh,
            'khmer_last_name': o.khmer_last_name,
            'khmer_first_name': o.khmer_first_name,
            'name_latin': o.full_name_latin,
            'latin_last_name': o.latin_last_name,
            'latin_first_name': o.latin_first_name,
            'officer_id_number': o.officer_id_number or '',
            'national_id_number': o.national_id_number or '',
            'phone': o.phone or '',
            'position': o.current_position_title or '-',
            'department_name': o.display_department_name,
            'current_photo_url': o.photo.url if o.photo else '',
            'has_photo': bool(o.photo),
        })

    return JsonResponse({'success': True, 'officers': data})


@login_required
def officer_batch_upload_photo_single(request):
    """
    AJAX handler to upload and update a photo for an individual officer in batch process.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST method required'}, status=405)

    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    officer_id = request.POST.get('officer_id')
    photo_file = request.FILES.get('photo')
    overwrite = request.POST.get('overwrite', 'true').lower() in ['true', '1', 'yes']

    if not officer_id:
        return JsonResponse({'success': False, 'error': 'ពុំបានបញ្ជាក់ ID មន្ត្រីឡើយ!'}, status=400)
    if not photo_file:
        return JsonResponse({'success': False, 'error': 'ពុំមាន File រូបថតឡើយ!'}, status=400)

    try:
        officer = CivilServantProfile.objects.get(pk=officer_id)
    except CivilServantProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': f'រកមិនឃើញទិន្នន័យមន្ត្រី ID #{officer_id} ឡើយ!'}, status=404)

    # Permission check: Admin, leadership or own department officer
    if not is_admin_or_lead and (not dept or officer.department != dept) and officer.user != request.user:
        return JsonResponse({'success': False, 'error': 'លោកអ្នកគ្មានសិទ្ធិកែប្រែជីវប្រវត្តិមន្ត្រីនៅក្រៅការិយាល័យរបស់ខ្លួនឡើយ!'}, status=403)

    # Edit window check for non-admin
    edit_window = OfficerEditWindowSetting.get_setting()
    if not is_system_admin and not edit_window.is_open_for_department(officer.department or dept):
        return JsonResponse({'success': False, 'error': f'ប្រព័ន្ធត្រូវបានបិទមិនឱ្យកែប្រែជីវប្រវត្តិមន្ត្រីឡើយ! ({edit_window.status_label_kh})'}, status=403)

    # Check if officer already has photo and overwrite is disabled
    if officer.photo and not overwrite:
        return JsonResponse({
            'success': True,
            'skipped': True,
            'officer_id': officer.pk,
            'officer_name_kh': officer.full_name_kh,
            'photo_url': officer.photo.url,
            'message': f'មន្ត្រី «{officer.full_name_kh}» មានរូបថតរួចហើយ (ត្រូវបានរំលង)'
        })

    # Validate image file
    valid_extensions = ['.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp']
    import os
    ext = os.path.splitext(photo_file.name)[1].lower()
    if ext not in valid_extensions:
        return JsonResponse({'success': False, 'error': f'ប្រភេទ File «{photo_file.name}» មិនត្រឹមត្រូវ (ទទួលតែ JPG, PNG, WEBP)'}, status=400)

    if photo_file.size > 10 * 1024 * 1024:  # 10MB
        return JsonResponse({'success': False, 'error': f'File រូបថត «{photo_file.name}» មានទំហំធំពេក (លើសពី 10MB)'}, status=400)

    # Save photo to officer
    officer.photo = photo_file
    officer.save(update_fields=['photo', 'updated_at'])

    # Audit log
    log_officer_action(
        request,
        officer=officer,
        action_type='PROFILE_EDIT',
        category='រូបថត 4x6',
        details=f"បាន Update រូបថតមន្ត្រីពីមុខងារ Batch Photo Upload តាមឯកសារ «{photo_file.name}»"
    )

    return JsonResponse({
        'success': True,
        'officer_id': officer.pk,
        'officer_name_kh': officer.full_name_kh,
        'officer_id_number': officer.officer_id_number or '-',
        'photo_url': officer.photo.url if officer.photo else '',
        'message': f"បាន Update រូបថតសម្រាប់ «{officer.full_name_kh}» ដោយជោគជ័យ!"
    })


@login_required
def officer_edit(request, pk):
    officer = get_object_or_404(CivilServantProfile, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    # Permission check: Admin, leadership or own department officer
    if not is_admin_or_lead and (not dept or officer.department != dept) and officer.user != request.user:
        messages.error(request, "⚠️ លោកអ្នកគ្មានសិទ្ធិកែប្រែជីវប្រវត្តិមន្ត្រីនៅក្រៅការិយាល័យរបស់ខ្លួនឡើយ!")
        return redirect('officer_list')

    # Edit window check for non-admin users
    edit_window = OfficerEditWindowSetting.get_setting()
    if not is_system_admin and not edit_window.is_open_for_department(officer.department or dept):
        messages.warning(request, f"🔒 ប្រព័ន្ធត្រូវបានបិទមិនឱ្យកែប្រែជីវប្រវត្តិមន្ត្រីឡើយ! ({edit_window.status_label_kh})")
        return redirect('officer_detail', pk=pk)

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh') if is_admin_or_lead else (Department.objects.filter(id=dept.id) if dept else [])
    users = User.objects.filter(is_active=True).order_by('username')

    if request.method == 'POST':
        try:
            old_dept = officer.department
            old_pos = officer.current_position_title
            pos_title = request.POST.get('current_position_title', '').strip()
            if pos_title in ['ប្រធានមន្ទីរ', 'អនុប្រធានមន្ទីរ']:
                lead_dept = Department.objects.filter(Q(code='LEAD') | Q(name_kh='ថ្នាក់ដឹកនាំមន្ទីរ')).first()
                if lead_dept:
                    officer.department = lead_dept
            elif is_admin_or_lead or is_system_admin:
                target_dept_id = request.POST.get('department')
                if target_dept_id:
                    officer.department = Department.objects.filter(id=target_dept_id).first()
                if request.POST.get('user'):
                    officer.user_id = request.POST.get('user')

            # ក- ព័ត៌មានផ្ទាល់ខ្លួន
            officer.khmer_last_name = request.POST.get('khmer_last_name', '').strip()
            officer.khmer_first_name = request.POST.get('khmer_first_name', '').strip()
            officer.latin_last_name = request.POST.get('latin_last_name', '').strip().upper()
            officer.latin_first_name = request.POST.get('latin_first_name', '').strip().upper()
            officer.gender = request.POST.get('gender', 'MALE')
            officer.dob = request.POST.get('dob', '').strip()
            officer.ethnicity = request.POST.get('ethnicity', 'ខ្មែរ').strip()
            officer.nationality = request.POST.get('nationality', 'ខ្មែរ').strip()

            # Resolve POB Address
            pob_pcode, pob_dcode, pob_ccode, pob_vcode, pob_pname, pob_dname, pob_cname, pob_vname = _resolve_geo_address(
                request.POST.get('pob_province_code'), request.POST.get('pob_district_code'),
                request.POST.get('pob_commune_code'), request.POST.get('pob_village_code'),
                request.POST.get('pob_province'), request.POST.get('pob_district'),
                request.POST.get('pob_commune'), request.POST.get('pob_village')
            )
            officer.pob_province_code = pob_pcode
            officer.pob_district_code = pob_dcode
            officer.pob_commune_code = pob_ccode
            officer.pob_village_code = pob_vcode
            officer.pob_village = pob_vname
            officer.pob_commune = pob_cname
            officer.pob_district = pob_dname
            officer.pob_province = pob_pname

            # Resolve Current Address
            cur_pcode, cur_dcode, cur_ccode, cur_vcode, cur_pname, cur_dname, cur_cname, cur_vname = _resolve_geo_address(
                request.POST.get('current_province_code'), request.POST.get('current_district_code'),
                request.POST.get('current_commune_code'), request.POST.get('current_village_code'),
                request.POST.get('current_province'), request.POST.get('current_district'),
                request.POST.get('current_commune'), request.POST.get('current_village')
            )
            officer.current_province_code = cur_pcode
            officer.current_district_code = cur_dcode
            officer.current_commune_code = cur_ccode
            officer.current_village_code = cur_vcode
            officer.current_house_no = request.POST.get('current_house_no', '').strip()
            officer.current_street = request.POST.get('current_street', '').strip()
            officer.current_village = cur_vname
            officer.current_commune = cur_cname
            officer.current_district = cur_dname
            officer.current_province = cur_pname

            # Resolve Permanent Address
            perm_same = (request.POST.get('perm_same_as_current') == 'on' or request.POST.get('perm_same_as_current') == 'true')
            officer.perm_same_as_current = perm_same
            if perm_same:
                officer.perm_province_code = cur_pcode
                officer.perm_district_code = cur_dcode
                officer.perm_commune_code = cur_ccode
                officer.perm_village_code = cur_vcode
                officer.perm_house_no = request.POST.get('current_house_no', '').strip()
                officer.perm_street = request.POST.get('current_street', '').strip()
                officer.perm_village = cur_vname
                officer.perm_commune = cur_cname
                officer.perm_district = cur_dname
                officer.perm_province = cur_pname
            else:
                perm_pcode, perm_dcode, perm_ccode, perm_vcode, perm_pname, perm_dname, perm_cname, perm_vname = _resolve_geo_address(
                    request.POST.get('perm_province_code'), request.POST.get('perm_district_code'),
                    request.POST.get('perm_commune_code'), request.POST.get('perm_village_code'),
                    request.POST.get('perm_province'), request.POST.get('perm_district'),
                    request.POST.get('perm_commune'), request.POST.get('perm_village')
                )
                officer.perm_province_code = perm_pcode
                officer.perm_district_code = perm_dcode
                officer.perm_commune_code = perm_ccode
                officer.perm_village_code = perm_vcode
                officer.perm_house_no = request.POST.get('perm_house_no', '').strip()
                officer.perm_street = request.POST.get('perm_street', '').strip()
                officer.perm_village = perm_vname
                officer.perm_commune = perm_cname
                officer.perm_district = perm_dname
                officer.perm_province = perm_pname

            officer.phone = request.POST.get('phone', '').strip()
            officer.email = request.POST.get('email', '').strip()
            officer.officer_id_number = request.POST.get('officer_id_number', '').strip()
            officer.national_id_number = request.POST.get('national_id_number', '').strip()
            officer.national_id_valid_from = request.POST.get('national_id_valid_from', '').strip()
            officer.national_id_valid_to = request.POST.get('national_id_valid_to', '').strip()
            officer.physical_condition = request.POST.get('physical_condition', 'គ្រប់គ្រាន់').strip()
            officer.disability_detail = request.POST.get('disability_detail', '').strip()

            # ខ- ព័ត៌មានគ្រួសារ
            officer.marital_status = request.POST.get('marital_status', 'SINGLE')
            officer.spouse_marriage_cert_no = request.POST.get('spouse_marriage_cert_no', '').strip()
            officer.spouse_name_kh = request.POST.get('spouse_name_kh', '').strip()
            officer.spouse_is_alive = (request.POST.get('spouse_is_alive', 'true') == 'true')
            officer.spouse_name_latin = request.POST.get('spouse_name_latin', '').strip().upper()
            officer.spouse_dob = request.POST.get('spouse_dob', '').strip()
            officer.spouse_national_id = request.POST.get('spouse_national_id', '').strip()
            officer.spouse_pob = request.POST.get('spouse_pob', '').strip()
            officer.spouse_occupation = request.POST.get('spouse_occupation', '').strip()
            officer.spouse_current_address = request.POST.get('spouse_current_address', '').strip()
            officer.spouse_organization = request.POST.get('spouse_organization', '').strip()

            officer.father_name = request.POST.get('father_name', '').strip()
            officer.father_is_alive = (request.POST.get('father_is_alive', 'true') == 'true')
            officer.father_pob = request.POST.get('father_pob', '').strip()
            officer.father_occupation = request.POST.get('father_occupation', '').strip()

            officer.mother_name = request.POST.get('mother_name', '').strip()
            officer.mother_is_alive = (request.POST.get('mother_is_alive', 'true') == 'true')
            officer.mother_pob = request.POST.get('mother_pob', '').strip()
            officer.mother_occupation = request.POST.get('mother_occupation', '').strip()

            # គ- ទំនាក់ទំនងអាសន្ន
            officer.emergency_last_name = request.POST.get('emergency_last_name', '').strip()
            officer.emergency_first_name = request.POST.get('emergency_first_name', '').strip()
            officer.emergency_gender = request.POST.get('emergency_gender', 'FEMALE')
            officer.emergency_relationship = request.POST.get('emergency_relationship', '').strip()
            officer.emergency_occupation = request.POST.get('emergency_occupation', '').strip()
            officer.emergency_address = request.POST.get('emergency_address', '').strip()
            officer.emergency_phone = request.POST.get('emergency_phone', '').strip()
            officer.emergency_email = request.POST.get('emergency_email', '').strip()

            # ច- ក្របខ័ណ្ឌការងារ
            officer.civil_service_start_date = request.POST.get('civil_service_start_date', '').strip()
            officer.civil_service_permanent_date = request.POST.get('civil_service_permanent_date', '').strip()
            officer.framework_name = request.POST.get('framework_name', '').strip()
            officer.framework_category = request.POST.get('framework_category', '').strip() or None
            officer.current_rank_and_step = request.POST.get('current_rank_and_step', '').strip()
            officer.current_position_title = request.POST.get('current_position_title', '').strip()
            officer.position_code = request.POST.get('position_code', '').strip() or CivilServantProfile.POSITION_CODE_MAP.get(officer.current_position_title, '')
            officer.highest_degree = request.POST.get('highest_degree', '').strip() or None
            officer.officer_status = request.POST.get('officer_status', 'ACTIVE').strip() or 'ACTIVE'

            # JSON Tables
            officer.children_data = _parse_tabular_json(request.POST, 'child', ['name', 'gender', 'dob', 'occupation'])
            officer.education_data = _parse_tabular_json(request.POST, 'edu', ['category', 'level_label', 'school', 'location', 'degree', 'skill', 'start_date', 'end_date'])
            officer.languages_data = _parse_tabular_json(request.POST, 'lang', ['language', 'reading', 'speaking', 'writing'])
            officer.history_public_sector = _parse_tabular_json(request.POST, 'pub', ['start_date', 'end_date', 'ministry', 'department', 'position', 'skill'])
            officer.history_private_sector = _parse_tabular_json(request.POST, 'priv', ['start_date', 'end_date', 'org', 'role', 'skill'])
            officer.promotions_by_seniority = _parse_tabular_json(request.POST, 'promo_sen', ['effective_date', 'ministry', 'department', 'office', 'old_rank_step', 'new_rank_step', 'promo_type'])
            officer.promotions_by_degree = _parse_tabular_json(request.POST, 'promo_deg', ['effective_date', 'school', 'location', 'degree', 'old_rank_step', 'new_rank_step'])
            officer.outside_framework_status = _parse_tabular_json(request.POST, 'out_fw', ['start_date', 'end_date', 'ministry', 'position'])
            officer.unpaid_leave_status = _parse_tabular_json(request.POST, 'unpaid', ['start_date', 'end_date', 'ministry', 'duration'])
            officer.awards_data = _parse_tabular_json(request.POST, 'award', ['doc_number', 'date', 'ministry', 'description', 'type'])
            officer.sanctions_data = _parse_tabular_json(request.POST, 'sanct', ['doc_number', 'date', 'ministry', 'description', 'type'])

            if 'photo' in request.FILES:
                officer.photo = request.FILES['photo']

            officer.save()

            # Automatic transfer history logging if department changed
            if old_dept != officer.department and officer.department is not None:
                OfficerDepartmentTransferHistory.objects.create(
                    officer=officer,
                    from_department=old_dept,
                    to_department=officer.department,
                    from_position_title=old_pos,
                    to_position_title=officer.current_position_title,
                    transfer_date=timezone.now().date(),
                    remarks='បានផ្លាស់ប្តូរការិយាល័យតាមរយៈការកែប្រែជីវប្រវត្តិ (Profile Edit)',
                    transferred_by=request.user,
                )

            # Process newly attached reference files from sections
            _handle_form_attachments(request, officer)

            # Record Audit Log
            log_officer_action(
                request,
                officer=officer,
                action_type='PROFILE_EDIT',
                details=f"បានកែប្រែព័ត៌មានជីវប្រវត្តិ «{officer.full_name_kh}»"
            )

            active_tab = request.POST.get('active_tab', 'tab-a')
            save_action = request.POST.get('save_action', 'save_and_continue')

            if save_action == 'save_and_continue':
                messages.success(request, f"✅ បានរក្សាទុកទិន្នន័យចុងក្រោយ «{officer.full_name_kh}» ដោយជោគជ័យ! លោកអ្នកអាចបន្តការងារបញ្ចូល ឬកែសម្រួលបានភ្លាមៗ។")
                return redirect(f"/officers/{officer.pk}/edit/?tab={active_tab}")
            else:
                messages.success(request, f"✅ បានកែប្រែ និងរក្សាទុកជីវប្រវត្តិ «{officer.full_name_kh}» ដោយជោគជ័យ!")
                return redirect('officer_detail', pk=officer.pk)
        except Exception as e:
            messages.error(request, f"⚠️ មានបញ្ហាក្នុងការកែប្រែទិន្នន័យ៖ {str(e)}")

    attachments = officer.attachments.all().order_by('category', '-created_at')
    attachments_by_category = {
        'PERSONAL': [a for a in attachments if a.category == 'PERSONAL'],
        'FAMILY': [a for a in attachments if a.category == 'FAMILY'],
        'EDUCATION': [a for a in attachments if a.category == 'EDUCATION'],
        'APPOINTMENT': [a for a in attachments if a.category == 'APPOINTMENT'],
        'PROMOTION': [a for a in attachments if a.category == 'PROMOTION'],
        'AWARD': [a for a in attachments if a.category == 'AWARD'],
        'OTHER': [a for a in attachments if a.category == 'OTHER'],
    }

    provinces = CambodiaProvince.objects.all().order_by('code')

    context = {
        'officer': officer,
        'departments': departments,
        'users': users,
        'provinces': provinces,
        'is_edit': True,
        'dept': dept,
        'is_admin_or_lead': is_admin_or_lead,
        'has_global_access': has_global_access,
        'is_system_admin': is_system_admin,
        'edit_window': edit_window,
        'attachments': attachments,
        'attachments_by_category': attachments_by_category,
        'attachment_categories': OfficerAttachment.CATEGORY_CHOICES,
    }
    return render(request, 'dms/officer_form.html', context)


@login_required
def officer_detail(request, pk):
    officer = get_object_or_404(CivilServantProfile.objects.select_related('department', 'user', 'created_by'), pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    # Permission check: Specialized offices cannot view officers from other departments
    if not is_admin_or_lead and (not dept or officer.department != dept) and officer.user != request.user:
        messages.error(request, "⚠️ លោកអ្នកគ្មានសិទ្ធិមើលជីវប្រវត្តិមន្ត្រីនៅក្រៅការិយាល័យរបស់ខ្លួនឡើយ!")
        return redirect('officer_list')

    edit_window = OfficerEditWindowSetting.get_setting()
    is_window_open = edit_window.is_open_for_department(officer.department or dept)
    can_dept_edit = is_admin_or_lead or (dept and officer.department == dept) or (officer.user == request.user)
    can_edit = is_system_admin or (can_dept_edit and is_window_open)

    children = officer.children_data or []
    female_children_count = sum(1 for c in children if c.get('gender') in ['ស្រី', 'FEMALE', 'Female'])
    total_children_count = len(children)

    # Attachments
    attachments = officer.attachments.select_related('uploaded_by').all().order_by('category', '-created_at')
    
    # Audit Logs (Admin & leadership can view)
    audit_logs = officer.audit_logs.select_related('user').all()[:40] if is_admin_or_lead else []

    # Department Transfer History & All Departments for Transfer Modal
    department_transfers = officer.department_transfers.select_related('from_department', 'to_department', 'transferred_by').all().order_by('-transfer_date', '-created_at')
    all_departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')

    context = {
        'officer': officer,
        'can_edit': can_edit,
        'is_admin_or_lead': is_admin_or_lead,
        'has_global_access': has_global_access,
        'is_system_admin': is_system_admin,
        'edit_window': edit_window,
        'is_window_open': is_window_open,
        'children_list': children,
        'total_children_count': total_children_count,
        'female_children_count': female_children_count,
        'education_list': officer.education_data or [],
        'languages_list': officer.languages_data or [],
        'pub_history_list': officer.history_public_sector or [],
        'priv_history_list': officer.history_private_sector or [],
        'promo_seniority_list': officer.promotions_by_seniority or [],
        'promo_degree_list': officer.promotions_by_degree or [],
        'outside_fw_list': officer.outside_framework_status or [],
        'unpaid_leave_list': officer.unpaid_leave_status or [],
        'awards_list': officer.awards_data or [],
        'sanctions_list': officer.sanctions_data or [],
        'attachments': attachments,
        'attachment_categories': OfficerAttachment.CATEGORY_CHOICES,
        'audit_logs': audit_logs,
        'department_transfers': department_transfers,
        'all_departments': all_departments,
    }
    return render(request, 'dms/officer_detail.html', context)


@login_required
def officer_transfer(request, pk):
    """
    Direct / Modal Department/Canton Transfer for Civil Servants (Admin / Leadership).
    Records a full audit and history entry of the transfer.
    """
    officer = get_object_or_404(CivilServantProfile, pk=pk)
    profile = getattr(request.user, 'profile', None)
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access or request.user.is_superuser
    if not is_admin_or_lead:
        messages.error(request, "⚠️ មានតែថ្នាក់ដឹកនាំ និងរដ្ឋបាល-បុគ្គលិក (ADMIN) ប៉ុណ្ណោះដែលអាចផ្លាស់ប្តូរការិយាល័យ/ខណ្ឌបាន!")
        return redirect('officer_detail', pk=pk)

    if request.method == 'POST':
        target_dept_id = request.POST.get('target_department')
        transfer_date_str = request.POST.get('transfer_date', '').strip()
        ref_letter = request.POST.get('reference_letter_number', '').strip()
        new_position = request.POST.get('current_position_title', '').strip()
        remarks = request.POST.get('remarks', '').strip()
        transfer_doc = request.FILES.get('transfer_document')

        target_dept = Department.objects.filter(id=target_dept_id, is_active=True).first() if target_dept_id else None
        if not target_dept:
            messages.error(request, "⚠️ សូមជ្រើសរើសការិយាល័យ/ខណ្ឌថ្មី!")
            return redirect('officer_detail', pk=pk)

        old_dept = officer.department
        old_pos = officer.current_position_title

        transfer_date = None
        if transfer_date_str:
            try:
                transfer_date = datetime.strptime(transfer_date_str, '%Y-%m-%d').date()
            except Exception:
                transfer_date = timezone.now().date()
        else:
            transfer_date = timezone.now().date()

        # Update officer department & position
        officer.department = target_dept
        if new_position:
            officer.current_position_title = new_position
        officer.save()

        # Append to public sector history table if not duplicate
        try:
            hist = list(officer.history_public_sector or [])
            hist.append({
                'start_date': transfer_date.strftime('%d/%m/%Y'),
                'end_date': 'បច្ចុប្បន្ន',
                'ministry': 'មន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទ ខេត្តប៉ៃលិន',
                'department': target_dept.name_kh,
                'position': officer.current_position_title,
                'skill': ref_letter or 'លិខិតផ្លាស់ប្តូរការិយាល័យ'
            })
            officer.history_public_sector = hist
            officer.save(update_fields=['history_public_sector'])
        except Exception:
            pass

        # Create history record
        OfficerDepartmentTransferHistory.objects.create(
            officer=officer,
            from_department=old_dept,
            to_department=target_dept,
            from_position_title=old_pos,
            to_position_title=officer.current_position_title,
            transfer_date=transfer_date,
            reference_letter_number=ref_letter,
            transfer_document=transfer_doc,
            remarks=remarks,
            transferred_by=request.user,
        )

        log_officer_action(
            request,
            officer=officer,
            action_type='EDIT',
            category='ច- ក្របខ័ណ្ឌ និងការិយាល័យ',
            details=f"បានផ្លាស់ប្តូរការិយាល័យពី «{old_dept.name_kh if old_dept else 'គ្មាន'}» ➔ «{target_dept.name_kh}» (លិខិតផ្ទេរ: {ref_letter or '-'})"
        )

        messages.success(request, f"✅ បានផ្លាស់ប្តូរការិយាល័យសម្រាប់មន្ត្រី «{officer.full_name_kh}» ទៅកាន់ «{target_dept.name_kh}» ដោយជោគជ័យ!")

    return redirect('officer_detail', pk=pk)


@login_required
def officer_transfer_delete(request, pk, transfer_id):
    """
    Delete a department transfer history record (Superadmin or Admin only).
    """
    officer = get_object_or_404(CivilServantProfile, pk=pk)
    profile = getattr(request.user, 'profile', None)
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access or request.user.is_superuser
    if not is_admin_or_lead:
        messages.error(request, "⚠️ មានតែ ADMIN ប៉ុណ្ណោះដែលអាចលុបកំណត់ត្រាប្រវត្តិផ្ទេរការិយាល័យបាន!")
        return redirect('officer_detail', pk=pk)

    if request.method == 'POST':
        transfer_record = get_object_or_404(OfficerDepartmentTransferHistory, pk=transfer_id, officer=officer)
        from_name = transfer_record.from_department.name_kh if transfer_record.from_department else "គ្មាន"
        to_name = transfer_record.to_department.name_kh if transfer_record.to_department else "គ្មាន"
        transfer_record.delete()
        messages.success(request, f"🗑️ បានលុបកំណត់ត្រាប្រវត្តិផ្លាស់ប្តូរការិយាល័យ «{from_name} ➔ {to_name}» រួចរាល់។")

    return redirect('officer_detail', pk=pk)


KHMER_DIGITS_MAP = {'0': '០', '1': '១', '2': '២', '3': '៣', '4': '៤', '5': '៥', '6': '៦', '7': '៧', '8': '៨', '9': '៩'}
KHMER_MONTHS_MAP = {
    1: 'មករា', 2: 'កុម្ភៈ', 3: 'មីនា', 4: 'មេសា', 5: 'ឧសភា', 6: 'មិថុនា',
    7: 'កក្កដា', 8: 'សីហា', 9: 'កញ្ញា', 10: 'តុលា', 11: 'វិច្ឆិកា', 12: 'ធ្នូ'
}

def to_khmer_digits(s):
    if s is None:
        return ''
    return ''.join(KHMER_DIGITS_MAP.get(ch, ch) for ch in str(s))

def format_khmer_date_str(val):
    if not val:
        return '-'
    val_str = str(val).strip()
    import re
    m = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$', val_str)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        month_name = KHMER_MONTHS_MAP.get(month, str(month))
        return f"ថ្ងៃទី{to_khmer_digits(day)} ខែ{month_name} ឆ្នាំ{to_khmer_digits(year)}"
    m2 = re.match(r'^(\d{4})[/-](\d{1,2})[/-](\d{1,2})$', val_str)
    if m2:
        year, month, day = int(m2.group(1)), int(m2.group(2)), int(m2.group(3))
        month_name = KHMER_MONTHS_MAP.get(month, str(month))
        return f"ថ្ងៃទី{to_khmer_digits(day)} ខែ{month_name} ឆ្នាំ{to_khmer_digits(year)}"
    return to_khmer_digits(val_str)


@login_required
def officer_print(request, pk):
    officer = get_object_or_404(CivilServantProfile.objects.select_related('department', 'user'), pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    if not is_admin_or_lead and (not dept or officer.department != dept) and officer.user != request.user:
        messages.error(request, "⚠️ លោកអ្នកគ្មានសិទ្ធិមើល ឬបោះពុម្ពជីវប្រវត្តិមន្ត្រីនៅក្រៅការិយាល័យរបស់ខ្លួនឡើយ!")
        return redirect('officer_list')

    children = officer.children_data or []
    formatted_children = []
    female_children_count = 0
    for idx, c in enumerate(children, 1):
        g = c.get('gender', '')
        is_female = g in ['ស្រី', 'FEMALE', 'Female']
        if is_female:
            female_children_count += 1
        formatted_children.append({
            'counter_kh': to_khmer_digits(idx),
            'name': c.get('name', ''),
            'gender_kh': 'ស្រី' if is_female else 'ប្រុស',
            'dob_kh': format_khmer_date_str(c.get('dob', '')),
            'occupation': c.get('occupation', 'ក្នុងបន្ទុក'),
        })

    edu_list = officer.education_data or []
    edu_general = [e for e in edu_list if e.get('category') == '1' or e.get('level_type') == 'GENERAL']
    edu_vocational = [e for e in edu_list if e.get('category') == '2' or e.get('level_type') == 'VOCATIONAL']
    edu_continuous = [e for e in edu_list if e.get('category') == '3' or e.get('level_type') == 'CONTINUOUS']

    now = timezone.now()
    context = {
        'officer': officer,
        'officer_dob_kh': format_khmer_date_str(officer.dob),
        'officer_nid_valid_from': to_khmer_digits(officer.national_id_valid_from) if officer.national_id_valid_from else '...',
        'officer_nid_valid_to': to_khmer_digits(officer.national_id_valid_to) if officer.national_id_valid_to else '...',
        'spouse_dob_kh': format_khmer_date_str(officer.spouse_dob),
        'children_list': formatted_children,
        'total_children_kh': to_khmer_digits(len(children)),
        'female_children_kh': to_khmer_digits(female_children_count),
        'edu_general': edu_general,
        'edu_vocational': edu_vocational,
        'edu_continuous': edu_continuous,
        'languages_list': officer.languages_data or [],
        'pub_history_list': officer.history_public_sector or [],
        'priv_history_list': officer.history_private_sector or [],
        'promo_seniority_list': officer.promotions_by_seniority or [],
        'promo_degree_list': officer.promotions_by_degree or [],
        'outside_fw_list': officer.outside_framework_status or [],
        'unpaid_leave_list': officer.unpaid_leave_status or [],
        'awards_list': officer.awards_data or [],
        'sanctions_list': officer.sanctions_data or [],
        'print_date': now,
        'print_day_kh': to_khmer_digits(now.day),
        'print_month_kh': KHMER_MONTHS_MAP.get(now.month, ''),
        'print_year_kh': to_khmer_digits(now.year),
    }
    return render(request, 'dms/officer_print.html', context)


@login_required
def officer_delete(request, pk):
    """
    Deletes a single civil servant profile (ADMIN only).
    Specialized offices are strictly forbidden from deleting officers.
    """
    officer = get_object_or_404(CivilServantProfile, pk=pk)
    is_admin = check_is_system_admin(request.user, request)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('format') == 'json'
    
    if not is_admin:
        if is_ajax:
            return JsonResponse({'success': False, 'error': '⚠️ ការិយាល័យជំនាញមិនត្រូវបានអនុញ្ញាតឱ្យលុបជីវប្រវត្តិមន្ត្រីឡើយ! មានតែ ADMIN ប៉ុណ្ណោះដែលអាចលុបបាន។'}, status=403)
        messages.error(request, "⚠️ ការិយាល័យជំនាញមិនត្រូវបានអនុញ្ញាតឱ្យលុបជីវប្រវត្តិមន្ត្រីឡើយ! មានតែ ADMIN ប៉ុណ្ណោះដែលអាចលុបបាន។")
        return redirect('officer_list')

    if request.method == 'POST':
        name = officer.full_name_kh
        officer_id = officer.officer_id_number or '-'
        
        # Log Audit before deleting
        log_officer_action(
            request,
            officer=officer,
            action_type='PROFILE_DELETE',
            details=f"Admin «{request.user.username}» បានលុបជីវប្រវត្តិមន្ត្រី «{name}» (អត្តលេខ: {officer_id}) ចេញពីប្រព័ន្ធ"
        )
        
        officer.delete()
        if is_ajax:
            return JsonResponse({'success': True, 'message': f"បានលុបជីវប្រវត្តិមន្ត្រី «{name}» (អត្តលេខ: {officer_id}) ចេញពីប្រព័ន្ធដោយជោគជ័យ!"})
        messages.success(request, f"🗑️ បានលុបជីវប្រវត្តិមន្ត្រី «{name}» (អត្តលេខ: {officer_id}) ចេញពីប្រព័ន្ធដោយជោគជ័យ!")
        return redirect('officer_list')
        
    return redirect('officer_detail', pk=pk)


@login_required
def officer_bulk_delete(request):
    """
    Deletes multiple selected officer profiles at once (ADMIN only).
    """
    is_admin = check_is_system_admin(request.user, request)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('format') == 'json'

    if not is_admin:
        if is_ajax:
            return JsonResponse({'success': False, 'error': '⚠️ ការិយាល័យជំនាញមិនត្រូវបានអនុញ្ញាតឱ្យលុបជីវប្រវត្តិមន្ត្រីឡើយ! មានតែ ADMIN ប៉ុណ្ណោះដែលអាចលុបបាន។'}, status=403)
        messages.error(request, "⚠️ ការិយាល័យជំនាញមិនត្រូវបានអនុញ្ញាតឱ្យលុបជីវប្រវត្តិមន្ត្រីឡើយ! មានតែ ADMIN ប៉ុណ្ណោះដែលអាចលុបបាន។")
        return redirect('officer_list')

    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_ids')
        if not selected_ids:
            import json
            try:
                data = json.loads(request.body.decode('utf-8'))
                selected_ids = data.get('selected_ids', [])
            except Exception:
                pass

        if not selected_ids:
            if is_ajax:
                return JsonResponse({'success': False, 'error': 'សូមជ្រើសរើសមន្ត្រីយ៉ាងហោចណាស់ម្នាក់ដើម្បីលុប!'}, status=400)
            messages.warning(request, "⚠️ សូមជ្រើសរើសមន្ត្រីយ៉ាងហោចណាស់ម្នាក់ដើម្បីលុប!")
            return redirect('officer_list')

        officers = CivilServantProfile.objects.filter(id__in=selected_ids)
        deleted_count = officers.count()
        officer_names = [f"{o.full_name_kh} ({o.officer_id_number or '-'})" for o in officers[:5]]
        preview_names = ", ".join(officer_names)
        if deleted_count > 5:
            preview_names += f" និង {deleted_count - 5} នាក់ទៀត"

        # Log audit
        log_officer_action(
            request,
            officer=None,
            action_type='PROFILE_BULK_DELETE',
            details=f"Admin «{request.user.username}» បានលុបមន្ត្រីចំនួន {deleted_count} នាក់៖ [{preview_names}]"
        )

        officers.delete()

        msg = f"🗑️ បានលុបជីវប្រវត្តិមន្ត្រីចំនួន {deleted_count} នាក់ចេញពីប្រព័ន្ធដោយជោគជ័យ!"
        if is_ajax:
            return JsonResponse({'success': True, 'message': msg, 'deleted_count': deleted_count})
        messages.success(request, msg)
        return redirect('officer_list')

    return redirect('officer_list')


@login_required
def officer_delete_all(request):
    """
    Deletes ALL officer profiles in the system or filtered department (ADMIN only).
    """
    is_admin = check_is_system_admin(request.user, request)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('format') == 'json'

    if not is_admin:
        if is_ajax:
            return JsonResponse({'success': False, 'error': '⚠️ ការិយាល័យជំនាញមិនត្រូវបានអនុញ្ញាតឱ្យលុបជីវប្រវត្តិមន្ត្រីឡើយ! មានតែ ADMIN ប៉ុណ្ណោះដែលអាចលុបបាន។'}, status=403)
        messages.error(request, "⚠️ ការិយាល័យជំនាញមិនត្រូវបានអនុញ្ញាតឱ្យលុបជីវប្រវត្តិមន្ត្រីឡើយ! មានតែ ADMIN ប៉ុណ្ណោះដែលអាចលុបបាន។")
        return redirect('officer_list')

    if request.method == 'POST':
        confirm_text = request.POST.get('confirm_text', '').strip()
        dept_id = request.POST.get('department_id', '').strip()

        if confirm_text not in ['DELETE', 'DELETE ALL', 'DELETEALL', 'លុបទាំងអស់']:
            msg = "⚠️ សូមវាយពាក្យបញ្ជាក់ «DELETE ALL» ឬ «លុបទាំងអស់» ដើម្បីផ្ទៀងផ្ទាត់ការលុបទាំងអស់!"
            if is_ajax:
                return JsonResponse({'success': False, 'error': msg}, status=400)
            messages.error(request, msg)
            return redirect('officer_list')

        queryset = CivilServantProfile.objects.all()
        target_desc = "ទាំងអស់ក្នុងប្រព័ន្ធ"
        if dept_id:
            queryset = queryset.filter(department_id=dept_id)
            dept_obj = Department.objects.filter(id=dept_id).first()
            if dept_obj:
                target_desc = f"ក្នុង «{dept_obj.name_kh}»"

        total_to_delete = queryset.count()
        if total_to_delete == 0:
            msg = "ℹ️ ពុំមានទិន្នន័យមន្ត្រីសម្រាប់លុបឡើយ!"
            if is_ajax:
                return JsonResponse({'success': False, 'error': msg}, status=400)
            messages.info(request, msg)
            return redirect('officer_list')

        # Log audit
        log_officer_action(
            request,
            officer=None,
            action_type='PROFILE_BULK_DELETE',
            details=f"Admin «{request.user.username}» បានលុបទិន្នន័យជីវប្រវត្តិមន្ត្រីទាំងអស់ ({target_desc}) សរុប {total_to_delete} នាក់"
        )

        queryset.delete()

        msg = f"🗑️ បានលុបទិន្នន័យជីវប្រវត្តិមន្ត្រី ({target_desc}) សរុប {total_to_delete} នាក់ចេញពីប្រព័ន្ធដោយជោគជ័យ!"
        if is_ajax:
            return JsonResponse({'success': True, 'message': msg, 'deleted_count': total_to_delete})
        messages.success(request, msg)
        return redirect('officer_list')

    return redirect('officer_list')


# =========================================================================
# 📎 OFFICER ATTACHMENTS & AUDIT TRAIL (ឯកសារយោង & កំណត់ត្រាសកម្មភាព)
# =========================================================================

@login_required
def officer_attachment_add(request, pk):
    officer = get_object_or_404(CivilServantProfile, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    if not is_admin_or_lead and officer.department != dept and officer.user != request.user:
        messages.error(request, "⚠️ លោកអ្នកគ្មានសិទ្ធិបញ្ចូលឯកសារយោងសម្រាប់មន្ត្រីនេះឡើយ!")
        return redirect('officer_detail', pk=pk)

    # Edit window check for non-admin users
    edit_window = OfficerEditWindowSetting.get_setting()
    if not is_system_admin and not edit_window.is_open_for_department(officer.department or dept):
        messages.warning(request, f"🔒 ប្រព័ន្ធត្រូវបានបិទមិនឱ្យបញ្ចូលឯកសារយោងឡើយ! ({edit_window.status_label_kh})")
        return redirect('officer_detail', pk=pk)

    if request.method == 'POST':
        category = request.POST.get('category', 'OTHER')
        title = request.POST.get('title', '').strip()
        doc_number = request.POST.get('doc_number', '').strip()
        issued_date = request.POST.get('issued_date', '').strip()
        description = request.POST.get('description', '').strip()
        uploaded_file = request.FILES.get('file')

        if not uploaded_file:
            messages.error(request, "⚠️ សូមជ្រើសរើស File ឯកសារយោងដើម្បីបញ្ចូល!")
            return redirect('officer_detail', pk=pk)

        # 2MB maximum limit check
        if uploaded_file.size > 2 * 1024 * 1024:
            file_size_mb = uploaded_file.size / (1024 * 1024)
            messages.error(request, f"⚠️ ទំហំ File «{uploaded_file.name}» គឺ {file_size_mb:.2f}MB ដែលលើសពី 2MB! ប្រព័ន្ធអនុញ្ញាតទំហំអតិបរមាត្រឹម 2MB ប៉ុណ្ណោះ។")
            return redirect('officer_detail', pk=pk)

        if not title:
            title = uploaded_file.name

        att = OfficerAttachment.objects.create(
            officer=officer,
            category=category,
            title=title,
            doc_number=doc_number,
            issued_date=issued_date,
            file=uploaded_file,
            file_size=uploaded_file.size,
            description=description,
            uploaded_by=request.user
        )

        log_officer_action(
            request,
            officer=officer,
            action_type='UPLOAD',
            attachment_title=att.title,
            category=att.get_category_display(),
            details=f"បានបញ្ចូលឯកសារយោងថ្មី៖ «{att.title}» (លេខ៖ {att.doc_number or '-'}, ទំហំ៖ {att.formatted_file_size})"
        )

        messages.success(request, f"✅ បានបញ្ចូលឯកសារយោង «{att.title}» ដោយជោគជ័យ!")
        return redirect('officer_detail', pk=pk)

    return redirect('officer_detail', pk=pk)


@login_required
def officer_attachment_edit(request, pk, attachment_id):
    officer = get_object_or_404(CivilServantProfile, pk=pk)
    att = get_object_or_404(OfficerAttachment, pk=attachment_id, officer=officer)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    if not is_admin_or_lead and officer.department != dept and officer.user != request.user:
        messages.error(request, "⚠️ លោកអ្នកគ្មានសិទ្ធិកែប្រែឯកសារយោងនេះឡើយ!")
        return redirect('officer_detail', pk=pk)

    # Edit window check for non-admin users
    edit_window = OfficerEditWindowSetting.get_setting()
    if not is_system_admin and not edit_window.is_open_for_department(officer.department or dept):
        messages.warning(request, f"🔒 ប្រព័ន្ធត្រូវបានបិទមិនឱ្យកែប្រែឯកសារយោងឡើយ! ({edit_window.status_label_kh})")
        return redirect('officer_detail', pk=pk)

    if request.method == 'POST':
        old_title = att.title
        
        att.category = request.POST.get('category', att.category)
        att.title = request.POST.get('title', att.title).strip()
        att.doc_number = request.POST.get('doc_number', '').strip()
        att.issued_date = request.POST.get('issued_date', '').strip()
        att.description = request.POST.get('description', '').strip()

        replaced_file = False
        if 'file' in request.FILES and request.FILES['file']:
            new_file = request.FILES['file']
            # 2MB maximum limit check
            if new_file.size > 2 * 1024 * 1024:
                file_size_mb = new_file.size / (1024 * 1024)
                messages.error(request, f"⚠️ ទំហំ File ថ្មី «{new_file.name}» គឺ {file_size_mb:.2f}MB ដែលលើសពី 2MB! ប្រព័ន្ធអនុញ្ញាតទំហំអតិបរមាត្រឹម 2MB ប៉ុណ្ណោះ។")
                return redirect('officer_detail', pk=pk)

            att.file = new_file
            att.file_size = new_file.size
            replaced_file = True

        att.save()

        detail_msg = f"បានកែប្រែឯកសារយោង៖ «{att.title}»"
        if replaced_file:
            detail_msg += f" (បានប្តូរ File ថ្មី, ទំហំ៖ {att.formatted_file_size})"
        if old_title != att.title:
            detail_msg += f" (ប្តូរឈ្មោះពី: «{old_title}»)"

        log_officer_action(
            request,
            officer=officer,
            action_type='EDIT',
            attachment_title=att.title,
            category=att.get_category_display(),
            details=detail_msg
        )

        messages.success(request, f"✅ បានកែប្រែ និងរក្សាទុកឯកសារយោង «{att.title}» ដោយជោគជ័យ!")
        return redirect('officer_detail', pk=pk)

    return redirect('officer_detail', pk=pk)


@login_required
def officer_attachment_delete(request, pk, attachment_id):
    officer = get_object_or_404(CivilServantProfile, pk=pk)
    att = get_object_or_404(OfficerAttachment, pk=attachment_id, officer=officer)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    if not is_admin_or_lead and officer.department != dept and officer.user != request.user:
        messages.error(request, "⚠️ លោកអ្នកគ្មានសិទ្ធិលុបឯកសារយោងនេះឡើយ!")
        return redirect('officer_detail', pk=pk)

    # Edit window check for non-admin users
    edit_window = OfficerEditWindowSetting.get_setting()
    if not is_system_admin and not edit_window.is_open_for_department(officer.department or dept):
        messages.warning(request, f"🔒 ប្រព័ន្ធត្រូវបានបិទមិនឱ្យលុបឯកសារយោងឡើយ! ({edit_window.status_label_kh})")
        return redirect('officer_detail', pk=pk)

    if request.method == 'POST':
        title = att.title
        category = att.get_category_display()
        file_name = os.path.basename(att.file.name) if att.file else '-'
        doc_no = att.doc_number or '-'
        
        att.delete()

        log_officer_action(
            request,
            officer=officer,
            action_type='DELETE',
            attachment_title=title,
            category=category,
            details=f"បានលុបឯកសារយោង៖ «{title}» (លេខលិខិត៖ {doc_no}, File: {file_name})"
        )

        messages.success(request, f"🗑️ បានលុបឯកសារយោង «{title}» ដោយជោគជ័យ!")
        return redirect('officer_detail', pk=pk)

    return redirect('officer_detail', pk=pk)


@login_required
def officer_edit_window_setting_view(request):
    """
    Configures and updates the submission/edit window for specialized offices (ADMIN only).
    """
    is_admin = check_is_system_admin(request.user, request)
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('format') == 'json'

    if not is_admin:
        if is_ajax:
            return JsonResponse({'success': False, 'error': '⚠️ មានតែ ADMIN ប៉ុណ្ណោះដែលអាចកំណត់កាលវិភាគប្រព័ន្ធបាន!'}, status=403)
        messages.error(request, "⚠️ មានតែ ADMIN ប៉ុណ្ណោះដែលអាចកំណត់កាលវិភាគប្រព័ន្ធបាន!")
        return redirect('officer_list')

    setting = OfficerEditWindowSetting.get_setting()

    if request.method == 'POST':
        try:
            # In HTML form POST, an unchecked switch is absent from request.POST
            setting.is_active = (request.POST.get('is_active') in ['1', 'true', 'on', True])
            setting.allow_specialized_edit = (request.POST.get('allow_specialized_edit') in ['1', 'true', 'on', True])

            title = request.POST.get('title', '').strip()
            if title:
                setting.title = title

            instruction_note = request.POST.get('instruction_note', '').strip()
            setting.instruction_note = instruction_note

            start_dt_raw = request.POST.get('start_datetime', '').strip()
            if start_dt_raw:
                try:
                    start_dt_raw_clean = start_dt_raw.replace('T', ' ')
                    if len(start_dt_raw_clean) == 16:
                        dt = datetime.strptime(start_dt_raw_clean, '%Y-%m-%d %H:%M')
                    else:
                        dt = datetime.strptime(start_dt_raw_clean, '%Y-%m-%d %H:%M:%S')
                    setting.start_datetime = timezone.make_aware(dt) if timezone.is_naive(dt) else dt
                except Exception:
                    pass
            else:
                setting.start_datetime = None

            end_dt_raw = request.POST.get('end_datetime', '').strip()
            if end_dt_raw:
                try:
                    end_dt_raw_clean = end_dt_raw.replace('T', ' ')
                    if len(end_dt_raw_clean) == 16:
                        dt = datetime.strptime(end_dt_raw_clean, '%Y-%m-%d %H:%M')
                    else:
                        dt = datetime.strptime(end_dt_raw_clean, '%Y-%m-%d %H:%M:%S')
                    setting.end_datetime = timezone.make_aware(dt) if timezone.is_naive(dt) else dt
                except Exception:
                    pass
            else:
                setting.end_datetime = None

            setting.updated_by = request.user
            setting.save()

            # Handle Allowed Departments selection
            dept_ids = request.POST.getlist('allowed_departments')
            if dept_ids:
                setting.allowed_departments.set(dept_ids)
            else:
                setting.allowed_departments.clear()

            allowed_dept_names = ", ".join([d.name_kh for d in setting.allowed_departments.all()]) or "គ្រប់ការិយាល័យទាំងអស់"

            # Record Audit Log
            log_officer_action(
                request,
                officer=None,
                action_type='SETTING_UPDATE',
                details=f"Admin «{request.user.username}» បានកំណត់កាលវិភាគ៖ ស្ថានភាព={setting.status_label_kh}, ចាប់ផ្តើម={setting.start_datetime_formatted or 'គ្មាន'}, បញ្ចប់={setting.end_datetime_formatted or 'គ្មាន'}, វិសាលភាព={allowed_dept_names}"
            )

            msg = f"✅ បានកែប្រែកាលវិភាគកែប្រែ និងបញ្ចូលជីវប្រវត្តិមន្ត្រីដោយជោគជ័យ! (ស្ថានភាពបច្ចុប្បន្ន៖ {setting.status_label_kh})"
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'message': msg,
                    'is_open': setting.is_open_for_specialized(),
                    'status_code': setting.status_code,
                    'status_label': setting.status_label_kh,
                })
            messages.success(request, msg)
        except Exception as e:
            if is_ajax:
                return JsonResponse({'success': False, 'error': f"មានបញ្ហា៖ {str(e)}"}, status=400)
            messages.error(request, f"⚠️ មានបញ្ហាក្នុងការរក្សាទុកការកំណត់៖ {str(e)}")

        return redirect(request.META.get('HTTP_REFERER') or 'officer_list')

    if is_ajax:
        return JsonResponse({
            'is_active': setting.is_active,
            'allow_specialized_edit': setting.allow_specialized_edit,
            'title': setting.title,
            'instruction_note': setting.instruction_note,
            'start_datetime': setting.start_datetime_input,
            'end_datetime': setting.end_datetime_input,
            'is_open': setting.is_open_for_specialized(),
            'status_code': setting.status_code,
            'status_label': setting.status_label_kh,
        })

    return redirect('officer_list')


@xframe_options_exempt
@login_required
def officer_attachment_view(request, pk, attachment_id):
    officer = get_object_or_404(CivilServantProfile, pk=pk)
    att = get_object_or_404(OfficerAttachment, pk=attachment_id, officer=officer)

    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    if not is_admin_or_lead and (not dept or officer.department != dept) and officer.user != request.user:
        raise Http404("លោកអ្នកគ្មានសិទ្ធិមើលឯកសារយោងរបស់មន្ត្រីនេះឡើយ!")
    
    if not att.file or not os.path.exists(att.file.path):
        raise Http404("File ឯកសារមិនត្រូវបានរកឃើញឡើយ!")

    # Record view log
    log_officer_action(
        request,
        officer=officer,
        action_type='VIEW',
        attachment_title=att.title,
        category=att.get_category_display(),
        details=f"បានបើកមើលឯកសារយោង៖ «{att.title}»"
    )

    filename = os.path.basename(att.file.name)
    file_path = att.file.path
    content_type, _ = mimetypes.guess_type(file_path)

    if not content_type:
        fn_lower = filename.lower()
        if fn_lower.endswith('.pdf'):
            content_type = 'application/pdf'
        elif fn_lower.endswith(('.jpg', '.jpeg')):
            content_type = 'image/jpeg'
        elif fn_lower.endswith('.png'):
            content_type = 'image/png'
        elif fn_lower.endswith('.webp'):
            content_type = 'image/webp'
        elif fn_lower.endswith('.gif'):
            content_type = 'image/gif'
        elif fn_lower.endswith(('.bmp', '.tiff')):
            content_type = 'image/bmp'
        else:
            content_type = 'application/octet-stream'

    response = FileResponse(open(file_path, 'rb'), content_type=content_type)
    response['X-Frame-Options'] = 'ALLOWALL'
    response['Content-Disposition'] = 'inline'
    response['Accept-Ranges'] = 'bytes'
    response['Cache-Control'] = 'private, max-age=3600'
    return response


@login_required
def officer_attachment_download(request, pk, attachment_id):
    officer = get_object_or_404(CivilServantProfile, pk=pk)
    att = get_object_or_404(OfficerAttachment, pk=attachment_id, officer=officer)

    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    if not is_admin_or_lead and (not dept or officer.department != dept) and officer.user != request.user:
        raise Http404("លោកអ្នកគ្មានសិទ្ធិទាញយកឯកសារយោងរបស់មន្ត្រីនេះឡើយ!")

    if not att.file or not os.path.exists(att.file.path):
        raise Http404("File ឯកសារមិនត្រូវបានរកឃើញឡើយ!")

    log_officer_action(
        request,
        officer=officer,
        action_type='VIEW',
        attachment_title=att.title,
        category=att.get_category_display(),
        details=f"បានទាញយកឯកសារយោង៖ «{att.title}»"
    )

    content_type, _ = mimetypes.guess_type(att.file.path)
    response = FileResponse(open(att.file.path, 'rb'), content_type=content_type or 'application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{os.path.basename(att.file.name)}"'
    return response


@login_required
def officer_audit_trail_view(request):
    profile = getattr(request.user, 'profile', None)
    is_admin = request.user.is_superuser or request.user.username.upper() == 'ADMIN'

    if not is_admin:
        messages.error(request, "⚠️ មានតែ ADMIN ប៉ុណ្ណោះដែលអាចមើលផ្ទាំង Audit Logs នេះបាន!")
        return redirect('officer_list')

    search_q = request.GET.get('q', '').strip()
    action_filter = request.GET.get('action_type', '').strip()
    user_filter = request.GET.get('user', '').strip()

    logs = OfficerAuditLog.objects.select_related('officer', 'user', 'officer__department').all()

    if search_q:
        logs = logs.filter(
            Q(officer__khmer_last_name__icontains=search_q) |
            Q(officer__khmer_first_name__icontains=search_q) |
            Q(officer__officer_id_number__icontains=search_q) |
            Q(attachment_title__icontains=search_q) |
            Q(details__icontains=search_q) |
            Q(user__username__icontains=search_q)
        )

    if action_filter:
        logs = logs.filter(action_type=action_filter)

    if user_filter:
        logs = logs.filter(user_id=user_filter)

    paginator = Paginator(logs, 35)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    all_users = User.objects.filter(is_active=True).order_by('username')

    context = {
        'page_obj': page_obj,
        'logs': page_obj.object_list,
        'search_q': search_q,
        'action_filter': action_filter,
        'user_filter': user_filter,
        'all_users': all_users,
        'total_logs': logs.count(),
        'action_choices': OfficerAuditLog.ACTION_CHOICES,
    }
    return render(request, 'dms/officer_audit_logs.html', context)


@login_required
def officer_audit_log_delete(request, log_id):
    is_admin = request.user.is_superuser or request.user.username.upper() == 'ADMIN'
    if not is_admin:
        messages.error(request, "⚠️ មានតែ ADMIN ប៉ុណ្ណោះដែលអាចលុបកំណត់ត្រា Audit Log បាន!")
        return redirect('officer_audit_trail')

    log_entry = get_object_or_404(OfficerAuditLog, pk=log_id)
    if request.method == 'POST':
        action_name = log_entry.get_action_type_display()
        log_entry.delete()
        messages.success(request, f"🗑️ បានលុបកំណត់ត្រាសកម្មភាព «{action_name}» ដោយជោគជ័យ!")
    
    ref_url = request.META.get('HTTP_REFERER')
    if ref_url and 'audit-logs' in ref_url:
        return redirect(ref_url)
    return redirect('officer_audit_trail')


@login_required
def officer_audit_trail_clear_all(request):
    is_admin = request.user.is_superuser or request.user.username.upper() == 'ADMIN'
    if not is_admin:
        messages.error(request, "⚠️ មានតែ ADMIN ប៉ុណ្ណោះដែលអាចសម្អាត Audit Log បាន!")
        return redirect('officer_audit_trail')

    if request.method == 'POST':
        total = OfficerAuditLog.objects.count()
        OfficerAuditLog.objects.all().delete()
        messages.success(request, f"🧹 បានសម្អាត និងលុបកំណត់ត្រាសកម្មភាពទាំងអស់សរុប {total} កំណត់ត្រា ដោយជោគជ័យ!")
    
    return redirect('officer_audit_trail')


@login_required
def officer_audit_trail_bulk_delete(request):
    is_admin = request.user.is_superuser or request.user.username.upper() == 'ADMIN'
    if not is_admin:
        messages.error(request, "⚠️ មានតែ ADMIN ប៉ុណ្ណោះដែលអាចលុប Audit Log បាន!")
        return redirect('officer_audit_trail')

    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_logs')
        if selected_ids:
            count = OfficerAuditLog.objects.filter(id__in=selected_ids).delete()[0]
            messages.success(request, f"🗑️ បានលុបកំណត់ត្រាដែលបានជ្រើសរើសចំនួន {count} ដោយជោគជ័យ!")
        else:
            messages.warning(request, "⚠️ សូមជ្រើសរើសកំណត់ត្រាយ៉ាងហោចណាស់មួយដើម្បីលុប!")

    ref_url = request.META.get('HTTP_REFERER')
    if ref_url and 'audit-logs' in ref_url:
        return redirect(ref_url)
    return redirect('officer_audit_trail')


# =========================================================================
# 📊 Officer Export Excel (E-1 សម្រាប់ក្រសួង & E-2 សម្រាប់មុខងារសាធារណៈ)
# =========================================================================

KHMER_MONTHS_NAMES = ['', 'មករា', 'កុម្ភៈ', 'មីនា', 'មេសា', 'ឧសភា', 'មិថុនា', 'កក្កដា', 'សីហា', 'កញ្ញា', 'តុលា', 'វិច្ឆិកា', 'ធ្នូ']
KHMER_NUM_MAP = {'0': '០', '1': '១', '2': '២', '3': '៣', '4': '៤', '5': '៥', '6': '៦', '7': '៧', '8': '៨', '9': '៩'}

def _to_khmer_digits(num_val):
    return ''.join(KHMER_NUM_MAP.get(c, c) for c in str(num_val))

def _format_khmer_date_standard(val):
    if not val:
        return ''
    import re
    val_str = str(val).strip()
    KHMER_DIGITS_REV = {'០': '0', '១': '1', '២': '2', '៣': '3', '៤': '4', '៥': '5', '៦': '6', '៧': '7', '៨': '8', '៩': '9'}
    for kh, ar in KHMER_DIGITS_REV.items():
        val_str = val_str.replace(kh, ar)
    KHMER_MONTHS_REV = {
        'មករា': '01', 'កុម្ភៈ': '02', 'មីនា': '03', 'មេសា': '04', 'ឧសភា': '05', 'មិថុនា': '06',
        'កក្កដា': '07', 'សីហា': '08', 'កញ្ញា': '09', 'តុលា': '10', 'វិច្ឆិកា': '11', 'ធ្នូ': '12'
    }
    m = re.search(r'(\d{1,2})\s*[-/ខែ\s]+\s*([^\s/-]+)\s*[-/ឆ្នាំ\s]+\s*(\d{4})', val_str)
    if m:
        d, mon, y = m.group(1).zfill(2), m.group(2).strip(), m.group(3)
        mon = KHMER_MONTHS_REV.get(mon, mon.zfill(2))
        return f'{d}-{mon}-{y}'
    m_dmy = re.search(r'(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})', val_str)
    if m_dmy:
        return f'{m_dmy.group(1).zfill(2)}-{m_dmy.group(2).zfill(2)}-{m_dmy.group(3)}'
    return val_str

def _extract_officer_degree_and_skill(o):
    DEGREE_MAP = {
        'DOCTOR': 'បណ្ឌិត',
        'MASTER': 'បរិញ្ញាបត្រជាន់ខ្ពស់',
        'BACHELOR': 'បរិញ្ញាបត្រ',
        'ASSOCIATE': 'បរិញ្ញាបត្ររង',
        'HIGHSCHOOL': 'ទុតិយភូមិ',
        'HIGH_SCHOOL': 'ទុតិយភូមិ',
        'SECONDARY': 'បឋមភូមិ',
        'PRIMARY': 'បឋមសិក្សា',
        'OTHER': 'វិញ្ញាបនបត្រ័ផ្សេងៗ',
        'NONE': 'គ្មាន',
    }
    degree = DEGREE_MAP.get(o.highest_degree, o.highest_degree or '')
    skill = ''
    edu_list = o.education_data if isinstance(o.education_data, list) else []
    
    for edu in edu_list:
        if isinstance(edu, dict):
            lvl = str(edu.get('level_type', '')).upper()
            edu_deg = edu.get('degree', '') or edu.get('level_label', '')
            edu_sk = edu.get('skill', '')
            if lvl in ['VOCATIONAL', 'HIGHER', 'UNIVERSITY'] or 'បរិញ្ញា' in str(edu_deg) or 'បណ្ឌិត' in str(edu_deg):
                if edu_sk and not skill:
                    skill = edu_sk
                if edu_deg and (not degree or degree in ['វិញ្ញាបនបត្រ័ផ្សេងៗ', 'ផ្សេងៗ', 'HIGHSCHOOL']):
                    degree = edu_deg
                    
    if not skill:
        for edu in edu_list:
            if isinstance(edu, dict) and edu.get('skill'):
                skill = edu.get('skill')
                break
                
    if not degree and edu_list and isinstance(edu_list[0], dict):
        degree = edu_list[0].get('degree') or edu_list[0].get('level_label') or ''
        
    return degree or 'គ្មាន', skill or ''

def _extract_officer_promotion_info(o):
    prom_date = ''
    legal_doc = ''
    deg_prom_date = ''
    deg_doc_no = ''
    
    proms = o.promotions_by_seniority if isinstance(o.promotions_by_seniority, list) else []
    if proms and isinstance(proms[0], dict):
        p = proms[0]
        p_raw = p.get('effective_date') or p.get('promo_date') or ''
        prom_date = _format_khmer_date_standard(p_raw)
        doc_no = p.get('doc_number') or p.get('legal_doc_number') or p.get('number') or ''
        doc_type = p.get('doc_type') or p.get('legal_doc_type') or ''
        if doc_no and doc_type:
            legal_doc = f'{doc_no} {doc_type}'
        elif doc_no:
            legal_doc = doc_no
        elif doc_type:
            legal_doc = doc_type

    deg_proms = o.promotions_by_degree if isinstance(o.promotions_by_degree, list) else []
    if deg_proms and isinstance(deg_proms[0], dict):
        dp = deg_proms[0]
        deg_prom_date = _format_khmer_date_standard(dp.get('effective_date') or dp.get('promo_date') or '')
        deg_doc_no = dp.get('degree_number') or dp.get('doc_number') or dp.get('number') or ''
        
    return prom_date, legal_doc, deg_prom_date, deg_doc_no


def _get_khmer_lunar_year_info(year):
    """
    គណនាឆ្នាំចន្ទគតិខ្មែរ (សត្វទាំង១២, ស័កទាំង១០, និងពុទ្ធសករាជ ព.ស) ដោយស្វ័យប្រវត្តិតាមប្រតិទិនខ្មែរពិតប្រាកដ
    - ឆ្នាំ ២០២៤: ឆ្នាំរោង ឆស័ក ព.ស ២៥៦៨
    - ឆ្នាំ ២០២៥: ឆ្នាំម្សាញ់ សប្តស័ក ព.ស ២៥៦៩
    - ឆ្នាំ ២០២៦: ឆ្នាំមមី អដ្ឋស័ក ព.ស ២៥៧០
    - ឆ្នាំ ២០២៧: ឆ្នាំមមែ នព្វស័ក ព.ស ២៥៧១
    - ឆ្នាំ ២០២៨: ឆ្នាំវក សំរឹទ្ធិស័ក ព.ស ២៥៧២
    """
    animals = ['ជូត', 'ឆ្លូវ', 'ខាល', 'ថោះ', 'រោង', 'ម្សាញ់', 'មមី', 'មមែ', 'វក', 'រកា', 'ចរ', 'កុរ']
    saks = ['សំរឹទ្ធិស័ក', 'ឯកស័ក', 'ទោស័ក', 'ត្រីស័ក', 'ចត្វាស័ក', 'បញ្ចស័ក', 'ឆស័ក', 'សប្តស័ក', 'អដ្ឋស័ក', 'នព្វស័ក']
    
    animal = animals[(year - 4) % 12]
    sak = saks[(year + 2) % 10]
    be_year = year + 544
    be_kh = _to_khmer_digits(str(be_year))
    
    return f"ឆ្នាំ{animal} {sak} ព.ស.{be_kh}"


def _calculate_e1_cadre_stats(officers_list):
    total = len(officers_list)
    female_total = sum(1 for o in officers_list if o.gender == 'FEMALE' or str(o.gender).upper() in ['F', 'FEMALE', 'ស្រី', 'ស'])
    
    cadre_a_total = 0
    cadre_a_female = 0
    cadre_b_total = 0
    cadre_b_female = 0
    cadre_c_total = 0
    cadre_c_female = 0
    
    for o in officers_list:
        rank = (o.current_rank_and_step or '').strip()
        is_f = o.gender == 'FEMALE' or str(o.gender).upper() in ['F', 'FEMALE', 'ស្រី', 'ស']
        
        if rank.startswith(('ក', 'A', 'a')):
            cadre_a_total += 1
            if is_f:
                cadre_a_female += 1
        elif rank.startswith(('ខ', 'B', 'b')):
            cadre_b_total += 1
            if is_f:
                cadre_b_female += 1
        else:
            cadre_c_total += 1
            if is_f:
                cadre_c_female += 1
                
    return {
        'total': total,
        'female_total': female_total,
        'cadre_a_total': cadre_a_total,
        'cadre_a_female': cadre_a_female,
        'cadre_b_total': cadre_b_total,
        'cadre_b_female': cadre_b_female,
        'cadre_c_total': cadre_c_total,
        'cadre_c_female': cadre_c_female,
    }


def _add_excel_khmer_divider(ws, col_idx, col_offset_px, row_idx, row_offset_px=3, width_px=65, height_px=18):
    """Adds the royal ornamental flourish divider under ជាតិ សាសនា ព្រះមហាក្សត្រ in Excel worksheets."""
    divider_path = os.path.join(settings.BASE_DIR, 'dms', 'static', 'dms', 'img', 'khmer_divider.png')
    if not os.path.exists(divider_path):
        return
    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker, XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU
        div_img = OpenpyxlImage(divider_path)
        div_size = XDRPositiveSize2D(pixels_to_EMU(width_px), pixels_to_EMU(height_px))
        marker = AnchorMarker(col=col_idx, colOff=pixels_to_EMU(col_offset_px), row=row_idx, rowOff=pixels_to_EMU(row_offset_px))
        div_img.anchor = OneCellAnchor(_from=marker, ext=div_size)
        ws.add_image(div_img)
    except Exception:
        pass


def _add_excel_centered_logo(ws, start_col_letter, end_col_letter, start_row_idx=0, row_offset_px=2, logo_size_px=60):
    """
    Precisely anchors the official circular MAFF logo so that its horizontal center
    perfectly matches the center axis of the organization header merged across start_col to end_col.
    """
    logo_path = os.path.join(settings.BASE_DIR, 'dms', 'static', 'dms', 'img', 'image1.jpeg')
    if not os.path.exists(logo_path):
        return
    try:
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker, XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU
        import openpyxl.utils

        def _get_col_px(col_letter):
            cd = ws.column_dimensions.get(col_letter)
            w = cd.width if (cd and cd.width is not None) else 8.43
            return int(w * 7 + 0.5) + 5

        start_idx = openpyxl.utils.column_index_from_string(start_col_letter) - 1
        end_idx = openpyxl.utils.column_index_from_string(end_col_letter) - 1

        left_of_start = 0
        for idx in range(start_idx):
            letter = openpyxl.utils.get_column_letter(idx + 1)
            left_of_start += _get_col_px(letter)

        range_width = 0
        for idx in range(start_idx, end_idx + 1):
            letter = openpyxl.utils.get_column_letter(idx + 1)
            range_width += _get_col_px(letter)

        center_x = left_of_start + (range_width / 2.0)
        logo_left_x = center_x - (logo_size_px / 2.0)

        # Find which column logo_left_x falls into
        cum = 0
        anchor_col = 0
        col_off_px = 0
        for idx in range(max(ws.max_column + 10, end_idx + 5)):
            letter = openpyxl.utils.get_column_letter(idx + 1)
            w = _get_col_px(letter)
            if cum + w > logo_left_x:
                anchor_col = idx
                col_off_px = max(0, logo_left_x - cum)
                break
            cum += w

        logo_img = OpenpyxlImage(logo_path)
        logo_ext = XDRPositiveSize2D(pixels_to_EMU(logo_size_px), pixels_to_EMU(logo_size_px))
        marker = AnchorMarker(col=anchor_col, colOff=pixels_to_EMU(col_off_px), row=start_row_idx, rowOff=pixels_to_EMU(row_offset_px))
        logo_img.anchor = OneCellAnchor(_from=marker, ext=logo_ext)
        ws.add_image(logo_img)
    except Exception:
        pass


def _build_e1_ministry_workbook(officers_list, department=None):
    from openpyxl.drawing.image import Image as OpenpyxlImage
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "01"

    # Page Setup: A4 Landscape Print Format
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    # Page Numbering: -1-, -2-, ...
    ws.oddFooter.center.text = "-&P-"
    ws.evenFooter.center.text = "-&P-"

    # Fonts
    font_title = Font(name='Khmer OS Muol Light', size=11, bold=False)
    font_org = Font(name='Khmer OS Muol Light', size=10, bold=False)
    font_tbl_header = Font(name='Khmer OS Muol', size=10, bold=False)
    font_dept_header = Font(name='Khmer OS Muol Light', size=10, bold=False)
    font_data = Font(name='Khmer OS Battambang', size=10, bold=False)
    font_data_sm = Font(name='Khmer OS Battambang', size=9, bold=False)
    font_num = Font(name='Arial', size=10, bold=False)
    font_id = Font(name='Arial', size=9, bold=False)

    # Borders
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_title = Alignment(horizontal='center', vertical='center')

    align_data_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    align_data_left = Alignment(horizontal='left', vertical='center', wrap_text=False)

    col_widths = {
        'A': 4.5,
        'B': 13.0,
        'C': 17.0,
        'D': 6.0,
        'E': 13.0,
        'F': 13.0,
        'G': 10.0,
        'H': 22.0,
        'I': 16.0,
        'J': 42.0
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row dimensions for Logo and Organization
    ws.row_dimensions[1].height = 24.0
    ws.row_dimensions[2].height = 24.0
    ws.row_dimensions[3].height = 20.0
    ws.row_dimensions[4].height = 20.0
    ws.row_dimensions[5].height = 8.0

    # Add Official Logo centered ABOVE Organization (A to E)
    _add_excel_centered_logo(ws, 'A', 'E', start_row_idx=0, row_offset_px=2, logo_size_px=58)

    is_specialized = _is_specialized_department(department)

    # Header Left (Organization - Below Logo)
    ws.merge_cells('A3:E3')
    ws.merge_cells('A4:E4')
    if is_specialized and department:
        # Specialized Office / Canton: Line 1 = Provincial Department, Line 2 = Office/Canton
        ws['A3'] = "មន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទខេត្តប៉ៃលិន"
        ws['A4'] = department.name_kh
    else:
        # Full Department / Ministry: Line 1 = Ministry, Line 2 = Provincial Department
        ws['A3'] = "ក្រសួងកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទ"
        ws['A4'] = "មន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទខេត្តប៉ៃលិន"

    ws['A3'].font = font_org
    ws['A3'].alignment = align_center
    ws['A4'].font = font_org
    ws['A4'].alignment = align_center

    # Header Right (Kingdom)
    ws.merge_cells('G2:J2')
    ws['G2'] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws['G2'].font = font_title
    ws['G2'].alignment = align_center

    ws.merge_cells('G3:J3')
    ws['G3'] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws['G3'].font = font_title
    ws['G3'].alignment = align_center

    # Royal Divider Flourish under ជាតិ សាសនា ព្រះមហាក្សត្រ
    ws.merge_cells('G4:J4')
    ws['G4'] = ""
    _add_excel_khmer_divider(ws, col_idx=8, col_offset_px=78, row_idx=3, row_offset_px=2, width_px=65, height_px=18)

    # Title Block
    now = datetime.now()
    month_kh = KHMER_MONTHS_NAMES[now.month] if 1 <= now.month <= 12 else str(now.month)
    year_kh = _to_khmer_digits(str(now.year))

    ws.merge_cells('A6:J6')
    ws['A6'] = "បញ្ជីរាយនាមមន្ត្រីរាជការ"
    ws['A6'].font = font_title
    ws['A6'].alignment = align_title

    ws.merge_cells('A7:J7')
    if is_specialized and department:
        ws['A7'] = f"របស់{department.name_kh}"
    else:
        ws['A7'] = "ក្នុងរចនាសម្ព័ន្ធមន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទខេត្តប៉ៃលិន"
    ws['A7'].font = font_title
    ws['A7'].alignment = align_title

    ws.merge_cells('A8:J8')
    ws['A8'] = f"ប្រចាំខែ{month_kh} ឆ្នាំ{year_kh}"
    ws['A8'].font = font_title
    ws['A8'].alignment = align_title

    ws.row_dimensions[6].height = 22.0
    ws.row_dimensions[7].height = 22.0
    ws.row_dimensions[8].height = 22.0
    ws.row_dimensions[10].height = 36.0

    headers = [
        'ល.រ', 'អត្តលេខ', 'ឈ្មោះ', 'ភេទ',
        'ថ្ងៃ-ខែ-ឆ្នាំ\nកំណើត', 'ថ្ងៃខែឆ្នាំ\nបម្រើការងារ',
        'កាំបៀវត្ស', 'មុខដំណែង', 'សញ្ញាប័ត្រ', 'ជំនាញ'
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_idx, value=h)
        cell.font = font_tbl_header
        cell.alignment = align_center
        cell.border = thin_border

    current_row = 11
    leadership_officers = []
    dept_map = {}

    for o in officers_list:
        pos = (o.current_position_title or '').strip()
        dept_name = o.department.name_kh if o.department else 'ផ្សេងៗ'
        is_lead = (o.department and (o.department.code in ['LEAD', 'LEADERSHIP'] or 'ថ្នាក់ដឹកនាំ' in o.department.name_kh)) or \
                  any(lead_title in pos for lead_title in ['ប្រធានមន្ទីរ', 'អនុប្រធានមន្ទីរ'])
        
        if is_lead:
            leadership_officers.append(o)
        else:
            if dept_name not in dept_map:
                dept_map[dept_name] = []
            dept_map[dept_name].append(o)

    if leadership_officers:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        dept_cell = ws.cell(row=current_row, column=1, value="ថ្នាក់ដឹកនាំមន្ទីរ")
        dept_cell.font = font_dept_header
        dept_cell.alignment = align_left
        for c in range(1, 11):
            ws.cell(row=current_row, column=c).border = thin_border
        ws.row_dimensions[current_row].height = 22.0
        current_row += 1

        for idx, o in enumerate(leadership_officers, 1):
            deg, skill = _extract_officer_degree_and_skill(o)
            gender_kh = 'ប្រុស' if o.gender == 'MALE' or str(o.gender).upper() in ['M', 'ប្រុស', 'ប'] else 'ស្រី'
            row_data = [
                (idx, font_num, align_data_center),
                (o.officer_id_number or '', font_id, align_data_center),
                (o.full_name_kh, font_data, align_data_left),
                (gender_kh, font_data, align_data_center),
                (_format_khmer_date_standard(o.dob), font_data, align_data_center),
                (_format_khmer_date_standard(o.civil_service_start_date), font_data, align_data_center),
                (o.current_rank_and_step or '', font_data, align_data_center),
                (o.current_position_title or '', font_data_sm, align_data_left),
                (deg, font_data_sm, align_data_left),
                (skill, font_data_sm, align_data_left),
            ]
            for col_idx, (val, fnt, aln) in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = fnt
                cell.alignment = aln
                cell.border = thin_border
            ws.row_dimensions[current_row].height = 22.0
            current_row += 1

    for dept_name, off_list in dept_map.items():
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        dept_cell = ws.cell(row=current_row, column=1, value=dept_name)
        dept_cell.font = font_dept_header
        dept_cell.alignment = align_left
        for c in range(1, 11):
            ws.cell(row=current_row, column=c).border = thin_border
        ws.row_dimensions[current_row].height = 22.0
        current_row += 1

        for idx, o in enumerate(off_list, 1):
            deg, skill = _extract_officer_degree_and_skill(o)
            gender_kh = 'ប្រុស' if o.gender == 'MALE' or str(o.gender).upper() in ['M', 'ប្រុស', 'ប'] else 'ស្រី'
            row_data = [
                (idx, font_num, align_data_center),
                (o.officer_id_number or '', font_id, align_data_center),
                (o.full_name_kh, font_data, align_data_left),
                (gender_kh, font_data, align_data_center),
                (_format_khmer_date_standard(o.dob), font_data, align_data_center),
                (_format_khmer_date_standard(o.civil_service_start_date), font_data, align_data_center),
                (o.current_rank_and_step or '', font_data, align_data_center),
                (o.current_position_title or '', font_data_sm, align_data_left),
                (deg, font_data_sm, align_data_left),
                (skill, font_data_sm, align_data_left),
            ]
            for col_idx, (val, fnt, aln) in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = fnt
                cell.alignment = aln
                cell.border = thin_border
            ws.row_dimensions[current_row].height = 22.0
            current_row += 1

    # Signatures & Cadre Summary at Bottom (Matching Image 2)
    cadre_stats = _calculate_e1_cadre_stats(officers_list)
    lunar_info = _get_khmer_lunar_year_info(now.year)

    current_row += 1  # 1 blank row
    
    font_stat = Font(name='Khmer OS Battambang', size=10, bold=False)
    align_stat_left = Alignment(horizontal='left', vertical='center')
    align_sig_center = Alignment(horizontal='center', vertical='center')

    # Row 1
    # Left: សរុបរួម
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    c_s1 = ws.cell(row=current_row, column=1, value=f"សរុបរួម៖ ...{cadre_stats['total']}..នាក់ , ស្រី...{cadre_stats['female_total']}..នាក់")
    c_s1.font = font_stat
    c_s1.alignment = align_stat_left

    # Right: ថ្ងៃ..................... ខែ.................. ឆ្នាំ...
    ws.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=10)
    c_r1 = ws.cell(row=current_row, column=7, value=f"ថ្ងៃ..................... ខែ.................. {lunar_info}")
    c_r1.font = font_stat
    c_r1.alignment = align_sig_center

    # Row 2
    # Left: ក្របខ័ណ្ឌ ក
    ws.merge_cells(start_row=current_row + 1, start_column=1, end_row=current_row + 1, end_column=2)
    c_s2 = ws.cell(row=current_row + 1, column=1, value=f"      ក្របខ័ណ្ឌ ក...{cadre_stats['cadre_a_total']}..នាក់ , ស្រី...{cadre_stats['cadre_a_female']}..នាក់")
    c_s2.font = font_stat
    c_s2.alignment = align_stat_left

    # Right: ប៉ៃលិន, ថ្ងៃទី....... ខែ................ ឆ្នាំ២០២៥
    ws.merge_cells(start_row=current_row + 1, start_column=7, end_row=current_row + 1, end_column=10)
    c_r2 = ws.cell(row=current_row + 1, column=7, value=f"ប៉ៃលិន, ថ្ងៃទី....... ខែ................ ឆ្នាំ{year_kh}")
    c_r2.font = font_stat
    c_r2.alignment = align_sig_center

    # Row 3
    # Left: ក្របខ័ណ្ឌ ខ
    ws.merge_cells(start_row=current_row + 2, start_column=1, end_row=current_row + 2, end_column=2)
    c_s3 = ws.cell(row=current_row + 2, column=1, value=f"      ក្របខ័ណ្ឌ ខ...{cadre_stats['cadre_b_total']}..នាក់ , ស្រី...{cadre_stats['cadre_b_female']}..នាក់")
    c_s3.font = font_stat
    c_s3.alignment = align_stat_left

    # Right: Head of Office Title
    ws.merge_cells(start_row=current_row + 2, start_column=7, end_row=current_row + 2, end_column=10)
    sig_right_title = _get_department_head_title(department) if (is_specialized and department) else "ប្រធានការិយាល័យរដ្ឋបាល-បុគ្គលិក"
    c_r3 = ws.cell(row=current_row + 2, column=7, value=sig_right_title)
    c_r3.font = font_dept_header
    c_r3.alignment = align_sig_center

    # Row 4
    # Left: ក្របខ័ណ្ឌ គ
    ws.merge_cells(start_row=current_row + 3, start_column=1, end_row=current_row + 3, end_column=2)
    c_s4 = ws.cell(row=current_row + 3, column=1, value=f"      ក្របខ័ណ្ឌ គ...{cadre_stats['cadre_c_total']}..នាក់ , ស្រី...{cadre_stats['cadre_c_female']}..នាក់")
    c_s4.font = font_stat
    c_s4.alignment = align_stat_left

    # Row 5 (Center: បានឃើញ និងឯកភាព dropped down below stats)
    ws.merge_cells(start_row=current_row + 4, start_column=3, end_row=current_row + 4, end_column=5)
    c_c1 = ws.cell(row=current_row + 4, column=3, value="បានឃើញ និងឯកភាព")
    c_c1.font = font_dept_header
    c_c1.alignment = align_sig_center

    # Row 6 (Center: ប្រធាន)
    ws.merge_cells(start_row=current_row + 5, start_column=3, end_row=current_row + 5, end_column=5)
    c_c2 = ws.cell(row=current_row + 5, column=3, value="ប្រធាន")
    c_c2.font = font_dept_header
    c_c2.alignment = align_sig_center

    return wb


def _build_e2_provincial_workbook(officers_list):
    from openpyxl.drawing.image import Image as OpenpyxlImage
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MAFFpl Update"

    # Page Setup: A4 Landscape Print Format
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    # Page Numbering: -1-, -2-, ...
    ws.oddFooter.center.text = "-&P-"
    ws.evenFooter.center.text = "-&P-"

    font_title = Font(name='Khmer OS Muol Light', size=11, bold=False)
    font_org = Font(name='Khmer OS Muol Light', size=10, bold=False)
    font_header_8 = Font(name='Khmer OS Muol Light', size=8, bold=False)
    font_header_7 = Font(name='Khmer OS Muol Light', size=7, bold=False)
    font_data = Font(name='Khmer OS Siemreap', size=8.5, bold=False)

    border_header_mid = Border(
        top=Side(style='double', color='000000'),
        bottom=Side(style='thin', color='000000'),
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000')
    )
    border_header_left = Border(
        top=Side(style='double', color='000000'),
        bottom=Side(style='thin', color='000000'),
        left=Side(style='double', color='000000'),
        right=Side(style='thin', color='000000')
    )
    border_header_right = Border(
        top=Side(style='double', color='000000'),
        bottom=Side(style='thin', color='000000'),
        left=Side(style='thin', color='000000'),
        right=Side(style='double', color='000000')
    )

    border_data_mid = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000')
    )
    border_data_left = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
        left=Side(style='double', color='000000'),
        right=Side(style='thin', color='000000')
    )
    border_data_right = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
        left=Side(style='thin', color='000000'),
        right=Side(style='double', color='000000')
    )

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_data_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    align_data_left = Alignment(horizontal='left', vertical='center', wrap_text=False)

    col_widths = {
        'A': 4.5,
        'B': 14.0,
        'C': 11.5,
        'D': 12.0,
        'E': 5.0,
        'F': 12.5,
        'G': 12.0,
        'H': 12.0,
        'I': 8.5,
        'J': 14.5,
        'K': 24.0,
        'L': 16.0,
        'M': 13.0,
        'N': 18.0,
        'O': 13.0,
        'P': 10.5,
        'Q': 24.0
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row dimensions for Logo and Organization
    ws.row_dimensions[1].height = 24.0
    ws.row_dimensions[2].height = 24.0
    ws.row_dimensions[3].height = 20.0
    ws.row_dimensions[4].height = 20.0
    ws.row_dimensions[5].height = 8.0

    # Add Official Logo centered ABOVE Organization (A to F)
    _add_excel_centered_logo(ws, 'A', 'F', start_row_idx=0, row_offset_px=2, logo_size_px=58)

    # Header Left (Organization - Below Logo)
    ws.merge_cells('A3:F3')
    ws['A3'] = "ក្រសួងកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទ"
    ws['A3'].font = font_org
    ws['A3'].alignment = align_center

    ws.merge_cells('A4:F4')
    ws['A4'] = "មន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទខេត្តប៉ៃលិន"
    ws['A4'].font = font_org
    ws['A4'].alignment = align_center

    # Header Right (Kingdom)
    ws.merge_cells('K2:P2')
    ws['K2'] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws['K2'].font = font_title
    ws['K2'].alignment = align_center

    ws.merge_cells('K3:P3')
    ws['K3'] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws['K3'].font = font_title
    ws['K3'].alignment = align_center

    # Royal Divider Flourish under ជាតិ សាសនា ព្រះមហាក្សត្រ
    ws.merge_cells('K4:P4')
    ws['K4'] = ""
    _add_excel_khmer_divider(ws, col_idx=12, col_offset_px=21, row_idx=3, row_offset_px=2, width_px=65, height_px=18)

    now = datetime.now()
    month_kh = KHMER_MONTHS_NAMES[now.month] if 1 <= now.month <= 12 else str(now.month)
    year_kh = _to_khmer_digits(str(now.year))

    # Title Block
    ws.merge_cells('A6:P6')
    ws['A6'] = "បញ្ជីរាយនាមមន្ត្រីរាជការស៊ីវិល និងស្ថានភាពបច្ចុប្បន្នភាព"
    ws['A6'].font = font_title
    ws['A6'].alignment = align_center

    ws.merge_cells('A7:P7')
    ws['A7'] = f"បញ្ជូនទៅកាន់៖ មន្ទីរមុខងារសាធារណៈខេត្តប៉ៃលិន (ប្រចាំខែ{month_kh} ឆ្នាំ{year_kh})"
    ws['A7'].font = font_title
    ws['A7'].alignment = align_center

    ws.row_dimensions[6].height = 22.0
    ws.row_dimensions[7].height = 22.0

    headers = [
        ('ល.រ', font_header_8),
        ('គោត្តនាម និងនាម', font_header_8),
        ('អត្តលេខ', font_header_8),
        ('លេខ អត្ដសញ្ញាណបណ្ណ', font_header_7),
        ('ភេទ', font_header_8),
        ('ថ្ងៃខែឆ្នាំកំណើត', font_header_8),
        ('ថ្ងៃខែឆ្នាំចូលបម្រើការងារ', font_header_7),
        ('តួនាទី', font_header_8),
        ('កាំប្រាក់', font_header_8),
        ('សញ្ញាបត្រ', font_header_8),
        ('ជំនាញឯកទេស', font_header_8),
        ('លេខទូរស័ព្ទ', font_header_8),
        ('ថ្ងៃខែឆ្នាំ ឡើងថ្នាក់តាមចុងក្រោយ', font_header_7),
        ('លេខ ប្រកាស អនុក្រឹត្យ ព្រះរាជក្រឹត្យ', font_header_7),
        ('ថ្ងៃខែឆ្នាំ ឡើងថ្នាក់ តាមសញ្ញាបត្រ', font_header_7),
        ('លេខសញ្ញាបត្រ', font_header_7),
        ('ការិយាល័យ/អង្គភាព', font_header_8),
    ]

    header_row = 9
    ws.row_dimensions[header_row].height = 42.0

    for col_idx, (h_text, fnt) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h_text)
        cell.font = fnt
        cell.alignment = align_center
        if col_idx == 1:
            cell.border = border_header_left
        elif col_idx == 16:
            cell.border = border_header_right
        else:
            cell.border = border_header_mid

    for idx, o in enumerate(officers_list, 1):
        r = header_row + idx
        ws.row_dimensions[r].height = 22.0

        gender_code = 'ប' if o.gender == 'MALE' or str(o.gender).upper() in ['M', 'ប្រុស', 'ប'] else 'ស'
        deg, skill = _extract_officer_degree_and_skill(o)
        prom_date, legal_doc, deg_prom_date, deg_doc_no = _extract_officer_promotion_info(o)
        dept_name = o.department.name_kh if o.department else ''

        row_vals = [
            (idx, align_data_center),
            (o.full_name_kh, align_data_left),
            (o.officer_id_number or '', align_data_center),
            (o.national_id_number or '', align_data_center),
            (gender_code, align_data_center),
            (_format_khmer_date_standard(o.dob), align_data_center),
            (_format_khmer_date_standard(o.civil_service_start_date), align_data_center),
            (o.current_position_title or '', align_data_left),
            (o.current_rank_and_step or '', align_data_center),
            (deg, align_data_center),
            (skill, align_data_left),
            (o.phone or '', align_data_center),
            (prom_date, align_data_center),
            (legal_doc, align_data_center),
            (deg_prom_date, align_data_center),
            (deg_doc_no, align_data_center),
            (dept_name, align_data_left),
        ]

        for col_idx, (val, aln) in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = font_data
            cell.alignment = aln
            if col_idx == 1:
                cell.border = border_data_left
            elif col_idx == 16:
                cell.border = border_data_right
            elif col_idx < 16:
                cell.border = border_data_mid
            else:
                cell.border = Border(top=Side(style='none'), bottom=Side(style='none'), left=Side(style='none'), right=Side(style='none'))

    # Signatures at Bottom
    r_sig = header_row + len(officers_list) + 2
    ws.merge_cells(start_row=r_sig, start_column=2, end_row=r_sig, end_column=5)
    ws.cell(row=r_sig, column=2, value="បានឃើញ និងបញ្ជាក់ត្រឹមត្រូវ").font = font_data
    ws.cell(row=r_sig, column=2).alignment = align_center

    ws.merge_cells(start_row=r_sig, start_column=11, end_row=r_sig, end_column=16)
    ws.cell(row=r_sig, column=11, value=f"ប៉ៃលិន, ថ្ងៃទី....... ខែ{month_kh} ឆ្នាំ{year_kh}").font = font_data
    ws.cell(row=r_sig, column=11).alignment = align_center

    r_sig += 1
    ws.merge_cells(start_row=r_sig, start_column=2, end_row=r_sig, end_column=5)
    ws.cell(row=r_sig, column=2, value="ប្រធានមន្ទីរ").font = font_org
    ws.cell(row=r_sig, column=2).alignment = align_center

    ws.merge_cells(start_row=r_sig, start_column=11, end_row=r_sig, end_column=16)
    ws.cell(row=r_sig, column=11, value="ប្រធានការិយាល័យរដ្ឋបាល បុគ្គលិក").font = font_org
    ws.cell(row=r_sig, column=11).alignment = align_center

    return wb


def can_export_civil_servants_to_excel(user, profile=None):
    """
    Strict Excel Export Access Control for Civil Servants:
    Specialized offices and cantons (ខណ្ឌ) are FORBIDDEN from exporting civil servants to Excel,
    EXCEPT for Administration-Personnel Office (ការិយាល័យរដ្ឋបាល-បុគ្គលិក) OR if authorized by ADMIN.
    """
    if not user or not user.is_authenticated:
        return False

    # 1. System Admin / Superuser / Staff / Admin username / Admin role
    if user.is_superuser or user.is_staff or (user.username or '').upper() in ['ADMIN', 'ADMINISTRATOR', 'ROOT']:
        return True

    if not profile:
        profile = getattr(user, 'profile', None)
    if not profile:
        return False

    if getattr(profile, 'is_admin', False) or getattr(profile, 'role', '') == 'ADMIN':
        return True

    # 2. Leadership (ប្រធានមន្ទីរ, អនុប្រធានមន្ទីរ)
    if getattr(profile, 'is_leadership', False) or getattr(profile, 'role', '') in ['LEADERSHIP', 'DIRECTOR', 'DEPUTY_DIRECTOR']:
        return True

    # 3. Administration-Personnel Office (ការិយាល័យរដ្ឋបាល-បុគ្គលិក / កិច្ចការទូទៅ)
    dept = profile.department
    if dept:
        name_kh = (dept.name_kh or '').strip().lower()
        code = (dept.code or '').strip().upper()
        # Strictly exclude Cantons (ខណ្ឌរដ្ឋបាលព្រៃឈើ, ខណ្ឌរដ្ឋបាលជលផល)
        if not (code.startswith('CANTON') or name_kh.startswith('ខណ្ឌ') or 'ខណ្ឌ' in name_kh):
            if code in ['ADMIN', 'ADMIN_PERS', 'ADMIN_PERSONNEL', 'ADMIN_DEPT', 'LEAD', 'GEN_AFFAIRS', 'GENERAL_AFFAIRS']:
                return True
            if ('រដ្ឋបាល' in name_kh and 'បុគ្គលិក' in name_kh) or ('កិច្ចការទូទៅ' in name_kh):
                return True
            if name_kh in ['ការិយាល័យរដ្ឋបាល បុគ្គលិក', 'ការិយាល័យរដ្ឋបាល-បុគ្គលិក', 'ការិយាល័យរដ្ឋបាល', 'ការិយាល័យកិច្ចការរដ្ឋបាលទូទៅ']:
                return True

    # 4. Explicit authorization granted by ADMIN to this profile
    if getattr(profile, 'can_export_civil_servant_excel', False):
        return True

    # Specialized offices and cantons are strictly forbidden by default
    return False


def can_access_officer_e2_report(user, profile=None):
    """
    Form E-2 Access Control:
    Form E-2 (ថ្នាក់ខេត្ត) is STRICTLY DISABLED for Specialized offices and Cantons (ខណ្ឌ).
    ONLY Administration-Personnel Office, Leadership, and ADMIN (or users explicitly granted can_view_e2_report) can access Form E-2.
    """
    if not user or not user.is_authenticated:
        return False

    if user.is_superuser or user.is_staff or (user.username or '').upper() in ['ADMIN', 'ADMINISTRATOR', 'ROOT']:
        return True

    if not profile:
        profile = getattr(user, 'profile', None)
    if not profile:
        return False

    if getattr(profile, 'is_admin', False) or getattr(profile, 'role', '') == 'ADMIN':
        return True

    if getattr(profile, 'is_leadership', False) or getattr(profile, 'role', '') in ['LEADERSHIP', 'DIRECTOR', 'DEPUTY_DIRECTOR']:
        return True

    dept = profile.department
    if dept:
        name_kh = (dept.name_kh or '').strip().lower()
        code = (dept.code or '').strip().upper()
        # Strictly exclude Cantons
        if not (code.startswith('CANTON') or name_kh.startswith('ខណ្ឌ') or 'ខណ្ឌ' in name_kh):
            if code in ['ADMIN', 'ADMIN_PERS', 'ADMIN_PERSONNEL', 'ADMIN_DEPT', 'LEAD', 'GEN_AFFAIRS', 'GENERAL_AFFAIRS']:
                return True
            if ('រដ្ឋបាល' in name_kh and 'បុគ្គលិក' in name_kh) or ('កិច្ចការទូទៅ' in name_kh):
                return True
            if name_kh in ['ការិយាល័យរដ្ឋបាល បុគ្គលិក', 'ការិយាល័យរដ្ឋបាល-បុគ្គលិក', 'ការិយាល័យរដ្ឋបាល', 'ការិយាល័យកិច្ចការរដ្ឋបាលទូទៅ']:
                return True

    if getattr(profile, 'can_view_e2_report', False):
        return True

    return False


@login_required
def officer_export_excel(request):
    """
    Export Civil Servants list to Excel (Form E-1 or Form E-2).
    Strict Permission: Specialized offices and cantons are FORBIDDEN from exporting to Excel,
    EXCEPT for Administration-Personnel Office OR if authorized by ADMIN.
    Form E-2 is additionally forbidden unless authorized for E-2.
    """
    profile = getattr(request.user, 'profile', None)
    if not can_export_civil_servants_to_excel(request.user, profile):
        messages.error(
            request,
            "ការិយាល័យជំនាញ និងខណ្ឌ មិនត្រូវបានអនុញ្ញាតឱ្យទាញចេញជា Excel ឡើយ! លើកលែងតែការិយាល័យរដ្ឋបាល-បុគ្គលិក ឬមានការអនុញ្ញាតដោយ ADMIN។"
        )
        return HttpResponseForbidden(
            "<div style='font-family: Khmer OS Battambang, sans-serif; text-align: center; margin-top: 80px;'>"
            "<h2 style='color: #dc2626;'>⛔ គ្មានសិទ្ធិទាញយកឯកសារជា Excel ឡើយ</h2>"
            "<p style='font-size: 16px; color: #475569;'>ការិយាល័យជំនាញ និងខណ្ឌ មិនត្រូវបានអនុញ្ញាតឱ្យទាញចេញជា Excel ដាច់ខាត លើកលែងតែការិយាល័យរដ្ឋបាល-បុគ្គលិក ឬមានការអនុញ្ញាតដោយ ADMIN។</p>"
            "<a href='javascript:history.back()' style='display: inline-block; margin-top: 15px; padding: 8px 20px; background: #2563eb; color: #fff; text-decoration: none; border-radius: 20px;'>ត្រឡប់ក្រោយ</a>"
            "</div>"
        )

    export_format = request.GET.get('format', 'e1').strip().lower()
    if export_format == 'e2' and not can_access_officer_e2_report(request.user, profile):
        messages.error(
            request,
            "ទម្រង់ E-2 (ថ្នាក់ខេត្ត) ត្រូវបានបិទសម្រាប់ការិយាល័យជំនាញ និងខណ្ឌ! សម្រាប់តែការិយាល័យរដ្ឋបាល-បុគ្គលិក ឬ ADMIN ប៉ុណ្ណោះ។"
        )
        return HttpResponseForbidden(
            "<div style='font-family: Khmer OS Battambang, sans-serif; text-align: center; margin-top: 80px;'>"
            "<h2 style='color: #dc2626;'>⛔ គ្មានសិទ្ធិទាញយកទម្រង់ E-2 ឡើយ</h2>"
            "<p style='font-size: 16px; color: #475569;'>ទម្រង់ E-2 (ថ្នាក់ខេត្ត) ត្រូវបានបិទសម្រាប់ការិយាល័យជំនាញ និងខណ្ឌ។ អនុញ្ញាតសម្រាប់តែការិយាល័យរដ្ឋបាល-បុគ្គលិក ឬ ADMIN ប៉ុណ្ណោះ។</p>"
            "<a href='javascript:history.back()' style='display: inline-block; margin-top: 15px; padding: 8px 20px; background: #2563eb; color: #fff; text-decoration: none; border-radius: 20px;'>ត្រឡប់ក្រោយ</a>"
            "</div>"
        )

    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    search_q = request.GET.get('q', '').strip()
    dept_filter = request.GET.get('department', '').strip()
    gender_filter = request.GET.get('gender', '').strip()
    marital_filter = request.GET.get('marital_status', '').strip()

    queryset = CivilServantProfile.objects.select_related('department').all()
    selected_dept = None
    if not is_admin_or_lead:
        if dept:
            queryset = queryset.filter(department=dept)
            selected_dept = dept
        else:
            queryset = queryset.none()
    elif dept_filter:
        queryset = queryset.filter(department_id=dept_filter)
        selected_dept = Department.objects.filter(id=dept_filter).first()

    if search_q:
        q_arabic = to_arabic_digits(search_q)
        queryset = queryset.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_last_name__icontains=search_q) |
            Q(latin_first_name__icontains=search_q) |
            Q(officer_id_number__icontains=search_q) |
            Q(officer_id_number__icontains=q_arabic) |
            Q(national_id_number__icontains=search_q) |
            Q(national_id_number__icontains=q_arabic) |
            Q(phone__icontains=search_q) |
            Q(phone__icontains=q_arabic) |
            Q(email__icontains=search_q) |
            Q(current_rank_and_step__icontains=search_q) |
            Q(current_position_title__icontains=search_q)
        )

    if gender_filter:
        queryset = queryset.filter(gender=gender_filter)
    if marital_filter:
        queryset = queryset.filter(marital_status=marital_filter)

    from .models import officer_sort_key
    officers_list = list(queryset)
    officers_list.sort(key=officer_sort_key)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    date_suffix = datetime.now().strftime('%Y%m%d_%H%M%S')

    if export_format in ['e2', 'provincial', 'mcs', '2']:
        wb = _build_e2_provincial_workbook(officers_list)
        filename = f"officers_E2_provincial_mcs_{date_suffix}.xlsx"
    else:
        wb = _build_e1_ministry_workbook(officers_list, department=selected_dept)
        filename = f"officers_E1_ministry_{date_suffix}.xlsx"

    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _paginate_preview_items(all_items, first_page_cap=20, mid_page_cap=26, last_page_cap=18, single_page_cap=14):
    """
    រៀបចំទំព័រ A4 Landscape ដាច់ៗពីគ្នាសម្រាប់ PDF Preview
    """
    pages = []
    if len(all_items) <= single_page_cap:
        pages.append({
            'page_num': 1,
            'is_first': True,
            'is_last': True,
            'items': all_items,
        })
        return pages

    # Page 1 (has full Header and Table Header)
    p1_take = min(len(all_items), first_page_cap)
    p1_items = all_items[:p1_take]
    pages.append({
        'page_num': 1,
        'is_first': True,
        'is_last': False,
        'items': p1_items,
    })

    remaining = all_items[p1_take:]
    page_num = 2
    while remaining:
        if len(remaining) <= last_page_cap:
            pages.append({
                'page_num': page_num,
                'is_first': False,
                'is_last': True,
                'items': remaining,
            })
            break
        elif len(remaining) <= (mid_page_cap + last_page_cap):
            half = len(remaining) // 2
            take = max(half, len(remaining) - last_page_cap)
            pages.append({
                'page_num': page_num,
                'is_first': False,
                'is_last': False,
                'items': remaining[:take],
            })
            remaining = remaining[take:]
            page_num += 1
        else:
            pages.append({
                'page_num': page_num,
                'is_first': False,
                'is_last': False,
                'items': remaining[:mid_page_cap],
            })
            remaining = remaining[mid_page_cap:]
            page_num += 1

    total_pages = len(pages)
    for p in pages:
        p['total_pages'] = total_pages
    return pages


@login_required
def officer_preview_pdf_e1(request):
    """
    មើលជា PDF / Print Preview សម្រាប់ទម្រង់ E-1 (បញ្ជូនទៅក្រសួងកសិកម្ម)
    """
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    search_q = request.GET.get('q', '').strip()
    dept_filter = request.GET.get('department', '').strip()
    gender_filter = request.GET.get('gender', '').strip()
    marital_filter = request.GET.get('marital_status', '').strip()

    queryset = CivilServantProfile.objects.select_related('department').all()
    selected_dept = None
    if not is_admin_or_lead:
        if dept:
            queryset = queryset.filter(department=dept)
            selected_dept = dept
        else:
            queryset = queryset.none()
    elif dept_filter:
        queryset = queryset.filter(department_id=dept_filter)
        selected_dept = Department.objects.filter(id=dept_filter).first()

    is_specialized = _is_specialized_department(selected_dept)
    office_head_title = _get_department_head_title(selected_dept) if (is_specialized and selected_dept) else "ប្រធានការិយាល័យរដ្ឋបាល-បុគ្គលិក"

    if search_q:
        q_arabic = to_arabic_digits(search_q)
        queryset = queryset.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_last_name__icontains=search_q) |
            Q(latin_first_name__icontains=search_q) |
            Q(officer_id_number__icontains=search_q) |
            Q(officer_id_number__icontains=q_arabic) |
            Q(national_id_number__icontains=search_q) |
            Q(national_id_number__icontains=q_arabic) |
            Q(phone__icontains=search_q) |
            Q(phone__icontains=q_arabic) |
            Q(email__icontains=search_q) |
            Q(current_rank_and_step__icontains=search_q) |
            Q(current_position_title__icontains=search_q)
        )

    if gender_filter:
        queryset = queryset.filter(gender=gender_filter)
    if marital_filter:
        queryset = queryset.filter(marital_status=marital_filter)

    from .models import officer_sort_key
    officers_list = list(queryset)
    officers_list.sort(key=officer_sort_key)

    leadership_officers = []
    dept_map = {}

    for o in officers_list:
        deg, skill = _extract_officer_degree_and_skill(o)
        o.e1_degree = deg
        o.e1_skill = skill
        o.e1_dob = _format_khmer_date_standard(o.dob)
        o.e1_start_date = _format_khmer_date_standard(o.civil_service_start_date)
        o.e1_gender_kh = 'ប្រុស' if o.gender == 'MALE' or str(o.gender).upper() in ['M', 'ប្រុស', 'ប'] else 'ស្រី'

        pos = (o.current_position_title or '').strip()
        dept_name = o.department.name_kh if o.department else 'ផ្សេងៗ'
        is_lead = (o.department and (o.department.code in ['LEAD', 'LEADERSHIP'] or 'ថ្នាក់ដឹកនាំ' in o.department.name_kh)) or \
                  any(lead_title in pos for lead_title in ['ប្រធានមន្ទីរ', 'អនុប្រធានមន្ទីរ'])
        
        if is_lead:
            leadership_officers.append(o)
        else:
            if dept_name not in dept_map:
                dept_map[dept_name] = []
            dept_map[dept_name].append(o)

    # Build sequential items for A4 pages
    table_items = []
    if leadership_officers:
        table_items.append({'is_header': True, 'title': 'ថ្នាក់ដឹកនាំមន្ទីរ'})
        for idx, o in enumerate(leadership_officers, 1):
            table_items.append({'is_header': False, 'officer': o, 'num': idx})

    for dept_name, off_list in dept_map.items():
        table_items.append({'is_header': True, 'title': dept_name})
        for idx, o in enumerate(off_list, 1):
            table_items.append({'is_header': False, 'officer': o, 'num': idx})

    pages = _paginate_preview_items(table_items, first_page_cap=23, mid_page_cap=24, last_page_cap=14, single_page_cap=12)

    now = datetime.now()
    month_kh = KHMER_MONTHS_NAMES[now.month] if 1 <= now.month <= 12 else str(now.month)
    year_kh = _to_khmer_digits(str(now.year))
    cadre_stats = _calculate_e1_cadre_stats(officers_list)
    lunar_year_text = _get_khmer_lunar_year_info(now.year)

    context = {
        'pages': pages,
        'total_count': len(officers_list),
        'female_count': cadre_stats['female_total'],
        'cadre_stats': cadre_stats,
        'lunar_year_text': lunar_year_text,
        'month_kh': month_kh,
        'year_kh': year_kh,
        'today': date.today(),
        'query_params': request.GET.urlencode(),
        'can_export_officer_excel': can_export_civil_servants_to_excel(request.user, profile),
        'is_specialized': is_specialized,
        'selected_department': selected_dept,
        'office_head_title': office_head_title,
    }
    return render(request, 'dms/officer_preview_e1_pdf.html', context)


@login_required
def officer_preview_pdf_e2(request):
    """
    មើលជា PDF / Print Preview សម្រាប់ទម្រង់ E-2 (បញ្ជូនទៅមន្ទីរមុខងារសាធារណៈខេត្ត)
    បិទសម្រាប់ការិយាល័យជំនាញ និងខណ្ឌ (Specialized offices & Cantons) លើកលែងតែរដ្ឋបាល-បុគ្គលិក ឬ ADMIN
    """
    profile = getattr(request.user, 'profile', None)
    if not can_access_officer_e2_report(request.user, profile):
        messages.error(
            request,
            "ទម្រង់ E-2 (ថ្នាក់ខេត្ត) ត្រូវបានបិទសម្រាប់ការិយាល័យជំនាញ និងខណ្ឌ! សម្រាប់តែការិយាល័យរដ្ឋបាល-បុគ្គលិក ឬ ADMIN ប៉ុណ្ណោះ។"
        )
        return HttpResponseForbidden(
            "<div style='font-family: Khmer OS Battambang, sans-serif; text-align: center; margin-top: 80px;'>"
            "<h2 style='color: #dc2626;'>⛔ ទម្រង់ E-2 ត្រូវបានបិទ</h2>"
            "<p style='font-size: 16px; color: #475569;'>ទម្រង់ E-2 (ថ្នាក់ខេត្ត) ត្រូវបានបិទសម្រាប់ការិយាល័យជំនាញ និងខណ្ឌ។ អនុញ្ញាតសម្រាប់តែការិយាល័យរដ្ឋបាល-បុគ្គលិក ឬ ADMIN ប៉ុណ្ណោះ។</p>"
            "<a href='javascript:history.back()' style='display: inline-block; margin-top: 15px; padding: 8px 20px; background: #2563eb; color: #fff; text-decoration: none; border-radius: 20px;'>ត្រឡប់ក្រោយ</a>"
            "</div>"
        )

    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    search_q = request.GET.get('q', '').strip()
    dept_filter = request.GET.get('department', '').strip()
    gender_filter = request.GET.get('gender', '').strip()
    marital_filter = request.GET.get('marital_status', '').strip()

    queryset = CivilServantProfile.objects.select_related('department').all()
    if not is_admin_or_lead:
        if dept:
            queryset = queryset.filter(department=dept)
        else:
            queryset = queryset.none()
    elif dept_filter:
        queryset = queryset.filter(department_id=dept_filter)

    if search_q:
        q_arabic = to_arabic_digits(search_q)
        queryset = queryset.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_last_name__icontains=search_q) |
            Q(latin_first_name__icontains=search_q) |
            Q(officer_id_number__icontains=search_q) |
            Q(officer_id_number__icontains=q_arabic) |
            Q(national_id_number__icontains=search_q) |
            Q(national_id_number__icontains=q_arabic) |
            Q(phone__icontains=search_q) |
            Q(phone__icontains=q_arabic) |
            Q(email__icontains=search_q) |
            Q(current_rank_and_step__icontains=search_q) |
            Q(current_position_title__icontains=search_q)
        )

    if gender_filter:
        queryset = queryset.filter(gender=gender_filter)
    if marital_filter:
        queryset = queryset.filter(marital_status=marital_filter)

    from .models import officer_sort_key
    officers_list = list(queryset)
    officers_list.sort(key=officer_sort_key)

    for o in officers_list:
        deg, skill = _extract_officer_degree_and_skill(o)
        prom_date, legal_doc, deg_prom_date, deg_doc_no = _extract_officer_promotion_info(o)
        o.e2_gender_code = 'ប' if o.gender == 'MALE' or str(o.gender).upper() in ['M', 'ប្រុស', 'ប'] else 'ស'
        o.e2_dob = _format_khmer_date_standard(o.dob)
        o.e2_start_date = _format_khmer_date_standard(o.civil_service_start_date)
        o.e2_degree = deg
        o.e2_skill = skill
        o.e2_prom_date = prom_date
        o.e2_legal_doc = legal_doc
        o.e2_deg_prom_date = deg_prom_date
        o.e2_deg_doc_no = deg_doc_no

    # Build sequential items for E-2
    e2_items = []
    for idx, o in enumerate(officers_list, 1):
        e2_items.append({'officer': o, 'num': idx})

    pages = _paginate_preview_items(e2_items, first_page_cap=22, mid_page_cap=24, last_page_cap=22, single_page_cap=12)

    now = datetime.now()
    month_kh = KHMER_MONTHS_NAMES[now.month] if 1 <= now.month <= 12 else str(now.month)
    year_kh = _to_khmer_digits(str(now.year))

    context = {
        'pages': pages,
        'total_count': len(officers_list),
        'female_count': sum(1 for o in officers_list if o.gender == 'FEMALE'),
        'male_count': sum(1 for o in officers_list if o.gender == 'MALE'),
        'month_kh': month_kh,
        'year_kh': year_kh,
        'today': date.today(),
        'query_params': request.GET.urlencode(),
        'can_export_officer_excel': can_export_civil_servants_to_excel(request.user, profile),
    }
    return render(request, 'dms/officer_preview_e2_pdf.html', context)




# =========================================================================
# 📊 HR Tracking & Forecast System (ការតាមដាននិវត្តន៍ ដំឡើងថ្នាក់ និងមេដាយ)
# =========================================================================

def parse_date_flexible(val):
    if not val:
        return None
    val_str = str(val).strip()
    
    KHMER_DIGITS_REV = {'០': '0', '១': '1', '២': '2', '៣': '3', '៤': '4', '៥': '5', '៦': '6', '៧': '7', '៨': '8', '៩': '9'}
    for kh, ar in KHMER_DIGITS_REV.items():
        val_str = val_str.replace(kh, ar)
        
    KHMER_MONTHS_REV = {
        'មករា': 1, 'កុម្ភៈ': 2, 'មីនា': 3, 'មេសា': 4, 'ឧសភា': 5, 'មិថុនា': 6,
        'កក្កដា': 7, 'សីហា': 8, 'កញ្ញា': 9, 'តុលា': 10, 'វិច្ឆិកា': 11, 'ធ្នូ': 12
    }
    
    import re, datetime
    m_kh = re.search(r'(\d{1,2})\s*ខែ\s*([^\s]+)\s*ឆ្នាំ\s*(\d{4})', val_str)
    if m_kh:
        day = int(m_kh.group(1))
        month_name = m_kh.group(2).strip()
        year = int(m_kh.group(3))
        month = KHMER_MONTHS_REV.get(month_name, 1)
        try:
            return datetime.date(year, month, day)
        except ValueError:
            return datetime.date(year, month, 1)
            
    m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', val_str)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime.date(year, month, day)
        except ValueError:
            return None
            
    m2 = re.search(r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})', val_str)
    if m2:
        year, month, day = int(m2.group(1)), int(m2.group(2)), int(m2.group(3))
        try:
            return datetime.date(year, month, day)
        except ValueError:
            return None
            
    m3 = re.search(r'(\d{4})', val_str)
    if m3:
        year = int(m3.group(1))
        if 1940 <= year <= 2060:
            return datetime.date(year, 1, 1)
            
    return None


def is_c1_rank(rank_str):
    """
    ពិនិត្យថាតើមន្ត្រីស្ថិតក្នុងកម្រិតថ្នាក់ «គ.១» (កម្រិតកំពូលនៃក្របខ័ណ្ឌ គ) ដែរឬទេ
    """
    if not rank_str:
        return False
    clean = str(rank_str).strip().replace(' ', '').replace('\u200b', '').replace('\ufeff', '')
    if clean in ['គ.១', 'គ.1', 'គ១', 'គ1', 'C.1', 'C1', 'c.1', 'c1', 'កម្រិតគ.១', 'កម្រិតគ.1']:
        return True
    if clean.startswith('គ.១') or clean.startswith('គ.1') or clean.startswith('គ១'):
        return True
    return False


def is_b1_1_rank(rank_str):
    """
    ពិនិត្យថាតើមន្ត្រីស្ថិតក្នុងកម្រិតថ្នាក់ «ខ.១.១» (កម្រិតកំពូលនៃក្របខ័ណ្ឌ ខ) ដែរឬទេ
    """
    if not rank_str:
        return False
    clean = str(rank_str).strip().replace(' ', '').replace('\u200b', '').replace('\ufeff', '')
    if clean in ['ខ.១.១', 'ខ.1.1', 'ខ.1.១', 'ខ.១.1', 'ខ១.១', 'ខ1.1', 'ខ១១', 'ខ11', 'B.1.1', 'B1.1', 'b.1.1', 'b1.1', 'កម្រិតខ.១.១', 'កម្រិតខ.1.1']:
        return True
    if clean.startswith('ខ.១.១') or clean.startswith('ខ.1.1') or clean.startswith('ខ១.១') or clean.startswith('ខ1.1'):
        return True
    return False


def is_max_framework_ceiling_rank(rank_str):
    """
    ពិនិត្យថាតើមន្ត្រីស្ថិតក្នុងកម្រិតកំពូលនៃក្របខ័ណ្ឌ (គ.១ ឬ ខ.១.១)
    """
    return is_c1_rank(rank_str) or is_b1_1_rank(rank_str)


def get_next_rank_suggestion(current_rank):
    if is_c1_rank(current_rank):
        return 'ប្រឡងប្តូរក្របខ័ណ្ឌ (ខ)'
    if is_b1_1_rank(current_rank):
        return 'ប្រឡងប្តូរក្របខ័ណ្ឌ (ក)'
    NEXT_RANK_MAP = {
        'គ.១០': 'គ.៩', 'គ.៩': 'គ.៨', 'គ.៨': 'គ.៧', 'គ.៧': 'គ.៦',
        'គ.៦': 'គ.៥', 'គ.៥': 'គ.៤', 'គ.៤': 'គ.៣', 'គ.៣': 'គ.២', 'គ.២': 'គ.១',
        'ខ.៣.៤': 'ខ.៣.៣', 'ខ.៣.៣': 'ខ.៣.២', 'ខ.៣.២': 'ខ.៣.១', 'ខ.៣.១': 'ខ.២.៤',
        'ខ.២.៤': 'ខ.២.៣', 'ខ.២.៣': 'ខ.២.២', 'ខ.២.២': 'ខ.២.១', 'ខ.២.១': 'ខ.១.៦',
        'ខ.១.៦': 'ខ.១.៥', 'ខ.១.៥': 'ខ.១.៤', 'ខ.១.៤': 'ខ.១.៣', 'ខ.១.៣': 'ខ.១.២', 'ខ.១.២': 'ខ.១.១',
        'ក.៣.៤': 'ក.៣.៣', 'ក.៣.៣': 'ក.៣.២', 'ក.៣.២': 'ក.៣.១', 'ក.៣.១': 'ក.២.៤',
        'ក.២.៤': 'ក.២.៣', 'ក.២.៣': 'ក.២.២', 'ក.២.២': 'ក.២.១', 'ក.២.១': 'ក.១.៦',
        'ក.១.៦': 'ក.១.៥', 'ក.១.៥': 'ក.១.៤', 'ក.១.៤': 'ក.១.៣', 'ក.១.៣': 'ក.១.២', 'ក.១.២': 'ក.១.១',
    }
    return NEXT_RANK_MAP.get(current_rank, '-')


def process_officer_retirement(officer, today=None):
    import datetime
    if today is None:
        today = datetime.date.today()
        
    birth_date = parse_date_flexible(officer.dob)
    if not birth_date:
        return {
            'officer': officer,
            'birth_date': None,
            'birth_date_str': officer.dob or '-',
            'age': None,
            'retirement_date': None,
            'retirement_date_str': '-',
            'retirement_year': None,
            'status': 'UNKNOWN',
            'status_label': 'មិនទាន់មានថ្ងៃខែឆ្នាំកំណើត',
            'status_badge': 'secondary',
            'days_left': 99999,
            'years_left': 999,
        }
        
    retirement_date = datetime.date(birth_date.year + 60, birth_date.month, min(birth_date.day, 28))
    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
    
    days_left = (retirement_date - today).days
    years_left = days_left / 365.25
    
    if days_left <= 0:
        status = 'RETIRED'
        status_label = 'ចូលនិវត្តន៍រួចរាល់'
        status_badge = 'dark'
    elif retirement_date.year == today.year:
        status = 'THIS_YEAR'
        status_label = 'ត្រូវចូលនិវត្តន៍ក្នុងឆ្នាំនេះ'
        status_badge = 'danger'
    elif retirement_date.year == today.year + 1:
        status = 'NEXT_YEAR'
        status_label = 'ត្រូវចូលនិវត្តន៍ឆ្នាំក្រោយ (១ ឆ្នាំទៀត)'
        status_badge = 'warning'
    elif retirement_date.year == today.year + 2:
        status = 'IN_2_YEARS'
        status_label = 'ត្រូវចូលនិវត្តន៍ក្នុងរយៈពេល ២ ឆ្នាំ'
        status_badge = 'info'
    else:
        status = 'ACTIVE'
        status_label = f'នៅសល់ {max(1, int(years_left))} ឆ្នាំទៀត'
        status_badge = 'success'
        
    return {
        'officer': officer,
        'birth_date': birth_date,
        'birth_date_str': birth_date.strftime('%d/%m/%Y'),
        'age': age,
        'retirement_date': retirement_date,
        'retirement_date_str': retirement_date.strftime('%d/%m/%Y'),
        'retirement_year': retirement_date.year,
        'status': status,
        'status_label': status_label,
        'status_badge': status_badge,
        'days_left': days_left,
        'years_left': round(years_left, 1),
    }


def process_officer_promotion(officer, today=None, target_year=None, rejected_req=None):
    """
    គណនាវេនស្នើសុំដំឡើងឋានន្តរស័ក្តិ និងថ្នាក់ (តាមវេនជ្រើសរើស / អតីតភាពការងារ ២ ឆ្នាំ)
    ច្បាប់ និងបទដ្ឋានគតិយុត្ត៖
    1. កាលបរិច្ឆេទដល់វេនជ្រើសរើសឡើងថ្នាក់ គឺគិតត្រឹមថ្ងៃទី ១៣ ខែមេសា នៃឆ្នាំនីមួយៗ (Khmer New Year Cycle)
    2. មន្ត្រីដែលឡើងឋានន្តរស័ក្តិតាមសញ្ញាបត្រ គឺមិនប៉ះពាល់ ឬកាត់ផ្តាច់វេនជ្រើសរើសឡើយ
    3. ករណីមន្ត្រីធ្លាក់ពីឆ្នាំមុន ត្រូវបានដាក់បញ្ចូលក្នុងវេនអតីតភាពស្វ័យប្រវត្តិសម្រាប់ឆ្នាំបន្ទាប់
    4. ករណីមន្ត្រីមាន គ.១ ឬ ខ.១.១ (កម្រិតកំពូលនៃក្របខ័ណ្ឌ) មិនមានមុខងារស្នើសុំដំឡើងថ្នាក់តាមវេនជ្រើសរើសឡើយ លុះត្រាតែមានការស្នើសុំប្រឡងផ្ទៃក្នុងថ្នាក់ក្រសួងប្តូរក្របខ័ណ្ឌតែមួយគត់
    """
    import datetime
    if today is None:
        today = datetime.date.today()
        
    if target_year is None:
        target_year = today.year
        
    last_promo_date = None
    
    # 🎯 គិតតែលើកាលបរិច្ឆេទដំឡើងតាមអតីតភាព/វេនជ្រើសរើសប៉ុណ្ណោះ (ការឡើងតាមសញ្ញាបត្រ មិនប៉ះពាល់ដល់វេនជ្រើសរើសឡើយ)
    if officer.promotions_by_seniority:
        for p in officer.promotions_by_seniority:
            d = parse_date_flexible(p.get('effective_date'))
            if d and (last_promo_date is None or d > last_promo_date):
                last_promo_date = d
                
    if not last_promo_date:
        last_promo_date = parse_date_flexible(officer.civil_service_permanent_date) or parse_date_flexible(officer.civil_service_start_date)
        
    has_degree_promo = bool(officer.promotions_by_degree)
    degree_promos_count = len(officer.promotions_by_degree or [])
    is_c1 = is_c1_rank(officer.current_rank_and_step)
    is_b1_1 = is_b1_1_rank(officer.current_rank_and_step)
    is_max_rank = is_c1 or is_b1_1

    # គណនាអតីតភាពគិតត្រឹមថ្ងៃទី ១៣ មេសា នៃ target_year
    target_cutoff_date = datetime.date(target_year, 4, 13)
    if last_promo_date:
        days_since_target = (target_cutoff_date - last_promo_date).days
        years_since = max(0.0, days_since_target / 365.25)
        months_since = max(0, int(days_since_target / 30.4375))
    else:
        years_since = 0.0
        months_since = 0

    # 🎯 រៀបចំប្រវត្តិឡើងថ្នាក់ និងឋានន្តរស័ក្តិកន្លងមក (Past Promotions History)
    raw_sen_promos = officer.promotions_by_seniority or []
    raw_deg_promos = officer.promotions_by_degree or []
    
    past_promotions = []
    
    # 1. បន្ថែមប្រវត្តិឡើងតាមអតីតភាព/វេនជ្រើសរើស
    for p in raw_sen_promos:
        if isinstance(p, dict):
            d_eff = p.get('effective_date', '') or '-'
            import re
            s_arabic = to_arabic_digits(str(d_eff))
            y_match = re.search(r'(19\d{2}|20\d{2})', s_arabic)
            yr = y_match.group(1) if y_match else ''
            inst = f"{p.get('ministry', '')} {p.get('department', '')}".strip() or '-'
            past_promotions.append({
                'date': d_eff,
                'year': yr or '-',
                'type_category': 'SENIORITY',
                'type_label': p.get('promo_type') or 'តាមវេនជ្រើសរើស/អតីតភាព',
                'old_rank_step': p.get('old_rank_step') or '-',
                'new_rank_step': p.get('new_rank_step') or '-',
                'institution': inst,
                'degree_info': '',
            })

    # 2. បន្ថែមប្រវត្តិឡើងតាមសញ្ញាបត្រ
    for d in raw_deg_promos:
        if isinstance(d, dict):
            d_eff = d.get('effective_date', '') or '-'
            import re
            s_arabic = to_arabic_digits(str(d_eff))
            y_match = re.search(r'(19\d{2}|20\d{2})', s_arabic)
            yr = y_match.group(1) if y_match else ''
            deg_title = d.get('degree', '') or ''
            sch = d.get('school', '') or ''
            deg_full = f"{deg_title} ({sch})" if (sch and deg_title) else (deg_title or sch or '-')
            past_promotions.append({
                'date': d_eff,
                'year': yr or '-',
                'type_category': 'DEGREE',
                'type_label': 'ដំឡើងតាមសញ្ញាបត្រ',
                'old_rank_step': d.get('old_rank_step') or '-',
                'new_rank_step': d.get('new_rank_step') or '-',
                'institution': d.get('location') or '-',
                'degree_info': deg_full,
            })

    # 3. បន្ថែមកាលបរិច្ឆេទតាំងស៊ប់ដំបូងក្នុងក្របខ័ណ្ឌ (Initial Permanent Date)
    if officer.civil_service_permanent_date:
        d_perm = officer.civil_service_permanent_date
        import re
        s_arabic = to_arabic_digits(str(d_perm))
        y_match = re.search(r'(19\d{2}|20\d{2})', s_arabic)
        yr = y_match.group(1) if y_match else ''
        past_promotions.append({
            'date': d_perm,
            'year': yr or '-',
            'type_category': 'INITIAL',
            'type_label': 'តាំងស៊ប់ក្នុងក្របខ័ណ្ឌរដ្ឋ',
            'old_rank_step': 'សាកល្បង',
            'new_rank_step': officer.current_rank_and_step if not (raw_sen_promos or raw_deg_promos) else (raw_sen_promos[0].get('old_rank_step') if raw_sen_promos else '-'),
            'institution': officer.department.name_kh if officer.department else 'ក្រសួងកសិកម្ម',
            'degree_info': officer.framework_name or '',
        })

    # 🎯 ករណីពិសេស៖ មន្ត្រីថ្នាក់ គ.១ ឬ ខ.១.១ (កម្រិតកំពូលនៃក្របខ័ណ្ឌ គ ឬ ខ)
    # មិនមានមុខងារស្នើសុំតាមវេនជ្រើសរើសឡើយ លុះត្រាតែមានការស្នើសុំប្រឡងផ្ទៃក្នុងថ្នាក់ក្រសួងប្តូរក្របខ័ណ្ឌតែមួយគត់
    if is_max_rank:
        target_fw = 'ខ' if is_c1 else 'ក'
        rank_name = 'គ.១' if is_c1 else 'ខ.១.១'
        fw_name = 'គ' if is_c1 else 'ខ'
        return {
            'officer': officer,
            'last_promo_date': last_promo_date,
            'last_promo_date_str': last_promo_date.strftime('%d/%m/%Y') if last_promo_date else '-',
            'years_in_rank': round(years_since, 1),
            'months_in_rank': months_since,
            'due_date': None,
            'due_date_str': 'ប្រឡងប្តូរក្របខ័ណ្ឌ',
            'due_cycle_year': None,
            'status': 'MAX_RANK_EXAM_ONLY',
            'status_label': f'🏆 ដល់កម្រិត {rank_name} (ប្រឡងប្តូរក្របខ័ណ្ឌ)',
            'status_badge': 'secondary',
            'next_suggested_rank': f'ប្រឡងប្តូរក្របខ័ណ្ឌ ({target_fw})',
            'has_degree_promo': has_degree_promo,
            'degree_promos_count': degree_promos_count,
            'has_previous_rejected': False,
            'previous_rejected_year': None,
            'previous_rejected_reason': '',
            'is_carried_over': False,
            'is_max_rank': True,
            'is_max_rank_c': is_c1,
            'is_max_rank_b': is_b1_1,
            'max_rank_name': rank_name,
            'max_framework_name': fw_name,
            'target_framework_name': target_fw,
            'can_regular_promote': False,
            'exam_only': True,
            'past_promotions_count': len(past_promotions),
            'past_promotions': past_promotions,
        }

    if not last_promo_date:
        return {
            'officer': officer,
            'last_promo_date': None,
            'last_promo_date_str': '-',
            'years_in_rank': 0,
            'months_in_rank': 0,
            'due_date': None,
            'due_date_str': '-',
            'due_cycle_year': None,
            'status': 'UNKNOWN',
            'status_label': 'មិនទាន់មានកាលបរិច្ឆេទតាំងស៊ប់/ដំឡើងថ្នាក់',
            'status_badge': 'secondary',
            'next_suggested_rank': get_next_rank_suggestion(officer.current_rank_and_step),
            'has_degree_promo': has_degree_promo,
            'degree_promos_count': degree_promos_count,
            'has_previous_rejected': False,
            'previous_rejected_year': None,
            'previous_rejected_reason': '',
            'is_carried_over': False,
            'is_max_rank': False,
            'is_max_rank_c': False,
            'is_max_rank_b': False,
            'max_rank_name': '',
            'max_framework_name': '',
            'target_framework_name': '',
            'can_regular_promote': True,
            'exam_only': False,
            'past_promotions_count': len(past_promotions),
            'past_promotions': past_promotions,
        }
        
    # 🎯 គណនាថ្ងៃដល់វេនផ្លូវការ៖ គិតត្រឹមថ្ងៃទី ១៣ ខែមេសា នៃឆ្នាំដែលគ្រប់អតីតភាព ២ ឆ្នាំ
    if last_promo_date <= datetime.date(last_promo_date.year, 4, 13):
        due_cycle_year = last_promo_date.year + 2
    else:
        due_cycle_year = last_promo_date.year + 3

    due_date = datetime.date(due_cycle_year, 4, 13)

    # 🔍 ពិនិត្យប្រវត្តិធ្លាក់/មិនទាន់អនុម័តពីឆ្នាំមុនៗ
    has_previous_rejected = False
    previous_rejected_year = None
    previous_rejected_reason = ''
    if rejected_req and rejected_req.request_year < target_year:
        has_previous_rejected = True
        previous_rejected_year = rejected_req.request_year
        previous_rejected_reason = rejected_req.ministry_decision_notes or rejected_req.meeting_decision or rejected_req.admin_notes or 'ធ្លាក់/មិនទាន់អនុម័ត'

    is_carried_over = False

    if due_cycle_year == target_year:
        status = 'DUE_NEW'
        status_label = f'🔔 ដល់វេនថ្មី (១៣ មេសា {target_year})'
        status_badge = 'danger'
    elif due_cycle_year < target_year:
        is_carried_over = True
        if has_previous_rejected:
            status = 'CARRIED_OVER'
            status_label = f'⚠️ ធ្លាក់ឆ្នាំ {previous_rejected_year} (ចូលវេនអតីតភាព {target_year})'
            status_badge = 'warning'
        else:
            status = 'CARRIED_OVER'
            status_label = f'⚠️ ដល់វេនអតីតភាពសេសសល់ ({round(years_since, 1)} ឆ្នាំ)'
            status_badge = 'warning'
    elif due_cycle_year == target_year + 1:
        status = 'DUE_SOON'
        status_label = f'⏳ ជិតដល់វេន (១៣ មេសា {due_cycle_year})'
        status_badge = 'primary'
    else:
        status = 'NORMAL'
        status_label = f'✅ មិនទាន់ដល់វេន (១៣ មេសា {due_cycle_year})'
        status_badge = 'secondary'
        
    return {
        'officer': officer,
        'last_promo_date': last_promo_date,
        'last_promo_date_str': last_promo_date.strftime('%d/%m/%Y'),
        'years_in_rank': round(years_since, 1),
        'months_in_rank': months_since,
        'due_date': due_date,
        'due_date_str': f"13/04/{due_date.year}",
        'due_cycle_year': due_cycle_year,
        'status': status,
        'status_label': status_label,
        'status_badge': status_badge,
        'next_suggested_rank': get_next_rank_suggestion(officer.current_rank_and_step),
        'has_degree_promo': has_degree_promo,
        'degree_promos_count': degree_promos_count,
        'has_previous_rejected': has_previous_rejected,
        'previous_rejected_year': previous_rejected_year,
        'previous_rejected_reason': previous_rejected_reason,
        'is_carried_over': is_carried_over,
        'is_max_rank': False,
        'is_max_rank_c': False,
        'is_max_rank_b': False,
        'max_rank_name': '',
        'max_framework_name': '',
        'target_framework_name': '',
        'can_regular_promote': True,
        'exam_only': False,
        'past_promotions_count': len(past_promotions),
        'past_promotions': past_promotions,
    }


def process_officer_medals(officer, today=None):
    """
    គណនា និងណែនាំមេដាយ/គ្រឿងឥស្សរិយយសសម្រាប់មន្ត្រីរាជការ
    គោលការណ៍៖ 
    - ការស្នើសុំ និងផ្តល់មេដាយ/គ្រឿងឥស្សរិយយស មិនកំណត់កំហិតលើអតីតភាពការងារឡើយ
    - អតីតភាពការងារ (ឆ្នាំបម្រើការងារ) គឺសម្រាប់ជាការណែនាំកម្រិតមេដាយស្តង់ដារ
    - ការិយាល័យជំនាញអាចស្នើសុំជូនមន្ត្រីគ្រប់រូបផ្អែកលើគុណបំណាច់ ស្នាដៃ និងការរួមចំណែក
    - ការសម្រេចចុងក្រោយគឺស្ថិតលើការវាយតម្លៃ និងឯកភាពរបស់គណៈកម្មការមន្ទីរ
    """
    import datetime
    if today is None:
        today = datetime.date.today()
        
    start_date = parse_date_flexible(officer.civil_service_start_date)
    if not start_date:
        start_date = parse_date_flexible(officer.civil_service_permanent_date)

    if not start_date:
        years_served = 0.0
        start_date_str = '-'
    else:
        days_served = (today - start_date).days
        years_served = max(0.0, days_served / 365.25)
        start_date_str = start_date.strftime('%d/%m/%Y')
        
    # ការណែនាំប្រភេទមេដាយតាមកម្រិតអតីតភាពស្តង់ដារ (សម្រាប់ជាឯកសារយោង)
    if years_served >= 30:
        recommended_medal = '👑 គ្រឿងឥស្សរិយយស (ថ្នាក់មហាសេរីវឌ្ឍន៍ / មហាធិបឌិន្ទ)'
        medal_tier = 5
        badge_color = 'danger'
    elif years_served >= 25:
        recommended_medal = '🎖️ គ្រឿងឥស្សរិយយស (ថ្នាក់ធិបឌិន្ទ / សេនា / មុនីសារាភ័ណ្ឌ)'
        medal_tier = 4
        badge_color = 'warning'
    elif years_served >= 20:
        recommended_medal = '🥇 មេដាយការងារ (ថ្នាក់មាស)'
        medal_tier = 3
        badge_color = 'warning'
    elif years_served >= 15:
        recommended_medal = '🥈 មេដាយការងារ (ថ្នាក់ប្រាក់)'
        medal_tier = 2
        badge_color = 'info'
    elif years_served >= 10:
        recommended_medal = '🥉 មេដាយការងារ (ថ្នាក់សំរឹទ្ធ)'
        medal_tier = 1
        badge_color = 'success'
    else:
        recommended_medal = '🏅 មេដាយស្ថាបនាជាតិ / លិខិតសរសើរ (ឬតាមការសម្រេចរបស់គណៈកម្មការ)'
        medal_tier = 0
        badge_color = 'primary'

    raw_awards = officer.awards_data or []
    past_awards = []
    for a in raw_awards:
        if isinstance(a, dict):
            d_str = a.get('date', '') or ''
            # Extract 4-digit year if available
            import re
            s_arabic = to_arabic_digits(str(d_str))
            y_match = re.search(r'(19\d{2}|20\d{2})', s_arabic)
            yr = y_match.group(1) if y_match else ''
            past_awards.append({
                'doc_number': a.get('doc_number', '') or '-',
                'date': d_str or '-',
                'year': yr or '-',
                'ministry': a.get('ministry', '') or '-',
                'description': a.get('description', '') or '-',
                'type': a.get('type', '') or '-',
            })
        
    return {
        'officer': officer,
        'start_date': start_date,
        'start_date_str': start_date_str,
        'years_of_service': round(years_served, 1),
        'recommended_medal': recommended_medal,
        'medal_tier': medal_tier,
        'badge_color': badge_color,
        'status': 'ELIGIBLE',
        'status_label': 'អាចស្នើសុំបាន (រង់ចាំការសម្រេចពីគណៈកម្មការ)',
        'past_awards_count': len(past_awards),
        'past_awards': past_awards,
    }


def check_has_global_hr_tracking_access(user, profile=None):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff or (user.username or '').upper() in ['ADMIN', 'ADMINISTRATOR', 'ROOT']:
        return True
    if not profile:
        profile = getattr(user, 'profile', None)
    if not profile:
        return False
    if getattr(profile, 'is_leadership', False) or getattr(profile, 'is_admin', False) or getattr(profile, 'role', '') in ['ADMIN', 'DIRECTOR', 'DEPUTY_DIRECTOR', 'LEADERSHIP', 'HR_ADMIN']:
        return True
    # Only Administration/Personnel, General Affairs, or Leadership roles have global access
    dept = profile.department
    if dept:
        name_kh = (dept.name_kh or '').strip().lower()
        code = (dept.code or '').strip().upper()
        # Strictly exclude Cantons (ខណ្ឌរដ្ឋបាលព្រៃឈើ, ខណ្ឌរដ្ឋបាលជលផល)
        if code.startswith('CANTON') or name_kh.startswith('ខណ្ឌ') or 'ខណ្ឌ' in name_kh:
            return False
        if code in ['ADMIN', 'ADMIN_PERS', 'ADMIN_PERSONNEL', 'GEN_AFFAIRS', 'GENERAL_AFFAIRS', 'GAD', 'ADMIN_DEPT', 'LEAD']:
            return True
        if 'កិច្ចការទូទៅ' in name_kh or 'កិច្ចការរដ្ឋបាលទូទៅ' in name_kh:
            return True
        if ('រដ្ឋបាល' in name_kh and 'បុគ្គលិក' in name_kh) or ('រដ្ឋបាល' in name_kh and 'ទូទៅ' in name_kh):
            return True
        if name_kh in ['ការិយាល័យរដ្ឋបាល បុគ្គលិក', 'ការិយាល័យរដ្ឋបាល-បុគ្គលិក', 'ការិយាល័យរដ្ឋបាល', 'ការិយាល័យកិច្ចការរដ្ឋបាលទូទៅ', 'ការិយាល័យកិច្ចការទូទៅ']:
            return True
    return False


@login_required
def officer_retirement_view(request):
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    
    dept_id = request.GET.get('department')
    search_q = request.GET.get('q', '').strip()
    year_param = request.GET.get('year', '').strip()
    status_param = request.GET.get('status', '').strip()
    
    officers = CivilServantProfile.objects.select_related('department').all().order_by('khmer_last_name', 'khmer_first_name')
    
    if not has_global_access:
        if user_dept:
            officers = officers.filter(department=user_dept)
        else:
            officers = officers.filter(created_by=request.user)
    elif dept_id:
        officers = officers.filter(department_id=dept_id)
        
    if search_q:
        q_arabic = to_arabic_digits(search_q)
        officers = officers.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_last_name__icontains=search_q) |
            Q(latin_first_name__icontains=search_q) |
            Q(officer_id_number__icontains=search_q) |
            Q(officer_id_number__icontains=q_arabic)
        )
        
    import datetime
    today = datetime.date.today()
    
    retirement_list = []
    due_this_year_count = 0
    due_next_year_count = 0
    already_retired_count = 0
    
    for off in officers:
        ret_data = process_officer_retirement(off, today)
        retirement_list.append(ret_data)
        if ret_data['status'] == 'THIS_YEAR':
            due_this_year_count += 1
        elif ret_data['status'] == 'NEXT_YEAR':
            due_next_year_count += 1
        elif ret_data['status'] == 'RETIRED':
            already_retired_count += 1
            
    retirement_list.sort(key=lambda x: (x['days_left'] < 0, x['days_left']))
    
    # 🎯 Dynamic Available Retirement Years from DB + 10 Years Range
    all_ret_years = set([r['retirement_year'] for r in retirement_list if r.get('retirement_year') and r['retirement_year'] >= 2026])
    all_ret_years.update([today.year + i for i in range(11)]) # 2026 to 2036+
    available_years = sorted(list(all_ret_years))

    selected_year = None
    if year_param:
        try:
            selected_year = int(to_arabic_digits(year_param))
        except Exception:
            selected_year = None

    # Filter list if specific year or status is requested
    displayed_list = retirement_list
    if selected_year:
        displayed_list = [r for r in displayed_list if r.get('retirement_year') == selected_year]
    if status_param and status_param != 'ALL':
        displayed_list = [r for r in displayed_list if r.get('status') == status_param]

    selected_year_count = len([r for r in retirement_list if r.get('retirement_year') == (selected_year or today.year)])

    if has_global_access:
        departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')
    else:
        departments = Department.objects.filter(id=user_dept.id) if user_dept else []
        
    context = {
        'has_global_access': has_global_access,
        'user_dept': user_dept,
        'departments': departments,
        'selected_dept': dept_id,
        'search_q': search_q,
        'today': today,
        'today_year': today.year,
        'selected_year': selected_year,
        'selected_year_str': str(selected_year) if selected_year else '',
        'selected_year_count': selected_year_count,
        'available_years': available_years,
        'retirement_list': displayed_list,
        'total_officers_count': len(retirement_list),
        'displayed_count': len(displayed_list),
        'due_this_year_count': due_this_year_count,
        'due_next_year_count': due_next_year_count,
        'already_retired_count': already_retired_count,
        'total_retirement_attention': due_this_year_count + due_next_year_count,
        'status_param': status_param,
    }
    return render(request, 'dms/officer_retirement.html', context)


@login_required
def officer_promotion_view(request):
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    
    dept_id = request.GET.get('department')
    search_q = request.GET.get('q', '').strip()
    active_view_tab = request.GET.get('view_tab', 'forecast') # 'forecast', 'requests', 'master'
    
    import datetime
    today = datetime.date.today()
    
    forecast_year_param = request.GET.get('forecast_year') or request.GET.get('year')
    if forecast_year_param:
        try:
            forecast_year = int(to_arabic_digits(forecast_year_param))
        except Exception:
            forecast_year = today.year
    else:
        forecast_year = today.year

    status_param = request.GET.get('status', 'ALL').strip()

    officers = CivilServantProfile.objects.select_related('department').all().order_by('khmer_last_name', 'khmer_first_name')
    
    if not has_global_access:
        if user_dept:
            officers = officers.filter(department=user_dept)
        else:
            officers = officers.filter(created_by=request.user)
    elif dept_id:
        officers = officers.filter(department_id=dept_id)
        
    if search_q:
        q_arabic = to_arabic_digits(search_q)
        officers = officers.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_last_name__icontains=search_q) |
            Q(latin_first_name__icontains=search_q) |
            Q(officer_id_number__icontains=search_q) |
            Q(officer_id_number__icontains=q_arabic)
        )
        
    # 🎯 Bulk-query previous rejected/disapproved promotion requests
    rejected_reqs = {}
    for r in OfficerPromotionRequest.objects.filter(status__in=['DEPT_REJECTED', 'MINISTRY_REJECTED', 'REJECTED']).order_by('request_year'):
        rejected_reqs[r.officer_id] = r

    promotion_list = []
    due_new_promo_count = 0
    due_carried_over_count = 0
    due_soon_promo_count = 0
    normal_promo_count = 0
    max_c1_count = 0
    
    for off in officers:
        pro_data = process_officer_promotion(off, today, target_year=forecast_year, rejected_req=rejected_reqs.get(off.id))
        promotion_list.append(pro_data)
        if pro_data['status'] == 'DUE_NEW':
            due_new_promo_count += 1
        elif pro_data['status'] == 'CARRIED_OVER':
            due_carried_over_count += 1
        elif pro_data['status'] == 'DUE_SOON':
            due_soon_promo_count += 1
        elif pro_data['status'] == 'MAX_RANK_EXAM_ONLY':
            max_c1_count += 1
        else:
            normal_promo_count += 1
            
    promotion_list.sort(key=lambda x: (x['status'] not in ['DUE_NEW', 'CARRIED_OVER'], x['status'] == 'MAX_RANK_EXAM_ONLY', -x['years_in_rank']))
    
    # 🎯 Filter by active status pill if requested
    displayed_promotion_list = promotion_list
    if status_param == 'DUE_ALL':
        displayed_promotion_list = [p for p in promotion_list if p['status'] in ['DUE_NEW', 'CARRIED_OVER']]
    elif status_param and status_param != 'ALL':
        displayed_promotion_list = [p for p in promotion_list if p['status'] == status_param]

    total_due_in_year = due_new_promo_count + due_carried_over_count
    available_years = [2026, 2027, 2028, 2029, 2030, 2031, 2032, 2033, 2034, 2035]

    if has_global_access:
        departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')
        dept_officers = CivilServantProfile.objects.select_related('department').all().order_by('khmer_last_name', 'khmer_first_name')
    else:
        departments = Department.objects.filter(id=user_dept.id) if user_dept else []
        dept_officers = CivilServantProfile.objects.filter(department=user_dept).order_by('khmer_last_name', 'khmer_first_name') if user_dept else []
        
    # Query Online Promotion Requests
    promo_requests_qs = OfficerPromotionRequest.objects.select_related('officer', 'department', 'submitted_by', 'reviewed_by').all().order_by('-created_at')
    if not has_global_access:
        if user_dept:
            promo_requests_qs = promo_requests_qs.filter(department=user_dept)
        else:
            promo_requests_qs = promo_requests_qs.filter(submitted_by=request.user)
    elif dept_id:
        promo_requests_qs = promo_requests_qs.filter(department_id=dept_id)

    pending_requests_count = promo_requests_qs.filter(status='PENDING').count()
    approved_requests_count = promo_requests_qs.filter(status='APPROVED').count()
    rejected_requests_count = promo_requests_qs.filter(status='REJECTED').count()
    submitted_to_ministry_count = promo_requests_qs.filter(is_submitted_to_ministry=True).count()

    # Filter for Master Dossier Tab
    master_year = request.GET.get('master_year') or request.GET.get('year') or today.year
    master_status = request.GET.get('req_status', 'ALL')
    master_requests_qs = promo_requests_qs.filter(request_year=int(master_year) if str(master_year).isdigit() else today.year)
    if master_status == 'SUBMITTED':
        master_requests_qs = master_requests_qs.filter(is_submitted_to_ministry=True)
    elif master_status != 'ALL':
        master_requests_qs = master_requests_qs.filter(status=master_status)

    context = {
        'has_global_access': has_global_access,
        'user_dept': user_dept,
        'departments': departments,
        'dept_officers': dept_officers,
        'selected_dept': dept_id,
        'search_q': search_q,
        'active_view_tab': active_view_tab,
        'today': today,
        'today_year': today.year,
        'forecast_year': forecast_year,
        'forecast_year_str': str(forecast_year),
        'available_years': available_years,
        'master_year': int(master_year) if str(master_year).isdigit() else today.year,
        'master_status': master_status,
        'master_requests': master_requests_qs,
        'promotion_list': displayed_promotion_list,
        'all_promotion_count': len(promotion_list),
        'displayed_promo_count': len(displayed_promotion_list),
        'total_due_in_year': total_due_in_year,
        'due_new_promo_count': due_new_promo_count,
        'due_carried_over_count': due_carried_over_count,
        'due_soon_promo_count': due_soon_promo_count,
        'normal_promo_count': normal_promo_count,
        'max_c1_count': max_c1_count,
        'total_promo_attention': total_due_in_year + due_soon_promo_count,
        'promotion_requests': promo_requests_qs,
        'pending_requests_count': pending_requests_count,
        'approved_requests_count': approved_requests_count,
        'rejected_requests_count': rejected_requests_count,
        'submitted_to_ministry_count': submitted_to_ministry_count,
        'total_requests_count': promo_requests_qs.count(),
        'status_param': status_param,
    }
    return render(request, 'dms/officer_promotion.html', context)


@login_required
def officer_promotion_request_create(request):
    if request.method != 'POST':
        return redirect('officer_promotion')

    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)

    officer_id = request.POST.get('officer_id')
    if not officer_id:
        messages.error(request, "⚠️ សូមជ្រើសរើសមន្ត្រីដែលត្រូវស្នើសុំដំឡើងថ្នាក់!")
        return redirect('officer_promotion')

    officer = get_object_or_404(CivilServantProfile, pk=officer_id)

    # Permission check: specialized office can only request for their officers
    if not has_global_access and officer.department != user_dept and officer.user != request.user:
        messages.error(request, "⚠️ លោកអ្នកពុំមានសិទ្ធិបង្កើតសំណើសុំដំឡើងថ្នាក់សម្រាប់មន្ត្រីនៃការិយាល័យផ្សេងឡើយ!")
        return redirect('officer_promotion')

    promotion_type = request.POST.get('promotion_type', 'SENIORITY')
    proposed_rank_and_step = request.POST.get('proposed_rank_and_step', '').strip()
    reason = request.POST.get('reason', '').strip()
    request_year = request.POST.get('request_year') or date.today().year

    # 🎯 ពិនិត្យករណីពិសេស៖ មន្ត្រីថ្នាក់ គ.១ ឬ ខ.១.១ មិនអនុញ្ញាតស្នើសុំតាមវេនជ្រើសរើសឡើយ
    if is_c1_rank(officer.current_rank_and_step):
        if promotion_type != 'EXAM':
            messages.error(request, "⚠️ មន្ត្រីថ្នាក់ គ.១ បានដល់កម្រិតកំពូលនៃក្របខ័ណ្ឌ គ ហើយ! ពុំមានមុខងារស្នើសុំដំឡើងថ្នាក់តាមវេនជ្រើសរើស/អតីតភាពឡើយ លុះត្រាតែមានការស្នើសុំប្រឡងផ្ទៃក្នុងថ្នាក់ក្រសួងដើម្បីប្តូរក្របខ័ណ្ឌតែមួយគត់!")
            return redirect('officer_promotion')
    elif is_b1_1_rank(officer.current_rank_and_step):
        if promotion_type != 'EXAM':
            messages.error(request, "⚠️ មន្ត្រីថ្នាក់ ខ.១.១ បានដល់កម្រិតកំពូលនៃក្របខ័ណ្ឌ ខ ហើយ! ពុំមានមុខងារស្នើសុំដំឡើងថ្នាក់តាមវេនជ្រើសរើស/អតីតភាពឡើយ លុះត្រាតែមានការស្នើសុំប្រឡងផ្ទៃក្នុងថ្នាក់ក្រសួងដើម្បីប្តូរក្របខ័ណ្ឌតែមួយគត់!")
            return redirect('officer_promotion')

    if not proposed_rank_and_step:
        messages.error(request, "⚠️ សូមបញ្ជាក់ថ្នាក់ដែលស្នើសុំដំឡើងទៅ!")
        return redirect('officer_promotion')

    req_obj = OfficerPromotionRequest(
        officer=officer,
        department=officer.department or user_dept,
        request_year=int(request_year),
        promotion_type=promotion_type,
        current_rank_and_step=officer.current_rank_and_step or '',
        proposed_rank_and_step=proposed_rank_and_step,
        years_in_current_rank=request.POST.get('years_in_current_rank', ''),
        reason=reason,
        submitted_by=request.user,
        status='PENDING'
    )

    if 'attachment' in request.FILES and request.FILES['attachment']:
        att_file = request.FILES['attachment']
        if att_file.size > 5 * 1024 * 1024:
            messages.error(request, "⚠️ ទំហំ File ឯកសារយោងលើសពី 5MB!")
            return redirect('officer_promotion')
        req_obj.attachment = att_file

    req_obj.save()

    # Log action
    log_officer_action(
        request,
        officer=officer,
        action_type='EDIT',
        category='ស្នើសុំដំឡើងថ្នាក់ Online',
        details=f"បានដាក់សំណើសុំដំឡើងថ្នាក់តាម Online សម្រាប់ «{officer.full_name_kh}» ទៅថ្នាក់ «{proposed_rank_and_step}» ({req_obj.get_promotion_type_display()})"
    )

    messages.success(request, f"✅ បានដាក់សំណើសុំដំឡើងថ្នាក់តាម Online សម្រាប់មន្ត្រី «{officer.full_name_kh}» ដោយជោគជ័យ! សំណើត្រូវបានបញ្ជូនទៅរង់ចាំការពិនិត្យពីថ្នាក់ដឹកនាំ/Admin។")
    return redirect('officer_promotion')


@login_required
def officer_promotion_request_edit(request, pk):
    """
    Allows Admin or Submitter to edit promotion request details (e.g. correcting wrong rank, type, or reason)
    """
    has_global_access = check_has_global_hr_tracking_access(request.user, getattr(request.user, 'profile', None))
    req_obj = get_object_or_404(OfficerPromotionRequest, pk=pk)

    if not has_global_access and req_obj.submitted_by != request.user:
        messages.error(request, "⚠️ លោកអ្នកពុំមានសិទ្ធិកែប្រែសំណើនេះឡើយ!")
        return redirect('officer_promotion')

    if request.method == 'POST':
        proposed_rank_and_step = request.POST.get('proposed_rank_and_step', '').strip()
        promotion_type = request.POST.get('promotion_type', req_obj.promotion_type)
        current_rank_and_step = request.POST.get('current_rank_and_step', req_obj.current_rank_and_step).strip()
        years_in_current_rank = request.POST.get('years_in_current_rank', req_obj.years_in_current_rank).strip()
        reason = request.POST.get('reason', '').strip()
        request_year = request.POST.get('request_year') or req_obj.request_year
        meeting_decision = request.POST.get('meeting_decision', req_obj.meeting_decision).strip()

        # 🎯 ពិនិត្យករណីពិសេស៖ មន្ត្រីថ្នាក់ គ.១ ឬ ខ.១.១ មិនអនុញ្ញាតស្នើសុំតាមវេនជ្រើសរើសឡើយ
        if is_c1_rank(current_rank_and_step):
            if promotion_type != 'EXAM':
                messages.error(request, "⚠️ មន្ត្រីថ្នាក់ គ.១ បានដល់កម្រិតកំពូលនៃក្របខ័ណ្ឌ គ ហើយ! ពុំមានមុខងារស្នើសុំដំឡើងថ្នាក់តាមវេនជ្រើសរើស/អតីតភាពឡើយ លុះត្រាតែមានការស្នើសុំប្រឡងផ្ទៃក្នុងថ្នាក់ក្រសួងដើម្បីប្តូរក្របខ័ណ្ឌតែមួយគត់!")
                return redirect(request.META.get('HTTP_REFERER', 'officer_promotion'))
        elif is_b1_1_rank(current_rank_and_step):
            if promotion_type != 'EXAM':
                messages.error(request, "⚠️ មន្ត្រីថ្នាក់ ខ.១.១ បានដល់កម្រិតកំពូលនៃក្របខ័ណ្ឌ ខ ហើយ! ពុំមានមុខងារស្នើសុំដំឡើងថ្នាក់តាមវេនជ្រើសរើស/អតីតភាពឡើយ លុះត្រាតែមានការស្នើសុំប្រឡងផ្ទៃក្នុងថ្នាក់ក្រសួងដើម្បីប្តូរក្របខ័ណ្ឌតែមួយគត់!")
                return redirect(request.META.get('HTTP_REFERER', 'officer_promotion'))

        if not proposed_rank_and_step:
            messages.error(request, "⚠️ សូមបញ្ជាក់ថ្នាក់ដែលស្នើសុំដំឡើងទៅ!")
            return redirect('officer_promotion')

        req_obj.proposed_rank_and_step = proposed_rank_and_step
        req_obj.promotion_type = promotion_type
        req_obj.current_rank_and_step = current_rank_and_step
        req_obj.years_in_current_rank = years_in_current_rank
        req_obj.reason = reason
        req_obj.request_year = int(request_year) if str(request_year).isdigit() else req_obj.request_year
        req_obj.meeting_decision = meeting_decision

        dept_id = request.POST.get('department')
        if has_global_access and dept_id:
            target_d = Department.objects.filter(id=dept_id).first()
            if target_d:
                req_obj.department = target_d

        if 'attachment' in request.FILES and request.FILES['attachment']:
            req_obj.attachment = request.FILES['attachment']

        req_obj.save()

        log_officer_action(
            request,
            officer=req_obj.officer,
            action_type='EDIT',
            category='កែតម្រូវសំណើសុំដំឡើងថ្នាក់',
            details=f"Admin/អ្នកប្រើប្រាស់បានកែតម្រូវទិន្នន័យសំណើសុំដំឡើងថ្នាក់របស់ «{req_obj.officer.full_name_kh}» ទៅជា «{proposed_rank_and_step}»"
        )
        messages.success(request, f"✅ បានកែតម្រូវសំណើសុំដំឡើងថ្នាក់សម្រាប់ «{req_obj.officer.full_name_kh}» ដោយជោគជ័យ!")

    return redirect(request.META.get('HTTP_REFERER', 'officer_promotion'))


@login_required
def officer_promotion_request_review(request, pk):
    """
    Multi-Stage Review Handler:
    1. DEPT_APPROVE / DEPT_REJECT: Department Leadership / Meeting level (Does NOT update officer profile)
    2. SUBMIT_MINISTRY: Records official submission letter to ministry (Does NOT update officer profile)
    3. MINISTRY_APPROVE: Official Ministry/Government decree/prakas approval -> UPDATES OFFICER PROFILE
    4. MINISTRY_REJECT: Ministry rejection -> Keeps officer profile intact with original rank
    """
    has_global_access = check_has_global_hr_tracking_access(request.user, getattr(request.user, 'profile', None))
    if not has_global_access:
        messages.error(request, "⚠️ មានតែ Admin/ថ្នាក់ដឹកនាំ ប៉ុណ្ណោះដែលអាចពិនិត្យ និងអនុម័តសំណើបាន!")
        return redirect('officer_promotion')

    req_obj = get_object_or_404(OfficerPromotionRequest, pk=pk)
    if request.method == 'POST':
        action = request.POST.get('action') # 'DEPT_APPROVE', 'DEPT_REJECT', 'SUBMIT_MINISTRY', 'MINISTRY_APPROVE', 'MINISTRY_REJECT', 'APPROVE', 'REJECT'
        admin_notes = request.POST.get('admin_notes', '').strip()
        meeting_decision = request.POST.get('meeting_decision', '').strip()

        req_obj.admin_notes = admin_notes
        req_obj.reviewed_by = request.user
        req_obj.reviewed_at = timezone.now()

        if meeting_decision:
            req_obj.meeting_decision = meeting_decision

        # Stage 1: Department Leadership Decision
        if action in ['DEPT_APPROVE', 'APPROVE']:
            req_obj.status = 'DEPT_APPROVED'
            log_officer_action(
                request,
                officer=req_obj.officer,
                action_type='EDIT',
                category='កិច្ចប្រជុំមន្ទីរឯកភាពដំឡើងថ្នាក់',
                details=f"អង្គប្រជុំថ្នាក់ដឹកនាំមន្ទីរបានឯកភាពលើសំណើសុំដំឡើងថ្នាក់របស់ «{req_obj.officer.full_name_kh}» (មតិ៖ {meeting_decision or 'ឯកភាព'})"
            )
            messages.success(request, f"🟢 អង្គប្រជុំថ្នាក់ដឹកនាំមន្ទីរបានឯកភាពលើសំណើសុំដំឡើងថ្នាក់របស់ «{req_obj.officer.full_name_kh}» (ត្រៀមរៀបចំឯកសារបញ្ជូនទៅក្រសួង)!")

        elif action in ['DEPT_REJECT', 'REJECT']:
            req_obj.status = 'DEPT_REJECTED'
            log_officer_action(
                request,
                officer=req_obj.officer,
                action_type='EDIT',
                category='មន្ទីរមិនឯកភាពដំឡើងថ្នាក់',
                details=f"ថ្នាក់ដឹកនាំមន្ទីរមិនទាន់ឯកភាពសំណើសុំដំឡើងថ្នាក់របស់ «{req_obj.officer.full_name_kh}» (មូលហេតុ៖ {admin_notes})"
            )
            messages.info(request, f"🔴 បានកត់ត្រាមិនទាន់ឯកភាពកម្រិតមន្ទីរលើសំណើរបស់ «{req_obj.officer.full_name_kh}»។")

        # Stage 2: Submission to Ministry
        elif action == 'SUBMIT_MINISTRY':
            doc_no = request.POST.get('ministry_doc_number', '').strip()
            sub_date_str = request.POST.get('ministry_submission_date', '').strip()
            req_obj.status = 'SUBMITTED'
            req_obj.is_submitted_to_ministry = True
            req_obj.ministry_doc_number = doc_no or req_obj.ministry_doc_number
            if sub_date_str:
                try:
                    req_obj.ministry_submission_date = datetime.strptime(sub_date_str, '%Y-%m-%d').date()
                except Exception:
                    req_obj.ministry_submission_date = date.today()
            else:
                req_obj.ministry_submission_date = date.today()

            log_officer_action(
                request,
                officer=req_obj.officer,
                action_type='EDIT',
                category='បញ្ជូនសំណើទៅក្រសួង',
                details=f"បានកត់ត្រាការបញ្ជូនសំណើដំឡើងថ្នាក់របស់ «{req_obj.officer.full_name_kh}» ទៅក្រសួង តាមលិខិតលេខ៖ {req_obj.ministry_doc_number}"
            )
            messages.success(request, f"🏛️ បានកត់ត្រាការបញ្ជូនសំណើរបស់ «{req_obj.officer.full_name_kh}» ទៅក្រសួងរួចរាល់!")

        # Stage 3: Official Ministry Decision (PRAKAS / ANUKRET / REACH KRET)
        elif action == 'MINISTRY_APPROVE':
            legal_type = request.POST.get('legal_doc_type', 'PRAKAS')
            legal_num = request.POST.get('legal_doc_number', '').strip()
            legal_date_str = request.POST.get('legal_doc_date', '').strip()
            effective_date_str = request.POST.get('effective_date', '').strip()
            min_notes = request.POST.get('ministry_decision_notes', '').strip()

            req_obj.status = 'MINISTRY_APPROVED'
            req_obj.legal_doc_type = legal_type
            req_obj.legal_doc_number = legal_num
            req_obj.ministry_decision_notes = min_notes

            parsed_legal_date = None
            parsed_effective_date = None
            if legal_date_str:
                try:
                    parsed_legal_date = datetime.strptime(legal_date_str, '%Y-%m-%d').date()
                except Exception:
                    parsed_legal_date = date.today()
            else:
                parsed_legal_date = date.today()

            if effective_date_str:
                try:
                    parsed_effective_date = datetime.strptime(effective_date_str, '%Y-%m-%d').date()
                except Exception:
                    parsed_effective_date = parsed_legal_date
            else:
                parsed_effective_date = parsed_legal_date

            req_obj.legal_doc_date = parsed_legal_date
            req_obj.effective_date = parsed_effective_date
            req_obj.is_profile_updated = True

            # 💥 NOW UPDATE OFFICER PROFILE & RANK
            officer = req_obj.officer
            old_rank = officer.current_rank_and_step
            officer.current_rank_and_step = req_obj.proposed_rank_and_step
            officer.last_promotion_date = parsed_effective_date or date.today()

            doc_title = f"{req_obj.get_legal_doc_type_display()} លេខ {legal_num}" if legal_num else "ប្រកាស/អនុក្រឹត្យផ្លូវការ"
            effective_str = parsed_effective_date.strftime('%d/%m/%Y') if parsed_effective_date else date.today().strftime('%d/%m/%Y')
            promo_entry = {
                'effective_date': effective_str,
                'ministry': 'ក្រសួងកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទ',
                'department': officer.display_department_name,
                'office': officer.display_department_name,
                'old_rank_step': old_rank or '',
                'new_rank_step': req_obj.proposed_rank_and_step,
                'promo_type': req_obj.get_promotion_type_display(),
                'doc_number': doc_title
            }
            history = officer.promotions_by_seniority or []
            history.append(promo_entry)
            officer.promotions_by_seniority = history
            officer.save()

            log_officer_action(
                request,
                officer=officer,
                action_type='EDIT',
                category='ក្រសួងអនុម័តដំឡើងថ្នាក់ផ្លូវការ',
                details=f"ក្រសួងបានអនុម័តជាផ្លូវការ ({doc_title}) ដំឡើងថ្នាក់របស់ «{officer.full_name_kh}» ពី «{old_rank}» ទៅជា «{req_obj.proposed_rank_and_step}»"
            )
            messages.success(request, f"👑 ក្រសួងបានអនុម័តជាផ្លូវការ ({doc_title})! ប្រព័ន្ធបាន Update ថ្នាក់ថ្មី «{req_obj.proposed_rank_and_step}» ទៅក្នុងប្រវត្តិរូបមន្ត្រីរួចរាល់។")

        elif action == 'MINISTRY_REJECT':
            min_notes = request.POST.get('ministry_decision_notes', '').strip()
            req_obj.status = 'MINISTRY_REJECTED'
            req_obj.ministry_decision_notes = min_notes

            log_officer_action(
                request,
                officer=req_obj.officer,
                action_type='EDIT',
                category='ក្រសួងបដិសេធដំឡើងថ្នាក់',
                details=f"ក្រសួងមិនបានអនុម័តសំណើសុំដំឡើងថ្នាក់របស់ «{req_obj.officer.full_name_kh}» (កំណត់សម្គាល់៖ {min_notes})។ រក្សាថ្នាក់ដើមដដែល។"
            )
            messages.warning(request, f"❌ បានកត់ត្រាការបដិសេធ/ធ្លាក់ពីក្រសួងសម្រាប់សំណើរបស់ «{req_obj.officer.full_name_kh}»។ ប្រព័ន្ធបានរក្សាថ្នាក់បច្ចុប្បន្នរបស់មន្ត្រីដដែល។")

        req_obj.save()

    return redirect(request.META.get('HTTP_REFERER', 'officer_promotion'))


@login_required
def officer_promotion_request_delete(request, pk):
    profile = getattr(request.user, 'profile', None)
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    req_obj = get_object_or_404(OfficerPromotionRequest, pk=pk)

    if not has_global_access and req_obj.submitted_by != request.user:
        messages.error(request, "⚠️ លោកអ្នកគ្មានសិទ្ធិលុបសំណើនេះឡើយ!")
        return redirect('officer_promotion')

    if req_obj.status == 'APPROVED' and not has_global_access:
        messages.error(request, "⚠️ សំណើដែលបានអនុម័តរួច មិនអាចលុបបានឡើយ!")
        return redirect('officer_promotion')

    if request.method == 'POST':
        officer_name = req_obj.officer.full_name_kh
        req_obj.delete()
        messages.success(request, f"🗑️ បានលុបសំណើសុំដំឡើងថ្នាក់របស់ «{officer_name}» ដោយជោគជ័យ!")

    return redirect('officer_promotion')


@login_required
def officer_medals_view(request):
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    
    dept_id = request.GET.get('department')
    search_q = request.GET.get('q', '').strip()
    active_view_tab = request.GET.get('view_tab', 'forecast') # 'forecast' or 'requests'
    
    officers = CivilServantProfile.objects.select_related('department').all().order_by('khmer_last_name', 'khmer_first_name')
    
    if not has_global_access:
        if user_dept:
            officers = officers.filter(department=user_dept)
        else:
            officers = officers.filter(created_by=request.user)
    elif dept_id:
        officers = officers.filter(department_id=dept_id)
        
    if search_q:
        officers = officers.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_last_name__icontains=search_q) |
            Q(latin_first_name__icontains=search_q) |
            Q(officer_id_number__icontains=search_q)
        )
        
    import datetime
    today = datetime.date.today()
    
    medals_list = []
    grand_orders_count = 0
    gold_medals_count = 0
    silver_count = 0
    bronze_count = 0
    merit_count = 0
    
    for off in officers:
        med_data = process_officer_medals(off, today)
        medals_list.append(med_data)
        if med_data['medal_tier'] >= 4:
            grand_orders_count += 1
        elif med_data['medal_tier'] == 3:
            gold_medals_count += 1
        elif med_data['medal_tier'] == 2:
            silver_count += 1
        elif med_data['medal_tier'] == 1:
            bronze_count += 1
        else:
            merit_count += 1
                
    medals_list.sort(key=lambda x: x['years_of_service'], reverse=True)
    eligible_medals_count = len(medals_list)
    
    if has_global_access:
        departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')
        dept_officers = CivilServantProfile.objects.select_related('department').all().order_by('khmer_last_name', 'khmer_first_name')
    else:
        departments = Department.objects.filter(id=user_dept.id) if user_dept else []
        dept_officers = CivilServantProfile.objects.filter(department=user_dept).order_by('khmer_last_name', 'khmer_first_name') if user_dept else []
        
    # Query Online Medal Requests
    medal_requests_qs = OfficerMedalRequest.objects.select_related('officer', 'department', 'submitted_by', 'reviewed_by').all().order_by('-created_at')
    if not has_global_access:
        if user_dept:
            medal_requests_qs = medal_requests_qs.filter(department=user_dept)
        else:
            medal_requests_qs = medal_requests_qs.filter(submitted_by=request.user)
    elif dept_id:
        medal_requests_qs = medal_requests_qs.filter(department_id=dept_id)

    pending_medal_requests_count = medal_requests_qs.filter(status='PENDING').count()
    approved_medal_requests_count = medal_requests_qs.filter(status='APPROVED').count()
    rejected_medal_requests_count = medal_requests_qs.filter(status='REJECTED').count()
    submitted_to_ministry_medals_count = medal_requests_qs.filter(is_submitted_to_ministry=True).count()

    # Filter for Master Dossier Tab
    master_year = request.GET.get('year') or today.year
    master_status = request.GET.get('req_status', 'ALL')
    master_requests_qs = medal_requests_qs.filter(request_year=int(master_year) if str(master_year).isdigit() else today.year)
    if master_status == 'SUBMITTED':
        master_requests_qs = master_requests_qs.filter(is_submitted_to_ministry=True)
    elif master_status != 'ALL':
        master_requests_qs = master_requests_qs.filter(status=master_status)

    context = {
        'has_global_access': has_global_access,
        'user_dept': user_dept,
        'departments': departments,
        'dept_officers': dept_officers,
        'selected_dept': dept_id,
        'search_q': search_q,
        'active_view_tab': active_view_tab,
        'today': today,
        'today_year': today.year,
        'master_year': int(master_year) if str(master_year).isdigit() else today.year,
        'master_status': master_status,
        'master_requests': master_requests_qs,
        'medals_list': medals_list,
        'eligible_medals_count': eligible_medals_count,
        'grand_orders_count': grand_orders_count,
        'gold_medals_count': gold_medals_count,
        'silver_count': silver_count,
        'bronze_count': bronze_count,
        'merit_count': merit_count,
        'silver_bronze_count': silver_count + bronze_count,
        'medal_requests': medal_requests_qs,
        'pending_medal_requests_count': pending_medal_requests_count,
        'approved_medal_requests_count': approved_medal_requests_count,
        'rejected_medal_requests_count': rejected_medal_requests_count,
        'submitted_to_ministry_medals_count': submitted_to_ministry_medals_count,
        'total_medal_requests_count': medal_requests_qs.count(),
    }
    return render(request, 'dms/officer_medals.html', context)


@login_required
def officer_medal_request_create(request):
    if request.method != 'POST':
        return redirect('officer_medals')

    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)

    officer_id = request.POST.get('officer_id')
    if not officer_id:
        messages.error(request, "⚠️ សូមជ្រើសរើសមន្ត្រីដែលត្រូវស្នើសុំមេដាយ!")
        return redirect('officer_medals')

    officer = get_object_or_404(CivilServantProfile, pk=officer_id)

    # Permission check: specialized office can only request for their officers
    if not has_global_access and officer.department != user_dept and officer.user != request.user:
        messages.error(request, "⚠️ លោកអ្នកពុំមានសិទ្ធិបង្កើតសំណើសុំមេដាយសម្រាប់មន្ត្រីនៃការិយាល័យផ្សេងឡើយ!")
        return redirect('officer_medals')

    proposed_medal = request.POST.get('proposed_medal', '').strip()
    achievements = request.POST.get('achievements', '').strip()
    years_of_service = request.POST.get('years_of_service', '').strip()
    request_year = request.POST.get('request_year') or date.today().year

    if not proposed_medal:
        messages.error(request, "⚠️ សូមបញ្ជាក់ប្រភេទមេដាយ ឬគ្រឿងឥស្សរិយយសដែលស្នើសុំ!")
        return redirect('officer_medals')

    req_obj = OfficerMedalRequest(
        officer=officer,
        department=officer.department or user_dept,
        request_year=int(request_year),
        proposed_medal=proposed_medal,
        years_of_service=years_of_service,
        achievements=achievements,
        submitted_by=request.user,
        status='PENDING'
    )

    if 'attachment' in request.FILES and request.FILES['attachment']:
        att_file = request.FILES['attachment']
        if att_file.size > 5 * 1024 * 1024:
            messages.error(request, "⚠️ ទំហំ File ឯកសារយោងលើសពី 5MB!")
            return redirect('officer_medals')
        req_obj.attachment = att_file

    req_obj.save()

    # Log action
    log_officer_action(
        request,
        officer=officer,
        action_type='EDIT',
        category='ស្នើសុំមេដាយ Online',
        details=f"បានដាក់សំណើសុំមេដាយតាម Online សម្រាប់ «{officer.full_name_kh}»៖ «{proposed_medal}» (អតីតភាព {years_of_service} ឆ្នាំ)"
    )

    messages.success(request, f"✅ បានដាក់សំណើសុំមេដាយតាម Online សម្រាប់មន្ត្រី «{officer.full_name_kh}» ដោយជោគជ័យ! សំណើត្រូវបានបញ្ជូនទៅរង់ចាំការពិនិត្យពីថ្នាក់ដឹកនាំ/Admin។")
    return redirect('officer_medals')


@login_required
def officer_medal_request_edit(request, pk):
    """
    Allows Admin or Submitter to edit medal request details (e.g. correcting wrong medal type, achievements, or years)
    """
    has_global_access = check_has_global_hr_tracking_access(request.user, getattr(request.user, 'profile', None))
    req_obj = get_object_or_404(OfficerMedalRequest, pk=pk)

    if not has_global_access and req_obj.submitted_by != request.user:
        messages.error(request, "⚠️ លោកអ្នកពុំមានសិទ្ធិកែប្រែសំណើនេះឡើយ!")
        return redirect('officer_medals')

    if request.method == 'POST':
        proposed_medal = request.POST.get('proposed_medal', '').strip()
        years_of_service = request.POST.get('years_of_service', req_obj.years_of_service).strip()
        achievements = request.POST.get('achievements', '').strip()
        request_year = request.POST.get('request_year') or req_obj.request_year
        meeting_decision = request.POST.get('meeting_decision', req_obj.meeting_decision).strip()

        if not proposed_medal:
            messages.error(request, "⚠️ សូមបញ្ជាក់ប្រភេទមេដាយដែលត្រូវស្នើសុំ!")
            return redirect('officer_medals')

        req_obj.proposed_medal = proposed_medal
        req_obj.years_of_service = years_of_service
        req_obj.achievements = achievements
        req_obj.request_year = int(request_year) if str(request_year).isdigit() else req_obj.request_year
        req_obj.meeting_decision = meeting_decision

        dept_id = request.POST.get('department')
        if has_global_access and dept_id:
            target_d = Department.objects.filter(id=dept_id).first()
            if target_d:
                req_obj.department = target_d

        if 'attachment' in request.FILES and request.FILES['attachment']:
            req_obj.attachment = request.FILES['attachment']

        req_obj.save()

        log_officer_action(
            request,
            officer=req_obj.officer,
            action_type='EDIT',
            category='កែតម្រូវសំណើសុំមេដាយ',
            details=f"Admin/អ្នកប្រើប្រាស់បានកែតម្រូវទិន្នន័យសំណើសុំមេដាយរបស់ «{req_obj.officer.full_name_kh}» ទៅជា «{proposed_medal}»"
        )
        messages.success(request, f"✅ បានកែតម្រូវសំណើសុំមេដាយសម្រាប់ «{req_obj.officer.full_name_kh}» ដោយជោគជ័យ!")

    return redirect(request.META.get('HTTP_REFERER', 'officer_medals'))


@login_required
def officer_medal_request_review(request, pk):
    """
    Multi-Stage Review Handler for Medals:
    1. DEPT_APPROVE / DEPT_REJECT: Department Leadership / Meeting level (Does NOT update officer awards)
    2. SUBMIT_MINISTRY: Records official submission letter to ministry (Does NOT update officer awards)
    3. MINISTRY_APPROVE: Official Ministry/Royal Decree approval -> UPDATES OFFICER AWARDS DATA
    4. MINISTRY_REJECT: Ministry rejection -> Keeps officer profile intact
    """
    has_global_access = check_has_global_hr_tracking_access(request.user, getattr(request.user, 'profile', None))
    if not has_global_access:
        messages.error(request, "⚠️ មានតែ Admin/ថ្នាក់ដឹកនាំ ប៉ុណ្ណោះដែលអាចពិនិត្យ និងអនុម័តសំណើបាន!")
        return redirect('officer_medals')

    req_obj = get_object_or_404(OfficerMedalRequest, pk=pk)
    if request.method == 'POST':
        action = request.POST.get('action') # 'DEPT_APPROVE', 'DEPT_REJECT', 'SUBMIT_MINISTRY', 'MINISTRY_APPROVE', 'MINISTRY_REJECT', 'APPROVE', 'REJECT'
        admin_notes = request.POST.get('admin_notes', '').strip()
        meeting_decision = request.POST.get('meeting_decision', '').strip()

        req_obj.admin_notes = admin_notes
        req_obj.reviewed_by = request.user
        req_obj.reviewed_at = timezone.now()

        if meeting_decision:
            req_obj.meeting_decision = meeting_decision

        # Stage 1: Department Leadership Decision
        if action in ['DEPT_APPROVE', 'APPROVE']:
            req_obj.status = 'DEPT_APPROVED'
            log_officer_action(
                request,
                officer=req_obj.officer,
                action_type='EDIT',
                category='កិច្ចប្រជុំមន្ទីរឯកភាពស្នើសុំមេដាយ',
                details=f"អង្គប្រជុំថ្នាក់ដឹកនាំមន្ទីរបានឯកភាពលើសំណើសុំមេដាយរបស់ «{req_obj.officer.full_name_kh}» (មតិ៖ {meeting_decision or 'ឯកភាព'})"
            )
            messages.success(request, f"🟢 អង្គប្រជុំថ្នាក់ដឹកនាំមន្ទីរបានឯកភាពលើសំណើសុំមេដាយរបស់ «{req_obj.officer.full_name_kh}» (ត្រៀមរៀបចំឯកសារបញ្ជូនទៅក្រសួង)!")

        elif action in ['DEPT_REJECT', 'REJECT']:
            req_obj.status = 'DEPT_REJECTED'
            log_officer_action(
                request,
                officer=req_obj.officer,
                action_type='EDIT',
                category='មន្ទីរមិនឯកភាពស្នើសុំមេដាយ',
                details=f"ថ្នាក់ដឹកនាំមន្ទីរមិនទាន់ឯកភាពសំណើសុំមេដាយរបស់ «{req_obj.officer.full_name_kh}» (មូលហេតុ៖ {admin_notes})"
            )
            messages.info(request, f"🔴 បានកត់ត្រាមិនទាន់ឯកភាពកម្រិតមន្ទីរលើសំណើរបស់ «{req_obj.officer.full_name_kh}»។")

        # Stage 2: Submission to Ministry
        elif action == 'SUBMIT_MINISTRY':
            doc_no = request.POST.get('ministry_doc_number', '').strip()
            sub_date_str = request.POST.get('ministry_submission_date', '').strip()
            req_obj.status = 'SUBMITTED'
            req_obj.is_submitted_to_ministry = True
            req_obj.ministry_doc_number = doc_no or req_obj.ministry_doc_number
            if sub_date_str:
                try:
                    req_obj.ministry_submission_date = datetime.strptime(sub_date_str, '%Y-%m-%d').date()
                except Exception:
                    req_obj.ministry_submission_date = date.today()
            else:
                req_obj.ministry_submission_date = date.today()

            log_officer_action(
                request,
                officer=req_obj.officer,
                action_type='EDIT',
                category='បញ្ជូនសំណើមេដាយទៅក្រសួង',
                details=f"បានកត់ត្រាការបញ្ជូនសំណើមេដាយរបស់ «{req_obj.officer.full_name_kh}» ទៅក្រសួង តាមលិខិតលេខ៖ {req_obj.ministry_doc_number}"
            )
            messages.success(request, f"🏛️ បានកត់ត្រាការបញ្ជូនសំណើរបស់ «{req_obj.officer.full_name_kh}» ទៅក្រសួងរួចរាល់!")

        # Stage 3: Official Ministry / Royal Decree Decision
        elif action == 'MINISTRY_APPROVE':
            legal_type = request.POST.get('legal_doc_type', 'ROYAL_DECREE')
            legal_num = request.POST.get('legal_doc_number', '').strip()
            legal_date_str = request.POST.get('legal_doc_date', '').strip()
            min_notes = request.POST.get('ministry_decision_notes', '').strip()

            req_obj.status = 'MINISTRY_APPROVED'
            req_obj.legal_doc_type = legal_type
            req_obj.legal_doc_number = legal_num
            req_obj.ministry_decision_notes = min_notes

            parsed_legal_date = None
            if legal_date_str:
                try:
                    parsed_legal_date = datetime.strptime(legal_date_str, '%Y-%m-%d').date()
                except Exception:
                    parsed_legal_date = date.today()
            else:
                parsed_legal_date = date.today()

            req_obj.legal_doc_date = parsed_legal_date
            req_obj.is_profile_updated = True

            # 💥 NOW UPDATE OFFICER AWARDS DATA
            officer = req_obj.officer
            doc_title = f"{req_obj.get_legal_doc_type_display()} លេខ {legal_num}" if legal_num else f"ព្រះរាជក្រឹត្យ/អនុក្រឹត្យ ឆ្នាំ {req_obj.request_year}"
            date_str = parsed_legal_date.strftime('%d/%m/%Y') if parsed_legal_date else date.today().strftime('%d/%m/%Y')
            award_entry = {
                'doc_number': doc_title,
                'date': date_str,
                'ministry': 'រាជរដ្ឋាភិបាលកម្ពុជា / ព្រះមហាក្សត្រ',
                'description': req_obj.proposed_medal,
                'type': 'គ្រឿងឥស្សរិយយស/មេដាយការងារ'
            }
            awards = officer.awards_data or []
            awards.append(award_entry)
            officer.awards_data = awards
            officer.save()

            log_officer_action(
                request,
                officer=officer,
                action_type='EDIT',
                category='ក្រសួង/ព្រះមហាក្សត្រប្រទានមេដាយផ្លូវការ',
                details=f"បានទទួលស្គាល់ការអនុម័តជាផ្លូវការ ({doc_title}) ប្រទានមេដាយ «{req_obj.proposed_medal}» ជូនមន្ត្រី «{officer.full_name_kh}»"
            )
            messages.success(request, f"👑 បានអនុម័តជាផ្លូវការ ({doc_title})! ប្រព័ន្ធបានកត់ត្រាមេដាយ «{req_obj.proposed_medal}» ចូលក្នុងប្រវត្តិរូបមន្ត្រីរួចរាល់។")

        elif action == 'MINISTRY_REJECT':
            min_notes = request.POST.get('ministry_decision_notes', '').strip()
            req_obj.status = 'MINISTRY_REJECTED'
            req_obj.ministry_decision_notes = min_notes

            log_officer_action(
                request,
                officer=req_obj.officer,
                action_type='EDIT',
                category='ក្រសួងបដិសេធសំណើមេដាយ',
                details=f"ក្រសួងមិនបានអនុម័តសំណើសុំមេដាយរបស់ «{req_obj.officer.full_name_kh}» (កំណត់សម្គាល់៖ {min_notes})។"
            )
            messages.warning(request, f"❌ បានកត់ត្រាការបដិសេធ/ធ្លាក់ពីក្រសួងសម្រាប់សំណើមេដាយរបស់ «{req_obj.officer.full_name_kh}»។")

        req_obj.save()

    return redirect(request.META.get('HTTP_REFERER', 'officer_medals'))


@login_required
def officer_medal_request_delete(request, pk):
    profile = getattr(request.user, 'profile', None)
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    req_obj = get_object_or_404(OfficerMedalRequest, pk=pk)

    if not has_global_access and req_obj.submitted_by != request.user:
        messages.error(request, "⚠️ លោកអ្នកគ្មានសិទ្ធិលុបសំណើនេះឡើយ!")
        return redirect('officer_medals')

    if req_obj.status == 'APPROVED' and not has_global_access:
        messages.error(request, "⚠️ សំណើដែលបានអនុម័តរួច មិនអាចលុបបានឡើយ!")
        return redirect('officer_medals')

    if request.method == 'POST':
        officer_name = req_obj.officer.full_name_kh
        req_obj.delete()
        messages.success(request, f"🗑️ បានលុបសំណើសុំមេដាយរបស់ «{officer_name}» ដោយជោគជ័យ!")

    return redirect('officer_medals')


# ==============================================================================
# 📑 MASTER CONSOLIDATED PROPOSAL DOSSIER & MINISTRY EXPORTS (តារាងស្នើសុំសរុប ជូនក្រសួង)
# ==============================================================================

@login_required
def officer_promotion_master_export_excel(request):
    """
    Excel Export សម្រាប់តារាងស្នើសុំដំឡើងថ្នាក់សរុប (សម្រាប់កិច្ចប្រជុំ និងបញ្ជូនទៅក្រសួង)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    
    req_year = request.GET.get('year') or date.today().year
    dept_id = request.GET.get('department')
    status_filter = request.GET.get('status')
    
    qs = OfficerPromotionRequest.objects.select_related('officer', 'department').filter(request_year=req_year)
    if not has_global_access:
        if user_dept:
            qs = qs.filter(department=user_dept)
        else:
            qs = qs.filter(submitted_by=request.user)
    elif dept_id:
        qs = qs.filter(department_id=dept_id)
        
    if status_filter and status_filter != 'ALL':
        qs = qs.filter(status=status_filter)
        
    qs = qs.order_by('department__order_index', 'officer__khmer_last_name', 'officer__khmer_first_name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"តារាងដំឡើងថ្នាក់ {req_year}"

    # Styling
    font_royal_muol = Font(name='Khmer OS Muol Light', size=11, bold=True)
    font_title_muol = Font(name='Khmer OS Muol Light', size=13, bold=True, color='B25900')
    font_body = Font(name='Khmer OS Battambang', size=10)
    font_body_bold = Font(name='Khmer OS Battambang', size=10, bold=True)
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='A0A0A0'),
        right=Side(style='thin', color='A0A0A0'),
        top=Side(style='thin', color='A0A0A0'),
        bottom=Side(style='thin', color='A0A0A0')
    )
    header_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
    header_font = Font(name='Khmer OS Muol Light', size=9.5, bold=True, color='FFFFFF')

    # Row 1-3: Royal Header
    ws.merge_cells('A1:D1')
    ws['A1'] = "ក្រសួងកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទ"
    ws['A1'].font = font_royal_muol

    ws.merge_cells('A2:D2')
    ws['A2'] = "មន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទ ខេត្តប៉ៃលិន"
    ws['A2'].font = font_royal_muol

    ws.merge_cells('K1:N1')
    ws['K1'] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws['K1'].font = font_royal_muol
    ws['K1'].alignment = align_center

    ws.merge_cells('K2:N2')
    ws['K2'] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws['K2'].font = font_royal_muol
    ws['K2'].alignment = align_center

    # Royal Divider Flourish
    ws.merge_cells('K3:N3')
    ws['K3'] = ""
    _add_excel_khmer_divider(ws, col_idx=11, col_offset_px=40, row_idx=2, row_offset_px=2, width_px=65, height_px=18)

    # Title Row
    ws.merge_cells('A4:N4')
    ws['A4'] = f"តារាងបូកសរុបសំណើសុំដំឡើងឋានន្តរស័ក្តិ និងថ្នាក់មន្ត្រីរាជការ ប្រចាំឆ្នាំ {req_year}"
    ws['A4'].font = font_title_muol
    ws['A4'].alignment = align_center

    ws.merge_cells('A5:N5')
    ws['A5'] = "(សម្រាប់ដាក់ឆ្លងកិច្ចប្រជុំគណៈកម្មការវាយតម្លៃ និងរៀបចំឯកសារបញ្ជូនទៅក្រសួង)"
    ws['A5'].font = font_body
    ws['A5'].alignment = align_center

    headers = [
        'ល.រ', 'អត្តលេខ', 'គោត្តនាម-នាម', 'ភេទ', 'ថ្ងៃខែឆ្នាំកំណើត',
        'មុខតំណែង & អង្គភាព', 'ថ្ងៃចូលបម្រើ', 'ថ្ងៃតាំងស៊ប់', 'ថ្នាក់បច្ចុប្បន្ន',
        'ថ្នាក់ស្នើសុំដំឡើង', 'ប្រភេទដំឡើង', 'អតីតភាពក្នុងថ្នាក់', 'មូលហេតុ & សមិទ្ធផល',
        'មតិអង្គប្រជុំ / គណៈកម្មការ'
    ]
    
    ws.append([]) # Row 6 empty
    ws.append(headers) # Row 7
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=7, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[7].height = 28

    current_row = 8
    for idx, r in enumerate(qs, 1):
        o = r.officer
        row_data = [
            idx,
            o.officer_id_number or '-',
            o.full_name_kh,
            o.get_gender_display(),
            o.birth_date or '-',
            f"{o.current_position_title or '-'} / {r.department.name_kh if r.department else '-'}",
            o.civil_service_start_date or '-',
            o.civil_service_permanent_date or '-',
            r.current_rank_and_step or '-',
            r.proposed_rank_and_step,
            r.get_promotion_type_display(),
            r.years_in_current_rank or '-',
            r.reason or '-',
            r.meeting_decision or 'ឯកភាពតាមការស្នើសុំ'
        ]
        ws.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            c = ws.cell(row=current_row, column=col_idx)
            c.font = font_body
            c.border = thin_border
            if col_idx in [1, 2, 4, 5, 7, 8, 9, 10, 11, 12]:
                c.alignment = align_center
            else:
                c.alignment = align_left
        current_row += 1

    # Signature rows
    current_row += 2
    ws.cell(row=current_row, column=2, value="អ្នករៀបចំតារាង").font = font_royal_muol
    ws.cell(row=current_row, column=6, value="បានឃើញ និងពិនិត្យត្រឹមត្រូវ\nប្រធានការិយាល័យរដ្ឋបាល និងបុគ្គលិក").font = font_royal_muol
    ws.cell(row=current_row, column=6).alignment = align_center
    ws.cell(row=current_row, column=12, value=f"ថ្ងៃ................... ខែ............. ឆ្នាំ{date.today().year}\nបានឃើញ និងឯកភាព\nប្រធានមន្ទីរ").font = font_royal_muol
    ws.cell(row=current_row, column=12).alignment = align_center

    # Column widths
    col_widths = [6, 12, 22, 8, 14, 28, 14, 14, 15, 16, 20, 15, 30, 28]
    for i, w in enumerate(col_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"master_promotion_dossier_{req_year}_{date.today().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def officer_promotion_master_print_view(request):
    """
    Print View សម្រាប់តារាងស្នើសុំដំឡើងថ្នាក់សរុប (A4 Landscape Formatted)
    """
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    
    req_year = request.GET.get('year') or date.today().year
    dept_id = request.GET.get('department')
    status_filter = request.GET.get('status')
    
    qs = OfficerPromotionRequest.objects.select_related('officer', 'department').filter(request_year=req_year)
    if not has_global_access:
        if user_dept:
            qs = qs.filter(department=user_dept)
        else:
            qs = qs.filter(submitted_by=request.user)
    elif dept_id:
        qs = qs.filter(department_id=dept_id)
        
    if status_filter and status_filter != 'ALL':
        qs = qs.filter(status=status_filter)
        
    qs = qs.order_by('department__order_index', 'officer__khmer_last_name', 'officer__khmer_first_name')
    
    female_count = qs.filter(officer__gender='FEMALE').count()

    context = {
        'req_year': req_year,
        'requests_list': qs,
        'female_count': female_count,
        'today': date.today(),
    }
    return render(request, 'dms/officer_promotion_master_print.html', context)


@login_required
def officer_promotion_batch_status_update(request):
    """
    Admin Bulk Update: Supports Department Approval, Ministry Submission, and Official Ministry Approval / Rejection
    """
    has_global_access = check_has_global_hr_tracking_access(request.user, getattr(request.user, 'profile', None))
    if not has_global_access:
        messages.error(request, "⚠️ មានតែ Admin/ថ្នាក់ដឹកនាំ ប៉ុណ្ណោះដែលអាចកែប្រែទិន្នន័យរួមបាន!")
        return redirect('officer_promotion')

    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_request_ids')
        action_type = request.POST.get('batch_action_type')
        meeting_decision = request.POST.get('batch_meeting_decision', '').strip()
        ministry_doc_number = request.POST.get('batch_ministry_doc_number', '').strip()
        legal_doc_type = request.POST.get('batch_legal_doc_type', 'PRAKAS')
        legal_doc_number = request.POST.get('batch_legal_doc_number', '').strip()
        legal_doc_date_str = request.POST.get('batch_legal_doc_date', '').strip()

        if not selected_ids:
            messages.warning(request, "⚠️ សូមជ្រើសរើសយ៉ាងហោចណាស់មួយសំណើ!")
            return redirect('officer_promotion')

        reqs = OfficerPromotionRequest.objects.select_related('officer').filter(id__in=selected_ids)
        updated_count = 0

        parsed_legal_date = None
        if legal_doc_date_str:
            try:
                parsed_legal_date = datetime.strptime(legal_doc_date_str, '%Y-%m-%d').date()
            except Exception:
                parsed_legal_date = date.today()
        else:
            parsed_legal_date = date.today()

        for r in reqs:
            if action_type in ['DEPT_APPROVE_ALL', 'APPROVE_ALL']:
                r.status = 'DEPT_APPROVED'
                r.reviewed_by = request.user
                r.reviewed_at = timezone.now()
                if meeting_decision:
                    r.meeting_decision = meeting_decision
            elif action_type == 'DEPT_REJECT_ALL':
                r.status = 'DEPT_REJECTED'
                r.reviewed_by = request.user
                r.reviewed_at = timezone.now()
            elif action_type == 'SUBMIT_MINISTRY':
                r.status = 'SUBMITTED'
                r.is_submitted_to_ministry = True
                r.ministry_doc_number = ministry_doc_number or r.ministry_doc_number
                r.ministry_submission_date = date.today()
            elif action_type == 'MINISTRY_APPROVE_ALL':
                r.status = 'MINISTRY_APPROVED'
                r.legal_doc_type = legal_doc_type
                r.legal_doc_number = legal_doc_number or r.legal_doc_number
                r.legal_doc_date = parsed_legal_date
                r.effective_date = parsed_legal_date
                r.is_profile_updated = True

                # 💥 UPDATE OFFICER PROFILE
                officer = r.officer
                old_rank = officer.current_rank_and_step
                officer.current_rank_and_step = r.proposed_rank_and_step
                officer.last_promotion_date = parsed_legal_date

                doc_title = f"{r.get_legal_doc_type_display()} លេខ {r.legal_doc_number}" if r.legal_doc_number else "ប្រកាស/អនុក្រឹត្យផ្លូវការ"
                promo_entry = {
                    'effective_date': parsed_legal_date.strftime('%d/%m/%Y'),
                    'ministry': 'ក្រសួងកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទ',
                    'department': officer.display_department_name,
                    'office': officer.display_department_name,
                    'old_rank_step': old_rank or '',
                    'new_rank_step': r.proposed_rank_and_step,
                    'promo_type': r.get_promotion_type_display(),
                    'doc_number': doc_title
                }
                history = officer.promotions_by_seniority or []
                history.append(promo_entry)
                officer.promotions_by_seniority = history
                officer.save()

            elif action_type == 'MINISTRY_REJECT_ALL':
                r.status = 'MINISTRY_REJECTED'
                r.ministry_decision_notes = request.POST.get('batch_ministry_notes', 'ក្រសួងមិនអនុម័ត')
            elif action_type == 'UPDATE_MEETING_DECISION':
                if meeting_decision:
                    r.meeting_decision = meeting_decision

            r.save()
            updated_count += 1

        messages.success(request, f"🎉 បានអនុវត្តសកម្មភាពរួមលើសំណើសុំដំឡើងថ្នាក់សរុបចំនួន {updated_count} ដោយជោគជ័យ!")

    return redirect(request.META.get('HTTP_REFERER', 'officer_promotion'))


@login_required
def officer_medals_master_export_excel(request):
    """
    Excel Export សម្រាប់តារាងស្នើសុំគ្រឿងឥស្សរិយយស & មេដាយសរុប (សម្រាប់កិច្ចប្រជុំ និងបញ្ជូនទៅក្រសួង)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    
    req_year = request.GET.get('year') or date.today().year
    dept_id = request.GET.get('department')
    status_filter = request.GET.get('status')
    
    qs = OfficerMedalRequest.objects.select_related('officer', 'department').filter(request_year=req_year)
    if not has_global_access:
        if user_dept:
            qs = qs.filter(department=user_dept)
        else:
            qs = qs.filter(submitted_by=request.user)
    elif dept_id:
        qs = qs.filter(department_id=dept_id)
        
    if status_filter and status_filter != 'ALL':
        qs = qs.filter(status=status_filter)
        
    qs = qs.order_by('department__order_index', 'officer__khmer_last_name', 'officer__khmer_first_name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"តារាងមេដាយ {req_year}"

    # Styling
    font_royal_muol = Font(name='Khmer OS Muol Light', size=11, bold=True)
    font_title_muol = Font(name='Khmer OS Muol Light', size=13, bold=True, color='15803D')
    font_body = Font(name='Khmer OS Battambang', size=10)
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='A0A0A0'),
        right=Side(style='thin', color='A0A0A0'),
        top=Side(style='thin', color='A0A0A0'),
        bottom=Side(style='thin', color='A0A0A0')
    )
    header_fill = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')
    header_font = Font(name='Khmer OS Muol Light', size=9.5, bold=True, color='FFFFFF')

    # Row 1-3: Royal Header
    ws.merge_cells('A1:D1')
    ws['A1'] = "ក្រសួងកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទ"
    ws['A1'].font = font_royal_muol

    ws.merge_cells('A2:D2')
    ws['A2'] = "មន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទ ខេត្តប៉ៃលិន"
    ws['A2'].font = font_royal_muol

    ws.merge_cells('I1:L1')
    ws['I1'] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws['I1'].font = font_royal_muol
    ws['I1'].alignment = align_center

    ws.merge_cells('I2:L2')
    ws['I2'] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws['I2'].font = font_royal_muol
    ws['I2'].alignment = align_center

    # Royal Divider Flourish
    ws.merge_cells('I3:L3')
    ws['I3'] = ""
    _add_excel_khmer_divider(ws, col_idx=9, col_offset_px=40, row_idx=2, row_offset_px=2, width_px=65, height_px=18)

    # Title Row
    ws.merge_cells('A4:L4')
    ws['A4'] = f"តារាងបូកសរុបសំណើសុំគ្រឿងឥស្សរិយយស និងមេដាយការងារ ប្រចាំឆ្នាំ {req_year}"
    ws['A4'].font = font_title_muol
    ws['A4'].alignment = align_center

    ws.merge_cells('A5:L5')
    ws['A5'] = "(សម្រាប់ដាក់ឆ្លងកិច្ចប្រជុំគណៈកម្មការវាយតម្លៃ និងរៀបចំឯកសារបញ្ជូនទៅក្រសួង)"
    ws['A5'].font = font_body
    ws['A5'].alignment = align_center

    headers = [
        'ល.រ', 'អត្តលេខ', 'គោត្តនាម-នាម', 'ភេទ', 'ថ្ងៃខែឆ្នាំកំណើត',
        'មុខតំណែង & អង្គភាព', 'ថ្ងៃចូលបម្រើ', 'អតីតភាពការងារ',
        'គ្រឿងឥស្សរិយយស / មេដាយស្នើសុំ', 'មេដាយធ្លាប់ទទួលបាន', 'ស្នាដៃ & គុណសម្បត្តិការងារ',
        'មតិអង្គប្រជុំ / គណៈកម្មការ'
    ]
    
    ws.append([]) # Row 6 empty
    ws.append(headers) # Row 7
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=7, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[7].height = 28

    current_row = 8
    for idx, r in enumerate(qs, 1):
        o = r.officer
        # Past awards summary string
        past_awards = [a.get('description', '') for a in (o.awards_data or []) if a.get('description')]
        past_str = ", ".join(past_awards[:2]) if past_awards else '-'

        row_data = [
            idx,
            o.officer_id_number or '-',
            o.full_name_kh,
            o.get_gender_display(),
            o.birth_date or '-',
            f"{o.current_position_title or '-'} / {r.department.name_kh if r.department else '-'}",
            o.civil_service_start_date or '-',
            f"{r.years_of_service} ឆ្នាំ" if r.years_of_service else '-',
            r.proposed_medal,
            past_str,
            r.achievements or '-',
            r.meeting_decision or 'ឯកភាពតាមការស្នើសុំ'
        ]
        ws.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            c = ws.cell(row=current_row, column=col_idx)
            c.font = font_body
            c.border = thin_border
            if col_idx in [1, 2, 4, 5, 7, 8]:
                c.alignment = align_center
            else:
                c.alignment = align_left
        current_row += 1

    # Signature rows
    current_row += 2
    ws.cell(row=current_row, column=2, value="អ្នករៀបចំតារាង").font = font_royal_muol
    ws.cell(row=current_row, column=5, value="បានឃើញ និងពិនិត្យត្រឹមត្រូវ\nប្រធានការិយាល័យរដ្ឋបាល និងបុគ្គលិក").font = font_royal_muol
    ws.cell(row=current_row, column=5).alignment = align_center
    ws.cell(row=current_row, column=10, value=f"ថ្ងៃ................... ខែ............. ឆ្នាំ{date.today().year}\nបានឃើញ និងឯកភាព\nប្រធានមន្ទីរ").font = font_royal_muol
    ws.cell(row=current_row, column=10).alignment = align_center

    # Column widths
    col_widths = [6, 12, 22, 8, 14, 28, 14, 15, 28, 22, 30, 28]
    for i, w in enumerate(col_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"master_medals_dossier_{req_year}_{date.today().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def officer_medals_master_print_view(request):
    """
    Print View សម្រាប់តារាងស្នើសុំមេដាយសរុប (A4 Landscape Formatted)
    """
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    
    req_year = request.GET.get('year') or date.today().year
    dept_id = request.GET.get('department')
    status_filter = request.GET.get('status')
    
    qs = OfficerMedalRequest.objects.select_related('officer', 'department').filter(request_year=req_year)
    if not has_global_access:
        if user_dept:
            qs = qs.filter(department=user_dept)
        else:
            qs = qs.filter(submitted_by=request.user)
    elif dept_id:
        qs = qs.filter(department_id=dept_id)
        
    if status_filter and status_filter != 'ALL':
        qs = qs.filter(status=status_filter)
        
    qs = qs.order_by('department__order_index', 'officer__khmer_last_name', 'officer__khmer_first_name')
    
    for r in qs:
        past_awards = [a.get('description', '') for a in (r.officer.awards_data or []) if a.get('description')]
        r.past_awards_str = ", ".join(past_awards[:2]) if past_awards else '-'
        
    female_count = qs.filter(officer__gender='FEMALE').count()

    context = {
        'req_year': req_year,
        'requests_list': qs,
        'female_count': female_count,
        'today': date.today(),
    }
    return render(request, 'dms/officer_medals_master_print.html', context)


@login_required
def officer_medals_batch_status_update(request):
    """
    Admin Bulk Update: Update meeting decision or ministry submission for multiple medal requests
    """
    has_global_access = check_has_global_hr_tracking_access(request.user, getattr(request.user, 'profile', None))
    if not has_global_access:
        messages.error(request, "⚠️ មានតែ Admin/ថ្នាក់ដឹកនាំ ប៉ុណ្ណោះដែលអាចកែប្រែទិន្នន័យរួមបាន!")
        return redirect('officer_medals')

    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_request_ids')
        action_type = request.POST.get('batch_action_type') # 'APPROVE_ALL', 'SUBMIT_MINISTRY', 'UPDATE_MEETING_DECISION'
        meeting_decision = request.POST.get('batch_meeting_decision', '').strip()
        ministry_doc_number = request.POST.get('batch_ministry_doc_number', '').strip()

        if not selected_ids:
            messages.warning(request, "⚠️ សូមជ្រើសរើសយ៉ាងហោចណាស់មួយសំណើ!")
            return redirect('officer_medals')

        reqs = OfficerMedalRequest.objects.select_related('officer').filter(id__in=selected_ids)
        updated_count = 0

        legal_doc_type = request.POST.get('batch_legal_doc_type', 'ROYAL_DECREE')
        legal_doc_number = request.POST.get('batch_legal_doc_number', '').strip()
        legal_doc_date_str = request.POST.get('batch_legal_doc_date', '').strip()

        parsed_legal_date = None
        if legal_doc_date_str:
            try:
                parsed_legal_date = datetime.strptime(legal_doc_date_str, '%Y-%m-%d').date()
            except Exception:
                parsed_legal_date = date.today()
        else:
            parsed_legal_date = date.today()

        for r in reqs:
            if action_type in ['DEPT_APPROVE_ALL', 'APPROVE_ALL']:
                r.status = 'DEPT_APPROVED'
                r.reviewed_by = request.user
                r.reviewed_at = timezone.now()
                if meeting_decision:
                    r.meeting_decision = meeting_decision
            elif action_type == 'DEPT_REJECT_ALL':
                r.status = 'DEPT_REJECTED'
                r.reviewed_by = request.user
                r.reviewed_at = timezone.now()
            elif action_type == 'SUBMIT_MINISTRY':
                r.status = 'SUBMITTED'
                r.is_submitted_to_ministry = True
                r.ministry_doc_number = ministry_doc_number or r.ministry_doc_number
                r.ministry_submission_date = date.today()
            elif action_type == 'MINISTRY_APPROVE_ALL':
                r.status = 'MINISTRY_APPROVED'
                r.legal_doc_type = legal_doc_type
                r.legal_doc_number = legal_doc_number or r.legal_doc_number
                r.legal_doc_date = parsed_legal_date
                r.effective_date = parsed_legal_date
                r.is_profile_updated = True

                # 💥 UPDATE OFFICER AWARDS DATA
                officer = r.officer
                doc_title = f"{r.get_legal_doc_type_display()} លេខ {r.legal_doc_number}" if r.legal_doc_number else f"ព្រះរាជក្រឹត្យ/អនុក្រឹត្យ ឆ្នាំ {r.request_year}"
                date_str = parsed_legal_date.strftime('%d/%m/%Y')
                award_entry = {
                    'doc_number': doc_title,
                    'date': date_str,
                    'ministry': 'រាជរដ្ឋាភិបាលកម្ពុជា / ព្រះមហាក្សត្រ',
                    'description': r.proposed_medal,
                    'type': 'គ្រឿងឥស្សរិយយស/មេដាយការងារ'
                }
                awards = officer.awards_data or []
                awards.append(award_entry)
                officer.awards_data = awards
                officer.save()

            elif action_type == 'MINISTRY_REJECT_ALL':
                r.status = 'MINISTRY_REJECTED'
                r.ministry_decision_notes = request.POST.get('batch_ministry_notes', 'ក្រសួងមិនអនុម័ត')
            elif action_type == 'UPDATE_MEETING_DECISION':
                if meeting_decision:
                    r.meeting_decision = meeting_decision

            r.save()
            updated_count += 1

        messages.success(request, f"🎉 បានអនុវត្តសកម្មភាពរួមលើសំណើសុំមេដាយសរុបចំនួន {updated_count} ដោយជោគជ័យ!")

    return redirect(request.META.get('HTTP_REFERER', 'officer_medals'))


@login_required
def officer_tracking_view(request, default_tab='retirement'):
    tab = request.GET.get('tab', default_tab)
    if tab == 'promotion':
        return redirect('officer_promotion')
    elif tab == 'medals':
        return redirect('officer_medals')
    return redirect('officer_retirement')


@login_required
def officer_tracking_retirement_view(request):
    return redirect('officer_retirement')


@login_required
def officer_tracking_promotion_view(request):
    return redirect('officer_promotion')


@login_required
def officer_tracking_medals_view(request):
    return redirect('officer_medals')


@login_required
def officer_retirement_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import datetime
    
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    
    today = datetime.date.today()
    dept_id = request.GET.get('department')
    search_q = request.GET.get('q', '').strip()
    year_param = request.GET.get('year', '').strip()
    status_filter = request.GET.get('status', '').strip()

    officers = CivilServantProfile.objects.select_related('department').all().order_by('khmer_last_name', 'khmer_first_name')
    
    if not has_global_access:
        if user_dept:
            officers = officers.filter(department=user_dept)
        else:
            officers = officers.filter(created_by=request.user)
    elif dept_id:
        officers = officers.filter(department_id=dept_id)

    if search_q:
        q_arabic = to_arabic_digits(search_q)
        officers = officers.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_last_name__icontains=search_q) |
            Q(latin_first_name__icontains=search_q) |
            Q(officer_id_number__icontains=search_q) |
            Q(officer_id_number__icontains=q_arabic)
        )
    
    selected_year = None
    if year_param:
        try:
            selected_year = int(to_arabic_digits(year_param))
        except Exception:
            selected_year = None

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    header_font = Font(name='Khmer OS Battambang', size=11, bold=True, color='FFFFFF')
    align_center = Alignment(horizontal='center', vertical='center')
    
    dept_title_suffix = f" - {user_dept.name_kh}" if not has_global_access and user_dept else ""

    status_title_map = {
        'THIS_YEAR': f" (ត្រូវចូលនិវត្តន៍ក្នុងឆ្នាំនេះ {today.year})",
        'NEXT_YEAR': f" (ត្រូវចូលនិវត្តន៍ក្នុងឆ្នាំក្រោយ {today.year + 1})",
        'IN_2_YEARS': " (ត្រូវចូលនិវត្តន៍ក្នុងរយៈពេល ២ ឆ្នាំ)",
        'ACTIVE': " (នៅសល់លើសពី ៣ ឆ្នាំ)",
        'RETIRED': " (បានចូលនិវត្តន៍រួចរាល់)",
    }
    status_suffix = status_title_map.get(status_filter, "")
    year_title_suffix = f" ប្រចាំឆ្នាំ {selected_year}" if selected_year else ""
    
    ws = wb.active
    ws.title = f"មន្ត្រីចូលនិវត្តន៍ {selected_year}" if selected_year else "មន្ត្រីត្រូវចូលនិវត្តន៍"
    
    ws.merge_cells('A1:I1')
    ws['A1'] = f"របាយការណ៍ព្យាករណ៍មន្ត្រីរាជការត្រូវចូលនិវត្តន៍ (អាយុ ៦០ ឆ្នាំ){year_title_suffix}{status_suffix}{dept_title_suffix}"
    ws['A1'].font = Font(name='Khmer OS Muol Light', size=13, bold=True, color='C00000')
    ws['A1'].alignment = align_center
    
    headers = ['ល.រ', 'អត្តលេខ', 'គោត្តនាម-នាមខ្លួន', 'ភេទ', 'ថ្ងៃខែឆ្នាំកំណើត', 'អាយុ', 'អង្គភាព & មុខតំណែង', 'កាលបរិច្ឆេទចូលនិវត្តន៍', 'ស្ថានភាព']
    ws.append([])
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center
        
    ret_list = [process_officer_retirement(o, today) for o in officers]
    ret_list.sort(key=lambda x: (x['days_left'] < 0, x['days_left']))
    
    # 🎯 Filter by selected year if requested
    if selected_year:
        ret_list = [r for r in ret_list if r.get('retirement_year') == selected_year]

    # 🎯 Filter by active status if selected
    if status_filter and status_filter != 'ALL':
        ret_list = [r for r in ret_list if r['status'] == status_filter]
    
    for idx, r in enumerate(ret_list, 1):
        o = r['officer']
        age_val = r.get('age')
        ws.append([
            idx,
            o.officer_id_number or '-',
            o.full_name_kh,
            o.get_gender_display(),
            r.get('birth_date_str') or '-',
            f"{age_val} ឆ្នាំ" if age_val is not None else '-',
            f"{o.current_position_title or '-'} / {o.department.name_kh if o.department else '-'}",
            r.get('retirement_date_str') or '-',
            r.get('status_label') or '-'
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    year_tag = f"_{selected_year}" if selected_year else ""
    filter_tag = f"_{status_filter.lower()}" if status_filter and status_filter != 'ALL' else ""
    filename = f"retirement_report{year_tag}{filter_tag}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def officer_promotion_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import datetime
    
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    
    today = datetime.date.today()
    dept_id = request.GET.get('department')
    search_q = request.GET.get('q', '').strip()
    forecast_year_param = request.GET.get('forecast_year') or request.GET.get('year')
    if forecast_year_param:
        try:
            forecast_year = int(to_arabic_digits(forecast_year_param))
        except Exception:
            forecast_year = today.year
    else:
        forecast_year = today.year

    status_filter = request.GET.get('status', '').strip()

    officers = CivilServantProfile.objects.select_related('department').all().order_by('khmer_last_name', 'khmer_first_name')
    
    if not has_global_access:
        if user_dept:
            officers = officers.filter(department=user_dept)
        else:
            officers = officers.filter(created_by=request.user)
    elif dept_id:
        officers = officers.filter(department_id=dept_id)

    if search_q:
        q_arabic = to_arabic_digits(search_q)
        officers = officers.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_last_name__icontains=search_q) |
            Q(latin_first_name__icontains=search_q) |
            Q(officer_id_number__icontains=search_q) |
            Q(officer_id_number__icontains=q_arabic)
        )
    
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color='B25900', end_color='B25900', fill_type='solid')
    header_font = Font(name='Khmer OS Battambang', size=11, bold=True, color='FFFFFF')
    align_center = Alignment(horizontal='center', vertical='center')
    
    dept_title_suffix = f" - {user_dept.name_kh}" if not has_global_access and user_dept else ""

    status_title_map = {
        'DUE_ALL': " (ដល់វេនសរុបក្នុងឆ្នាំ)",
        'DUE_NEW': f" (ដល់វេនថ្មីក្នុងឆ្នាំ {forecast_year})",
        'CARRIED_OVER': " (ធ្លាក់ឆ្នាំមុន / ដល់វេនអតីតភាពស្វ័យប្រវត្តិ)",
        'DUE_SOON': " (ជិតដល់វេន - ឆ្នាំបន្ទាប់)",
        'MAX_RANK_EXAM_ONLY': " (ដល់កម្រិតកំពូល គ.១ / ខ.១.១ - ប្រឡងប្តូរក្របខ័ណ្ឌ)",
        'NORMAL': " (មិនទាន់ដល់វេន)",
    }
    status_suffix = status_title_map.get(status_filter, "")
    
    ws = wb.active
    ws.title = f"ដំឡើងថ្នាក់ {forecast_year}"
    
    ws.merge_cells('A1:I1')
    ws['A1'] = f"បញ្ជីមន្ត្រីរាជការដល់វេនត្រូវស្នើសុំដំឡើងឋានន្តរស័ក្តិ និងថ្នាក់ ប្រចាំឆ្នាំ {forecast_year} (គិត ២ ឆ្នាំម្តង){status_suffix}{dept_title_suffix}"
    ws['A1'].font = Font(name='Khmer OS Muol Light', size=13, bold=True, color='B25900')
    ws['A1'].alignment = align_center
    
    headers = ['ល.រ', 'អត្តលេខ', 'គោត្តនាម-នាមខ្លួន', 'មុខតំណែង & អង្គភាព', 'ថ្នាក់បច្ចុប្បន្ន', 'កាលបរិច្ឆេទថ្នាក់ចុងក្រោយ', f'អតីតភាពក្នុងថ្នាក់ (គិតត្រឹម {forecast_year})', 'ថ្នាក់ត្រូវស្នើសុំបន្ទាប់', 'ស្ថានភាពវេន & ប្រវត្តិធ្លាក់ឆ្នាំមុន']
    ws.append([])
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center
        
    rejected_reqs = {}
    for r in OfficerPromotionRequest.objects.filter(status__in=['DEPT_REJECTED', 'MINISTRY_REJECTED', 'REJECTED']).order_by('request_year'):
        rejected_reqs[r.officer_id] = r

    promo_list = [process_officer_promotion(o, today, target_year=forecast_year, rejected_req=rejected_reqs.get(o.id)) for o in officers]
    promo_list.sort(key=lambda x: (x['status'] not in ['DUE_NEW', 'CARRIED_OVER'], -x['years_in_rank']))
    
    # 🎯 Filter by active status if selected
    if status_filter == 'DUE_ALL':
        promo_list = [p for p in promo_list if p['status'] in ['DUE_NEW', 'CARRIED_OVER']]
    elif status_filter and status_filter != 'ALL':
        promo_list = [p for p in promo_list if p['status'] == status_filter]

    for idx, p in enumerate(promo_list, 1):
        o = p['officer']
        ws.append([
            idx,
            o.officer_id_number or '-',
            o.full_name_kh,
            f"{o.current_position_title or '-'} / {o.department.name_kh if o.department else '-'}",
            o.current_rank_and_step or '-',
            p.get('last_promo_date_str') or '-',
            f"{p.get('years_in_rank', 0)} ឆ្នាំ ({p.get('months_in_rank', 0)} ខែ)",
            p.get('next_suggested_rank') or '-',
            p.get('status_label') or '-'
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filter_tag = f"_{status_filter.lower()}" if status_filter and status_filter != 'ALL' else ""
    filename = f"promotion_report_{forecast_year}{filter_tag}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def officer_medals_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import datetime
    
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    
    today = datetime.date.today()
    dept_id = request.GET.get('department')
    search_q = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()

    officers = CivilServantProfile.objects.select_related('department').all().order_by('khmer_last_name', 'khmer_first_name')
    
    if not has_global_access:
        if user_dept:
            officers = officers.filter(department=user_dept)
        else:
            officers = officers.filter(created_by=request.user)
    elif dept_id:
        officers = officers.filter(department_id=dept_id)

    if search_q:
        officers = officers.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_last_name__icontains=search_q) |
            Q(latin_first_name__icontains=search_q) |
            Q(officer_id_number__icontains=search_q)
        )
    
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color='1A5D1A', end_color='1A5D1A', fill_type='solid')
    header_font = Font(name='Khmer OS Battambang', size=11, bold=True, color='FFFFFF')
    align_center = Alignment(horizontal='center', vertical='center')
    
    dept_title_suffix = f" - {user_dept.name_kh}" if not has_global_access and user_dept else ""
    
    status_title_map = {
        'ELIGIBLE': " (គ្រប់លក្ខខណ្ឌស្នើសុំមេដាយ)",
        'NOT_YET': " (មិនទាន់គ្រប់លក្ខខណ្ឌ)",
    }
    status_suffix = status_title_map.get(status_filter, "")

    ws = wb.active
    ws.title = "ស្នើសុំមេដាយ"
    
    ws.merge_cells('A1:G1')
    ws['A1'] = f"បញ្ជីមន្ត្រីរាជការគ្រប់លក្ខខណ្ឌស្នើសុំមេដាយ និងគ្រឿងឥស្សរិយយសការងារ{status_suffix}{dept_title_suffix}"
    ws['A1'].font = Font(name='Khmer OS Muol Light', size=13, bold=True, color='1A5D1A')
    ws['A1'].alignment = align_center
    
    headers = ['ល.រ', 'អត្តលេខ', 'គោត្តនាម-នាមខ្លួន', 'អង្គភាព & មុខតំណែង', 'ថ្ងៃចូលបម្រើការងារ', 'អតីតភាពការងារសរុប', 'មេដាយ/គ្រឿងឥស្សរិយយសត្រូវស្នើសុំ']
    ws.append([])
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center
        
    med_list = [process_officer_medals(o, today) for o in officers]
    med_list.sort(key=lambda x: x['years_of_service'], reverse=True)
    
    # 🎯 Filter by active status if selected
    if status_filter and status_filter != 'ALL':
        med_list = [m for m in med_list if m['status'] == status_filter]

    for idx, m in enumerate(med_list, 1):
        o = m['officer']
        ws.append([
            idx,
            o.officer_id_number or '-',
            o.full_name_kh,
            f"{o.current_position_title or '-'} / {o.department.name_kh if o.department else '-'}",
            m.get('start_date_str') or '-',
            f"{m.get('years_of_service', 0)} ឆ្នាំ",
            m.get('recommended_medal') or '-'
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filter_tag = f"_{status_filter.lower()}" if status_filter and status_filter != 'ALL' else ""
    filename = f"medals_report{filter_tag}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def officer_tracking_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import datetime
    
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    
    today = datetime.date.today()
    dept_id = request.GET.get('department')
    officers = CivilServantProfile.objects.select_related('department').all().order_by('khmer_last_name', 'khmer_first_name')
    
    if not has_global_access:
        if user_dept:
            officers = officers.filter(department=user_dept)
        else:
            officers = officers.filter(created_by=request.user)
    elif dept_id:
        officers = officers.filter(department_id=dept_id)
    
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(name='Khmer OS Battambang', size=11, bold=True, color='FFFFFF')
    align_center = Alignment(horizontal='center', vertical='center')
    
    dept_title_suffix = f" - {user_dept.name_kh}" if not has_global_access and user_dept else ""
    
    # Sheet 1: ចូលនិវត្តន៍ (Retirement)
    ws1 = wb.active
    ws1.title = "មន្ត្រីត្រូវចូលនិវត្តន៍"
    
    ws1.merge_cells('A1:H1')
    ws1['A1'] = f"របាយការណ៍ព្យាករណ៍មន្ត្រីរាជការត្រូវចូលនិវត្តន៍ (អាយុ ៦០ ឆ្នាំ){dept_title_suffix}"
    ws1['A1'].font = Font(name='Khmer OS Muol Light', size=13, bold=True, color='1F4E78')
    ws1['A1'].alignment = align_center
    
    headers1 = ['ល.រ', 'អត្តលេខ', 'គោត្តនាម-នាមខ្លួន', 'ភេទ', 'ថ្ងៃខែឆ្នាំកំណើត', 'អាយុ', 'អង្គភាព & មុខតំណែង', 'ស្ថានភាពចូលនិវត្តន៍']
    ws1.append([])
    ws1.append(headers1)
    for col_idx in range(1, len(headers1) + 1):
        c = ws1.cell(row=3, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center
        
    ret_list = [process_officer_retirement(o, today) for o in officers]
    ret_list.sort(key=lambda x: (x['days_left'] < 0, x['days_left']))
    
    for idx, r in enumerate(ret_list, 1):
        o = r['officer']
        age_val = r.get('age')
        ws1.append([
            idx,
            o.officer_id_number or '-',
            o.full_name_kh,
            o.get_gender_display(),
            r.get('birth_date_str') or '-',
            f"{age_val} ឆ្នាំ" if age_val is not None else '-',
            f"{o.current_position_title or '-'} / {o.department.name_kh if o.department else '-'}",
            r.get('status_label') or '-'
        ])
        
    # Sheet 2: ដំឡើងថ្នាក់ ២ឆ្នាំម្តង (Promotions)
    ws2 = wb.create_sheet(title="ស្នើសុំដំឡើងថ្នាក់ (២ឆ្នាំម្តង)")
    ws2.merge_cells('A1:H1')
    ws2['A1'] = f"បញ្ជីមន្ត្រីរាជការដល់វេនត្រូវស្នើសុំដំឡើងឋានន្តរស័ក្តិ និងថ្នាក់ (គិត ២ ឆ្នាំម្តង){dept_title_suffix}"
    ws2['A1'].font = Font(name='Khmer OS Muol Light', size=13, bold=True, color='1F4E78')
    ws2['A1'].alignment = align_center
    
    headers2 = ['ល.រ', 'អត្តលេខ', 'គោត្តនាម-នាមខ្លួន', 'មុខតំណែង & អង្គភាព', 'ថ្នាក់បច្ចុប្បន្ន', 'កាលបរិច្ឆេទថ្នាក់ចុងក្រោយ', 'អតីតភាពក្នុងថ្នាក់', 'ថ្នាក់ត្រូវស្នើសុំបន្ទាប់', 'ស្ថានភាព']
    ws2.append([])
    ws2.append(headers2)
    for col_idx in range(1, len(headers2) + 1):
        c = ws2.cell(row=3, column=col_idx)
        c.fill = PatternFill(start_color='B25900', end_color='B25900', fill_type='solid')
        c.font = header_font
        c.alignment = align_center
        
    promo_list = [process_officer_promotion(o, today) for o in officers]
    promo_list.sort(key=lambda x: x['years_in_rank'], reverse=True)
    
    for idx, p in enumerate(promo_list, 1):
        o = p['officer']
        ws2.append([
            idx,
            o.officer_id_number or '-',
            o.full_name_kh,
            f"{o.current_position_title or '-'} / {o.department.name_kh if o.department else '-'}",
            o.current_rank_and_step or '-',
            p['last_promo_date_str'],
            f"{p['years_in_rank']} ឆ្នាំ ({p['months_in_rank']} ខែ)",
            p['next_suggested_rank'],
            p['status_label']
        ])
        
    # Sheet 3: ស្នើសុំមេដាយ (Medals & Honors)
    ws3 = wb.create_sheet(title="ស្នើសុំមេដាយ-គ្រឿងឥស្សរិយយស")
    ws3.merge_cells('A1:G1')
    ws3['A1'] = f"បញ្ជីមន្ត្រីរាជការគ្រប់លក្ខខណ្ឌស្នើសុំមេដាយ និងគ្រឿងឥស្សរិយយសការងារ{dept_title_suffix}"
    ws3['A1'].font = Font(name='Khmer OS Muol Light', size=13, bold=True, color='1F4E78')
    ws3['A1'].alignment = align_center
    
    headers3 = ['ល.រ', 'អត្តលេខ', 'គោត្តនាម-នាមខ្លួន', 'អង្គភាព & មុខតំណែង', 'ថ្ងៃចូលបម្រើការងារ', 'អតីតភាពការងារសរុប', 'មេដាយ/គ្រឿងឥស្សរិយយសត្រូវស្នើសុំ']
    ws3.append([])
    ws3.append(headers3)
    for col_idx in range(1, len(headers3) + 1):
        c = ws3.cell(row=3, column=col_idx)
        c.fill = PatternFill(start_color='1A5D1A', end_color='1A5D1A', fill_type='solid')
        c.font = header_font
        c.alignment = align_center
        
    med_list = [process_officer_medals(o, today) for o in officers]
    med_list.sort(key=lambda x: x['years_of_service'], reverse=True)
    
    for idx, m in enumerate(med_list, 1):
        o = m['officer']
        ws3.append([
            idx,
            o.officer_id_number or '-',
            o.full_name_kh,
            f"{o.current_position_title or '-'} / {o.department.name_kh if o.department else '-'}",
            m['start_date_str'],
            f"{m['years_of_service']} ឆ្នាំ",
            m['recommended_medal']
        ])
        
    for sheet in [ws1, ws2, ws3]:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 15)
            
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"HR_Tracking_Report_{today.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ==============================================================================
# 📊 OFFICER STATUS REPORT & OVERVIEW (ផ្ទាំងរបាយការណ៍ និងស្ថានភាពមន្ត្រី)
# ==============================================================================

def get_khmer_lunar_and_solar_date(d=None):
    """
    Computes official formatted solar date and lunar calendar text for formal Cambodian administrative reports.
    """
    import datetime
    if d is None:
        d = datetime.date.today()

    kh_days = {
        0: 'ចន្ទ', 1: 'អង្គារ', 2: 'ពុធ', 3: 'ព្រហស្បតិ៍', 4: 'សុក្រ', 5: 'សៅរ៍', 6: 'អាទិត្យ'
    }
    kh_months = {
        1: 'មករា', 2: 'កុម្ភៈ', 3: 'មីនា', 4: 'មេសា', 5: 'ឧសភា', 6: 'មិថុនា',
        7: 'កក្កដា', 8: 'សីហា', 9: 'កញ្ញា', 10: 'តុលា', 11: 'វិច្ឆិកា', 12: 'ធ្នូ'
    }
    zodiac_animals = ['ជូត', 'ឆ្លូវ', 'ខាល', 'ថោះ', 'រោង', 'ម្សាញ់', 'មមី', 'មមែ', 'វក', 'រកា', 'ចរ', 'កុរ']
    sak_names = ['ឯកស័ក', 'ទោស័ក', 'ត្រីស័ក', 'ចត្វាស័ក', 'បញ្ចស័ក', 'ឆស័ក', 'សប្តស័ក', 'អដ្ឋស័ក', 'នព្វស័ក', 'សំរឹទ្ធិស័ក']

    be_year = d.year + 544 if d.month >= 5 else d.year + 543
    zodiac = zodiac_animals[(d.year - 4) % 12]
    sak = sak_names[(d.year - 4) % 10]
    day_name = kh_days.get(d.weekday(), 'ចន្ទ')
    month_name = kh_months.get(d.month, 'មករា')

    solar_date_str = f"ថ្ងៃទី{to_khmer_digits(d.day)} ខែ{month_name} ឆ្នាំ{to_khmer_digits(d.year)}"
    lunar_date_str = f"ថ្ងៃ{day_name} ...កើត/រោច ខែ... ឆ្នាំ{zodiac} {sak} ព.ស {to_khmer_digits(be_year)}"

    return {
        'solar_date_str': solar_date_str,
        'lunar_date_str': lunar_date_str,
        'day_kh': to_khmer_digits(d.day),
        'month_kh': month_name,
        'year_kh': to_khmer_digits(d.year),
        'be_year_kh': to_khmer_digits(be_year),
        'day_name_kh': day_name,
        'zodiac_kh': zodiac,
        'sak_kh': sak,
    }


@login_required
def officer_status_report_view(request):
    """
    Dedicated Interactive Dashboard & Overview for Civil Servant Status (ស្ថានភាពមន្ត្រីរាជការ)
    Categorizes officers into:
    ក. ព័ត៌មានមន្ត្រីតាមក្របខ័ណ្ឌ (Framework A, B, C, D)
    ខ. ព័ត៌មានមន្ត្រីតាមសញ្ញាប័ត្រ (Doctorate, Master, Bachelor, Associate, High School, Other)
    គ. ស្ថានភាពមន្ត្រី គិតមកដល់បច្ចុប្បន្ន (Active, Unpaid, Outside FW, Dismissed, Retired, Trainee, Transferred)
    """
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    dept_id = request.GET.get('department', '').strip()
    fw_filter = request.GET.get('framework', '').strip()
    degree_filter = request.GET.get('degree', '').strip()
    status_filter = request.GET.get('status', '').strip()
    gender_filter = request.GET.get('gender', '').strip()
    search_q = request.GET.get('q', '').strip()

    base_qs = CivilServantProfile.objects.select_related('department', 'user').all()

    if not has_global_access:
        if user_dept:
            base_qs = base_qs.filter(department=user_dept)
        else:
            base_qs = base_qs.filter(created_by=request.user)
    elif dept_id:
        base_qs = base_qs.filter(department_id=dept_id)

    # Calculate Aggregated Statistics for the Scope (Department or Whole Organization)
    total_officers = base_qs.count()
    female_count = base_qs.filter(gender='FEMALE').count()
    male_count = base_qs.filter(gender='MALE').count()

    # ក. តាមក្របខ័ណ្ឌ
    cat_a_count = base_qs.filter(framework_category='A').count()
    cat_b_count = base_qs.filter(framework_category='B').count()
    cat_c_count = base_qs.filter(framework_category='C').count()
    cat_d_count = base_qs.filter(framework_category='D').count()

    # ខ. តាមសញ្ញាប័ត្រ
    phd_count = base_qs.filter(highest_degree='DOCTORATE').count()
    master_count = base_qs.filter(highest_degree='MASTER').count()
    bachelor_count = base_qs.filter(highest_degree='BACHELOR').count()
    associate_count = base_qs.filter(highest_degree='ASSOCIATE').count()
    highschool_count = base_qs.filter(highest_degree='HIGHSCHOOL').count()
    other_degree_count = base_qs.filter(Q(highest_degree='OTHER') | Q(highest_degree__isnull=True) | Q(highest_degree='')).count()

    # គ. តាមស្ថានភាពមន្ត្រី
    active_count = base_qs.filter(officer_status='ACTIVE').count()
    unpaid_leave_count = base_qs.filter(officer_status='UNPAID_LEAVE').count()
    outside_framework_count = base_qs.filter(officer_status='OUTSIDE_FRAMEWORK').count()
    dismissed_count = base_qs.filter(officer_status='DISMISSED').count()
    retired_count = base_qs.filter(officer_status='RETIRED').count()
    trainee_count = base_qs.filter(officer_status='TRAINEE').count()
    transferred_out_count = base_qs.filter(officer_status='TRANSFERRED_OUT').count()

    # Filter queryset for the table display
    table_qs = base_qs
    if fw_filter and fw_filter != 'ALL':
        table_qs = table_qs.filter(framework_category=fw_filter)
    if degree_filter and degree_filter != 'ALL':
        if degree_filter == 'OTHER':
            table_qs = table_qs.filter(Q(highest_degree='OTHER') | Q(highest_degree__isnull=True) | Q(highest_degree=''))
        else:
            table_qs = table_qs.filter(highest_degree=degree_filter)
    if status_filter and status_filter != 'ALL':
        table_qs = table_qs.filter(officer_status=status_filter)
    if gender_filter and gender_filter != 'ALL':
        table_qs = table_qs.filter(gender=gender_filter)
    if search_q:
        q_arabic = to_arabic_digits(search_q)
        table_qs = table_qs.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_last_name__icontains=search_q) |
            Q(latin_first_name__icontains=search_q) |
            Q(officer_id_number__icontains=search_q) |
            Q(officer_id_number__icontains=q_arabic) |
            Q(phone__icontains=search_q) |
            Q(current_position_title__icontains=search_q)
        )

    # Sort using standard sorting key
    from .models import officer_sort_key
    officers_list = list(table_qs)
    officers_list.sort(key=officer_sort_key)

    paginator = Paginator(officers_list, 30)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh') if has_global_access else (Department.objects.filter(id=user_dept.id) if user_dept else [])

    selected_dept_obj = None
    if dept_id:
        selected_dept_obj = Department.objects.filter(id=dept_id).first()

    today = date.today()
    date_info = get_khmer_lunar_and_solar_date(today)

    edit_window = OfficerEditWindowSetting.get_setting()
    is_window_open = edit_window.is_open_for_specialized()

    context = {
        'page_obj': page_obj,
        'officers': page_obj.object_list,
        'total_in_table': len(officers_list),
        'has_global_access': has_global_access,
        'is_admin_or_lead': is_admin_or_lead,
        'is_system_admin': is_system_admin,
        'user_dept': user_dept,
        'departments': departments,
        'selected_dept': dept_id,
        'selected_dept_obj': selected_dept_obj,
        'fw_filter': fw_filter,
        'degree_filter': degree_filter,
        'status_filter': status_filter,
        'gender_filter': gender_filter,
        'search_q': search_q,

        # Khmer Digits & Stats for Section ក
        'total_officers': total_officers,
        'total_officers_kh': to_khmer_digits(total_officers),
        'female_count': female_count,
        'female_count_kh': to_khmer_digits(female_count),
        'male_count': male_count,
        'male_count_kh': to_khmer_digits(male_count),
        'cat_a_count': cat_a_count,
        'cat_a_count_kh': to_khmer_digits(cat_a_count),
        'cat_b_count': cat_b_count,
        'cat_b_count_kh': to_khmer_digits(cat_b_count),
        'cat_c_count': cat_c_count,
        'cat_c_count_kh': to_khmer_digits(cat_c_count),
        'cat_d_count': cat_d_count,
        'cat_d_count_kh': to_khmer_digits(cat_d_count),

        # Section ខ
        'phd_count': phd_count,
        'phd_count_kh': to_khmer_digits(phd_count),
        'master_count': master_count,
        'master_count_kh': to_khmer_digits(master_count),
        'bachelor_count': bachelor_count,
        'bachelor_count_kh': to_khmer_digits(bachelor_count),
        'associate_count': associate_count,
        'associate_count_kh': to_khmer_digits(associate_count),
        'highschool_count': highschool_count,
        'highschool_count_kh': to_khmer_digits(highschool_count),
        'other_degree_count': other_degree_count,
        'other_degree_count_kh': to_khmer_digits(other_degree_count),

        # Section គ
        'active_count': active_count,
        'active_count_kh': to_khmer_digits(active_count),
        'unpaid_leave_count': unpaid_leave_count,
        'unpaid_leave_count_kh': to_khmer_digits(unpaid_leave_count),
        'outside_framework_count': outside_framework_count,
        'outside_framework_count_kh': to_khmer_digits(outside_framework_count),
        'dismissed_count': dismissed_count,
        'dismissed_count_kh': to_khmer_digits(dismissed_count),
        'retired_count': retired_count,
        'retired_count_kh': to_khmer_digits(retired_count),
        'trainee_count': trainee_count,
        'trainee_count_kh': to_khmer_digits(trainee_count),
        'transferred_out_count': transferred_out_count,
        'transferred_out_count_kh': to_khmer_digits(transferred_out_count),

        'status_choices': CivilServantProfile.OFFICER_STATUS_CHOICES,
        'degree_choices': CivilServantProfile.DEGREE_CHOICES,
        'framework_choices': CivilServantProfile.FRAMEWORK_CHOICES,
        'date_info': date_info,
        'today': today,
        'is_window_open': is_window_open,
        'can_export_officer_excel': can_export_civil_servants_to_excel(request.user, profile),
        'can_view_e2_report': can_access_officer_e2_report(request.user, profile),
    }
    return render(request, 'dms/officer_status_report.html', context)


@login_required
def officer_status_report_print_view(request):
    """
    Official Printable Format of Civil Servant Status Report (ស្ថានភាពមន្ត្រីរាជការ)
    Designed 1-to-1 to perfectly match the formal Cambodian administration document layout.
    """
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)

    dept_id = request.GET.get('department', '').strip()
    custom_date_str = request.GET.get('date', '').strip()
    custom_lunar_str = request.GET.get('lunar_text', '').strip()
    location_str = request.GET.get('location', 'ប៉ៃលិន').strip()
    include_roster = request.GET.get('include_roster') == '1'

    base_qs = CivilServantProfile.objects.select_related('department').all()

    if not has_global_access:
        if user_dept:
            base_qs = base_qs.filter(department=user_dept)
        else:
            base_qs = base_qs.filter(created_by=request.user)
    elif dept_id:
        base_qs = base_qs.filter(department_id=dept_id)

    selected_dept_obj = Department.objects.filter(id=dept_id).first() if dept_id else None

    # Aggregations
    total_officers = base_qs.count()
    female_count = base_qs.filter(gender='FEMALE').count()
    male_count = base_qs.filter(gender='MALE').count()

    cat_a_count = base_qs.filter(framework_category='A').count()
    cat_b_count = base_qs.filter(framework_category='B').count()
    cat_c_count = base_qs.filter(framework_category='C').count()
    cat_d_count = base_qs.filter(framework_category='D').count()

    phd_count = base_qs.filter(highest_degree='DOCTORATE').count()
    master_count = base_qs.filter(highest_degree='MASTER').count()
    bachelor_count = base_qs.filter(highest_degree='BACHELOR').count()
    associate_count = base_qs.filter(highest_degree='ASSOCIATE').count()
    highschool_count = base_qs.filter(highest_degree='HIGHSCHOOL').count()
    other_degree_count = base_qs.filter(Q(highest_degree='OTHER') | Q(highest_degree__isnull=True) | Q(highest_degree='')).count()

    active_count = base_qs.filter(officer_status='ACTIVE').count()
    unpaid_leave_count = base_qs.filter(officer_status='UNPAID_LEAVE').count()
    outside_framework_count = base_qs.filter(officer_status='OUTSIDE_FRAMEWORK').count()
    dismissed_count = base_qs.filter(officer_status='DISMISSED').count()
    retired_count = base_qs.filter(officer_status='RETIRED').count()
    trainee_count = base_qs.filter(officer_status='TRAINEE').count()
    transferred_out_count = base_qs.filter(officer_status='TRANSFERRED_OUT').count()

    today = date.today()
    date_info = get_khmer_lunar_and_solar_date(today)

    from .models import officer_sort_key
    officers_list = list(base_qs)
    officers_list.sort(key=officer_sort_key)

    context = {
        'selected_dept_obj': selected_dept_obj,
        'user_dept': user_dept,
        'has_global_access': has_global_access,
        'location_str': location_str,
        'custom_date_str': custom_date_str,
        'custom_lunar_str': custom_lunar_str,
        'include_roster': include_roster,
        'officers_list': officers_list,

        # Section ក
        'total_officers': total_officers,
        'total_officers_kh': to_khmer_digits(total_officers),
        'female_count': female_count,
        'female_count_kh': to_khmer_digits(female_count),
        'male_count': male_count,
        'male_count_kh': to_khmer_digits(male_count),
        'cat_a_count': cat_a_count,
        'cat_a_count_kh': to_khmer_digits(cat_a_count),
        'cat_b_count': cat_b_count,
        'cat_b_count_kh': to_khmer_digits(cat_b_count),
        'cat_c_count': cat_c_count,
        'cat_c_count_kh': to_khmer_digits(cat_c_count),
        'cat_d_count': cat_d_count,
        'cat_d_count_kh': to_khmer_digits(cat_d_count),

        # Section ខ
        'phd_count': phd_count,
        'phd_count_kh': to_khmer_digits(phd_count),
        'master_count': master_count,
        'master_count_kh': to_khmer_digits(master_count),
        'bachelor_count': bachelor_count,
        'bachelor_count_kh': to_khmer_digits(bachelor_count),
        'associate_count': associate_count,
        'associate_count_kh': to_khmer_digits(associate_count),
        'highschool_count': highschool_count,
        'highschool_count_kh': to_khmer_digits(highschool_count),
        'other_degree_count': other_degree_count,
        'other_degree_count_kh': to_khmer_digits(other_degree_count),

        # Section គ
        'active_count': active_count,
        'active_count_kh': to_khmer_digits(active_count),
        'unpaid_leave_count': unpaid_leave_count,
        'unpaid_leave_count_kh': to_khmer_digits(unpaid_leave_count),
        'outside_framework_count': outside_framework_count,
        'outside_framework_count_kh': to_khmer_digits(outside_framework_count),
        'dismissed_count': dismissed_count,
        'dismissed_count_kh': to_khmer_digits(dismissed_count),
        'retired_count': retired_count,
        'retired_count_kh': to_khmer_digits(retired_count),
        'trainee_count': trainee_count,
        'trainee_count_kh': to_khmer_digits(trainee_count),
        'transferred_out_count': transferred_out_count,
        'transferred_out_count_kh': to_khmer_digits(transferred_out_count),

        'date_info': date_info,
        'today': today,
    }
    return render(request, 'dms/officer_status_report_print.html', context)


@login_required
def officer_status_export_excel(request):
    """
    Excel Export of Officer Cadre Status Report with 2 sheets:
    1. របាយការណ៍ស្ថានភាពមន្ត្រី (Executive 3-Part Summary)
    2. បញ្ជីមន្ត្រីរាជការលម្អិត (Detailed officer list with Framework, Highest Degree, Status)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import datetime

    profile = getattr(request.user, 'profile', None)
    if not can_export_civil_servants_to_excel(request.user, profile):
        messages.error(
            request,
            "ការិយាល័យជំនាញ និងខណ្ឌ មិនត្រូវបានអនុញ្ញាតឱ្យទាញចេញជា Excel ឡើយ! លើកលែងតែការិយាល័យរដ្ឋបាល-បុគ្គលិក ឬមានការអនុញ្ញាតដោយ ADMIN។"
        )
        return HttpResponseForbidden(
            "<div style='font-family: Khmer OS Battambang, sans-serif; text-align: center; margin-top: 80px;'>"
            "<h2 style='color: #dc2626;'>⛔ គ្មានសិទ្ធិទាញយកឯកសារជា Excel ឡើយ</h2>"
            "<p style='font-size: 16px; color: #475569;'>ការិយាល័យជំនាញ និងខណ្ឌ មិនត្រូវបានអនុញ្ញាតឱ្យទាញចេញជា Excel ដាច់ខាត លើកលែងតែការិយាល័យរដ្ឋបាល-បុគ្គលិក ឬមានការអនុញ្ញាតដោយ ADMIN។</p>"
            "<a href='javascript:history.back()' style='display: inline-block; margin-top: 15px; padding: 8px 20px; background: #2563eb; color: #fff; text-decoration: none; border-radius: 20px;'>ត្រឡប់ក្រោយ</a>"
            "</div>"
        )

    user_dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)

    dept_id = request.GET.get('department', '').strip()
    base_qs = CivilServantProfile.objects.select_related('department').all()

    if not has_global_access:
        if user_dept:
            base_qs = base_qs.filter(department=user_dept)
        else:
            base_qs = base_qs.filter(created_by=request.user)
    elif dept_id:
        base_qs = base_qs.filter(department_id=dept_id)

    selected_dept_obj = Department.objects.filter(id=dept_id).first() if dept_id else None
    dept_label = selected_dept_obj.name_kh if selected_dept_obj else (user_dept.name_kh if not has_global_access and user_dept else "មន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទខេត្តប៉ៃលិន")

    # Aggregations
    total_officers = base_qs.count()
    female_count = base_qs.filter(gender='FEMALE').count()
    male_count = base_qs.filter(gender='MALE').count()

    cat_a_count = base_qs.filter(framework_category='A').count()
    cat_b_count = base_qs.filter(framework_category='B').count()
    cat_c_count = base_qs.filter(framework_category='C').count()
    cat_d_count = base_qs.filter(framework_category='D').count()

    phd_count = base_qs.filter(highest_degree='DOCTORATE').count()
    master_count = base_qs.filter(highest_degree='MASTER').count()
    bachelor_count = base_qs.filter(highest_degree='BACHELOR').count()
    associate_count = base_qs.filter(highest_degree='ASSOCIATE').count()
    highschool_count = base_qs.filter(highest_degree='HIGHSCHOOL').count()
    other_degree_count = base_qs.filter(Q(highest_degree='OTHER') | Q(highest_degree__isnull=True) | Q(highest_degree='')).count()

    active_count = base_qs.filter(officer_status='ACTIVE').count()
    unpaid_leave_count = base_qs.filter(officer_status='UNPAID_LEAVE').count()
    outside_framework_count = base_qs.filter(officer_status='OUTSIDE_FRAMEWORK').count()
    dismissed_count = base_qs.filter(officer_status='DISMISSED').count()
    retired_count = base_qs.filter(officer_status='RETIRED').count()
    trainee_count = base_qs.filter(officer_status='TRAINEE').count()
    transferred_out_count = base_qs.filter(officer_status='TRANSFERRED_OUT').count()

    wb = openpyxl.Workbook()

    # Fonts & Styles
    f_royal_muol = Font(name='Khmer OS Muol Light', size=11, bold=True)
    f_title_muol = Font(name='Khmer OS Muol Light', size=14, bold=True, color='1F4E78')
    f_section_muol = Font(name='Khmer OS Muol Light', size=11, bold=True, color='1F4E78')
    f_header = Font(name='Khmer OS Battambang', size=10.5, bold=True, color='FFFFFF')
    f_body = Font(name='Khmer OS Battambang', size=10)
    f_body_bold = Font(name='Khmer OS Battambang', size=10, bold=True)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    fill_sec = PatternFill(start_color='E8EEF5', end_color='E8EEF5', fill_type='solid')

    border_thin = Border(
        left=Side(style='thin', color='A0A0A0'),
        right=Side(style='thin', color='A0A0A0'),
        top=Side(style='thin', color='A0A0A0'),
        bottom=Side(style='thin', color='A0A0A0')
    )

    # -------------------------------------------------------------
    # SHEET 1: របាយការណ៍ស្ថានភាពមន្ត្រីរាជការ (Summary)
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "ស្ថានភាពមន្ត្រីរាជការ"

    ws1.merge_cells('A1:D1')
    ws1['A1'] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws1['A1'].font = f_royal_muol
    ws1['A1'].alignment = align_center

    ws1.merge_cells('A2:D2')
    ws1['A2'] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws1['A2'].font = f_royal_muol
    ws1['A2'].alignment = align_center

    ws1.merge_cells('A4:D4')
    ws1['A4'] = "ក្រសួងកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទ"
    ws1['A4'].font = f_royal_muol

    ws1.merge_cells('A5:D5')
    ws1['A5'] = dept_label
    ws1['A5'].font = f_body_bold

    ws1.merge_cells('A7:D7')
    ws1['A7'] = "ស្ថានភាពមន្ត្រីរាជការ"
    ws1['A7'].font = f_title_muol
    ws1['A7'].alignment = align_center

    # Section ក
    ws1.merge_cells('A9:D9')
    ws1['A9'] = "ក. ព័ត៌មានមន្ត្រីតាមក្របខ័ណ្ឌ"
    ws1['A9'].font = f_section_muol
    ws1['A9'].fill = fill_sec

    ws1.cell(row=10, column=1, value="- មន្ត្រីរាជការសរុបរបស់អង្គភាព").font = f_body_bold
    ws1.cell(row=10, column=2, value=f"{total_officers} នាក់ ( ស្រី៖ {female_count} នាក់  ប្រុស៖ {male_count} នាក់ )").font = f_body
    
    ws1.cell(row=11, column=1, value="- ចំនួនមន្ត្រីរាជការតាមក្របខណ្ឌ").font = f_body_bold
    ws1.cell(row=11, column=2, value=f"(ក) {cat_a_count} នាក់   (ខ) {cat_b_count} នាក់   (គ) {cat_c_count} នាក់   (ឃ) {cat_d_count} នាក់").font = f_body

    # Section ខ
    ws1.merge_cells('A13:D13')
    ws1['A13'] = "ខ. ព័ត៌មានមន្ត្រីតាមសញ្ញាប័ត្រ"
    ws1['A13'].font = f_section_muol
    ws1['A13'].fill = fill_sec

    degrees_rows = [
        ("(១). បណ្ឌិត", f"{phd_count} នាក់", "(២). បរិញ្ញាបត្រជាន់ខ្ពស់(អនុបណ្ឌិត)", f"{master_count} នាក់"),
        ("(៣). បរិញ្ញាប័ត្រ", f"{bachelor_count} នាក់", "(៤). បរិញ្ញាប័ត្ររង", f"{associate_count} នាក់"),
        ("(៥). មធ្យមសិក្សាទុតិយភូមិ", f"{highschool_count} នាក់", "(៦). ផ្សេងៗ", f"{other_degree_count} នាក់"),
    ]
    for r_idx, (d1_k, d1_v, d2_k, d2_v) in enumerate(degrees_rows, start=14):
        ws1.cell(row=r_idx, column=1, value=d1_k).font = f_body
        ws1.cell(row=r_idx, column=2, value=d1_v).font = f_body_bold
        ws1.cell(row=r_idx, column=3, value=d2_k).font = f_body
        ws1.cell(row=r_idx, column=4, value=d2_v).font = f_body_bold

    # Section គ
    ws1.merge_cells('A18:D18')
    ws1['A18'] = "គ. ស្ថានភាពមន្ត្រី (គិតមកដល់បច្ចុប្បន្ន)"
    ws1['A18'].font = f_section_muol
    ws1['A18'].fill = fill_sec

    status_rows = [
        ("១. មន្ត្រីសកម្ម", f"{active_count} នាក់", "៥. មន្ត្រីចូលនិវត្តន៍", f"{retired_count} នាក់"),
        ("២. មន្ត្រីស្ថិតក្នុងភាពទំនេរគ្មានបៀវត្ស", f"{unpaid_leave_count} នាក់", "៦. មន្ត្រីកម្មសិក្សា", f"{trainee_count} នាក់"),
        ("៣. មន្ត្រីស្ថិតនៅក្រៅក្របខណ្ឌដើម", f"{outside_framework_count} នាក់", "៧. មន្ត្រីផ្ទេរចេញ", f"{transferred_out_count} នាក់"),
        ("៤. មន្ត្រីត្រូវបានលុបឈ្មោះ", f"{dismissed_count} នាក់", "", ""),
    ]
    for r_idx, (s1_k, s1_v, s2_k, s2_v) in enumerate(status_rows, start=19):
        ws1.cell(row=r_idx, column=1, value=s1_k).font = f_body
        ws1.cell(row=r_idx, column=2, value=s1_v).font = f_body_bold
        ws1.cell(row=r_idx, column=3, value=s2_k).font = f_body
        ws1.cell(row=r_idx, column=4, value=s2_v).font = f_body_bold

    today = datetime.date.today()
    d_info = get_khmer_lunar_and_solar_date(today)

    ws1.merge_cells('C25:D25')
    ws1['C25'] = "បានឃើញ និងឯកភាព"
    ws1['C25'].font = f_royal_muol
    ws1['C25'].alignment = align_center

    ws1.merge_cells('C26:D26')
    ws1['C26'] = f"ធ្វើនៅប៉ៃលិន, {d_info['solar_date_str']}"
    ws1['C26'].font = f_body
    ws1['C26'].alignment = align_center

    ws1.merge_cells('C27:D27')
    ws1['C27'] = "ប្រធានអង្គភាព"
    ws1['C27'].font = f_royal_muol
    ws1['C27'].alignment = align_center

    ws1.column_dimensions['A'].width = 32
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 34
    ws1.column_dimensions['D'].width = 25

    # -------------------------------------------------------------
    # SHEET 2: បញ្ជីមន្ត្រីរាជការលម្អិត (Detailed List)
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="បញ្ជីមន្ត្រីរាជការលម្អិត")

    ws2.merge_cells('A1:J1')
    ws2['A1'] = f"បញ្ជីរាយនាមមន្ត្រីរាជការ - ស្ថានភាព ក្របខ័ណ្ឌ និងសញ្ញាបត្រ ({dept_label})"
    ws2['A1'].font = f_title_muol
    ws2['A1'].alignment = align_center

    ws2.merge_cells('A2:J2')
    ws2['A2'] = f"កាលបរិច្ឆេទ Export៖ {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')} | សរុប៖ {total_officers} នាក់ (ស្រី: {female_count} / ប្រុស: {male_count})"
    ws2['A2'].font = Font(name='Khmer OS Battambang', size=10, italic=True, color='555555')
    ws2['A2'].alignment = align_center

    headers2 = [
        'ល.រ', 'អត្តលេខ', 'គោត្តនាម-នាមខ្លួន (ខ្មែរ)', 'ភេទ', 'ថ្ងៃខែឆ្នាំកំណើត',
        'ការិយាល័យ/អង្គភាព', 'មុខតំណែង', 'ក្របខ័ណ្ឌ', 'សញ្ញាបត្រខ្ពស់បំផុត', 'ស្ថានភាពមន្ត្រី'
    ]
    ws2.append([])
    ws2.append(headers2)

    for col_num in range(1, len(headers2) + 1):
        c = ws2.cell(row=4, column=col_num)
        c.font = f_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_thin

    from .models import officer_sort_key
    officers_list = list(base_qs)
    officers_list.sort(key=officer_sort_key)

    for idx, o in enumerate(officers_list, start=1):
        row = [
            idx,
            o.officer_id_number or '-',
            o.full_name_kh,
            o.get_gender_display(),
            o.dob or '-',
            o.department.name_kh if o.department else 'មិនទាន់កំណត់',
            o.current_position_title or '-',
            o.framework_category_label,
            o.highest_degree_label,
            o.officer_status_label,
        ]
        ws2.append(row)
        curr_row = 4 + idx
        for col_num in range(1, len(row) + 1):
            cell = ws2.cell(row=curr_row, column=col_num)
            cell.font = f_body
            cell.border = border_thin
            if col_num in [1, 2, 4, 5, 8]:
                cell.alignment = align_center
            elif col_num in [9, 10]:
                cell.alignment = align_center
                cell.font = f_body_bold
            else:
                cell.alignment = align_left

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 14)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"officer_status_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def api_officer_status_update(request):
    """
    AJAX API for Quick Status, Degree, or Framework Category Update directly from the Status Report table.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    officer_id = request.POST.get('officer_id')
    if not officer_id:
        return JsonResponse({'success': False, 'error': 'Missing officer ID'}, status=400)

    officer = get_object_or_404(CivilServantProfile, pk=officer_id)

    if not is_admin_or_lead and (not dept or officer.department != dept):
        return JsonResponse({'success': False, 'error': 'លោកអ្នកគ្មានសិទ្ធិកែប្រែស្ថានភាពមន្ត្រីនៅក្រៅការិយាល័យរបស់ខ្លួនឡើយ!'}, status=403)

    edit_window = OfficerEditWindowSetting.get_setting()
    if not is_system_admin and not edit_window.is_open_for_department(officer.department or dept):
        return JsonResponse({'success': False, 'error': f'ប្រព័ន្ធត្រូវបានបិទមិនឱ្យកែប្រែឡើយ! ({edit_window.status_label_kh})'}, status=403)

    new_status = request.POST.get('officer_status')
    new_degree = request.POST.get('highest_degree')
    new_framework = request.POST.get('framework_category')

    changes = []
    if new_status and new_status in dict(CivilServantProfile.OFFICER_STATUS_CHOICES):
        old_st = officer.officer_status_label
        officer.officer_status = new_status
        changes.append(f"ស្ថានភាព៖ {old_st} ➔ {officer.officer_status_label}")

    if new_degree and new_degree in dict(CivilServantProfile.DEGREE_CHOICES):
        old_deg = officer.highest_degree_label
        officer.highest_degree = new_degree
        changes.append(f"សញ្ញាបត្រ៖ {old_deg} ➔ {officer.highest_degree_label}")

    if new_framework and new_framework in dict(CivilServantProfile.FRAMEWORK_CHOICES):
        old_fw = officer.framework_category_label
        officer.framework_category = new_framework
        changes.append(f"ក្របខ័ណ្ឌ៖ {old_fw} ➔ {officer.framework_category_label}")

    if changes:
        officer.save()
        log_officer_action(
            request,
            officer=officer,
            action_type='PROFILE_EDIT',
            category='ស្ថានភាព & កម្រិតសញ្ញាបត្រ',
            details=f"បានកែប្រែស្ថានភាពមន្ត្រី «{officer.full_name_kh}»៖ " + ", ".join(changes)
        )

        return JsonResponse({
            'success': True,
            'message': f"បានកែប្រែទិន្នន័យសម្រាប់ «{officer.full_name_kh}» ដោយជោគជ័យ!",
            'officer_id': officer.id,
            'officer_name_kh': officer.full_name_kh,
            'officer_status': officer.officer_status,
            'officer_status_label': officer.officer_status_label,
            'officer_status_badge_class': officer.officer_status_badge_class,
            'highest_degree': officer.highest_degree,
            'highest_degree_label': officer.highest_degree_label,
            'framework_category': officer.framework_category,
            'framework_category_label': officer.framework_category_label,
        })

    return JsonResponse({'success': False, 'error': 'No valid changes provided'}, status=400)


# ==============================================================================
# CAMBODIA GEOGRAPHY API (PROVINCES, DISTRICTS, COMMUNES, VILLAGES)
# ==============================================================================

@login_required
def api_geo_provinces(request):
    """
    Returns JSON list of all 25 provinces/cities in Cambodia.
    """
    provinces = CambodiaProvince.objects.all().order_by('code')
    data = [{'code': p.code, 'name_kh': p.name_kh, 'name_en': p.name_en} for p in provinces]
    return JsonResponse({'success': True, 'provinces': data})


@login_required
def api_geo_districts(request):
    """
    Returns JSON list of districts/khans/krongs filtered by province_code or province_name.
    """
    province_code = request.GET.get('province_code', '').strip()
    province_name = request.GET.get('province_name', '').strip()

    qs = CambodiaDistrict.objects.all().select_related('province').order_by('code')
    if province_code:
        qs = qs.filter(Q(province_id=province_code) | Q(province__code=province_code))
    elif province_name:
        qs = qs.filter(province__name_kh=province_name)

    data = [{
        'code': d.code,
        'province_code': d.province_id,
        'name_kh': d.name_kh,
        'name_en': d.name_en
    } for d in qs]
    return JsonResponse({'success': True, 'districts': data})


@login_required
def api_geo_communes(request):
    """
    Returns JSON list of communes/sangkats filtered by district_code or district_name.
    """
    district_code = request.GET.get('district_code', '').strip()
    district_name = request.GET.get('district_name', '').strip()
    province_code = request.GET.get('province_code', '').strip()

    qs = CambodiaCommune.objects.all().select_related('district', 'province').order_by('code')
    if district_code:
        qs = qs.filter(district_id=district_code)
    elif district_name:
        qs = qs.filter(district__name_kh=district_name)
        if province_code:
            qs = qs.filter(Q(province_id=province_code) | Q(province__code=province_code))

    data = [{
        'code': c.code,
        'district_code': c.district_id,
        'province_code': c.province_id,
        'name_kh': c.name_kh,
        'name_en': c.name_en
    } for c in qs]
    return JsonResponse({'success': True, 'communes': data})


@login_required
def api_geo_villages(request):
    """
    Returns JSON list of villages/phums filtered by commune_code or commune_name.
    """
    commune_code = request.GET.get('commune_code', '').strip()
    commune_name = request.GET.get('commune_name', '').strip()
    district_code = request.GET.get('district_code', '').strip()

    qs = CambodiaVillage.objects.all().select_related('commune', 'district', 'province').order_by('name_kh', 'code')
    if commune_code:
        qs = qs.filter(commune_id=commune_code)
    elif commune_name:
        qs = qs.filter(commune__name_kh=commune_name)
        if district_code:
            qs = qs.filter(district_id=district_code)

    data = [{
        'code': v.code,
        'commune_code': v.commune_id,
        'district_code': v.district_id,
        'province_code': v.province_id,
        'name_kh': v.name_kh,
        'name_en': v.name_en
    } for v in qs]
    return JsonResponse({'success': True, 'villages': data})


# ==============================================================================
# CAMBODIA GEOGRAPHY ADMIN MANAGEMENT (CREATE, UPDATE, DELETE DISTRICT/COMMUNE/VILLAGE)
# ==============================================================================

@login_required
def geography_manage_view(request):
    """
    Dedicated Geography Management Dashboard for ADMIN to browse, search, create, and update
    Cambodia Districts, Communes, and Villages.
    """
    if not check_is_system_admin(request.user, request):
        messages.error(request, "⚠️ មានតែ ADMIN មួយគត់ដែលត្រូវបានអនុញ្ញាតឱ្យគ្រប់គ្រងទិន្នន័យភូមិសាស្ត្ររដ្ឋបាល!")
        return redirect('officer_list')

    level = request.GET.get('level', 'district')  # 'district', 'commune', 'village'
    p_code = request.GET.get('province_code', '').strip()
    d_code = request.GET.get('district_code', '').strip()
    c_code = request.GET.get('commune_code', '').strip()
    q = request.GET.get('q', '').strip()

    # Overall stats
    total_provinces = CambodiaProvince.objects.count()
    total_districts = CambodiaDistrict.objects.count()
    total_communes = CambodiaCommune.objects.count()
    total_villages = CambodiaVillage.objects.count()

    provinces = CambodiaProvince.objects.all().order_by('code')

    page_obj = None
    if level == 'district':
        qs = CambodiaDistrict.objects.all().select_related('province').order_by('province__code', 'code')
        if p_code:
            qs = qs.filter(province_id=p_code)
        if q:
            qs = qs.filter(Q(name_kh__icontains=q) | Q(name_en__icontains=q) | Q(code__icontains=q))
        paginator = Paginator(qs, 30)
        page_num = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_num)

    elif level == 'commune':
        qs = CambodiaCommune.objects.all().select_related('district', 'province').order_by('district__code', 'code')
        if p_code:
            qs = qs.filter(province_id=p_code)
        if d_code:
            qs = qs.filter(district_id=d_code)
        if q:
            qs = qs.filter(Q(name_kh__icontains=q) | Q(name_en__icontains=q) | Q(code__icontains=q))
        paginator = Paginator(qs, 30)
        page_num = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_num)

    elif level == 'village':
        qs = CambodiaVillage.objects.all().select_related('commune', 'district', 'province').order_by('commune__code', 'code')
        if p_code:
            qs = qs.filter(province_id=p_code)
        if d_code:
            qs = qs.filter(district_id=d_code)
        if c_code:
            qs = qs.filter(commune_id=c_code)
        if q:
            qs = qs.filter(Q(name_kh__icontains=q) | Q(name_en__icontains=q) | Q(code__icontains=q))
        paginator = Paginator(qs, 40)
        page_num = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_num)

    context = {
        'level': level,
        'provinces': provinces,
        'selected_p_code': p_code,
        'selected_d_code': d_code,
        'selected_c_code': c_code,
        'q': q,
        'total_provinces': total_provinces,
        'total_districts': total_districts,
        'total_communes': total_communes,
        'total_villages': total_villages,
        'page_obj': page_obj,
    }
    return render(request, 'dms/geography_manage.html', context)


@login_required
def api_geo_district_save(request):
    """
    Creates or updates a Cambodia District (Admin Only).
    """
    if not check_is_system_admin(request.user, request):
        return JsonResponse({'success': False, 'error': 'គ្មានសិទ្ធិអនុវត្តប្រតិបត្តិការនេះឡើយ!'}, status=403)

    code = request.POST.get('code', '').strip()
    province_code = request.POST.get('province_code', '').strip()
    name_kh = request.POST.get('name_kh', '').strip()
    name_en = request.POST.get('name_en', '').strip()
    is_edit = request.POST.get('is_edit') == 'true'

    if not code or not province_code or not name_kh:
        return JsonResponse({'success': False, 'error': 'សូមបំពេញ លេខកូដស្រុក, រាជធានី/ខេត្ត និង ឈ្មោះជាភាសាខ្មែរ!'})

    province = CambodiaProvince.objects.filter(code=province_code).first()
    if not province:
        return JsonResponse({'success': False, 'error': 'មិនអាចស្វែងរករាជធានី/ខេត្តដែលបានជ្រើសរើសឡើយ!'})

    try:
        if is_edit and orig_code:
            district = CambodiaDistrict.objects.filter(code=orig_code).first()
            if not district:
                return JsonResponse({'success': False, 'error': 'រកមិនឃើញទិន្នន័យស្រុកដែលត្រូវកែប្រែឡើយ!'})
            if orig_code != code and CambodiaDistrict.objects.filter(code=code).exists():
                return JsonResponse({'success': False, 'error': f'លេខកូដស្រុក «{code}» នេះមានរួចហើយក្នុងប្រព័ន្ធ!'})

            district.code = code
            district.province = province
            district.name_kh = name_kh
            district.name_en = name_en
            district.save()
            action_desc = f"បានកែប្រែស្រុក/ក្រុង/ខណ្ឌ «{name_kh}» (កូដ: {code})"
        else:
            if CambodiaDistrict.objects.filter(code=code).exists():
                return JsonResponse({'success': False, 'error': f'លេខកូដស្រុក «{code}» នេះមានរួចហើយក្នុងប្រព័ន្ធ!'})
            district = CambodiaDistrict.objects.create(
                code=code,
                province=province,
                name_kh=name_kh,
                name_en=name_en
            )
            action_desc = f"បានបន្ថែមស្រុក/ក្រុង/ខណ្ឌថ្មី «{name_kh}» (កូដ: {code})"

        log_officer_action(request, None, 'SYSTEM_SETTING', action_desc)

        return JsonResponse({
            'success': True,
            'message': action_desc,
            'district': {
                'code': district.code,
                'province_code': district.province_id,
                'province_name': district.province.name_kh if district.province else '',
                'name_kh': district.name_kh,
                'name_en': district.name_en
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def api_geo_commune_save(request):
    """
    Creates or updates a Cambodia Commune (Admin Only).
    """
    if not check_is_system_admin(request.user, request):
        return JsonResponse({'success': False, 'error': 'គ្មានសិទ្ធិអនុវត្តប្រតិបត្តិការនេះឡើយ!'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)

    code = request.POST.get('code', '').strip()
    district_code = request.POST.get('district_code', '').strip()
    province_code = request.POST.get('province_code', '').strip()
    name_kh = request.POST.get('name_kh', '').strip()
    name_en = request.POST.get('name_en', '').strip()
    is_edit = request.POST.get('is_edit') == 'true'
    orig_code = request.POST.get('orig_code', '').strip()

    if not code or not district_code or not name_kh:
        return JsonResponse({'success': False, 'error': 'សូមបំពេញ លេខកូដឃុំ/សង្កាត់, ក្រុង/ស្រុក/ខណ្ឌ និង ឈ្មោះជាភាសាខ្មែរ!'})

    district = CambodiaDistrict.objects.filter(code=district_code).select_related('province').first()
    if not district:
        return JsonResponse({'success': False, 'error': 'មិនអាចស្វែងរកស្រុកដែលបានជ្រើសរើសឡើយ!'})

    province_id = province_code or district.province_id

    try:
        if is_edit and orig_code:
            commune = CambodiaCommune.objects.filter(code=orig_code).first()
            if not commune:
                return JsonResponse({'success': False, 'error': 'រកមិនឃើញទិន្នន័យឃុំដែលត្រូវកែប្រែឡើយ!'})
            if orig_code != code and CambodiaCommune.objects.filter(code=code).exists():
                return JsonResponse({'success': False, 'error': f'លេខកូដឃុំ/សង្កាត់ «{code}» នេះមានរួចហើយក្នុងប្រព័ន្ធ!'})

            commune.code = code
            commune.district = district
            commune.province_id = province_id
            commune.name_kh = name_kh
            commune.name_en = name_en
            commune.save()
            action_desc = f"បានកែប្រែឃុំ/សង្កាត់ «{name_kh}» (កូដ: {code})"
        else:
            if CambodiaCommune.objects.filter(code=code).exists():
                return JsonResponse({'success': False, 'error': f'លេខកូដឃុំ/សង្កាត់ «{code}» នេះមានរួចហើយក្នុងប្រព័ន្ធ!'})
            commune = CambodiaCommune.objects.create(
                code=code,
                district=district,
                province_id=province_id,
                name_kh=name_kh,
                name_en=name_en
            )
            action_desc = f"បានបន្ថែមឃុំ/សង្កាត់ថ្មី «{name_kh}» (កូដ: {code})"

        log_officer_action(request, None, 'SYSTEM_SETTING', action_desc)

        return JsonResponse({
            'success': True,
            'message': action_desc,
            'commune': {
                'code': commune.code,
                'district_code': commune.district_id,
                'province_code': commune.province_id,
                'name_kh': commune.name_kh,
                'name_en': commune.name_en
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def api_geo_village_save(request):
    """
    Creates or updates a Cambodia Village (Admin Only).
    """
    if not check_is_system_admin(request.user, request):
        return JsonResponse({'success': False, 'error': 'គ្មានសិទ្ធិអនុវត្តប្រតិបត្តិការនេះឡើយ!'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)

    code = request.POST.get('code', '').strip()
    commune_code = request.POST.get('commune_code', '').strip()
    district_code = request.POST.get('district_code', '').strip()
    province_code = request.POST.get('province_code', '').strip()
    name_kh = request.POST.get('name_kh', '').strip()
    name_en = request.POST.get('name_en', '').strip()
    is_edit = request.POST.get('is_edit') == 'true'
    village_id = request.POST.get('id', '').strip()

    if not code or not commune_code or not name_kh:
        return JsonResponse({'success': False, 'error': 'សូមបំពេញ លេខកូដភូមិ, ឃុំ/សង្កាត់ និង ឈ្មោះជាភាសាខ្មែរ!'})

    commune = CambodiaCommune.objects.filter(code=commune_code).select_related('district', 'province').first()
    if not commune:
        return JsonResponse({'success': False, 'error': 'មិនអាចស្វែងរកឃុំ/សង្កាត់ដែលបានជ្រើសរើសឡើយ!'})

    dist_id = district_code or commune.district_id
    prov_id = province_code or commune.province_id

    try:
        if is_edit and village_id:
            village = CambodiaVillage.objects.filter(id=village_id).first()
            if not village:
                return JsonResponse({'success': False, 'error': 'រកមិនឃើញទិន្នន័យភូមិដែលត្រូវកែប្រែឡើយ!'})

            village.code = code
            village.commune = commune
            village.district_id = dist_id
            village.province_id = prov_id
            village.name_kh = name_kh
            village.name_en = name_en
            village.save()
            action_desc = f"បានកែប្រែភូមិ «{name_kh}» (កូដ: {code})"
        else:
            village = CambodiaVillage.objects.create(
                code=code,
                commune=commune,
                district_id=dist_id,
                province_id=prov_id,
                name_kh=name_kh,
                name_en=name_en
            )
            action_desc = f"បានបន្ថែមភូមិថ្មី «{name_kh}» (កូដ: {code})"

        log_officer_action(request, None, 'SYSTEM_SETTING', action_desc)

        return JsonResponse({
            'success': True,
            'message': action_desc,
            'village': {
                'id': village.id,
                'code': village.code,
                'commune_code': village.commune_id,
                'district_code': village.district_id,
                'province_code': village.province_id,
                'name_kh': village.name_kh,
                'name_en': village.name_en
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def api_geo_item_delete(request):
    """
    Deletes a district, commune, or village (Admin Only).
    """
    if not check_is_system_admin(request.user, request):
        return JsonResponse({'success': False, 'error': 'គ្មានសិទ្ធិអនុវត្តប្រតិបត្តិការនេះឡើយ!'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)

    level = request.POST.get('level', '').strip()
    code = request.POST.get('code', '').strip()
    item_id = request.POST.get('id', '').strip()

    try:
        if level == 'district':
            item = CambodiaDistrict.objects.filter(code=code).first()
            if not item:
                return JsonResponse({'success': False, 'error': 'រកមិនឃើញទិន្នន័យស្រុក!'})
            name = item.name_kh
            item.delete()
            action_desc = f"បានលុបស្រុក/ក្រុង «{name}» (កូដ: {code})"
        elif level == 'commune':
            item = CambodiaCommune.objects.filter(code=code).first()
            if not item:
                return JsonResponse({'success': False, 'error': 'រកមិនឃើញទិន្នន័យឃុំ!'})
            name = item.name_kh
            item.delete()
            action_desc = f"បានលុបឃុំ/សង្កាត់ «{name}» (កូដ: {code})"
        elif level == 'village':
            if item_id:
                item = CambodiaVillage.objects.filter(id=item_id).first()
            else:
                item = CambodiaVillage.objects.filter(code=code).first()
            if not item:
                return JsonResponse({'success': False, 'error': 'រកមិនឃើញទិន្នន័យភូមិ!'})
            name = item.name_kh
            item.delete()
            action_desc = f"បានលុបភូមិ «{name}» (កូដ: {item.code})"
        else:
            return JsonResponse({'success': False, 'error': 'ប្រភេទរដ្ឋបាលមិនត្រឹមត្រូវ!'})

        log_officer_action(request, None, 'SYSTEM_SETTING', action_desc)
        return JsonResponse({'success': True, 'message': action_desc})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def can_export_contract_officers_to_excel(user, profile=None):
    """
    Strict Excel Export Access Control:
    Specialized offices are FORBIDDEN from exporting contract officers to Excel,
    EXCEPT for Administration-Personnel Office (ការិយាល័យរដ្ឋបាល-បុគ្គលិក) OR if authorized by ADMIN.
    """
    if not user or not user.is_authenticated:
        return False

    # 1. System Admin / Superuser / Staff / Admin username / Admin role
    if user.is_superuser or user.is_staff or (user.username or '').upper() in ['ADMIN', 'ADMINISTRATOR', 'ROOT']:
        return True

    if not profile:
        profile = getattr(user, 'profile', None)
    if not profile:
        return False

    if getattr(profile, 'is_admin', False) or getattr(profile, 'role', '') == 'ADMIN':
        return True

    # 2. Leadership (ប្រធានមន្ទីរ, អនុប្រធានមន្ទីរ)
    if getattr(profile, 'is_leadership', False) or getattr(profile, 'role', '') in ['LEADERSHIP', 'DIRECTOR', 'DEPUTY_DIRECTOR']:
        return True

    # 3. Administration-Personnel Office (ការិយាល័យរដ្ឋបាល-បុគ្គលិក / កិច្ចការទូទៅ)
    dept = profile.department
    if dept:
        name_kh = (dept.name_kh or '').strip().lower()
        code = (dept.code or '').strip().upper()
        # Strictly exclude Cantons (ខណ្ឌរដ្ឋបាលព្រៃឈើ, ខណ្ឌរដ្ឋបាលជលផល)
        if not (code.startswith('CANTON') or name_kh.startswith('ខណ្ឌ') or 'ខណ្ឌ' in name_kh):
            if code in ['ADMIN', 'ADMIN_PERS', 'ADMIN_PERSONNEL', 'ADMIN_DEPT', 'LEAD', 'GEN_AFFAIRS', 'GENERAL_AFFAIRS']:
                return True
            if ('រដ្ឋបាល' in name_kh and 'បុគ្គលិក' in name_kh) or ('កិច្ចការទូទៅ' in name_kh):
                return True
            if name_kh in ['ការិយាល័យរដ្ឋបាល បុគ្គលិក', 'ការិយាល័យរដ្ឋបាល-បុគ្គលិក', 'ការិយាល័យរដ្ឋបាល', 'ការិយាល័យកិច្ចការរដ្ឋបាលទូទៅ']:
                return True

    # 4. Explicit authorization granted by ADMIN to this profile
    if getattr(profile, 'can_export_contract_excel', False):
        return True

    # Specialized offices are strictly forbidden by default
    return False


@login_required
def contract_officer_list(request):
    """
    List of Contract Civil Servants / Contract Staff with year filtering, search, counters and permissions.
    Strict Department Isolation: Specialized offices can ONLY see their own department's contract staff.
    """
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    current_year = max(2026, timezone.now().year)

    search_q = request.GET.get('q', '').strip()
    year_filter = request.GET.get('year', '').strip()
    dept_filter = request.GET.get('department', '').strip()
    gender_filter = request.GET.get('gender', '').strip()
    status_filter = request.GET.get('contract_status', '').strip()
    id_type_filter = request.GET.get('id_type', '').strip()

    base_queryset = ContractOfficer.objects.select_related('department', 'created_by', 'last_renewed_by').all()

    # Strict Department-Level Access Control
    if not is_admin_or_lead:
        if dept:
            base_queryset = base_queryset.filter(department=dept)
        else:
            base_queryset = base_queryset.none()
    elif dept_filter:
        base_queryset = base_queryset.filter(department_id=dept_filter)

    # Dynamic available years strictly starting from 2026 onwards
    db_years = [y for y in base_queryset.values_list('contract_year', flat=True).distinct() if y and y >= 2026]
    year_set = set(db_years)
    year_set.update([2026, 2027, 2028, 2029, 2030])
    available_years = sorted(list(year_set))

    queryset = base_queryset

    if year_filter:
        try:
            year_int = int(to_arabic_digits(year_filter))
            queryset = queryset.filter(contract_year=year_int)
        except Exception:
            pass

    if search_q:
        q_arabic = to_arabic_digits(search_q)
        queryset = queryset.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_name__icontains=search_q) |
            Q(id_number__icontains=search_q) |
            Q(id_number__icontains=q_arabic) |
            Q(phone__icontains=search_q) |
            Q(phone__icontains=q_arabic) |
            Q(email__icontains=search_q) |
            Q(position_title__icontains=search_q) |
            Q(working_unit__icontains=search_q) |
            Q(skill_specialization__icontains=search_q) |
            Q(contract_number__icontains=search_q)
        )

    if gender_filter:
        queryset = queryset.filter(gender=gender_filter)

    if status_filter:
        queryset = queryset.filter(contract_status=status_filter)

    if id_type_filter:
        queryset = queryset.filter(id_type=id_type_filter)

    total_officers = queryset.count()
    male_count = queryset.filter(gender='MALE').count()
    female_count = queryset.filter(gender='FEMALE').count()
    active_count = queryset.filter(contract_status='ACTIVE').count()
    renewed_count = queryset.filter(contract_status='RENEWED').count()
    expired_count = queryset.filter(contract_status='EXPIRED').count()
    current_year_count = base_queryset.filter(contract_year=current_year).count()

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh') if is_admin_or_lead else (Department.objects.filter(id=dept.id) if dept else [])

    paginator = Paginator(queryset.order_by('-contract_year', '-id'), 25)
    page_number = request.GET.get('page')
    officers = paginator.get_page(page_number)

    context = {
        'officers': officers,
        'total_officers': total_officers,
        'male_count': male_count,
        'female_count': female_count,
        'active_count': active_count,
        'renewed_count': renewed_count,
        'expired_count': expired_count,
        'current_year_count': current_year_count,
        'current_year': current_year,
        'next_year': current_year + 1,
        'available_years': available_years,
        'year_filter': year_filter,
        'departments': departments,
        'search_q': search_q,
        'dept_filter': dept_filter,
        'gender_filter': gender_filter,
        'status_filter': status_filter,
        'id_type_filter': id_type_filter,
        'dept': dept,
        'is_admin_or_lead': is_admin_or_lead,
        'is_system_admin': is_system_admin,
        'can_export_contract_excel': can_export_contract_officers_to_excel(request.user, profile),
    }
    return render(request, 'dms/contract_officer_list.html', context)


@login_required
def contract_officer_create(request):
    """
    Create a new Contract Civil Servant Profile matching the standardized Khmer Curriculum Vitae format.
    """
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)
    current_year = max(2026, timezone.now().year)

    if request.method == 'POST':
        try:
            # 1. ព័ត៌មានផ្ទាល់ខ្លួន
            khmer_last_name = request.POST.get('khmer_last_name', '').strip()
            khmer_first_name = request.POST.get('khmer_first_name', '').strip()
            latin_name = request.POST.get('latin_name', '').strip()
            gender = request.POST.get('gender', 'MALE')
            nationality = request.POST.get('nationality', 'ខ្មែរ').strip() or 'ខ្មែរ'
            ethnicity = request.POST.get('ethnicity', 'ខ្មែរ').strip() or 'ខ្មែរ'
            dob = request.POST.get('dob', '').strip()

            # ទីកន្លែងកំណើត
            pob_province_code, pob_district_code, pob_commune_code, pob_village_code, \
            pob_province, pob_district, pob_commune, pob_village = _resolve_geo_address(
                request.POST.get('pob_province_code'), request.POST.get('pob_district_code'),
                request.POST.get('pob_commune_code'), request.POST.get('pob_village_code'),
                request.POST.get('pob_province'), request.POST.get('pob_district'),
                request.POST.get('pob_commune'), request.POST.get('pob_village')
            )
            place_of_birth = request.POST.get('place_of_birth', '').strip()

            # កម្រិតវប្បធម៌ & បណ្តុះបណ្តាល
            general_education = request.POST.get('general_education', '').strip()
            training_level = request.POST.get('training_level', '').strip()
            skill_specialization = request.POST.get('skill_specialization', '').strip()

            # អត្តសញ្ញាណប័ណ្ណ / លិខិតឆ្លងដែន
            id_type = request.POST.get('id_type', 'NATIONAL_ID')
            id_number = request.POST.get('id_number', '').strip()
            if id_number:
                id_number = to_arabic_digits(id_number)

            # អង្គភាព / ការិយាល័យ
            department_obj = None
            if is_admin_or_lead:
                department_id = request.POST.get('department')
                if department_id:
                    department_obj = Department.objects.filter(id=department_id).first()
            else:
                department_obj = dept

            working_unit = request.POST.get('working_unit', '').strip()

            # 2. ព័ត៌មានទំនាក់ទំនង
            current_province_code, current_district_code, current_commune_code, current_village_code, \
            current_province, current_district, current_commune, current_village = _resolve_geo_address(
                request.POST.get('current_province_code'), request.POST.get('current_district_code'),
                request.POST.get('current_commune_code'), request.POST.get('current_village_code'),
                request.POST.get('current_province'), request.POST.get('current_district'),
                request.POST.get('current_commune'), request.POST.get('current_village')
            )
            current_house_no = request.POST.get('current_house_no', '').strip()
            current_street = request.POST.get('current_street', '').strip()
            current_address = request.POST.get('current_address', '').strip()

            phone = request.POST.get('phone', '').strip()
            if phone:
                phone = to_arabic_digits(phone)
            email = request.POST.get('email', '').strip()

            # 3. ព័ត៌មានកិច្ចសន្យា
            contract_year_str = request.POST.get('contract_year', str(current_year)).strip()
            try:
                contract_year = int(to_arabic_digits(contract_year_str))
                if contract_year < 2026:
                    contract_year = 2026
            except Exception:
                contract_year = current_year

            position_title = request.POST.get('position_title', 'មន្ត្រីជាប់កិច្ចសន្យា').strip() or 'មន្ត្រីជាប់កិច្ចសន្យា'
            contract_number = request.POST.get('contract_number', '').strip()
            contract_start_date_str = request.POST.get('contract_start_date', '').strip()
            contract_end_date_str = request.POST.get('contract_end_date', '').strip()
            salary_str = request.POST.get('salary', '').strip()
            contract_status = request.POST.get('contract_status', 'ACTIVE')
            remarks = request.POST.get('remarks', '').strip()

            contract_start_date = None
            if contract_start_date_str:
                try:
                    contract_start_date = datetime.strptime(contract_start_date_str, '%Y-%m-%d').date()
                except Exception:
                    pass
            elif contract_year:
                contract_start_date = date(contract_year, 1, 1)

            contract_end_date = None
            if contract_end_date_str:
                try:
                    contract_end_date = datetime.strptime(contract_end_date_str, '%Y-%m-%d').date()
                except Exception:
                    pass
            elif contract_year:
                contract_end_date = date(contract_year, 12, 31)

            salary = None
            if salary_str:
                try:
                    salary = float(salary_str.replace(',', ''))
                except Exception:
                    pass

            if not khmer_last_name or not khmer_first_name:
                messages.error(request, 'សូមបញ្ចូលគោត្តនាម និងនាមខ្លួនជាភាសាខ្មែរ!')
                raise ValueError("Missing name")

            contract_officer = ContractOfficer(
                department=department_obj,
                created_by=request.user,
                contract_year=contract_year,
                original_join_year=contract_year,
                contract_count_years=1,
                khmer_last_name=khmer_last_name,
                khmer_first_name=khmer_first_name,
                latin_name=latin_name,
                gender=gender,
                nationality=nationality,
                ethnicity=ethnicity,
                dob=dob,
                pob_province_code=pob_province_code,
                pob_district_code=pob_district_code,
                pob_commune_code=pob_commune_code,
                pob_village_code=pob_village_code,
                pob_province=pob_province,
                pob_district=pob_district,
                pob_commune=pob_commune,
                pob_village=pob_village,
                place_of_birth=place_of_birth,
                general_education=general_education,
                training_level=training_level,
                skill_specialization=skill_specialization,
                id_type=id_type,
                id_number=id_number,
                working_unit=working_unit,
                current_province_code=current_province_code,
                current_district_code=current_district_code,
                current_commune_code=current_commune_code,
                current_village_code=current_village_code,
                current_province=current_province,
                current_district=current_district,
                current_commune=current_commune,
                current_village=current_village,
                current_house_no=current_house_no,
                current_street=current_street,
                current_address=current_address,
                phone=phone,
                email=email,
                position_title=position_title,
                contract_number=contract_number,
                contract_start_date=contract_start_date,
                contract_end_date=contract_end_date,
                salary=salary,
                contract_status=contract_status,
                remarks=remarks,
            )

            if 'photo' in request.FILES:
                contract_officer.photo = request.FILES['photo']
            if 'contract_file' in request.FILES:
                contract_officer.contract_file = request.FILES['contract_file']

            contract_officer.save()

            messages.success(request, f'បានបញ្ចូលជីវប្រវត្តិមន្ត្រីជាប់កិច្ចសន្យា «{contract_officer.full_name_kh}» (ឆ្នាំ {contract_officer.contract_year}) ដោយជោគជ័យ!')
            
            save_action = request.POST.get('save_action', 'save_and_close')
            if save_action == 'save_and_continue':
                return redirect('contract_officer_create')
            return redirect('contract_officer_detail', pk=contract_officer.pk)

        except Exception as e:
            if not any(m.message for m in messages.get_messages(request)):
                messages.error(request, f'មានបញ្ហាក្នុងការរក្សាទុក៖ {str(e)}')

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh') if is_admin_or_lead else (Department.objects.filter(id=dept.id) if dept else [])
    provinces = CambodiaProvince.objects.all().order_by('code')
    year_options = [2026, 2027, 2028, 2029, 2030]

    context = {
        'is_edit': False,
        'current_year': current_year,
        'year_options': year_options,
        'departments': departments,
        'provinces': provinces,
        'dept': dept,
        'is_admin_or_lead': is_admin_or_lead,
        'is_system_admin': is_system_admin,
    }
    return render(request, 'dms/contract_officer_form.html', context)


@login_required
def contract_officer_edit(request, pk):
    """
    Edit an existing Contract Civil Servant Profile.
    """
    contract_officer = get_object_or_404(ContractOfficer, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)
    current_year = max(2026, timezone.now().year)

    # Permission check: specialized offices can only edit their own department's staff
    if not is_admin_or_lead:
        if not dept or contract_officer.department != dept:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិកែប្រែព័ត៌មានមន្ត្រីជាប់កិច្ចសន្យារបស់ការិយាល័យផ្សេងឡើយ!')
            return redirect('contract_officer_list')

    if request.method == 'POST':
        try:
            old_dept = contract_officer.department
            old_unit = contract_officer.working_unit
            old_pos = contract_officer.position_title

            khmer_last_name = request.POST.get('khmer_last_name', '').strip()
            khmer_first_name = request.POST.get('khmer_first_name', '').strip()
            latin_name = request.POST.get('latin_name', '').strip()
            gender = request.POST.get('gender', 'MALE')
            nationality = request.POST.get('nationality', 'ខ្មែរ').strip() or 'ខ្មែរ'
            ethnicity = request.POST.get('ethnicity', 'ខ្មែរ').strip() or 'ខ្មែរ'
            dob = request.POST.get('dob', '').strip()

            # ទីកន្លែងកំណើត
            pob_province_code, pob_district_code, pob_commune_code, pob_village_code, \
            pob_province, pob_district, pob_commune, pob_village = _resolve_geo_address(
                request.POST.get('pob_province_code'), request.POST.get('pob_district_code'),
                request.POST.get('pob_commune_code'), request.POST.get('pob_village_code'),
                request.POST.get('pob_province'), request.POST.get('pob_district'),
                request.POST.get('pob_commune'), request.POST.get('pob_village')
            )
            place_of_birth = request.POST.get('place_of_birth', '').strip()

            general_education = request.POST.get('general_education', '').strip()
            training_level = request.POST.get('training_level', '').strip()
            skill_specialization = request.POST.get('skill_specialization', '').strip()

            id_type = request.POST.get('id_type', 'NATIONAL_ID')
            id_number = request.POST.get('id_number', '').strip()
            if id_number:
                id_number = to_arabic_digits(id_number)

            if is_admin_or_lead:
                department_id = request.POST.get('department')
                if department_id:
                    contract_officer.department = Department.objects.filter(id=department_id).first()
                else:
                    contract_officer.department = None
            else:
                contract_officer.department = dept

            working_unit = request.POST.get('working_unit', '').strip()

            # ព័ត៌មានទំនាក់ទំនង
            current_province_code, current_district_code, current_commune_code, current_village_code, \
            current_province, current_district, current_commune, current_village = _resolve_geo_address(
                request.POST.get('current_province_code'), request.POST.get('current_district_code'),
                request.POST.get('current_commune_code'), request.POST.get('current_village_code'),
                request.POST.get('current_province'), request.POST.get('current_district'),
                request.POST.get('current_commune'), request.POST.get('current_village')
            )
            current_house_no = request.POST.get('current_house_no', '').strip()
            current_street = request.POST.get('current_street', '').strip()
            current_address = request.POST.get('current_address', '').strip()

            phone = request.POST.get('phone', '').strip()
            if phone:
                phone = to_arabic_digits(phone)
            email = request.POST.get('email', '').strip()

            contract_year_str = request.POST.get('contract_year', str(contract_officer.contract_year)).strip()
            try:
                contract_officer.contract_year = int(to_arabic_digits(contract_year_str))
            except Exception:
                pass

            position_title = request.POST.get('position_title', 'មន្ត្រីជាប់កិច្ចសន្យា').strip() or 'មន្ត្រីជាប់កិច្ចសន្យា'
            contract_number = request.POST.get('contract_number', '').strip()
            contract_start_date_str = request.POST.get('contract_start_date', '').strip()
            contract_end_date_str = request.POST.get('contract_end_date', '').strip()
            salary_str = request.POST.get('salary', '').strip()
            contract_status = request.POST.get('contract_status', 'ACTIVE')
            remarks = request.POST.get('remarks', '').strip()

            if contract_start_date_str:
                try:
                    contract_officer.contract_start_date = datetime.strptime(contract_start_date_str, '%Y-%m-%d').date()
                except Exception:
                    pass
            else:
                contract_officer.contract_start_date = None

            if contract_end_date_str:
                try:
                    contract_officer.contract_end_date = datetime.strptime(contract_end_date_str, '%Y-%m-%d').date()
                except Exception:
                    pass
            else:
                contract_officer.contract_end_date = None

            if salary_str:
                try:
                    contract_officer.salary = float(salary_str.replace(',', ''))
                except Exception:
                    pass
            else:
                contract_officer.salary = None

            if not khmer_last_name or not khmer_first_name:
                messages.error(request, 'សូមបញ្ចូលគោត្តនាម និងនាមខ្លួនជាភាសាខ្មែរ!')
                raise ValueError("Missing name")

            contract_officer.khmer_last_name = khmer_last_name
            contract_officer.khmer_first_name = khmer_first_name
            contract_officer.latin_name = latin_name
            contract_officer.gender = gender
            contract_officer.nationality = nationality
            contract_officer.ethnicity = ethnicity
            contract_officer.dob = dob
            contract_officer.pob_province_code = pob_province_code
            contract_officer.pob_district_code = pob_district_code
            contract_officer.pob_commune_code = pob_commune_code
            contract_officer.pob_village_code = pob_village_code
            contract_officer.pob_province = pob_province
            contract_officer.pob_district = pob_district
            contract_officer.pob_commune = pob_commune
            contract_officer.pob_village = pob_village
            contract_officer.place_of_birth = place_of_birth
            contract_officer.general_education = general_education
            contract_officer.training_level = training_level
            contract_officer.skill_specialization = skill_specialization
            contract_officer.id_type = id_type
            contract_officer.id_number = id_number
            contract_officer.working_unit = working_unit
            contract_officer.current_province_code = current_province_code
            contract_officer.current_district_code = current_district_code
            contract_officer.current_commune_code = current_commune_code
            contract_officer.current_village_code = current_village_code
            contract_officer.current_province = current_province
            contract_officer.current_district = current_district
            contract_officer.current_commune = current_commune
            contract_officer.current_village = current_village
            contract_officer.current_house_no = current_house_no
            contract_officer.current_street = current_street
            contract_officer.current_address = current_address
            contract_officer.phone = phone
            contract_officer.email = email
            contract_officer.position_title = position_title
            contract_officer.contract_number = contract_number
            contract_officer.contract_status = contract_status
            contract_officer.remarks = remarks

            if 'photo' in request.FILES:
                contract_officer.photo = request.FILES['photo']
            if 'contract_file' in request.FILES:
                contract_officer.contract_file = request.FILES['contract_file']

            contract_officer.save()

            # Automatic transfer history logging if department changed
            if old_dept != contract_officer.department and contract_officer.department is not None:
                ContractOfficerTransferHistory.objects.create(
                    contract_officer=contract_officer,
                    from_department=old_dept,
                    to_department=contract_officer.department,
                    from_working_unit=old_unit or (old_dept.name_kh if old_dept else ''),
                    to_working_unit=working_unit or (contract_officer.department.name_kh if contract_officer.department else ''),
                    from_position_title=old_pos,
                    to_position_title=position_title,
                    transfer_date=timezone.now().date(),
                    remarks='បានផ្លាស់ប្តូរការិយាល័យតាមរយៈការកែប្រែជីវប្រវត្តិ (Profile Edit)',
                    transferred_by=request.user,
                )

            messages.success(request, f'បានកែប្រែជីវប្រវត្តិមន្ត្រីជាប់កិច្ចសន្យា «{contract_officer.full_name_kh}» (ឆ្នាំ {contract_officer.contract_year}) ដោយជោគជ័យ!')
            return redirect('contract_officer_detail', pk=contract_officer.pk)

        except Exception as e:
            if not any(m.message for m in messages.get_messages(request)):
                messages.error(request, f'មានបញ្ហាក្នុងការកែប្រែ៖ {str(e)}')

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh') if is_admin_or_lead else (Department.objects.filter(id=dept.id) if dept else [])
    provinces = CambodiaProvince.objects.all().order_by('code')
    year_options = sorted(list(set([2026, 2027, 2028, 2029, 2030, contract_officer.contract_year])))

    context = {
        'is_edit': True,
        'contract_officer': contract_officer,
        'current_year': current_year,
        'year_options': year_options,
        'departments': departments,
        'provinces': provinces,
        'dept': dept,
        'is_admin_or_lead': is_admin_or_lead,
        'is_system_admin': is_system_admin,
    }
    return render(request, 'dms/contract_officer_form.html', context)


@login_required
def contract_officer_renew(request, pk):
    """
    1-Click / Modal Renewal for an existing Contract Civil Servant to a new contract year.
    Strict Department Isolation: Specialized offices can only renew their own staff, Admin can renew any.
    """
    contract_officer = get_object_or_404(ContractOfficer, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    if not is_admin_or_lead:
        if not dept or contract_officer.department != dept:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិបន្តកិច្ចសន្យាសម្រាប់មន្ត្រីនៃការិយាល័យផ្សេងឡើយ!')
            return redirect('contract_officer_list')

    if request.method == 'POST':
        try:
            from_year = contract_officer.contract_year
            to_year_str = request.POST.get('to_year', str(from_year + 1)).strip()
            to_year = int(to_arabic_digits(to_year_str))

            contract_number = request.POST.get('contract_number', '').strip()
            start_date_str = request.POST.get('contract_start_date', '').strip()
            end_date_str = request.POST.get('contract_end_date', '').strip()
            salary_str = request.POST.get('salary', '').strip()
            position_title = request.POST.get('position_title', contract_officer.position_title).strip() or contract_officer.position_title
            working_unit = request.POST.get('working_unit', contract_officer.working_unit).strip() or contract_officer.working_unit
            remarks = request.POST.get('remarks', '').strip()

            parsed_start_date = None
            if start_date_str:
                try:
                    parsed_start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                except Exception:
                    pass
            if not parsed_start_date:
                parsed_start_date = date(to_year, 1, 1)

            parsed_end_date = None
            if end_date_str:
                try:
                    parsed_end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                except Exception:
                    pass
            if not parsed_end_date:
                parsed_end_date = date(to_year, 12, 31)

            parsed_salary = contract_officer.salary
            if salary_str:
                try:
                    parsed_salary = float(salary_str.replace(',', ''))
                except Exception:
                    pass

            renewal_doc = request.FILES.get('renewal_document')

            # Create renewal history record
            ContractOfficerRenewalHistory.objects.create(
                contract_officer=contract_officer,
                from_year=from_year,
                to_year=to_year,
                contract_number=contract_number or contract_officer.contract_number,
                contract_start_date=parsed_start_date,
                contract_end_date=parsed_end_date,
                salary=parsed_salary,
                position_title=position_title,
                working_unit=working_unit,
                renewal_document=renewal_doc,
                remarks=remarks,
                approved_by=request.user
            )

            # Update officer record
            if not contract_officer.original_join_year:
                contract_officer.original_join_year = from_year
            contract_officer.contract_year = to_year
            contract_officer.contract_count_years = (contract_officer.contract_count_years or 1) + 1
            contract_officer.contract_status = 'RENEWED'
            if contract_number:
                contract_officer.contract_number = contract_number
            contract_officer.contract_start_date = parsed_start_date
            contract_officer.contract_end_date = parsed_end_date
            contract_officer.salary = parsed_salary
            contract_officer.position_title = position_title
            contract_officer.working_unit = working_unit
            if renewal_doc:
                contract_officer.contract_file = renewal_doc
            contract_officer.last_renewed_at = timezone.now()
            contract_officer.last_renewed_by = request.user
            contract_officer.save()

            messages.success(request, f'🎉 បានអនុម័តបន្តកិច្ចសន្យាសម្រាប់ «{contract_officer.full_name_kh}» ទៅកាន់ឆ្នាំ {to_year} ដោយជោគជ័យ!')

        except Exception as e:
            messages.error(request, f'មានបញ្ហាក្នុងការបន្តកិច្ចសន្យា៖ {str(e)}')

    return redirect('contract_officer_detail', pk=contract_officer.pk)


@login_required
def contract_officer_batch_renew(request):
    """
    Batch / Multi-select Renewal for multiple Contract Staff at once (Admin / Leadership).
    """
    if request.method != 'POST':
        return redirect('contract_officer_list')

    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    selected_ids = request.POST.getlist('selected_officers')
    if not selected_ids:
        messages.warning(request, 'សូមជ្រើសរើសមន្ត្រីដែលត្រូវបន្តកិច្ចសន្យាជាមុនសិន!')
        return redirect('contract_officer_list')

    current_year = timezone.now().year
    to_year_str = request.POST.get('batch_to_year', str(current_year + 1)).strip()
    try:
        to_year = int(to_arabic_digits(to_year_str))
    except Exception:
        to_year = current_year + 1

    batch_remarks = request.POST.get('batch_remarks', '').strip()
    batch_contract_prefix = request.POST.get('batch_contract_prefix', '').strip()

    queryset = ContractOfficer.objects.filter(id__in=selected_ids)
    if not is_admin_or_lead and not is_system_admin:
        if dept:
            queryset = queryset.filter(department=dept)
        else:
            queryset = queryset.none()

    renewed_count = 0
    for officer in queryset:
        from_year = officer.contract_year
        if from_year == to_year:
            continue # already renewed to this year

        start_date = date(to_year, 1, 1)
        end_date = date(to_year, 12, 31)
        contract_num = f"{batch_contract_prefix}/{to_year}" if batch_contract_prefix else officer.contract_number

        ContractOfficerRenewalHistory.objects.create(
            contract_officer=officer,
            from_year=from_year,
            to_year=to_year,
            contract_number=contract_num,
            contract_start_date=start_date,
            contract_end_date=end_date,
            salary=officer.salary,
            position_title=officer.position_title,
            working_unit=officer.working_unit,
            remarks=batch_remarks or f"បន្តកិច្ចសន្យាជាក្រុមទៅឆ្នាំ {to_year}",
            approved_by=request.user
        )

        if not officer.original_join_year:
            officer.original_join_year = from_year
        officer.contract_year = to_year
        officer.contract_count_years = (officer.contract_count_years or 1) + 1
        officer.contract_status = 'RENEWED'
        if batch_contract_prefix:
            officer.contract_number = contract_num
        officer.contract_start_date = start_date
        officer.contract_end_date = end_date
        officer.last_renewed_at = timezone.now()
        officer.last_renewed_by = request.user
        officer.save()
        renewed_count += 1

    messages.success(request, f'🎉 បានអនុម័តបន្តកិច្ចសន្យាទៅឆ្នាំ {to_year} សម្រាប់មន្ត្រីចំនួន {renewed_count} នាក់ដោយជោគជ័យ!')
    return redirect('contract_officer_list')


@login_required
def contract_officer_detail(request, pk):
    """
    Detailed profile view of a Contract Civil Servant including renewal history, transfer history, attachments and print action.
    Strict Department Isolation: Specialized offices can ONLY view their own department's contract staff.
    """
    contract_officer = get_object_or_404(ContractOfficer, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    if not is_admin_or_lead:
        if not dept or contract_officer.department != dept:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិមើលព័ត៌មានមន្ត្រីជាប់កិច្ចសន្យារបស់ការិយាល័យផ្សេងឡើយ!')
            return redirect('contract_officer_list')

    attachments = contract_officer.attachments.all().order_by('-created_at')
    renewals = contract_officer.renewals.select_related('approved_by').all().order_by('-to_year', '-approved_at')
    transfers = contract_officer.transfers.select_related('from_department', 'to_department', 'transferred_by').all().order_by('-transfer_date', '-created_at')
    all_departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')
    can_edit = is_admin_or_lead or (dept and contract_officer.department == dept)
    current_year = timezone.now().year
    next_year = contract_officer.contract_year + 1

    context = {
        'contract_officer': contract_officer,
        'attachments': attachments,
        'renewals': renewals,
        'transfers': transfers,
        'all_departments': all_departments,
        'can_edit': can_edit,
        'current_year': current_year,
        'next_year': next_year,
        'is_admin_or_lead': is_admin_or_lead,
        'is_system_admin': is_system_admin,
    }
    return render(request, 'dms/contract_officer_detail.html', context)


@login_required
def contract_officer_transfer(request, pk):
    """
    Direct / Modal Department/Canton Transfer for Contract Officials (Admin / Leadership).
    Records a full audit and history entry of the transfer.
    """
    contract_officer = get_object_or_404(ContractOfficer, pk=pk)
    profile = getattr(request.user, 'profile', None)
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access or request.user.is_superuser
    if not is_admin_or_lead:
        messages.error(request, "⚠️ មានតែថ្នាក់ដឹកនាំ និងរដ្ឋបាល-បុគ្គលិក (ADMIN) ប៉ុណ្ណោះដែលអាចផ្លាស់ប្តូរការិយាល័យ/ខណ្ឌបាន!")
        return redirect('contract_officer_detail', pk=pk)

    if request.method == 'POST':
        target_dept_id = request.POST.get('target_department')
        transfer_date_str = request.POST.get('transfer_date', '').strip()
        ref_letter = request.POST.get('reference_letter_number', '').strip()
        new_position = request.POST.get('position_title', '').strip()
        new_working_unit = request.POST.get('working_unit', '').strip()
        remarks = request.POST.get('remarks', '').strip()
        transfer_doc = request.FILES.get('transfer_document')

        target_dept = Department.objects.filter(id=target_dept_id, is_active=True).first() if target_dept_id else None
        if not target_dept:
            messages.error(request, "⚠️ សូមជ្រើសរើសការិយាល័យ/ខណ្ឌថ្មី!")
            return redirect('contract_officer_detail', pk=pk)

        old_dept = contract_officer.department
        old_unit = contract_officer.working_unit
        old_pos = contract_officer.position_title

        transfer_date = None
        if transfer_date_str:
            try:
                transfer_date = datetime.strptime(transfer_date_str, '%Y-%m-%d').date()
            except Exception:
                transfer_date = timezone.now().date()
        else:
            transfer_date = timezone.now().date()

        # Update contract officer
        contract_officer.department = target_dept
        if new_position:
            contract_officer.position_title = new_position
        if new_working_unit:
            contract_officer.working_unit = new_working_unit
        else:
            contract_officer.working_unit = target_dept.name_kh
        contract_officer.save()

        # Create history record
        ContractOfficerTransferHistory.objects.create(
            contract_officer=contract_officer,
            from_department=old_dept,
            to_department=target_dept,
            from_working_unit=old_unit or (old_dept.name_kh if old_dept else ''),
            to_working_unit=contract_officer.working_unit,
            from_position_title=old_pos,
            to_position_title=contract_officer.position_title,
            transfer_date=transfer_date,
            reference_letter_number=ref_letter,
            transfer_document=transfer_doc,
            remarks=remarks,
            transferred_by=request.user,
        )

        messages.success(request, f"✅ បានផ្លាស់ប្តូរការិយាល័យសម្រាប់មន្ត្រីកិច្ចសន្យា «{contract_officer.full_name_kh}» ទៅកាន់ «{target_dept.name_kh}» ដោយជោគជ័យ!")

    return redirect('contract_officer_detail', pk=pk)


@login_required
def contract_officer_transfer_delete(request, pk, transfer_id):
    """
    Delete a contract officer department transfer history record (Superadmin or Admin only).
    """
    contract_officer = get_object_or_404(ContractOfficer, pk=pk)
    profile = getattr(request.user, 'profile', None)
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access or request.user.is_superuser
    if not is_admin_or_lead:
        messages.error(request, "⚠️ មានតែ ADMIN ប៉ុណ្ណោះដែលអាចលុបកំណត់ត្រាប្រវត្តិផ្ទេរការិយាល័យបាន!")
        return redirect('contract_officer_detail', pk=pk)

    if request.method == 'POST':
        transfer_record = get_object_or_404(ContractOfficerTransferHistory, pk=transfer_id, contract_officer=contract_officer)
        from_name = transfer_record.from_department.name_kh if transfer_record.from_department else (transfer_record.from_working_unit or "គ្មាន")
        to_name = transfer_record.to_department.name_kh if transfer_record.to_department else (transfer_record.to_working_unit or "គ្មាន")
        transfer_record.delete()
        messages.success(request, f"🗑️ បានលុបកំណត់ត្រាប្រវត្តិផ្លាស់ប្តូរការិយាល័យ «{from_name} ➔ {to_name}» រួចរាល់។")

    return redirect('contract_officer_detail', pk=pk)


@login_required
def contract_officer_print(request, pk):
    """
    Standard Official Khmer Curriculum Vitae (ជីវប្រវត្តិរូបសង្ខេប) Print Layout for Contract Officials.
    Strict Department Isolation: Specialized offices can ONLY print their own department's contract staff.
    """
    contract_officer = get_object_or_404(ContractOfficer, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)

    if not has_global_access:
        if not dept or contract_officer.department != dept:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិមើល ឬបោះពុម្ពព័ត៌មានមន្ត្រីជាប់កិច្ចសន្យារបស់ការិយាល័យផ្សេងឡើយ!')
            return redirect('contract_officer_list')

    return render(request, 'dms/contract_officer_print.html', {'officer': contract_officer})


@login_required
def contract_officer_delete(request, pk):
    """
    Delete a Contract Civil Servant.
    Strictly restricted: ONLY ADMIN is allowed to delete contract officers freely.
    Non-admin users are strictly forbidden from deleting contract officers.
    """
    contract_officer = get_object_or_404(ContractOfficer, pk=pk)
    profile = getattr(request.user, 'profile', None)
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access or request.user.is_superuser or request.user.is_staff or check_is_system_admin(request.user, request)

    if not is_admin_or_lead:
        messages.error(request, '⚠️ មានតែ ADMIN ប៉ុណ្ណោះដែលមានសិទ្ធិលុបមន្ត្រីជាប់កិច្ចសន្យាចេញពីប្រព័ន្ធ! ក្រៅពី ADMIN គឺមិនអនុញ្ញាតឡើយ។')
        return redirect('contract_officer_list')

    if request.method == 'POST':
        name = contract_officer.full_name_kh
        year = contract_officer.contract_year
        contract_officer.delete()
        messages.success(request, f'🗑️ បានលុបមន្ត្រីជាប់កិច្ចសន្យា «{name}» (ឆ្នាំ {year}) ចេញពីប្រព័ន្ធដោយជោគជ័យ!')

    return redirect('contract_officer_list')


@login_required
def contract_officer_bulk_delete(request):
    """
    Bulk delete selected Contract Civil Servants.
    Strictly restricted: ONLY ADMIN is allowed to delete contract officers.
    Non-admin users are strictly forbidden.
    """
    if request.method != 'POST':
        return redirect('contract_officer_list')

    profile = getattr(request.user, 'profile', None)
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access or request.user.is_superuser or request.user.is_staff or check_is_system_admin(request.user, request)

    if not is_admin_or_lead:
        messages.error(request, '⚠️ មានតែ ADMIN ប៉ុណ្ណោះដែលមានសិទ្ធិលុបមន្ត្រីជាប់កិច្ចសន្យាជាក្រុមចេញពីប្រព័ន្ធ! ក្រៅពី ADMIN គឺមិនអនុញ្ញាតឡើយ។')
        return redirect('contract_officer_list')

    selected_ids = request.POST.getlist('selected_officers')
    if not selected_ids:
        messages.warning(request, 'សូមជ្រើសរើសមន្ត្រីដែលត្រូវលុបជាមុនសិន!')
        return redirect('contract_officer_list')

    queryset = ContractOfficer.objects.filter(id__in=selected_ids)
    count = queryset.count()
    queryset.delete()
    messages.success(request, f'🗑️ បានលុបមន្ត្រីជាប់កិច្ចសន្យាចំនួន {count} នាក់ចេញពីប្រព័ន្ធដោយជោគជ័យ!')
    return redirect('contract_officer_list')


def _format_d1_dob(val):
    if not val:
        return '-'
    std_dmy = _format_khmer_date_standard(val)
    if '-' in std_dmy and len(std_dmy.split('-')) == 3:
        parts = std_dmy.split('-')
        dotted = f"{parts[0]}.{parts[1]}.{parts[2]}"
        return _to_khmer_digits(dotted)
    return _to_khmer_digits(str(val))


def _format_d1_phone(val):
    if not val:
        return '-'
    val_clean = str(val).strip()
    # Keep strictly on single line - if multiple numbers, take the first/primary one
    if ',' in val_clean:
        val_clean = val_clean.split(',')[0].strip()
    elif '/' in val_clean:
        val_clean = val_clean.split('/')[0].strip()
    return _to_khmer_digits(val_clean)


def _format_d1_education(val):
    """
    Cleans education string so it strictly displays on a single line.
    """
    if not val:
        return 'គ្មាន'
    clean = ' '.join(str(val).split()).strip()
    return clean or 'គ្មាន'


def _format_contract_officer_office(officer):
    """
    Returns the specific Office / Canton where the contract officer serves (ខណ្ឌ/ការិយាល័យបំពេញការងារ),
    strictly excluding the overarching organization/department name (មន្ទីរកសិកម្ម...).
    """
    # 1. If assigned to a specific department (which represents an Office or Canton in DCM):
    dept = getattr(officer, 'department', None)
    if dept and dept.name_kh:
        dname = dept.name_kh.strip()
        if 'មន្ទីរកសិកម្ម' not in dname:
            return dname

    # 2. Check working_unit field
    unit = (getattr(officer, 'working_unit', '') or '').strip()
    if unit:
        if 'មន្ទីរកសិកម្ម' in unit:
            # Check if an office name is inside it, e.g. "មន្ទីរកសិកម្ម... ការិយាល័យ..."
            for kw in ['ការិយាល័យ', 'ខណ្ឌ']:
                if kw in unit:
                    parts = unit.split(kw)
                    return (kw + ' ' + parts[-1]).strip(' -:/')
            if dept and dept.name_kh and 'មន្ទីរកសិកម្ម' not in dept.name_kh:
                return dept.name_kh.strip()
            return "ការិយាល័យរដ្ឋបាល-បុគ្គលិក"
        return unit

def _is_specialized_department(dept):
    """
    Checks if a department is a specialized office / canton (ការិយាល័យជំនាញផ្សេងៗ / ខណ្ឌ),
    as opposed to Provincial Leadership or Administration/Personnel.
    """
    if not dept:
        return False
    code = (getattr(dept, 'code', '') or '').upper()
    name = getattr(dept, 'name_kh', '') or ''
    admin_codes = ['LEAD', 'ADMIN_PERS', 'ADMIN_DEPT', 'ADMIN_PERSONNEL', 'GENERAL_AFFAIRS', 'ADMIN', 'ADMIN_NEW']
    if code in admin_codes:
        return False
    if 'ថ្នាក់ដឹកនាំ' in name:
        return False
    return True


def _get_department_head_title(dept):
    """
    Returns appropriate title for head of office / canton (ប្រធានការិយាល័យ... ឬ នាយខណ្ឌ...).
    """
    if not dept:
        return "ប្រធានការិយាល័យ"
    dname = dept.name_kh.strip()
    code = (dept.code or '').upper()
    if 'ខណ្ឌ' in dname or code.startswith('CANTON'):
        if dname.startswith('ខណ្ឌ'):
            return 'នាយ' + dname
        return 'នាយខណ្ឌ ' + dname
    if dname.startswith('ការិយាល័យ'):
        return 'ប្រធាន' + dname
    return 'ប្រធានការិយាល័យ ' + dname


def _get_specialized_office_leadership_info(dept, override_type=None):
    """
    Returns leadership configuration for a specialized office/canton:
    - 'head': Has Head of Office/Canton -> Center: 'ប្រធាន...' or 'នាយ...', Right: 'អ្នកធ្វើតារាង'
    - 'deputy': No Head, but has Deputy in charge -> Center: 'អនុប្រធានទទួលបន្ទុករួមការិយាល័យ' (or 'អនុប្រធានទទួលបន្ទុករួមខណ្ឌ'), Right: 'អ្នកធ្វើតារាង'
    - 'solo': Solo Officer (មន្ត្រីទោល) -> No Center signature, Right: 'មន្ត្រីទទួលបន្ទុកដឹកនាំរួម'
    """
    if not dept:
        return {
            'type': 'head',
            'center_title': 'ប្រធានការិយាល័យ',
            'right_title': 'អ្នកធ្វើតារាង',
            'has_center': True
        }

    is_canton = 'ខណ្ឌ' in (dept.name_kh or '') or (dept.code or '').upper().startswith('CANTON')

    if override_type in ['head', 'deputy', 'solo']:
        lead_type = override_type
    else:
        # Auto-detect from active staff
        cs_list = list(CivilServantProfile.objects.filter(department=dept))
        up_list = list(UserProfile.objects.filter(department=dept))

        has_head = False
        has_deputy = False

        for cs in cs_list:
            pos = (cs.current_position_title or '').strip()
            if ('ប្រធាន' in pos and 'អនុ' not in pos) or ('នាយខណ្ឌ' in pos and 'រង' not in pos):
                has_head = True
            elif 'អនុប្រធាន' in pos or 'នាយរង' in pos or 'រង' in pos:
                has_deputy = True

        if not has_head and not has_deputy:
            for up in up_list:
                pos = (up.position_title or '').strip()
                if ('ប្រធាន' in pos and 'អនុ' not in pos) or ('នាយខណ្ឌ' in pos and 'រង' not in pos):
                    has_head = True
                elif 'អនុប្រធាន' in pos or 'នាយរង' in pos or 'រង' in pos:
                    has_deputy = True

        if has_head:
            lead_type = 'head'
        elif has_deputy:
            lead_type = 'deputy'
        else:
            lead_type = 'solo'

    if lead_type == 'head':
        center_title = _get_department_head_title(dept)
        right_title = 'អ្នកធ្វើតារាង'
        has_center = True
    elif lead_type == 'deputy':
        center_title = 'អនុប្រធានទទួលបន្ទុករួមខណ្ឌ' if is_canton else 'អនុប្រធានទទួលបន្ទុករួមការិយាល័យ'
        right_title = 'អ្នកធ្វើតារាង'
        has_center = True
    else:  # solo
        center_title = ''
        right_title = 'មន្ត្រីទទួលបន្ទុកដឹកនាំរួម'
        has_center = False

    return {
        'type': lead_type,
        'center_title': center_title,
        'right_title': right_title,
        'has_center': has_center,
    }


def _build_d1_contract_workbook(officers_list, year=None, department=None, lead_type=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "06"

    # Page Setup: A4 Landscape Print Format strictly matching Excel/D-1.xlsx
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    # Column widths strictly matching D-1.xlsx
    col_widths = {
        'A': 5.14,
        'B': 15.71,
        'C': 25.14,
        'D': 6.14,
        'E': 16.29,
        'F': 28.43,
        'G': 15.86,
        'H': 12.43,
        'I': 18.43,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row heights
    ws.row_dimensions[1].height = 18.0
    ws.row_dimensions[2].height = 18.0
    ws.row_dimensions[3].height = 18.0
    ws.row_dimensions[4].height = 23.25
    ws.row_dimensions[5].height = 23.25
    ws.row_dimensions[6].height = 6.0
    ws.row_dimensions[7].height = 71.1
    ws.row_dimensions[8].height = 43.5

    # Fonts (Strictly bold=False as mandated)
    font_org = Font(name='Khmer OS Muol Light', size=10, color='1E3A8A', bold=False)
    font_org_sub = Font(name='Khmer OS Muol Light', size=9.5, color='1E3A8A', bold=False)
    font_kingdom = Font(name='Khmer OS Muol Light', size=10.5, color='1E3A8A', bold=False)
    font_stars = Font(name='Khmer OS Battambang', size=12, color='1E3A8A', bold=False)
    font_title = Font(name='Khmer OS Muol Light', size=11, color='1E3A8A', bold=False)
    font_header = Font(name='Khmer OS Battambang', size=11, bold=False)
    font_data = Font(name='Khmer OS Battambang', size=11, bold=False)
    font_latin = Font(name='Cambria', size=11, bold=False)
    font_footer_bold = Font(name='Khmer OS Battambang', size=10.5, bold=False)
    font_footer_normal = Font(name='Khmer OS Battambang', size=10, bold=False)
    font_footer_muol = Font(name='Khmer OS Muol Light', size=10.5, bold=False)

    # Thin Border
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_data_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    align_data_left = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # Logo Image on Left Header (precisely centered above Organization A to C)
    _add_excel_centered_logo(ws, 'A', 'C', start_row_idx=0, row_offset_px=2, logo_size_px=60)

    is_specialized = _is_specialized_department(department)
    lead_info = _get_specialized_office_leadership_info(department, override_type=lead_type) if is_specialized else None

    # Left Header: Organization under Logo (centered across A to C)
    ws.merge_cells('A4:C4')
    ws.merge_cells('A5:C5')
    if is_specialized and department:
        # Specialized Office: Line 1 = Provincial Department, Line 2 = Office/Canton
        ws['A4'] = "មន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទខេត្តប៉ៃលិន"
        ws['A4'].font = font_org
        ws['A4'].alignment = align_center

        ws['A5'] = department.name_kh
        ws['A5'].font = font_org_sub
        ws['A5'].alignment = align_center
    else:
        # Full Department / Admin: Line 1 = Ministry, Line 2 = Provincial Department
        ws['A4'] = "ក្រសួងកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទ"
        ws['A4'].font = font_org
        ws['A4'].alignment = align_center

        ws['A5'] = "មន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទខេត្តប៉ៃលិន"
        ws['A5'].font = font_org_sub
        ws['A5'].alignment = align_center

    # Right Header: Kingdom & Motto
    ws.merge_cells('G2:I2')
    ws['G2'] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws['G2'].font = font_kingdom
    ws['G2'].alignment = align_center

    ws.merge_cells('G3:I3')
    ws['G3'] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws['G3'].font = font_kingdom
    ws['G3'].alignment = align_center

    # Royal Divider Flourish under ជាតិ សាសនា ព្រះមហាក្សត្រ
    ws.merge_cells('G4:I4')
    ws['G4'] = ""
    _add_excel_khmer_divider(ws, col_idx=7, col_offset_px=23, row_idx=3, row_offset_px=5, width_px=65, height_px=18)

    # Row 7: Title block
    now = datetime.now()
    month_kh = KHMER_MONTHS_NAMES[now.month] if 1 <= now.month <= 12 else str(now.month)
    year_val = year if year else now.year
    year_kh = _to_khmer_digits(str(year_val))

    ws.merge_cells('A7:I7')
    if is_specialized and department:
        owner_name = f"របស់{department.name_kh}"
    else:
        owner_name = "របស់មន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទខេត្តប៉ៃលិន"

    title_text = (
        "បញ្ជីបច្ចុប្បន្នភាពមន្ត្រីជាប់កិច្ចសន្យា\n"
        f"{owner_name}\n"
        f"ប្រចាំខែ{month_kh} ឆ្នាំ{year_kh}"
    )
    ws['A7'] = title_text
    ws['A7'].font = font_title
    ws['A7'].alignment = align_center

    # Row 8: Headers (9 columns)
    headers = [
        ('A', 'ល.រ'),
        ('B', 'គោត្តនាម នាម'),
        ('C', 'អក្សរឡាតាំង'),
        ('D', 'ភេទ'),
        ('E', 'ថ្ងៃខែឆ្នាំកំណើត'),
        ('F', 'ខណ្ឌ/ការិ.បំពេញការងារ'),
        ('G', 'តួនាទី'),
        ('H', 'កម្រិតវប្បធម៌'),
        ('I', 'លេខទូរស័ព្ទ'),
    ]
    for col_letter, h_text in headers:
        cell = ws[f'{col_letter}8']
        cell.value = h_text
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    # Row 9+: Data Rows (Uniform height 25.0, single-line cell content)
    current_row = 9
    for idx, o in enumerate(officers_list, start=1):
        ws.row_dimensions[current_row].height = 25.0

        g_kh = 'ប្រុស' if o.gender == 'MALE' or str(o.gender).upper() in ['M', 'ប្រុស', 'ប'] else 'ស្រី'
        dob_display = _format_d1_dob(o.dob)
        office_display = _format_contract_officer_office(o)
        edu_display = _format_d1_education(o.general_education or o.training_level)
        phone_display = _format_d1_phone(o.phone)

        cells_data = [
            ('A', idx, font_data, align_data_center),
            ('B', o.full_name_kh, font_data, align_data_left),
            ('C', o.full_name_latin, font_latin, align_data_left),
            ('D', g_kh, font_data, align_data_center),
            ('E', dob_display, font_data, align_data_center),
            ('F', office_display, font_data, align_data_center),
            ('G', o.position_title or 'មន្ត្រីជាប់កិច្ចសន្យា', font_data, align_data_center),
            ('H', edu_display, font_data, align_data_center),
            ('I', phone_display, font_data, align_data_center),
        ]
        for col_letter, val, f_style, a_style in cells_data:
            c = ws[f'{col_letter}{current_row}']
            c.value = val
            c.font = f_style
            c.alignment = a_style
            c.border = thin_border

        current_row += 1

    # Bottom Summary and Hierarchical Signatures (strictly matching D-1 format)
    is_specialized = _is_specialized_department(department)
    head_title = _get_department_head_title(department) if is_specialized else ""
    last_officer = officers_list[-1] if officers_list else None
    total_count = len(officers_list)
    female_count = sum(1 for o in officers_list if o.gender == 'FEMALE' or str(o.gender).upper() in ['F', 'FEMALE', 'ស្រី', 'ស'])
    male_count = total_count - female_count
    total_kh = _to_khmer_digits(str(total_count))
    female_kh = _to_khmer_digits(str(female_count))
    male_kh = _to_khmer_digits(str(male_count))
    lunar_info = _get_khmer_lunar_year_info(year_val) or "ឆ្នាំមមី អដ្ឋស័ក ព.ស.២៥៧០"

    # Blank spacing row right after data rows (tightened)
    ws.row_dimensions[current_row].height = 6.0

    # Row 1: Closing Line (Col A:E) and Top-Right Lunar Date (Col G:I)
    current_row += 1
    ws.row_dimensions[current_row].height = 16.0
    if last_officer:
        ws[f'A{current_row}'] = f"- បិទបញ្ចប់ត្រឹមលេខរៀងទី{total_kh} ឈ្មោះ {last_officer.full_name_kh}"
    else:
        ws[f'A{current_row}'] = "- មិនមានទិន្នន័យមន្ត្រី"
    ws[f'A{current_row}'].font = Font(name='Khmer OS Battambang', size=10, bold=False)

    ws.merge_cells(f'G{current_row}:I{current_row}')
    ws[f'G{current_row}'] = f"ថ្ងៃ.....................ខែ...............{lunar_info}"
    ws[f'G{current_row}'].font = Font(name='Khmer OS Battambang', size=9.5, bold=False)
    ws[f'G{current_row}'].alignment = align_center

    # Row 2: Total count & Female count (Col A:E) and Top-Right Solar Date (Col G:I)
    current_row += 1
    ws.row_dimensions[current_row].height = 16.0
    ws[f'A{current_row}'] = f"សរុបរួម៖ .....{total_kh}..... នាក់ , ស្រី .....{female_kh}..... នាក់"
    ws[f'A{current_row}'].font = Font(name='Khmer OS Battambang', size=10, bold=False)

    ws.merge_cells(f'G{current_row}:I{current_row}')
    ws[f'G{current_row}'] = f"ប៉ៃលិនថ្ងៃទី................ខែ...............ឆ្នាំ{year_kh}"
    ws[f'G{current_row}'].font = Font(name='Khmer OS Battambang', size=9.5, bold=False)
    ws[f'G{current_row}'].alignment = align_center

    # Row 3: Male count (Col A:E) and Right Signer Title (Col G:I)
    current_row += 1
    ws.row_dimensions[current_row].height = 18.0
    ws[f'A{current_row}'] = f"( ប្រុស .....{male_kh}..... នាក់ )"
    ws[f'A{current_row}'].font = Font(name='Khmer OS Battambang', size=10, bold=False)

    ws.merge_cells(f'G{current_row}:I{current_row}')
    if is_specialized and lead_info:
        ws[f'G{current_row}'] = lead_info['right_title']
    else:
        ws[f'G{current_row}'] = "អ្នកធ្វើតារាង" if is_specialized else "មន្ត្រីទទួលបន្ទុក"
    ws[f'G{current_row}'].font = Font(name='Khmer OS Muol Light', size=10.5)
    ws[f'G{current_row}'].alignment = align_center

    # Small spacing row before bottom signatures (tightened)
    current_row += 1
    ws.row_dimensions[current_row].height = 6.0

    # Row 4: Signatures Section (Directly consecutive, zero gap)
    if is_specialized:
        if lead_info and lead_info['has_center']:
            # Specialized Office with Head or Deputy: In Center D:F
            current_row += 1
            ws.row_dimensions[current_row].height = 18.0
            ws.merge_cells(f'D{current_row}:F{current_row}')
            ws[f'D{current_row}'] = "បានឃើញ និងពិនិត្យត្រឹមត្រូវ"
            ws[f'D{current_row}'].font = Font(name='Khmer OS Muol Light', size=10.5, bold=False)
            ws[f'D{current_row}'].alignment = align_center

            # Title of Office Head / Deputy Head (dead center of worksheet in D:F, immediately next row!)
            current_row += 1
            ws.row_dimensions[current_row].height = 18.0
            ws.merge_cells(f'D{current_row}:F{current_row}')
            ws[f'D{current_row}'] = lead_info['center_title']
            ws[f'D{current_row}'].font = Font(name='Khmer OS Muol Light', size=10.0, bold=False)
            ws[f'D{current_row}'].alignment = align_center
        else:
            # Solo Officer: No center signature
            pass

    else:
        # Full 3-Level Hierarchy: Left (A:C), Center (D:F - dead center of worksheet)
        # Line 1: Action Headers
        current_row += 1
        ws.row_dimensions[current_row].height = 18.0
        # Left Header
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = "បានឃើញ និងឯកភាព"
        ws[f'A{current_row}'].font = Font(name='Khmer OS Muol Light', size=10.5, bold=False)
        ws[f'A{current_row}'].alignment = align_center
        # Center Header
        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = "បានឃើញ និងពិនិត្យត្រឹមត្រូវ"
        ws[f'D{current_row}'].font = Font(name='Khmer OS Muol Light', size=10.5, bold=False)
        ws[f'D{current_row}'].alignment = align_center

        # Line 2: Lunar Date on Left (A:C) & Head of Office Title on Center (D:F)
        current_row += 1
        ws.row_dimensions[current_row].height = 14.0
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = f"ថ្ងៃ.....................ខែ...............{lunar_info}"
        ws[f'A{current_row}'].font = Font(name='Khmer OS Battambang', size=9.0, bold=False)
        ws[f'A{current_row}'].alignment = align_center

        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = "ប្រធានការិយាល័យរដ្ឋបាល-បុគ្គលិក"
        ws[f'D{current_row}'].font = Font(name='Khmer OS Muol Light', size=10.0, bold=False)
        ws[f'D{current_row}'].alignment = align_center

        # Line 3: Solar Date on Left (A:C)
        current_row += 1
        ws.row_dimensions[current_row].height = 14.0
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = f"ប៉ៃលិនថ្ងៃទី................ខែ...............ឆ្នាំ{year_kh}"
        ws[f'A{current_row}'].font = Font(name='Khmer OS Battambang', size=9.0, bold=False)
        ws[f'A{current_row}'].alignment = align_center

        # Line 4: Provincial Department Director Title directly on Left (A:C) - NO EMPTY ROW!
        current_row += 1
        ws.row_dimensions[current_row].height = 18.0
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = "ប្រធានមន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទខេត្តប៉ៃលិន"
        ws[f'A{current_row}'].font = Font(name='Khmer OS Muol Light', size=9.5, bold=False)
        ws[f'A{current_row}'].alignment = align_center

    return wb


@login_required
def contract_officer_export_excel(request):
    """
    Export Contract Civil Servants list to Excel strictly formatted as Form D-1 (Excel/D-1.xlsx).
    Strict Permission: Specialized offices are FORBIDDEN from exporting to Excel,
    EXCEPT for Administration-Personnel Office (ការិយាល័យរដ្ឋបាល-បុគ្គលិក) OR if authorized by ADMIN.
    """
    profile = getattr(request.user, 'profile', None)
    if not can_export_contract_officers_to_excel(request.user, profile):
        messages.error(
            request,
            "ការិយាល័យជំនាញមិនត្រូវបានអនុញ្ញាតឱ្យទាញចេញជា Excel ឡើយ! លើកលែងតែការិយាល័យរដ្ឋបាល-បុគ្គលិក ឬមានការអនុញ្ញាតដោយ ADMIN។"
        )
        return HttpResponseForbidden(
            "<div style='font-family: Khmer OS Battambang, sans-serif; text-align: center; margin-top: 80px;'>"
            "<h2 style='color: #dc2626;'>⛔ គ្មានសិទ្ធិទាញយកឯកសារជា Excel ឡើយ</h2>"
            "<p style='font-size: 16px; color: #475569;'>ការិយាល័យជំនាញមិនត្រូវបានអនុញ្ញាតឱ្យទាញចេញជា Excel ដាច់ខាត លើកលែងតែការិយាល័យរដ្ឋបាល-បុគ្គលិក ឬមានការអនុញ្ញាតដោយ ADMIN។</p>"
            "<a href='javascript:history.back()' style='display: inline-block; margin-top: 15px; padding: 8px 20px; background: #2563eb; color: #fff; text-decoration: none; border-radius: 20px;'>ត្រឡប់ក្រោយ</a>"
            "</div>"
        )

    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    queryset = ContractOfficer.objects.select_related('department').all()

    selected_dept = None
    if not is_admin_or_lead:
        if dept:
            queryset = queryset.filter(department=dept)
            selected_dept = dept
        else:
            queryset = queryset.none()
    else:
        dept_filter = request.GET.get('department', '').strip()
        if dept_filter:
            queryset = queryset.filter(department_id=dept_filter)
            selected_dept = Department.objects.filter(pk=dept_filter).first()

    year_filter = request.GET.get('year', '').strip()
    year_int = None
    if year_filter:
        try:
            year_int = int(to_arabic_digits(year_filter))
            queryset = queryset.filter(contract_year=year_int)
        except Exception:
            pass

    search_q = request.GET.get('q', '').strip()
    if search_q:
        q_arabic = to_arabic_digits(search_q)
        queryset = queryset.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_name__icontains=search_q) |
            Q(id_number__icontains=search_q) |
            Q(id_number__icontains=q_arabic) |
            Q(phone__icontains=search_q) |
            Q(position_title__icontains=search_q) |
            Q(working_unit__icontains=search_q)
        )

    gender_filter = request.GET.get('gender', '').strip()
    if gender_filter:
        queryset = queryset.filter(gender=gender_filter)

    status_filter = request.GET.get('contract_status', '').strip()
    if status_filter:
        queryset = queryset.filter(contract_status=status_filter)

    officers_list = list(queryset.order_by('-contract_year', 'khmer_last_name', 'khmer_first_name'))

    lead_type_param = request.GET.get('lead_type', '').strip()
    now = datetime.now()
    year_val = year_int if year_int else now.year
    wb = _build_d1_contract_workbook(officers_list, year=year_val, department=selected_dept, lead_type=lead_type_param)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    month_str = str(now.month).zfill(2)
    response['Content-Disposition'] = f'attachment; filename="contract_officers_D1_{year_val}_{month_str}.xlsx"'
    wb.save(response)
    return response


@login_required
def contract_officer_preview_pdf_d1(request):
    """
    មើលជា PDF / Print Preview សម្រាប់ទម្រង់ D-1 (បញ្ជីបច្ចុប្បន្នភាពមន្ត្រីជាប់កិច្ចសន្យា)
    """
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    queryset = ContractOfficer.objects.select_related('department').all()

    selected_dept = None
    if not is_admin_or_lead:
        if dept:
            queryset = queryset.filter(department=dept)
            selected_dept = dept
        else:
            queryset = queryset.none()
    else:
        dept_filter = request.GET.get('department', '').strip()
        if dept_filter:
            queryset = queryset.filter(department_id=dept_filter)
            selected_dept = Department.objects.filter(pk=dept_filter).first()

    year_filter = request.GET.get('year', '').strip()
    year_int = None
    if year_filter:
        try:
            year_int = int(to_arabic_digits(year_filter))
            queryset = queryset.filter(contract_year=year_int)
        except Exception:
            pass

    search_q = request.GET.get('q', '').strip()
    if search_q:
        q_arabic = to_arabic_digits(search_q)
        queryset = queryset.filter(
            Q(khmer_last_name__icontains=search_q) |
            Q(khmer_first_name__icontains=search_q) |
            Q(latin_name__icontains=search_q) |
            Q(id_number__icontains=search_q) |
            Q(id_number__icontains=q_arabic) |
            Q(phone__icontains=search_q) |
            Q(position_title__icontains=search_q) |
            Q(working_unit__icontains=search_q)
        )

    gender_filter = request.GET.get('gender', '').strip()
    if gender_filter:
        queryset = queryset.filter(gender=gender_filter)

    status_filter = request.GET.get('contract_status', '').strip()
    if status_filter:
        queryset = queryset.filter(contract_status=status_filter)

    officers_list = list(queryset.order_by('-contract_year', 'khmer_last_name', 'khmer_first_name'))

    table_items = []
    for idx, o in enumerate(officers_list, 1):
        g_kh = 'ប្រុស' if o.gender == 'MALE' or str(o.gender).upper() in ['M', 'ប្រុស', 'ប'] else 'ស្រី'
        table_items.append({
            'num': idx,
            'officer': o,
            'full_name_kh': o.full_name_kh,
            'full_name_latin': o.full_name_latin,
            'gender_kh': g_kh,
            'dob_display': _format_d1_dob(o.dob),
            'working_unit': _format_contract_officer_office(o),
            'position': o.position_title or 'មន្ត្រីជាប់កិច្ចសន្យា',
            'education': _format_d1_education(o.general_education or o.training_level),
            'phone': _format_d1_phone(o.phone),
        })

    pages = _paginate_preview_items(table_items, first_page_cap=14, mid_page_cap=18, last_page_cap=11, single_page_cap=9)

    now = datetime.now()
    month_kh = KHMER_MONTHS_NAMES[now.month] if 1 <= now.month <= 12 else str(now.month)
    year_val = year_int if year_int else now.year
    year_kh = _to_khmer_digits(str(year_val))
    lunar_year_text = _get_khmer_lunar_year_info(year_val)

    female_count = sum(1 for item in table_items if item['gender_kh'] == 'ស្រី')
    male_count = len(table_items) - female_count
    female_count_kh = _to_khmer_digits(str(female_count))
    male_count_kh = _to_khmer_digits(str(male_count))

    is_specialized = _is_specialized_department(selected_dept)
    lead_type_param = request.GET.get('lead_type', '').strip()
    lead_info = _get_specialized_office_leadership_info(selected_dept, override_type=lead_type_param) if is_specialized else None

    office_head_title = lead_info['center_title'] if lead_info else ""
    office_right_title = lead_info['right_title'] if lead_info else ("អ្នកធ្វើតារាង" if is_specialized else "មន្ត្រីទទួលបន្ទុក")
    has_center_signature = lead_info['has_center'] if lead_info else True

    lead_type_display = ""
    if lead_info:
        if lead_info['type'] == 'head':
            lead_type_display = "មានប្រធានការិយាល័យ"
        elif lead_info['type'] == 'deputy':
            lead_type_display = "អនុប្រធានទទួលបន្ទុករួម"
        else:
            lead_type_display = "មន្ត្រីទោល"

    get_params = request.GET.copy()
    excel_query_string = get_params.urlencode()

    get_params_no_lead = get_params.copy()
    if 'lead_type' in get_params_no_lead:
        del get_params_no_lead['lead_type']
    base_query_string = get_params_no_lead.urlencode()

    last_officer = officers_list[-1] if officers_list else None

    context = {
        'pages': pages,
        'total_count': len(table_items),
        'total_count_kh': _to_khmer_digits(str(len(table_items))),
        'female_count': female_count,
        'female_count_kh': female_count_kh,
        'male_count': male_count,
        'male_count_kh': male_count_kh,
        'last_officer': last_officer,
        'is_specialized': is_specialized,
        'office_head_title': office_head_title,
        'office_right_title': office_right_title,
        'has_center_signature': has_center_signature,
        'request_lead_type': lead_type_param,
        'lead_type_display': lead_type_display,
        'excel_query_string': excel_query_string,
        'base_query_string': base_query_string,
        'selected_department': selected_dept,
        'can_export_contract_excel': can_export_contract_officers_to_excel(request.user, profile),
        'lunar_year_text': lunar_year_text,
        'month_kh': month_kh,
        'year_kh': year_kh,
        'today': date.today(),
        'query_params': request.GET.urlencode(),
    }
    return render(request, 'dms/contract_officer_preview_d1_pdf.html', context)


@login_required
def contract_officer_attachment_add(request, pk):
    """
    Add an attachment to a Contract Civil Servant profile.
    Strict Department Isolation: Specialized offices can only add to their own department's staff.
    """
    contract_officer = get_object_or_404(ContractOfficer, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)

    if not has_global_access:
        if not dept or contract_officer.department != dept:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិបន្ថែមឯកសារសម្រាប់មន្ត្រីនៃការិយាល័យផ្សេងឡើយ!')
            return redirect('contract_officer_list')

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        file_obj = request.FILES.get('file')
        file_type = request.POST.get('file_type', '').strip()
        notes = request.POST.get('notes', '').strip()

        if not title or not file_obj:
            messages.error(request, 'សូមបញ្ចូលចំណងជើង និងជ្រើសរើសឯកសារភ្ជាប់!')
        else:
            attachment = ContractOfficerAttachment.objects.create(
                contract_officer=contract_officer,
                title=title,
                file=file_obj,
                file_type=file_type or 'ឯកសារទូទៅ',
                notes=notes,
                uploaded_by=request.user
            )
            messages.success(request, f'បានបញ្ចូលឯកសារ «{attachment.title}» ដោយជោគជ័យ!')

    return redirect('contract_officer_detail', pk=contract_officer.pk)


@login_required
def contract_officer_attachment_delete(request, pk, attachment_id):
    """
    Delete an attachment from a Contract Civil Servant profile.
    Strict Department Isolation.
    """
    contract_officer = get_object_or_404(ContractOfficer, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)

    if not has_global_access:
        if not dept or contract_officer.department != dept:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិលុបឯកសាររបស់មន្ត្រីនៃការិយាល័យផ្សេងឡើយ!')
            return redirect('contract_officer_list')

    attachment = get_object_or_404(ContractOfficerAttachment, pk=attachment_id, contract_officer=contract_officer)
    title = attachment.title
    attachment.delete()
    messages.success(request, f'បានលុបឯកសារ «{title}» រួចរាល់!')
    return redirect('contract_officer_detail', pk=contract_officer.pk)


@login_required
def contract_officer_attachment_view(request, pk, attachment_id):
    """
    Inline preview of contract officer attachment.
    Strict Department Isolation.
    """
    contract_officer = get_object_or_404(ContractOfficer, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)

    if not has_global_access:
        if not dept or contract_officer.department != dept:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិមើលឯកសាររបស់មន្ត្រីនៃការិយាល័យផ្សេងឡើយ!')
            return redirect('contract_officer_list')

    attachment = get_object_or_404(ContractOfficerAttachment, pk=attachment_id, contract_officer=contract_officer)
    file_path = attachment.file.path
    if not os.path.exists(file_path):
        messages.error(request, "រកមិនឃើញឯកសារក្នុង Server ទេ!")
        return redirect('contract_officer_detail', pk=contract_officer.pk)
    content_type, _ = mimetypes.guess_type(file_path)
    return FileResponse(open(file_path, 'rb'), content_type=content_type or 'application/octet-stream')


@login_required
def contract_officer_attachment_download(request, pk, attachment_id):
    """
    Download contract officer attachment.
    Strict Department Isolation.
    """
    contract_officer = get_object_or_404(ContractOfficer, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)

    if not has_global_access:
        if not dept or contract_officer.department != dept:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិទាញយកឯកសាររបស់មន្ត្រីនៃការិយាល័យផ្សេងឡើយ!')
            return redirect('contract_officer_list')

    attachment = get_object_or_404(ContractOfficerAttachment, pk=attachment_id, contract_officer=contract_officer)
    file_path = attachment.file.path
    if not os.path.exists(file_path):
        messages.error(request, "រកមិនឃើញឯកសារក្នុង Server ទេ!")
        return redirect('contract_officer_detail', pk=contract_officer.pk)
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=os.path.basename(file_path))


@login_required
def contract_officer_import_docx(request):
    """
    Imports Contract Civil Servant data from uploaded Word (.docx) file(s).
    Supports single file submission and AJAX batch multi-file upload with progress bar.
    """
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.POST.get('format') == 'json' or 'application/json' in request.headers.get('Accept', '')
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_system_admin = check_is_system_admin(request.user, request)

    if request.method == 'POST':
        docx_file = request.FILES.get('docx_file')
        if not docx_file:
            if is_ajax:
                return JsonResponse({'success': False, 'error': 'សូមជ្រើសរើសឯកសារ Word (.docx) ដើម្បីនាំចូល!'})
            messages.error(request, "⚠️ សូមជ្រើសរើសឯកសារ Word (.docx) ដើម្បីនាំចូល!")
            return redirect('contract_officer_list')

        if not docx_file.name.lower().endswith('.docx'):
            if is_ajax:
                return JsonResponse({'success': False, 'error': 'ប្រព័ន្ធគាំទ្រតែឯកសារប្រភេទ Word (.docx) ប៉ុណ្ណោះ!'})
            messages.error(request, "⚠️ ប្រព័ន្ធគាំទ្រតែឯកសារប្រភេទ Word (.docx) ប៉ុណ្ណោះ!")
            return redirect('contract_officer_list')

        try:
            from .docx_parser import parse_docx_contract_officer, compare_contract_officer_data
            data = parse_docx_contract_officer(docx_file, original_filename=docx_file.name)

            if not data.get('khmer_last_name') and not data.get('khmer_first_name'):
                if is_ajax:
                    return JsonResponse({'success': False, 'error': 'មិនអាចស្វែងរកឈ្មោះមន្ត្រីក្នុងឯកសារ Word នេះបានឡើយ។'})
                messages.error(request, "⚠️ មិនអាចស្វែងរកឈ្មោះមន្ត្រីក្នុងឯកសារ Word នេះបានឡើយ។ សូមពិនិត្យមើលទម្រង់ឯកសារម្តងទៀត។")
                return redirect('contract_officer_list')

            if not is_admin_or_lead:
                if not dept:
                    err_msg = "គណនីរបស់លោកអ្នកមិនទាន់ត្រូវបានកំណត់ការិយាល័យឡើយ! មិនអាចនាំចូលទិន្នន័យបានទេ។"
                    if is_ajax:
                        return JsonResponse({'success': False, 'error': err_msg})
                    messages.error(request, f"⚠️ {err_msg}")
                    return redirect('contract_officer_list')
                target_dept = dept
            else:
                target_dept_id = request.POST.get('department')
                target_dept = None
                if target_dept_id:
                    target_dept = Department.objects.filter(id=target_dept_id).first()
                else:
                    unit_txt = data.get('working_unit', '')
                    if unit_txt:
                        for d in Department.objects.filter(is_active=True):
                            if d.name_kh in unit_txt or (d.name_en and d.name_en in unit_txt):
                                target_dept = d
                                break

            # Find matching existing contract officer by ID Number or Name
            officer = None
            id_num = data.get('id_number', '').strip()
            if id_num:
                officer = ContractOfficer.objects.filter(id_number=id_num).first()

            if not officer:
                officer = ContractOfficer.objects.filter(
                    khmer_last_name=data['khmer_last_name'],
                    khmer_first_name=data['khmer_first_name']
                ).first()

            is_new = officer is None

            if not is_new and not has_global_access and officer.department and officer.department != dept:
                dept_name = officer.department.name_kh
                err_msg = f"មន្ត្រី «{officer.full_name_kh}» ស្ថិតនៅក្នុង «{dept_name}» រួចហើយ! លោកអ្នកគ្មានសិទ្ធិកែប្រែ ឬបន្ថែមទិន្នន័យមន្ត្រីក្រៅការិយាល័យឡើយ។"
                if is_ajax:
                    return JsonResponse({'success': False, 'error': err_msg})
                messages.error(request, f"⚠️ {err_msg}")
                return redirect('contract_officer_list')

            # Auto-resolve Geographic administrative IDs
            p_pcode, p_dcode, p_ccode, p_vcode, p_pname, p_dname, p_cname, p_vname = _resolve_geo_address(
                data.get('pob_province_code'), data.get('pob_district_code'), data.get('pob_commune_code'), data.get('pob_village_code'),
                data.get('pob_province'), data.get('pob_district'), data.get('pob_commune'), data.get('pob_village')
            )
            data['pob_province_code'] = p_pcode
            data['pob_district_code'] = p_dcode
            data['pob_commune_code'] = p_ccode
            data['pob_village_code'] = p_vcode
            data['pob_province'] = p_pname
            data['pob_district'] = p_dname
            data['pob_commune'] = p_cname
            data['pob_village'] = p_vname

            c_pcode, c_dcode, c_ccode, c_vcode, c_pname, c_dname, c_cname, c_vname = _resolve_geo_address(
                data.get('current_province_code'), data.get('current_district_code'), data.get('current_commune_code'), data.get('current_village_code'),
                data.get('current_province'), data.get('current_district'), data.get('current_commune'), data.get('current_village')
            )
            data['current_province_code'] = c_pcode
            data['current_district_code'] = c_dcode
            data['current_commune_code'] = c_ccode
            data['current_village_code'] = c_vcode
            data['current_province'] = c_pname
            data['current_district'] = c_dname
            data['current_commune'] = c_cname
            data['current_village'] = c_vname

            duplicate_mode = request.POST.get('duplicate_mode', 'review')
            confirm_update = request.POST.get('confirm_update') == 'true' or request.POST.get('confirm_update') == '1'

            if not is_new:
                diffs = compare_contract_officer_data(officer, data)
                if not diffs:
                    if is_ajax:
                        return JsonResponse({
                            'success': True,
                            'action': 'identical',
                            'officer_id': officer.id,
                            'officer_name': officer.full_name_kh,
                            'message': f"មន្ត្រី «{officer.full_name_kh}» មានទិន្នន័យដូចគ្នាបេះបិទរួចហើយ (មិនចាំបាច់កែប្រែ)។"
                        })
                    messages.info(request, f"ℹ️ មន្ត្រី «{officer.full_name_kh}» មានទិន្នន័យដូចគ្នារួចហើយ។")
                    return redirect('contract_officer_detail', pk=officer.pk)

                if duplicate_mode == 'skip' and not confirm_update:
                    if is_ajax:
                        return JsonResponse({
                            'success': True,
                            'action': 'skipped',
                            'officer_id': officer.id,
                            'officer_name': officer.full_name_kh,
                            'message': f"បានរំលងមន្ត្រី «{officer.full_name_kh}» ព្រោះមានក្នុងប្រព័ន្ធរួចហើយ។"
                        })
                    messages.info(request, f"ℹ️ បានរំលងមន្ត្រី «{officer.full_name_kh}»។")
                    return redirect('contract_officer_list')

                if duplicate_mode == 'review' and not confirm_update:
                    if is_ajax:
                        return JsonResponse({
                            'success': True,
                            'action': 'needs_review',
                            'officer_id': officer.id,
                            'officer_name': officer.full_name_kh,
                            'diffs': diffs,
                            'message': f"រកឃើញទិន្នន័យខុសគ្នាចំនួន {len(diffs)} ចំណុច សម្រាប់មន្ត្រី «{officer.full_name_kh}»។"
                        })

                # Update officer fields
                for field in [
                    'khmer_last_name', 'khmer_first_name', 'latin_name', 'gender',
                    'nationality', 'ethnicity', 'dob', 'pob_village', 'pob_commune',
                    'pob_district', 'pob_province', 'place_of_birth',
                    'pob_province_code', 'pob_district_code', 'pob_commune_code', 'pob_village_code',
                    'general_education', 'training_level', 'skill_specialization',
                    'id_type', 'id_number', 'working_unit',
                    'current_house_no', 'current_street', 'current_village', 'current_commune',
                    'current_district', 'current_province', 'current_address',
                    'current_province_code', 'current_district_code', 'current_commune_code', 'current_village_code',
                    'phone', 'email', 'position_title', 'contract_number'
                ]:
                    if data.get(field):
                        setattr(officer, field, data[field])

                if target_dept:
                    officer.department = target_dept

                # Save photo if extracted and officer has no photo or updating
                if data.get('photo_bytes'):
                    fn = data.get('photo_filename') or f"photo_{officer.id}.jpg"
                    officer.photo.save(fn, ContentFile(data['photo_bytes']), save=False)

                officer.save()

                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'action': 'updated',
                        'officer_id': officer.id,
                        'officer_name': officer.full_name_kh,
                        'message': f"បានធ្វើបច្ចុប្បន្នភាពទិន្នន័យមន្ត្រី «{officer.full_name_kh}» ដោយជោគជ័យ!"
                    })
                messages.success(request, f"✅ បានធ្វើបច្ចុប្បន្នភាពទិន្នន័យមន្ត្រី «{officer.full_name_kh}» ដោយជោគជ័យ!")
                return redirect('contract_officer_detail', pk=officer.pk)

            else:
                # Create New Contract Officer
                new_officer = ContractOfficer(
                    department=target_dept,
                    created_by=request.user,
                    khmer_last_name=data.get('khmer_last_name', ''),
                    khmer_first_name=data.get('khmer_first_name', ''),
                    latin_name=data.get('latin_name', ''),
                    gender=data.get('gender', 'MALE'),
                    nationality=data.get('nationality', 'ខ្មែរ') or 'ខ្មែរ',
                    ethnicity=data.get('ethnicity', 'ខ្មែរ') or 'ខ្មែរ',
                    dob=data.get('dob', ''),
                    pob_village=data.get('pob_village', ''),
                    pob_commune=data.get('pob_commune', ''),
                    pob_district=data.get('pob_district', ''),
                    pob_province=data.get('pob_province', ''),
                    place_of_birth=data.get('place_of_birth', ''),
                    pob_province_code=data.get('pob_province_code'),
                    pob_district_code=data.get('pob_district_code'),
                    pob_commune_code=data.get('pob_commune_code'),
                    pob_village_code=data.get('pob_village_code'),
                    general_education=data.get('general_education', ''),
                    training_level=data.get('training_level', ''),
                    skill_specialization=data.get('skill_specialization', ''),
                    id_type=data.get('id_type', 'NATIONAL_ID'),
                    id_number=data.get('id_number', ''),
                    working_unit=data.get('working_unit', ''),
                    current_house_no=data.get('current_house_no', ''),
                    current_street=data.get('current_street', ''),
                    current_village=data.get('current_village', ''),
                    current_commune=data.get('current_commune', ''),
                    current_district=data.get('current_district', ''),
                    current_province=data.get('current_province', ''),
                    current_address=data.get('current_address', ''),
                    current_province_code=data.get('current_province_code'),
                    current_district_code=data.get('current_district_code'),
                    current_commune_code=data.get('current_commune_code'),
                    current_village_code=data.get('current_village_code'),
                    phone=data.get('phone', ''),
                    email=data.get('email', ''),
                    position_title=data.get('position_title', 'មន្ត្រីជាប់កិច្ចសន្យា') or 'មន្ត្រីជាប់កិច្ចសន្យា',
                    contract_number=data.get('contract_number', ''),
                    contract_status=data.get('contract_status', 'ACTIVE'),
                )

                if data.get('photo_bytes'):
                    fn = data.get('photo_filename') or "photo.jpg"
                    new_officer.photo.save(fn, ContentFile(data['photo_bytes']), save=False)

                new_officer.save()

                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'action': 'created',
                        'officer_id': new_officer.id,
                        'officer_name': new_officer.full_name_kh,
                        'message': f"បានបញ្ចូលជីវប្រវត្តិមន្ត្រីជាប់កិច្ចសន្យាថ្មី «{new_officer.full_name_kh}» ដោយជោគជ័យ!"
                    })
                messages.success(request, f"✅ បានបញ្ចូលជីវប្រវត្តិមន្ត្រីជាប់កិច្ចសន្យាថ្មី «{new_officer.full_name_kh}» ដោយជោគជ័យ!")
                return redirect('contract_officer_detail', pk=new_officer.pk)

        except Exception as e:
            if is_ajax:
                return JsonResponse({'success': False, 'error': f"កំហុសក្នុងការ Parse ឯកសារ Word៖ {str(e)}"})
            messages.error(request, f"⚠️ កំហុសក្នុងការ Parse ឯកសារ Word៖ {str(e)}")
            return redirect('contract_officer_list')

    return redirect('contract_officer_list')


# ==============================================================================
# 🤖 SMART AI & PDF/IMAGE CONTRACT OFFICER SCAN & REVIEW VIEWS
# ==============================================================================

@login_required
def api_contract_officer_scan_file(request):
    """
    Scans and parses uploaded PDF, Image, or Word documents for Contract Officer CVs.
    Returns structured data, preview images, and extracted 4x6 portrait photo.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    uploaded_file = request.FILES.get('file') or request.FILES.get('docx_file') or request.FILES.get('pdf_file') or request.FILES.get('image_file')
    if not uploaded_file:
        return JsonResponse({'success': False, 'error': 'សូមជ្រើសរើសឯកសារ PDF, រូបភាព ឬ Word ដើម្បីស្កេន!'})

    custom_key = request.POST.get('gemini_api_key', '').strip() or None

    try:
        from .pdf_image_parser import parse_contract_officer_document, compare_contract_officer_data
        
        file_bytes = uploaded_file.read()
        filename = uploaded_file.name
        
        scan_result = parse_contract_officer_document(file_bytes, filename, custom_api_key=custom_key)
        
        if not scan_result.get('success'):
            return JsonResponse({'success': False, 'error': scan_result.get('error', 'មិនអាចអានឯកសារបានឡើយ')})

        data = scan_result.get('data', {})

        # Resolve Geographic Codes and Names
        p_pcode, p_dcode, p_ccode, p_vcode, p_pname, p_dname, p_cname, p_vname = _resolve_geo_address(
            data.get('pob_province_code'), data.get('pob_district_code'), data.get('pob_commune_code'), data.get('pob_village_code'),
            data.get('pob_province'), data.get('pob_district'), data.get('pob_commune'), data.get('pob_village')
        )
        data['pob_province_code'] = p_pcode or ''
        data['pob_district_code'] = p_dcode or ''
        data['pob_commune_code'] = p_ccode or ''
        data['pob_village_code'] = p_vcode or ''
        data['pob_province'] = p_pname or data.get('pob_province', '')
        data['pob_district'] = p_dname or data.get('pob_district', '')
        data['pob_commune'] = p_cname or data.get('pob_commune', '')
        data['pob_village'] = p_vname or data.get('pob_village', '')

        c_pcode, c_dcode, c_ccode, c_vcode, c_pname, c_dname, c_cname, c_vname = _resolve_geo_address(
            data.get('current_province_code'), data.get('current_district_code'), data.get('current_commune_code'), data.get('current_village_code'),
            data.get('current_province'), data.get('current_district'), data.get('current_commune'), data.get('current_village')
        )
        data['current_province_code'] = c_pcode or ''
        data['current_district_code'] = c_dcode or ''
        data['current_commune_code'] = c_ccode or ''
        data['current_village_code'] = c_vcode or ''
        data['current_province'] = c_pname or data.get('current_province', '')
        data['current_district'] = c_dname or data.get('current_district', '')
        data['current_commune'] = c_cname or data.get('current_commune', '')
        data['current_village'] = c_vname or data.get('current_village', '')

        # Auto-match Working Unit to Department
        unit_text = data.get('working_unit', '')
        detected_dept_id = None
        detected_dept_name = None
        if unit_text:
            for d in Department.objects.filter(is_active=True):
                if d.name_kh in unit_text or (d.name_en and d.name_en in unit_text):
                    detected_dept_id = d.id
                    detected_dept_name = d.name_kh
                    break

        # Check existing officer in database
        existing_officer_info = None
        officer = None
        id_num = data.get('id_number', '').strip()
        if id_num:
            officer = ContractOfficer.objects.filter(id_number=id_num).first()
        if not officer and data.get('khmer_last_name') and data.get('khmer_first_name'):
            officer = ContractOfficer.objects.filter(
                khmer_last_name=data['khmer_last_name'],
                khmer_first_name=data['khmer_first_name']
            ).first()

        if officer:
            diffs = compare_contract_officer_data(officer, data)
            existing_officer_info = {
                'id': officer.id,
                'name': officer.full_name_kh,
                'department_name': officer.department.name_kh if officer.department else (officer.working_unit or 'គ្មាន'),
                'current_year': officer.contract_year,
                'diffs': diffs,
                'is_identical': len(diffs) == 0,
            }

        return JsonResponse({
            'success': True,
            'data': data,
            'photo_base64': scan_result.get('photo_base64'),
            'preview_images': scan_result.get('preview_images', []),
            'source': scan_result.get('source'),
            'has_api_key': scan_result.get('has_api_key'),
            'filename': filename,
            'detected_dept_id': detected_dept_id,
            'detected_dept_name': detected_dept_name,
            'existing_officer': existing_officer_info,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f"កំហុសក្នុងការស្កេន៖ {str(e)}"})


@login_required
def contract_officer_confirm_scan_import(request):
    """
    Saves the user-reviewed and edited Contract Officer data into the system,
    with photo attachment and optional original document file.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('format') == 'json'

    try:
        officer_id = request.POST.get('officer_id')
        confirm_update = request.POST.get('confirm_update') in ['true', '1', True]
        
        khmer_last_name = request.POST.get('khmer_last_name', '').strip()
        khmer_first_name = request.POST.get('khmer_first_name', '').strip()
        if not khmer_last_name or not khmer_first_name:
            err = 'សូមបញ្ចូលគោត្តនាម និងនាមខ្លួនជាភាសាខ្មែរ!'
            if is_ajax: return JsonResponse({'success': False, 'error': err})
            messages.error(request, err)
            return redirect('contract_officer_list')

        latin_name = request.POST.get('latin_name', '').strip().upper()
        gender = request.POST.get('gender', 'MALE')
        nationality = request.POST.get('nationality', 'ខ្មែរ').strip() or 'ខ្មែរ'
        ethnicity = request.POST.get('ethnicity', 'ខ្មែរ').strip() or 'ខ្មែរ'
        dob = request.POST.get('dob', '').strip()

        # Geography POB
        pob_pcode, pob_dcode, pob_ccode, pob_vcode, pob_pname, pob_dname, pob_cname, pob_vname = _resolve_geo_address(
            request.POST.get('pob_province_code'), request.POST.get('pob_district_code'),
            request.POST.get('pob_commune_code'), request.POST.get('pob_village_code'),
            request.POST.get('pob_province'), request.POST.get('pob_district'),
            request.POST.get('pob_commune'), request.POST.get('pob_village')
        )
        place_of_birth = request.POST.get('place_of_birth', '').strip()

        # Geography Current
        cur_pcode, cur_dcode, cur_ccode, cur_vcode, cur_pname, cur_dname, cur_cname, cur_vname = _resolve_geo_address(
            request.POST.get('current_province_code'), request.POST.get('current_district_code'),
            request.POST.get('current_commune_code'), request.POST.get('current_village_code'),
            request.POST.get('current_province'), request.POST.get('current_district'),
            request.POST.get('current_commune'), request.POST.get('current_village')
        )
        current_house_no = request.POST.get('current_house_no', '').strip()
        current_street = request.POST.get('current_street', '').strip()
        current_address = request.POST.get('current_address', '').strip()

        general_education = request.POST.get('general_education', '').strip()
        training_level = request.POST.get('training_level', '').strip()
        skill_specialization = request.POST.get('skill_specialization', '').strip()

        id_type = request.POST.get('id_type', 'NATIONAL_ID')
        id_number = to_arabic_digits(request.POST.get('id_number', '').strip())

        working_unit = request.POST.get('working_unit', '').strip()
        phone = to_arabic_digits(request.POST.get('phone', '').strip())
        email = request.POST.get('email', '').strip()

        position_title = request.POST.get('position_title', 'មន្ត្រីជាប់កិច្ចសន្យា').strip() or 'មន្ត្រីជាប់កិច្ចសន្យា'
        contract_number = request.POST.get('contract_number', '').strip()

        contract_year_val = request.POST.get('contract_year', '2026').strip()
        try:
            contract_year = int(to_arabic_digits(contract_year_val))
            if contract_year < 2026: contract_year = 2026
        except Exception:
            contract_year = 2026

        salary_val = request.POST.get('salary', '').strip()
        salary = None
        if salary_val:
            try: salary = float(salary_val.replace(',', ''))
            except Exception: pass

        contract_status = request.POST.get('contract_status', 'ACTIVE')
        remarks = request.POST.get('remarks', '').strip()

        # Department assignment
        target_dept = None
        if is_admin_or_lead:
            dept_id = request.POST.get('department')
            if dept_id:
                target_dept = Department.objects.filter(id=dept_id).first()
        else:
            target_dept = dept

        # Check existing record
        officer = None
        if officer_id:
            officer = ContractOfficer.objects.filter(pk=officer_id).first()
        elif id_number:
            officer = ContractOfficer.objects.filter(id_number=id_number).first()
        if not officer:
            officer = ContractOfficer.objects.filter(khmer_last_name=khmer_last_name, khmer_first_name=khmer_first_name).first()

        # Permission check
        if officer and not has_global_access and officer.department and officer.department != dept:
            err = f"មន្ត្រី «{officer.full_name_kh}» ស្ថិតនៅក្នុង «{officer.department.name_kh}» រួចហើយ! លោកអ្នកគ្មានសិទ្ធិកែប្រែទិន្នន័យមន្ត្រីក្រៅការិយាល័យឡើយ។"
            if is_ajax: return JsonResponse({'success': False, 'error': err})
            messages.error(request, err)
            return redirect('contract_officer_list')

        is_new = officer is None

        if not is_new and confirm_update:
            # Update existing
            officer.khmer_last_name = khmer_last_name
            officer.khmer_first_name = khmer_first_name
            officer.latin_name = latin_name
            officer.gender = gender
            officer.nationality = nationality
            officer.ethnicity = ethnicity
            officer.dob = dob
            officer.pob_province_code = pob_pcode
            officer.pob_district_code = pob_dcode
            officer.pob_commune_code = pob_ccode
            officer.pob_village_code = pob_vcode
            officer.pob_province = pob_pname
            officer.pob_district = pob_dname
            officer.pob_commune = pob_cname
            officer.pob_village = pob_vname
            officer.place_of_birth = place_of_birth
            officer.current_province_code = cur_pcode
            officer.current_district_code = cur_dcode
            officer.current_commune_code = cur_ccode
            officer.current_village_code = cur_vcode
            officer.current_province = cur_pname
            officer.current_district = cur_dname
            officer.current_commune = cur_cname
            officer.current_village = cur_vname
            officer.current_house_no = current_house_no
            officer.current_street = current_street
            officer.current_address = current_address
            officer.general_education = general_education
            officer.training_level = training_level
            officer.skill_specialization = skill_specialization
            officer.id_type = id_type
            officer.id_number = id_number
            officer.working_unit = working_unit
            officer.phone = phone
            officer.email = email
            officer.position_title = position_title
            officer.contract_number = contract_number
            officer.contract_year = contract_year
            if salary is not None: officer.salary = salary
            officer.contract_status = contract_status
            if target_dept: officer.department = target_dept
            if remarks: officer.remarks = remarks
        else:
            # Create new
            officer = ContractOfficer(
                department=target_dept,
                created_by=request.user,
                khmer_last_name=khmer_last_name,
                khmer_first_name=khmer_first_name,
                latin_name=latin_name,
                gender=gender,
                nationality=nationality,
                ethnicity=ethnicity,
                dob=dob,
                pob_province_code=pob_pcode,
                pob_district_code=pob_dcode,
                pob_commune_code=pob_ccode,
                pob_village_code=pob_vcode,
                pob_province=pob_pname,
                pob_district=pob_dname,
                pob_commune=pob_cname,
                pob_village=pob_vname,
                place_of_birth=place_of_birth,
                current_province_code=cur_pcode,
                current_district_code=cur_dcode,
                current_commune_code=cur_ccode,
                current_village_code=cur_vcode,
                current_province=cur_pname,
                current_district=cur_dname,
                current_commune=cur_cname,
                current_village=cur_vname,
                current_house_no=current_house_no,
                current_street=current_street,
                current_address=current_address,
                general_education=general_education,
                training_level=training_level,
                skill_specialization=skill_specialization,
                id_type=id_type,
                id_number=id_number,
                working_unit=working_unit,
                phone=phone,
                email=email,
                position_title=position_title,
                contract_number=contract_number,
                contract_year=contract_year,
                original_join_year=contract_year,
                contract_count_years=1,
                salary=salary,
                contract_status=contract_status,
                remarks=remarks,
            )

        # Handle Photo base64 or photo file upload
        photo_b64 = request.POST.get('photo_base64', '').strip()
        if photo_b64 and ',' in photo_b64:
            try:
                import base64
                format_prefix, imgstr = photo_b64.split(';base64,')
                ext = format_prefix.split('/')[-1]
                if ext == 'jpeg': ext = 'jpg'
                photo_bytes = base64.b64decode(imgstr)
                officer.photo.save(f"photo_{officer.full_name_latin or 'contract'}_{contract_year}.{ext}", ContentFile(photo_bytes), save=False)
            except Exception as e:
                print(f"Error saving photo from base64: {e}")

        if 'photo' in request.FILES:
            officer.photo = request.FILES['photo']

        officer.save()

        # Handle Scanned Document Attachment if provided
        if 'scanned_file' in request.FILES:
            try:
                scanned_f = request.FILES['scanned_file']
                ContractOfficerAttachment.objects.create(
                    contract_officer=officer,
                    title=f"ឯកសារជីវប្រវត្តិស្កេន (ឆ្នាំ {contract_year}) - {scanned_f.name}",
                    file=scanned_f,
                    file_type=scanned_f.name.split('.')[-1].upper(),
                    notes="ឯកសារនាំចូលដោយស្វ័យប្រវត្តិតាមរយៈ Smart AI Scanner",
                    uploaded_by=request.user
                )
            except Exception as e:
                print(f"Error saving scanned attachment: {e}")

        action_word = "បានធ្វើបច្ចុប្បន្នភាព" if not is_new else "បានបញ្ចូលជីវប្រវត្តិថ្មី"
        msg = f"✅ {action_word}មន្ត្រីជាប់កិច្ចសន្យា «{officer.full_name_kh}» (ឆ្នាំ {officer.contract_year}) ដោយជោគជ័យ!"

        if is_ajax:
            return JsonResponse({
                'success': True,
                'officer_id': officer.id,
                'officer_name': officer.full_name_kh,
                'action': 'updated' if not is_new else 'created',
                'message': msg,
                'detail_url': f"/contract-officers/{officer.id}/"
            })

        messages.success(request, msg)
        return redirect('contract_officer_detail', pk=officer.pk)

    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'error': f"មានបញ្ហាក្នុងការរក្សាទុក៖ {str(e)}"})
        messages.error(request, f"មានបញ្ហាក្នុងការរក្សាទុក៖ {str(e)}")
        return redirect('contract_officer_list')


@login_required
def api_save_gemini_api_key(request):
    """
    Saves or updates the Gemini API Key from frontend settings modal.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    is_system_admin = check_is_system_admin(request.user, request)
    profile = getattr(request.user, 'profile', None)
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)

    if not (is_system_admin or has_global_access or request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'success': False, 'error': '⚠️ មានតែ ADMIN ប៉ុណ្ណោះដែលអាចកំណត់ API Key បាន!'}, status=403)

    api_key = request.POST.get('gemini_api_key', '').strip()
    if not api_key:
        return JsonResponse({'success': False, 'error': 'សូមបញ្ចូល Google Gemini API Key!'})

    from .pdf_image_parser import save_gemini_api_key
    success = save_gemini_api_key(api_key)

    if success:
        return JsonResponse({'success': True, 'message': '✅ បានរក្សាទុក Google Gemini API Key ដោយជោគជ័យ! ប្រព័ន្ធត្រៀមស្កេនអក្សរខ្មែរ និងអក្សរសរសេរដៃបាន ១០០%។'})
    else:
        return JsonResponse({'success': False, 'error': 'មិនអាចរក្សាទុក API Key ទៅកាន់ File .env បានឡើយ។'})



# ==============================================================================
# 🚗 VEHICLE & ASSET USAGE MANAGEMENT (រថយន្ត និង ម៉ូតូរបស់អង្គភាព)
# ==============================================================================

@login_required
def vehicle_list(request):
    """
    Inventory list of official vehicles and motorbikes.
    """
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    queryset = Vehicle.objects.all().select_related('department')

    type_filter = request.GET.get('type', '').strip()
    status_filter = request.GET.get('status', '').strip()
    dept_filter = request.GET.get('department', '').strip()
    search_q = request.GET.get('q', '').strip()

    if type_filter:
        queryset = queryset.filter(vehicle_type=type_filter)
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    if dept_filter:
        queryset = queryset.filter(department_id=dept_filter)

    if search_q:
        q_arabic = to_arabic_digits(search_q)
        queryset = queryset.filter(
            Q(name__icontains=search_q) |
            Q(brand__icontains=search_q) |
            Q(plate_number__icontains=search_q) |
            Q(plate_number__icontains=q_arabic) |
            Q(chassis_number__icontains=search_q) |
            Q(engine_number__icontains=search_q) |
            Q(current_user_name__icontains=search_q) |
            Q(notes__icontains=search_q)
        )

    # Stats
    total_vehicles = Vehicle.objects.count()
    available_count = Vehicle.objects.filter(status='AVAILABLE').count()
    in_use_count = Vehicle.objects.filter(status='IN_USE').count()
    maintenance_count = Vehicle.objects.filter(status='MAINTENANCE').count()
    motorcycle_count = Vehicle.objects.filter(vehicle_type='MOTORCYCLE').count()
    car_count = Vehicle.objects.exclude(vehicle_type='MOTORCYCLE').count()

    paginator = Paginator(queryset.order_by('-created_at'), 20)
    page_number = request.GET.get('page', 1)
    vehicles = paginator.get_page(page_number)

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')

    context = {
        'vehicles': vehicles,
        'departments': departments,
        'total_vehicles': total_vehicles,
        'available_count': available_count,
        'in_use_count': in_use_count,
        'maintenance_count': maintenance_count,
        'motorcycle_count': motorcycle_count,
        'car_count': car_count,
        'type_filter': type_filter,
        'status_filter': status_filter,
        'dept_filter': dept_filter,
        'search_q': search_q,
        'is_admin_or_lead': is_admin_or_lead,
    }
    return render(request, 'dms/vehicle_list.html', context)


@login_required
def vehicle_create(request):
    """
    Create a new vehicle/motorbike entry (Admin / General Affairs only).
    """
    profile = getattr(request.user, 'profile', None)
    if not check_has_global_hr_tracking_access(request.user, profile):
        messages.error(request, '⚠️ មានតែ ADMIN ឬថ្នាក់ដឹកនាំ និងកិច្ចការទូទៅប៉ុណ្ណោះដែលអាចចុះបញ្ជីយានយន្ត!')
        return redirect('vehicle_list')

    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            vehicle_type = request.POST.get('vehicle_type', 'MOTORCYCLE')
            brand = request.POST.get('brand', '').strip()
            model_year = request.POST.get('model_year', '').strip()
            color = request.POST.get('color', '').strip()
            plate_number = request.POST.get('plate_number', '').strip()
            chassis_number = request.POST.get('chassis_number', '').strip()
            engine_number = request.POST.get('engine_number', '').strip()
            status = request.POST.get('status', 'AVAILABLE')
            department_id = request.POST.get('department')
            current_user_name = request.POST.get('current_user_name', '').strip()
            previous_user_name = request.POST.get('previous_user_name', '').strip()
            odometer = request.POST.get('odometer', '').strip()
            notes = request.POST.get('notes', '').strip()

            if not brand or not plate_number:
                messages.error(request, 'សូមបញ្ចូលម៉ាក និងស្លាកលេខយានយន្ត!')
                raise ValueError("Missing brand or plate")

            dept_obj = Department.objects.filter(id=department_id).first() if department_id else None

            vehicle = Vehicle(
                name=name or f"{brand} ({plate_number})",
                vehicle_type=vehicle_type,
                brand=brand,
                model_year=model_year,
                color=color,
                plate_number=plate_number,
                chassis_number=chassis_number,
                engine_number=engine_number,
                status=status,
                department=dept_obj,
                current_user_name=current_user_name,
                previous_user_name=previous_user_name,
                odometer=odometer,
                notes=notes,
                created_by=request.user,
            )

            if 'photo_front' in request.FILES:
                vehicle.photo_front = request.FILES['photo_front']
            if 'photo_side' in request.FILES:
                vehicle.photo_side = request.FILES['photo_side']
            if 'registration_card_photo' in request.FILES:
                vehicle.registration_card_photo = request.FILES['registration_card_photo']

            vehicle.save()
            messages.success(request, f"បានចុះបញ្ជីយានយន្ត «{vehicle.brand} ({vehicle.plate_number})» ដោយជោគជ័យ!")
            return redirect('vehicle_list')
        except Exception as e:
            if not any(m.message for m in messages.get_messages(request)):
                messages.error(request, f"មានបញ្ហាក្នុងការចុះបញ្ជីយានយន្ត៖ {str(e)}")

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')
    return render(request, 'dms/vehicle_form.html', {'is_edit': False, 'departments': departments})


@login_required
def vehicle_edit(request, pk):
    """
    Edit vehicle/motorbike entry (Admin / General Affairs only).
    """
    vehicle = get_object_or_404(Vehicle, pk=pk)
    profile = getattr(request.user, 'profile', None)
    if not check_has_global_hr_tracking_access(request.user, profile):
        messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិកែប្រែទិន្នន័យយានយន្តឡើយ!')
        return redirect('vehicle_list')

    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            vehicle_type = request.POST.get('vehicle_type', 'MOTORCYCLE')
            brand = request.POST.get('brand', '').strip()
            model_year = request.POST.get('model_year', '').strip()
            color = request.POST.get('color', '').strip()
            plate_number = request.POST.get('plate_number', '').strip()
            chassis_number = request.POST.get('chassis_number', '').strip()
            engine_number = request.POST.get('engine_number', '').strip()
            status = request.POST.get('status', 'AVAILABLE')
            department_id = request.POST.get('department')
            current_user_name = request.POST.get('current_user_name', '').strip()
            previous_user_name = request.POST.get('previous_user_name', '').strip()
            odometer = request.POST.get('odometer', '').strip()
            notes = request.POST.get('notes', '').strip()

            if not brand or not plate_number:
                messages.error(request, 'សូមបញ្ចូលម៉ាក និងស្លាកលេខយានយន្ត!')
                raise ValueError("Missing brand or plate")

            vehicle.name = name or f"{brand} ({plate_number})"
            vehicle.vehicle_type = vehicle_type
            vehicle.brand = brand
            vehicle.model_year = model_year
            vehicle.color = color
            vehicle.plate_number = plate_number
            vehicle.chassis_number = chassis_number
            vehicle.engine_number = engine_number
            vehicle.status = status
            vehicle.department = Department.objects.filter(id=department_id).first() if department_id else None
            vehicle.current_user_name = current_user_name
            vehicle.previous_user_name = previous_user_name
            vehicle.odometer = odometer
            vehicle.notes = notes

            if 'photo_front' in request.FILES:
                vehicle.photo_front = request.FILES['photo_front']
            if 'photo_side' in request.FILES:
                vehicle.photo_side = request.FILES['photo_side']
            if 'registration_card_photo' in request.FILES:
                vehicle.registration_card_photo = request.FILES['registration_card_photo']

            vehicle.save()
            messages.success(request, f"បានកែប្រែទិន្នន័យយានយន្ត «{vehicle.brand} ({vehicle.plate_number})» ដោយជោគជ័យ!")
            return redirect('vehicle_list')
        except Exception as e:
            if not any(m.message for m in messages.get_messages(request)):
                messages.error(request, f"មានបញ្ហាក្នុងការកែប្រែ៖ {str(e)}")

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')
    return render(request, 'dms/vehicle_form.html', {'is_edit': True, 'vehicle': vehicle, 'departments': departments})


@login_required
def vehicle_delete(request, pk):
    """
    Delete a vehicle entry.
    """
    vehicle = get_object_or_404(Vehicle, pk=pk)
    profile = getattr(request.user, 'profile', None)
    if not check_has_global_hr_tracking_access(request.user, profile):
        messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិលុបទិន្នន័យយានយន្តឡើយ!')
        return redirect('vehicle_list')

    if request.method == 'POST':
        name = f"{vehicle.brand} ({vehicle.plate_number})"
        vehicle.delete()
        messages.success(request, f"បានលុបយានយន្ត «{name}» ចេញពីប្រព័ន្ធរួចរាល់!")
    return redirect('vehicle_list')


# ------------------------------------------------------------------------------
# 📝 VEHICLE BORROWING REQUESTS (ពាក្យសុំខ្ចីទ្រព្យប្រើប្រាស់)
# ------------------------------------------------------------------------------

@login_required
def vehicle_request_list(request):
    """
    List of Vehicle Borrowing Requests.
    Scoped: Specialized offices see their own department requests. Admin/Leadership see all.
    """
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    queryset = VehicleRequest.objects.all().select_related('applicant_department', 'vehicle', 'applicant')

    if not is_admin_or_lead:
        if dept:
            queryset = queryset.filter(applicant_department=dept)
        else:
            queryset = queryset.filter(applicant=request.user)
    else:
        dept_filter = request.GET.get('department', '').strip()
        if dept_filter:
            queryset = queryset.filter(applicant_department_id=dept_filter)

    status_filter = request.GET.get('status', '').strip()
    type_filter = request.GET.get('vehicle_type', '').strip()
    search_q = request.GET.get('q', '').strip()

    if status_filter:
        queryset = queryset.filter(status=status_filter)
    if type_filter:
        queryset = queryset.filter(vehicle_type=type_filter)
    if search_q:
        q_arabic = to_arabic_digits(search_q)
        queryset = queryset.filter(
            Q(applicant_name__icontains=search_q) |
            Q(applicant_id_number__icontains=search_q) |
            Q(applicant_id_number__icontains=q_arabic) |
            Q(applicant_phone__icontains=search_q) |
            Q(request_number__icontains=search_q) |
            Q(vehicle_brand__icontains=search_q) |
            Q(vehicle_plate_number__icontains=search_q) |
            Q(vehicle_plate_number__icontains=q_arabic) |
            Q(purpose__icontains=search_q)
        )

    # Counters scoped
    base_qs = VehicleRequest.objects.all() if is_admin_or_lead else (VehicleRequest.objects.filter(applicant_department=dept) if dept else VehicleRequest.objects.filter(applicant=request.user))
    total_count = base_qs.count()
    pending_count = base_qs.filter(status='PENDING').count()
    approved_count = base_qs.filter(status='APPROVED').count()
    in_use_count = base_qs.filter(status='HANDED_OVER').count()
    returned_count = base_qs.filter(status='RETURNED').count()
    rejected_count = base_qs.filter(status='REJECTED').count()

    paginator = Paginator(queryset.order_by('-created_at'), 15)
    page_number = request.GET.get('page', 1)
    requests_page = paginator.get_page(page_number)

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh') if is_admin_or_lead else []

    context = {
        'requests': requests_page,
        'departments': departments,
        'total_count': total_count,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'in_use_count': in_use_count,
        'returned_count': returned_count,
        'rejected_count': rejected_count,
        'status_filter': status_filter,
        'type_filter': type_filter,
        'search_q': search_q,
        'dept_filter': request.GET.get('department', '').strip(),
        'dept': dept,
        'is_admin_or_lead': is_admin_or_lead,
    }
    return render(request, 'dms/vehicle_request_list.html', context)


@login_required
def vehicle_request_create(request):
    """
    Create a new vehicle/motorbike loan request (Open to all offices).
    """
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    if request.method == 'POST':
        try:
            applicant_name = request.POST.get('applicant_name', '').strip()
            applicant_gender = request.POST.get('applicant_gender', 'MALE')
            applicant_position = request.POST.get('applicant_position', '').strip()
            applicant_id_number = request.POST.get('applicant_id_number', '').strip()
            applicant_phone = request.POST.get('applicant_phone', '').strip()
            if applicant_id_number:
                applicant_id_number = to_arabic_digits(applicant_id_number)
            if applicant_phone:
                applicant_phone = to_arabic_digits(applicant_phone)

            # Department selection: admin can choose; non-admin locked to own department
            department_obj = None
            if is_admin_or_lead:
                department_id = request.POST.get('applicant_department')
                if department_id:
                    department_obj = Department.objects.filter(id=department_id).first()
                elif dept:
                    department_obj = dept
            else:
                department_obj = dept

            applicant_department_name = request.POST.get('applicant_department_name', '').strip()
            if not applicant_department_name and department_obj:
                applicant_department_name = department_obj.name_kh

            # Vehicle info
            vehicle_id = request.POST.get('vehicle')
            vehicle_obj = None
            if vehicle_id:
                vehicle_obj = Vehicle.objects.filter(id=vehicle_id).first()

            vehicle_type = request.POST.get('vehicle_type', 'MOTORCYCLE')
            vehicle_brand = request.POST.get('vehicle_brand', '').strip()
            vehicle_color = request.POST.get('vehicle_color', '').strip()
            vehicle_model_year = request.POST.get('vehicle_model_year', '').strip()
            vehicle_chassis_number = request.POST.get('vehicle_chassis_number', '').strip()
            vehicle_engine_number = request.POST.get('vehicle_engine_number', '').strip()
            vehicle_plate_number = request.POST.get('vehicle_plate_number', '').strip()

            if vehicle_obj:
                if not vehicle_brand:
                    vehicle_brand = vehicle_obj.brand
                if not vehicle_color:
                    vehicle_color = vehicle_obj.color
                if not vehicle_model_year:
                    vehicle_model_year = vehicle_obj.model_year
                if not vehicle_chassis_number:
                    vehicle_chassis_number = vehicle_obj.chassis_number
                if not vehicle_engine_number:
                    vehicle_engine_number = vehicle_obj.engine_number
                if not vehicle_plate_number:
                    vehicle_plate_number = vehicle_obj.plate_number
                vehicle_type = vehicle_obj.vehicle_type

            # Duration & Purpose
            duration_text = request.POST.get('duration_text', '១ឆ្នាំ').strip()
            start_date_str = request.POST.get('start_date', '').strip()
            end_date_str = request.POST.get('end_date', '').strip()
            purpose = request.POST.get('purpose', '').strip()

            start_date = None
            if start_date_str:
                try:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                except Exception:
                    pass

            end_date = None
            if end_date_str:
                try:
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                except Exception:
                    pass

            if not applicant_name:
                messages.error(request, 'សូមបញ្ចូលឈ្មោះអ្នកស្នើសុំ!')
                raise ValueError("Missing applicant name")

            # Try link with CivilServantProfile
            applicant_profile = None
            if applicant_id_number:
                applicant_profile = CivilServantProfile.objects.filter(officer_id_number=applicant_id_number).first()
            if not applicant_profile:
                parts = applicant_name.split()
                if len(parts) >= 2:
                    applicant_profile = CivilServantProfile.objects.filter(khmer_last_name=parts[0], khmer_first_name=' '.join(parts[1:])).first()

            v_req = VehicleRequest(
                applicant=request.user,
                applicant_profile=applicant_profile,
                applicant_name=applicant_name,
                applicant_gender=applicant_gender,
                applicant_position=applicant_position,
                applicant_department=department_obj,
                applicant_department_name=applicant_department_name,
                applicant_id_number=applicant_id_number,
                applicant_phone=applicant_phone,
                vehicle=vehicle_obj,
                vehicle_type=vehicle_type,
                vehicle_brand=vehicle_brand,
                vehicle_color=vehicle_color,
                vehicle_model_year=vehicle_model_year,
                vehicle_chassis_number=vehicle_chassis_number,
                vehicle_engine_number=vehicle_engine_number,
                vehicle_plate_number=vehicle_plate_number,
                duration_text=duration_text or '១ឆ្នាំ',
                start_date=start_date,
                end_date=end_date,
                purpose=purpose,
                status='PENDING',
            )

            last_user_name = request.POST.get('last_user_name', '').strip()
            if last_user_name:
                v_req.last_user_name = last_user_name

            if 'photo_front' in request.FILES:
                v_req.photo_front = request.FILES['photo_front']
            if 'photo_side' in request.FILES:
                v_req.photo_side = request.FILES['photo_side']
            if 'registration_card_photo' in request.FILES:
                v_req.registration_card_photo = request.FILES['registration_card_photo']

            # Auto request number
            current_year = datetime.now().year
            seq_count = VehicleRequest.objects.filter(created_at__year=current_year).count() + 1
            v_req.request_number = f"{seq_count:02d}/{current_year} មករន.បល"

            v_req.save()

            messages.success(request, f"បានបញ្ជូនពាក្យសុំខ្ចីមធ្យោបាយ «{v_req.display_vehicle_type_kh} {v_req.display_brand}» ដោយជោគជ័យ!")
            return redirect('vehicle_request_detail', pk=v_req.pk)
        except Exception as e:
            if not any(m.message for m in messages.get_messages(request)):
                messages.error(request, f"មានបញ្ហាក្នុងការបង្កើតពាក្យសុំ៖ {str(e)}")

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh') if is_admin_or_lead else (Department.objects.filter(id=dept.id) if dept else [])
    available_vehicles = Vehicle.objects.filter(status='AVAILABLE').order_by('vehicle_type', 'brand')

    officers = CivilServantProfile.objects.all()
    if not is_admin_or_lead and dept:
        officers = officers.filter(department=dept)
    officers = officers.order_by('khmer_last_name', 'khmer_first_name')

    context = {
        'is_edit': False,
        'departments': departments,
        'available_vehicles': available_vehicles,
        'officers': officers,
        'dept': dept,
        'is_admin_or_lead': is_admin_or_lead,
    }
    return render(request, 'dms/vehicle_request_form.html', context)


@login_required
def vehicle_request_detail(request, pk):
    """
    Detail view of a Vehicle Request including Handover & Return actions.
    Strict Department Isolation: Specialized offices can ONLY view requests of their department.
    """
    v_req = get_object_or_404(VehicleRequest, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    if not is_admin_or_lead:
        if v_req.applicant_department != dept and v_req.applicant != request.user:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិមើលព័ត៌មានពាក្យសុំរបស់ការិយាល័យផ្សេងឡើយ!')
            return redirect('vehicle_request_list')

    attachments = v_req.attachments.all().order_by('-created_at')
    can_edit = is_admin_or_lead or (v_req.status == 'PENDING' and (v_req.applicant_department == dept or v_req.applicant == request.user))
    can_approve = is_admin_or_lead
    available_vehicles = Vehicle.objects.filter(status='AVAILABLE')

    context = {
        'request_obj': v_req,
        'attachments': attachments,
        'can_edit': can_edit,
        'can_approve': can_approve,
        'available_vehicles': available_vehicles,
        'is_admin_or_lead': is_admin_or_lead,
    }
    return render(request, 'dms/vehicle_request_detail.html', context)


@login_required
def vehicle_request_edit(request, pk):
    """
    Edit a Vehicle Request.
    """
    v_req = get_object_or_404(VehicleRequest, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    if not is_admin_or_lead:
        if v_req.applicant_department != dept and v_req.applicant != request.user:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិកែប្រែពាក្យសុំរបស់ការិយាល័យផ្សេងឡើយ!')
            return redirect('vehicle_request_list')

    if request.method == 'POST':
        try:
            applicant_name = request.POST.get('applicant_name', '').strip()
            applicant_gender = request.POST.get('applicant_gender', 'MALE')
            applicant_position = request.POST.get('applicant_position', '').strip()
            applicant_id_number = request.POST.get('applicant_id_number', '').strip()
            applicant_phone = request.POST.get('applicant_phone', '').strip()
            if applicant_id_number:
                applicant_id_number = to_arabic_digits(applicant_id_number)
            if applicant_phone:
                applicant_phone = to_arabic_digits(applicant_phone)

            if is_admin_or_lead:
                department_id = request.POST.get('applicant_department')
                if department_id:
                    v_req.applicant_department = Department.objects.filter(id=department_id).first()

            applicant_department_name = request.POST.get('applicant_department_name', '').strip()
            if applicant_department_name:
                v_req.applicant_department_name = applicant_department_name

            vehicle_id = request.POST.get('vehicle')
            if vehicle_id:
                v_req.vehicle = Vehicle.objects.filter(id=vehicle_id).first()

            vehicle_type = request.POST.get('vehicle_type', 'MOTORCYCLE')
            vehicle_brand = request.POST.get('vehicle_brand', '').strip()
            vehicle_color = request.POST.get('vehicle_color', '').strip()
            vehicle_model_year = request.POST.get('vehicle_model_year', '').strip()
            vehicle_chassis_number = request.POST.get('vehicle_chassis_number', '').strip()
            vehicle_engine_number = request.POST.get('vehicle_engine_number', '').strip()
            vehicle_plate_number = request.POST.get('vehicle_plate_number', '').strip()

            last_user_name = request.POST.get('last_user_name', '').strip()
            if last_user_name:
                v_req.last_user_name = last_user_name

            if 'photo_front' in request.FILES:
                v_req.photo_front = request.FILES['photo_front']
            if 'photo_side' in request.FILES:
                v_req.photo_side = request.FILES['photo_side']
            if 'registration_card_photo' in request.FILES:
                v_req.registration_card_photo = request.FILES['registration_card_photo']

            duration_text = request.POST.get('duration_text', '១ឆ្នាំ').strip()
            start_date_str = request.POST.get('start_date', '').strip()
            end_date_str = request.POST.get('end_date', '').strip()
            purpose = request.POST.get('purpose', '').strip()

            if start_date_str:
                try:
                    v_req.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                except Exception:
                    pass
            else:
                v_req.start_date = None

            if end_date_str:
                try:
                    v_req.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                except Exception:
                    pass
            else:
                v_req.end_date = None

            if not applicant_name:
                messages.error(request, 'សូមបញ្ចូលឈ្មោះអ្នកស្នើសុំ!')
                raise ValueError("Missing name")

            v_req.applicant_name = applicant_name
            v_req.applicant_gender = applicant_gender
            v_req.applicant_position = applicant_position
            v_req.applicant_id_number = applicant_id_number
            v_req.applicant_phone = applicant_phone
            v_req.vehicle_type = vehicle_type
            v_req.vehicle_brand = vehicle_brand
            v_req.vehicle_color = vehicle_color
            v_req.vehicle_model_year = vehicle_model_year
            v_req.vehicle_chassis_number = vehicle_chassis_number
            v_req.vehicle_engine_number = vehicle_engine_number
            v_req.vehicle_plate_number = vehicle_plate_number
            v_req.duration_text = duration_text
            v_req.purpose = purpose

            v_req.save()

            messages.success(request, f"បានកែប្រែពាក្យសុំ «{v_req.display_vehicle_type_kh} {v_req.display_brand}» ដោយជោគជ័យ!")
            return redirect('vehicle_request_detail', pk=v_req.pk)
        except Exception as e:
            if not any(m.message for m in messages.get_messages(request)):
                messages.error(request, f"មានបញ្ហាក្នុងការកែប្រែ៖ {str(e)}")

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh') if is_admin_or_lead else (Department.objects.filter(id=dept.id) if dept else [])
    available_vehicles = Vehicle.objects.filter(Q(status='AVAILABLE') | Q(id=v_req.vehicle_id if v_req.vehicle else 0)).order_by('vehicle_type', 'brand')

    officers = CivilServantProfile.objects.all()
    if not is_admin_or_lead and dept:
        officers = officers.filter(department=dept)
    officers = officers.order_by('khmer_last_name', 'khmer_first_name')

    context = {
        'is_edit': True,
        'request_obj': v_req,
        'departments': departments,
        'available_vehicles': available_vehicles,
        'officers': officers,
        'dept': dept,
        'is_admin_or_lead': is_admin_or_lead,
    }
    return render(request, 'dms/vehicle_request_form.html', context)


@login_required
def vehicle_request_delete(request, pk):
    """
    Delete a Vehicle Request.
    """
    v_req = get_object_or_404(VehicleRequest, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    if not is_admin_or_lead:
        if v_req.applicant_department != dept and v_req.applicant != request.user:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិលុបពាក្យសុំរបស់ការិយាល័យផ្សេងឡើយ!')
            return redirect('vehicle_request_list')

    if request.method == 'POST':
        name = v_req.applicant_name
        v_req.delete()
        messages.success(request, f"បានលុបពាក្យសុំរបស់លោក/លោកស្រី «{name}» ចេញពីប្រព័ន្ធរួចរាល់!")
    return redirect('vehicle_request_list')


@login_required
def vehicle_request_upload_photos(request, pk):
    """
    Quick upload or update vehicle photos (Front, Side, Registration Card) and last user name.
    """
    v_req = get_object_or_404(VehicleRequest, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)

    if not has_global_access:
        if v_req.applicant_department != dept and v_req.applicant != request.user:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិកែប្រែពាក្យសុំរបស់ការិយាល័យផ្សេងឡើយ!')
            return redirect('vehicle_request_list')

    if request.method == 'POST':
        try:
            last_user_name = request.POST.get('last_user_name', '').strip()
            if last_user_name:
                v_req.last_user_name = last_user_name

            if 'photo_front' in request.FILES:
                v_req.photo_front = request.FILES['photo_front']
            if 'photo_side' in request.FILES:
                v_req.photo_side = request.FILES['photo_side']
            if 'registration_card_photo' in request.FILES:
                v_req.registration_card_photo = request.FILES['registration_card_photo']

            v_req.save()

            # Also update vehicle if linked
            if v_req.vehicle:
                v = v_req.vehicle
                if v_req.photo_front and not v.photo_front:
                    v.photo_front = v_req.photo_front
                if v_req.photo_side and not v.photo_side:
                    v.photo_side = v_req.photo_side
                if v_req.registration_card_photo and not v.registration_card_photo:
                    v.registration_card_photo = v_req.registration_card_photo
                if last_user_name and not v.previous_user_name:
                    v.previous_user_name = last_user_name
                v.save()

            messages.success(request, '✅ បានបញ្ចូល និងរក្សាទុករូបថតយានយន្តដោយជោគជ័យ!')
        except Exception as e:
            messages.error(request, f'មានបញ្ហាក្នុងការបញ្ចូលរូបថត៖ {str(e)}')

    return redirect('vehicle_request_detail', pk=v_req.pk)


@login_required
def vehicle_request_approve(request, pk):
    """
    Approve vehicle request (Admin / Leadership only).
    """
    v_req = get_object_or_404(VehicleRequest, pk=pk)
    profile = getattr(request.user, 'profile', None)
    if not check_has_global_hr_tracking_access(request.user, profile):
        messages.error(request, '⚠️ មានតែ ADMIN ឬថ្នាក់ដឹកនាំប៉ុណ្ណោះដែលអាចអនុម័តពាក្យសុំ!')
        return redirect('vehicle_request_detail', pk=v_req.pk)

    if request.method == 'POST':
        comments = request.POST.get('approval_comments', '').strip()
        vehicle_id = request.POST.get('assigned_vehicle')
        if vehicle_id:
            assigned_vehicle = Vehicle.objects.filter(id=vehicle_id).first()
            if assigned_vehicle:
                v_req.vehicle = assigned_vehicle
                v_req.vehicle_brand = assigned_vehicle.brand
                v_req.vehicle_color = assigned_vehicle.color
                v_req.vehicle_model_year = assigned_vehicle.model_year
                v_req.vehicle_chassis_number = assigned_vehicle.chassis_number
                v_req.vehicle_engine_number = assigned_vehicle.engine_number
                v_req.vehicle_plate_number = assigned_vehicle.plate_number
                v_req.vehicle_type = assigned_vehicle.vehicle_type

        v_req.status = 'APPROVED'
        v_req.approved_by = request.user
        v_req.approval_date = timezone.now().date()
        v_req.approval_comments = comments
        v_req.save()
        messages.success(request, f"✅ បានអនុម័តពាក្យសុំរបស់ «{v_req.applicant_name}» ដោយជោគជ័យ!")

    return redirect('vehicle_request_detail', pk=v_req.pk)


@login_required
def vehicle_request_reject(request, pk):
    """
    Reject vehicle request (Admin / Leadership only).
    """
    v_req = get_object_or_404(VehicleRequest, pk=pk)
    profile = getattr(request.user, 'profile', None)
    if not check_has_global_hr_tracking_access(request.user, profile):
        messages.error(request, '⚠️ មានតែ ADMIN ឬថ្នាក់ដឹកនាំប៉ុណ្ណោះដែលអាចបដិសេធពាក្យសុំ!')
        return redirect('vehicle_request_detail', pk=v_req.pk)

    if request.method == 'POST':
        comments = request.POST.get('approval_comments', '').strip()
        v_req.status = 'REJECTED'
        v_req.approved_by = request.user
        v_req.approval_date = timezone.now().date()
        v_req.approval_comments = comments
        v_req.save()
        messages.warning(request, f"បានបដិសេធពាក្យសុំរបស់ «{v_req.applicant_name}»!")

    return redirect('vehicle_request_detail', pk=v_req.pk)


@login_required
def vehicle_request_handover(request, pk):
    """
    Execute Handover Minutes (កំណត់ហេតុប្រគល់-ទទួល).
    Transitions request to 'HANDED_OVER' and vehicle to 'IN_USE'.
    """
    v_req = get_object_or_404(VehicleRequest, pk=pk)
    profile = getattr(request.user, 'profile', None)
    if not check_has_global_hr_tracking_access(request.user, profile):
        messages.error(request, '⚠️ មានតែ ADMIN ឬមន្ត្រីទទួលបន្ទុកទ្រព្យសម្បត្តិរដ្ឋប៉ុណ្ណោះដែលអាចបំពេញកំណត់ហេតុប្រគល់-ទទួល!')
        return redirect('vehicle_request_detail', pk=v_req.pk)

    if request.method == 'POST':
        try:
            handover_datetime_str = request.POST.get('handover_datetime', '').strip()
            handover_officer_name = request.POST.get('handover_officer_name', '').strip()
            handover_officer_position = request.POST.get('handover_officer_position', '').strip()
            recipient_name = request.POST.get('recipient_name', '').strip() or v_req.applicant_name
            last_user_name = request.POST.get('last_user_name', '').strip()
            vehicle_condition = request.POST.get('vehicle_condition_handover', '').strip()
            odometer = request.POST.get('odometer_at_handover', '').strip()

            handover_datetime = None
            if handover_datetime_str:
                try:
                    handover_datetime = datetime.strptime(handover_datetime_str, '%Y-%m-%dT%H:%M')
                except Exception:
                    try:
                        handover_datetime = datetime.strptime(handover_datetime_str, '%Y-%m-%d %H:%M')
                    except Exception:
                        handover_datetime = timezone.now()
            else:
                handover_datetime = timezone.now()

            v_req.handover_datetime = handover_datetime
            v_req.handover_officer_name = handover_officer_name
            v_req.handover_officer_position = handover_officer_position
            v_req.recipient_name = recipient_name
            v_req.last_user_name = last_user_name
            v_req.vehicle_condition_handover = vehicle_condition
            v_req.odometer_at_handover = odometer
            v_req.status = 'HANDED_OVER'

            if 'photo_front' in request.FILES:
                v_req.photo_front = request.FILES['photo_front']
            if 'photo_side' in request.FILES:
                v_req.photo_side = request.FILES['photo_side']
            if 'registration_card_photo' in request.FILES:
                v_req.registration_card_photo = request.FILES['registration_card_photo']

            v_req.save()

            # Update assigned vehicle status
            if v_req.vehicle:
                v = v_req.vehicle
                v.status = 'IN_USE'
                v.current_user_name = v_req.recipient_name or v_req.applicant_name
                if last_user_name:
                    v.previous_user_name = last_user_name
                if odometer:
                    v.odometer = odometer
                if v_req.photo_front and not v.photo_front:
                    v.photo_front = v_req.photo_front
                if v_req.photo_side and not v.photo_side:
                    v.photo_side = v_req.photo_side
                if v_req.registration_card_photo and not v.registration_card_photo:
                    v.registration_card_photo = v_req.registration_card_photo
                v.save()

            messages.success(request, f"✅ បានបំពេញកំណត់ហេតុប្រគល់-ទទួលជូន «{v_req.recipient_name}» ដោយជោគជ័យ!")
        except Exception as e:
            messages.error(request, f"មានបញ្ហាក្នុងការរក្សាទុកកំណត់ហេតុ៖ {str(e)}")

    return redirect('vehicle_request_detail', pk=v_req.pk)


@login_required
def vehicle_request_return(request, pk):
    """
    Record return of vehicle/motorbike.
    Transitions request to 'RETURNED' and vehicle to 'AVAILABLE'.
    """
    v_req = get_object_or_404(VehicleRequest, pk=pk)
    profile = getattr(request.user, 'profile', None)
    if not check_has_global_hr_tracking_access(request.user, profile):
        messages.error(request, '⚠️ មានតែ ADMIN ឬមន្ត្រីទទួលបន្ទុកប៉ុណ្ណោះដែលអាចកត់ត្រាការប្រគល់យានយន្តមកវិញ!')
        return redirect('vehicle_request_detail', pk=v_req.pk)

    if request.method == 'POST':
        try:
            return_datetime_str = request.POST.get('return_datetime', '').strip()
            return_condition = request.POST.get('return_condition', '').strip()
            odometer_at_return = request.POST.get('odometer_at_return', '').strip()
            return_notes = request.POST.get('return_notes', '').strip()

            return_datetime = None
            if return_datetime_str:
                try:
                    return_datetime = datetime.strptime(return_datetime_str, '%Y-%m-%dT%H:%M')
                except Exception:
                    return_datetime = timezone.now()
            else:
                return_datetime = timezone.now()

            v_req.return_datetime = return_datetime
            v_req.return_condition = return_condition
            v_req.odometer_at_return = odometer_at_return
            v_req.return_notes = return_notes
            v_req.received_by = request.user
            v_req.status = 'RETURNED'
            v_req.save()

            if v_req.vehicle:
                v = v_req.vehicle
                v.status = 'AVAILABLE'
                v.previous_user_name = v.current_user_name or v_req.applicant_name
                v.current_user_name = ''
                if odometer_at_return:
                    v.odometer = odometer_at_return
                v.save()

            messages.success(request, f"✅ បានកត់ត្រាការប្រគល់យានយន្តត្រឡប់មកវិញដោយជោគជ័យ!")
        except Exception as e:
            messages.error(request, f"មានបញ្ហាក្នុងការរក្សាទុក៖ {str(e)}")

    return redirect('vehicle_request_detail', pk=v_req.pk)


@login_required
def vehicle_request_print(request, pk):
    """
    Standard Official Khmer A4 Print Layout matching 'ពាក្យសុំខ្ចីទ្រព្យប្រើប្រាស់២០២៦.doc' 100%.
    Supports printing Application Letter (Page 1), Handover Minutes (Page 2), or Complete Dossier.
    """
    v_req = get_object_or_404(VehicleRequest, pk=pk)
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)

    if not has_global_access:
        if v_req.applicant_department != dept and v_req.applicant != request.user:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិមើល ឬបោះពុម្ពពាក្យសុំរបស់ការិយាល័យផ្សេងឡើយ!')
            return redirect('vehicle_request_list')

    page_mode = request.GET.get('page', 'all')

    context = {
        'req': v_req,
        'page_mode': page_mode,
    }
    return render(request, 'dms/vehicle_request_print.html', context)


@login_required
def vehicle_request_export_excel(request):
    """
    Export Vehicle Loan Requests to Excel formatted nicely.
    """
    profile = getattr(request.user, 'profile', None)
    dept = profile.department if profile else None
    has_global_access = check_has_global_hr_tracking_access(request.user, profile)
    is_admin_or_lead = has_global_access

    queryset = VehicleRequest.objects.all().select_related('applicant_department', 'vehicle')

    if not is_admin_or_lead:
        if dept:
            queryset = queryset.filter(applicant_department=dept)
        else:
            queryset = queryset.filter(applicant=request.user)
    else:
        dept_filter = request.GET.get('department', '').strip()
        if dept_filter:
            queryset = queryset.filter(applicant_department_id=dept_filter)

    status_filter = request.GET.get('status', '').strip()
    type_filter = request.GET.get('vehicle_type', '').strip()
    search_q = request.GET.get('q', '').strip()

    if status_filter:
        queryset = queryset.filter(status=status_filter)
    if type_filter:
        queryset = queryset.filter(vehicle_type=type_filter)
    if search_q:
        q_arabic = to_arabic_digits(search_q)
        queryset = queryset.filter(
            Q(applicant_name__icontains=search_q) |
            Q(applicant_id_number__icontains=search_q) |
            Q(applicant_id_number__icontains=q_arabic) |
            Q(vehicle_brand__icontains=search_q) |
            Q(vehicle_plate_number__icontains=search_q)
        )

    records = queryset.order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ពាក្យសុំខ្ចីរថយន្ត-ម៉ូតូ"

    font_title = Font(name="Khmer OS Muol Light", size=14, bold=True, color="1E3A8A")
    font_sub = Font(name="Khmer OS Battambang", size=11, italic=True)
    font_header = Font(name="Khmer OS Battambang", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Khmer OS Battambang", size=10)
    font_data_bold = Font(name="Khmer OS Battambang", size=10, bold=True)
    font_latin = Font(name="Arial", size=10)

    fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.merge_cells('A1:J1')
    ws['A1'] = "បញ្ជីពាក្យសុំខ្ចីប្រើប្រាស់រថយន្ត និងម៉ូតូរបស់អង្គភាព"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center

    ws.merge_cells('A2:J2')
    sub_title = f"ទាញយកនៅថ្ងៃទី {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    if not is_admin_or_lead and dept:
        sub_title += f" | {dept.name_kh}"
    ws['A2'] = sub_title
    ws['A2'].font = font_sub
    ws['A2'].alignment = align_center

    headers = [
        "ល.រ", "លេខលិខិតសុំ", "ឈ្មោះអ្នកស្នើសុំ", "តួនាទី & ការិយាល័យ",
        "ប្រភេទមធ្យោបាយ", "ម៉ាក & ស្លាកលេខ", "រយៈពេលស្នើសុំ", "កាលបរិច្ឆេទ", "ស្ថានភាព", "គោលបំណង"
    ]

    ws.row_dimensions[4].height = 28
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    row_num = 5
    for idx, r in enumerate(records, 1):
        ws.row_dimensions[row_num].height = 24
        is_even = (idx % 2 == 0)

        date_range = "-"
        if r.start_date and r.end_date:
            date_range = f"{r.start_date.strftime('%d/%m/%Y')} - {r.end_date.strftime('%d/%m/%Y')}"

        data_row = [
            (idx, align_center, font_data),
            (r.request_number or "-", align_center, font_latin),
            (r.applicant_name, align_left, font_data_bold),
            (f"{r.applicant_position or ''} - {r.applicant_department_name or ''}".strip(' -'), align_left, font_data),
            (r.display_vehicle_type_kh, align_center, font_data),
            (f"{r.display_brand} ({r.display_plate_number})", align_left, font_data),
            (r.duration_text or "-", align_center, font_data),
            (date_range, align_center, font_latin),
            (r.get_status_display(), align_center, font_data_bold),
            (r.purpose or "-", align_left, font_data),
        ]

        for col_num, (val, alignment, font_style) in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.alignment = alignment
            cell.font = font_style
            cell.border = thin_border
            if is_even:
                cell.fill = fill_zebra

        row_num += 1

    column_widths = [6, 18, 22, 30, 16, 26, 16, 24, 20, 35]
    for i, w in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="vehicle_requests_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@login_required
def vehicle_export_excel(request):
    """
    Export Vehicle Inventory list to Excel.
    """
    profile = getattr(request.user, 'profile', None)

    queryset = Vehicle.objects.all().select_related('department').order_by('vehicle_type', 'brand', 'plate_number')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "សារពើភណ្ឌយានយន្ត"

    font_title = Font(name="Khmer OS Muol Light", size=14, bold=True, color="1E3A8A")
    font_sub = Font(name="Khmer OS Battambang", size=11, italic=True)
    font_header = Font(name="Khmer OS Battambang", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Khmer OS Battambang", size=10)
    font_data_bold = Font(name="Khmer OS Battambang", size=10, bold=True)
    font_latin = Font(name="Arial", size=10)

    fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.merge_cells('A1:I1')
    ws['A1'] = "បញ្ជីសារពើភណ្ឌរថយន្ត និងម៉ូតូរបស់អង្គភាព"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center

    ws.merge_cells('A2:I2')
    ws['A2'] = f"ទាញយកនៅថ្ងៃទី {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = font_sub
    ws['A2'].alignment = align_center

    headers = [
        "ល.រ", "ប្រភេទ", "ម៉ាក (Model)", "ស្លាកលេខ", "ពណ៌ & ឆ្នាំផលិត",
        "លេខតួ & លេខម៉ាស៊ីន", "ស្ថានភាព", "អង្គភាពកាន់កាប់", "អ្នកប្រើប្រាស់បច្ចុប្បន្ន"
    ]

    ws.row_dimensions[4].height = 28
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    row_num = 5
    for idx, v in enumerate(queryset, 1):
        ws.row_dimensions[row_num].height = 24
        is_even = (idx % 2 == 0)

        data_row = [
            (idx, align_center, font_data),
            (v.get_vehicle_type_display(), align_center, font_data),
            (v.brand, align_left, font_data_bold),
            (v.plate_number, align_center, font_latin),
            (f"{v.color or '-'} ({v.model_year or '-'})", align_center, font_data),
            (f"តួ: {v.chassis_number or '-'} | ម៉ាស៊ីន: {v.engine_number or '-'}", align_left, font_latin),
            (v.get_status_display(), align_center, font_data_bold),
            (v.department.name_kh if v.department else "-", align_left, font_data),
            (v.current_user_name or "-", align_left, font_data),
        ]

        for col_num, (val, alignment, font_style) in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.alignment = alignment
            cell.font = font_style
            cell.border = thin_border
            if is_even:
                cell.fill = fill_zebra

        row_num += 1

    column_widths = [6, 18, 22, 18, 20, 32, 22, 28, 24]
    for i, w in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="vehicle_inventory_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@login_required
def api_get_officer_info(request):
    """
    AJAX helper to fetch officer details (MCS ID, position, department, phone, gender)
    for auto-populating vehicle request form.
    """
    officer_id = request.GET.get('officer_id', '').strip()
    if not officer_id:
        return JsonResponse({'success': False, 'error': 'Missing officer_id'})

    officer = CivilServantProfile.objects.filter(id=officer_id).select_related('department').first()
    if not officer:
        return JsonResponse({'success': False, 'error': 'Officer not found'})

    return JsonResponse({
        'success': True,
        'officer': {
            'id': officer.id,
            'name': officer.full_name_kh,
            'gender': officer.gender,
            'position': officer.current_position_title or '',
            'department_id': officer.department_id,
            'department_name': officer.department.name_kh if officer.department else '',
            'id_number': officer.officer_id_number or '',
            'phone': officer.phone or '',
        }
    })


# ==============================================================================
# 📋 ATTENDANCE MANAGEMENT MODULE (ម៉ូឌុលគ្រប់គ្រងវត្តមានមន្ត្រី និងមន្ត្រីជាប់កិច្ចសន្យា)
# ==============================================================================
from django.urls import reverse

def _is_admin_user(user):
    """
    Strict Admin Check: Only User 'ADMIN' or superusers have multi-department viewing rights.
    """
    if not user or not user.is_authenticated:
        return False
    return user.is_superuser or user.username.upper() == 'ADMIN'


def _get_leadership_roster():
    """
    Builds the Leadership & Heads Roster matching Attendend B.xlsx:
    Section I: ថ្នាក់ដឹកនាំមន្ទីរ (Department Leadership: ប្រធានមន្ទីរ, អនុប្រធានមន្ទីរ, នាយក, នាយករង, ប្រធានអង្គភាព, អនុប្រធានអង្គភាព or any non-bureau/non-canton leadership role)
    Section II: ខណ្ឌរដ្ឋបាល (Canton Chiefs; if none, Deputy Chief or sole/representative officer)
    Section III: ប្រធានការិយាល័យ (Office Chiefs; if none, Deputy Chief or sole/representative officer)
    """
    section_lead = []
    section_canton = []
    section_office = []

    # 1. Section I: Department Leadership (ថ្នាក់ដឹកនាំមន្ទីរ)
    lead_officers = list(CivilServantProfile.objects.filter(
        officer_status__in=['ACTIVE', 'IN_OFFICE', 'STUDYING_IN_COUNTRY', 'STUDYING_ABROAD']
    ).filter(
        current_position_title__iregex=r'(ប្រធានមន្ទីរ|អនុប្រធានមន្ទីរ|នាយក|នាយករង|ប្រធានអង្គភាព|អនុប្រធានអង្គភាព)'
    ).order_by('id'))

    lead_dept = Department.objects.filter(code='LEAD', is_active=True).first()
    if lead_dept:
        for o in CivilServantProfile.objects.filter(department=lead_dept).exclude(officer_status__in=['DISMISSED', 'RETIRED', 'TRANSFERRED_OUT']):
            if o not in lead_officers:
                lead_officers.append(o)

    # Any other officer with a position that is NOT a bureau / canton / regular staff position
    for o in CivilServantProfile.objects.exclude(officer_status__in=['DISMISSED', 'RETIRED', 'TRANSFERRED_OUT']).select_related('department'):
        pos = (o.current_position_title or '').strip()
        if not pos:
            continue
        is_bureau_canton_staff = (
            'ការិយាល័យ' in pos or 'ខណ្ឌ' in pos or 'ផ្នែក' in pos or 'មន្ត្រី' in pos or 'មន្រ្តី' in pos or 'មន្រ្ដី' in pos
        )
        if not is_bureau_canton_staff and o not in lead_officers:
            lead_officers.append(o)

    def _lead_sort_key(o):
        pos = normalize_khmer_role(o.current_position_title or '')
        if 'ប្រធានមន្ទីរ' in pos and 'អនុ' not in pos:
            return (1, get_rank_step_sort_weight(o.current_rank_and_step), o.full_name_kh)
        elif 'អនុប្រធានមន្ទីរ' in pos:
            return (2, get_rank_step_sort_weight(o.current_rank_and_step), o.full_name_kh)
        return (3, get_rank_step_sort_weight(o.current_rank_and_step), o.full_name_kh)

    sorted_lead = sorted(lead_officers, key=_lead_sort_key)
    for idx, o in enumerate(sorted_lead, 1):
        section_lead.append({
            'section': 'I. ថ្នាក់ដឹកនាំមន្ទីរ',
            'section_code': 'LEAD',
            'section_num': 1,
            'index': idx,
            'officer': o,
            'person_id': o.id,
            'person_type': 'CIVIL_SERVANT',
            'name': o.full_name_kh,
            'gender': 'ប' if o.gender == 'MALE' else 'ស',
            'position': o.current_position_title or 'ថ្នាក់ដឹកនាំ',
            'department': o.department,
            'is_acting': False,
        })

    # 2. Section II: Canton Heads (ខណ្ឌរដ្ឋបាល)
    canton_depts = Department.objects.filter(is_active=True, name_kh__contains='ខណ្ឌ').order_by('order_index', 'name_kh')
    canton_idx = 1
    for d in canton_depts:
        chief = CivilServantProfile.objects.filter(
            department=d
        ).exclude(officer_status__in=['DISMISSED', 'RETIRED', 'TRANSFERRED_OUT']).filter(
            current_position_title__iregex=r'នាយខណ្ឌ'
        ).exclude(
            current_position_title__iregex=r'នាយរង'
        ).first()

        is_acting = False
        head_officer = chief

        if not head_officer:
            head_officer = CivilServantProfile.objects.filter(
                department=d
            ).exclude(officer_status__in=['DISMISSED', 'RETIRED', 'TRANSFERRED_OUT']).filter(
                current_position_title__iregex=r'នាយរង'
            ).first()
            if head_officer:
                is_acting = True

        if not head_officer:
            head_officer = CivilServantProfile.objects.filter(
                department=d
            ).exclude(officer_status__in=['DISMISSED', 'RETIRED', 'TRANSFERRED_OUT']).first()
            if head_officer:
                is_acting = True

        if head_officer:
            pos_display = head_officer.current_position_title or d.name_kh
            if is_acting and 'នាយខណ្ឌ' not in pos_display:
                if 'នាយរង' in pos_display:
                    pos_display = f"{pos_display} (ទទួលបន្ទុករួម)"
                else:
                    pos_display = f"{pos_display} (តំណាង {d.name_kh})"

            section_canton.append({
                'section': 'II. ខណ្ឌរដ្ឋបាល',
                'section_code': 'CANTON',
                'section_num': 2,
                'index': canton_idx,
                'officer': head_officer,
                'person_id': head_officer.id,
                'person_type': 'CIVIL_SERVANT',
                'name': head_officer.full_name_kh,
                'gender': 'ប' if head_officer.gender == 'MALE' else 'ស',
                'position': pos_display,
                'department': d,
                'is_acting': is_acting,
            })
            canton_idx += 1

    # 3. Section III: Office Heads (ប្រធានការិយាល័យ)
    office_depts = Department.objects.filter(is_active=True).exclude(code='LEAD').exclude(name_kh__contains='ខណ្ឌ').order_by('order_index', 'name_kh')
    office_idx = 1
    for d in office_depts:
        chief = CivilServantProfile.objects.filter(
            department=d
        ).exclude(officer_status__in=['DISMISSED', 'RETIRED', 'TRANSFERRED_OUT']).filter(
            current_position_title__iregex=r'ប្រធានការិយាល័យ'
        ).exclude(
            current_position_title__iregex=r'អនុប្រធាន'
        ).first()

        is_acting = False
        head_officer = chief

        if not head_officer:
            head_officer = CivilServantProfile.objects.filter(
                department=d
            ).exclude(officer_status__in=['DISMISSED', 'RETIRED', 'TRANSFERRED_OUT']).filter(
                current_position_title__iregex=r'អនុប្រធាន'
            ).first()
            if head_officer:
                is_acting = True

        if not head_officer:
            head_officer = CivilServantProfile.objects.filter(
                department=d
            ).exclude(officer_status__in=['DISMISSED', 'RETIRED', 'TRANSFERRED_OUT']).first()
            if head_officer:
                is_acting = True

        if head_officer:
            pos_display = head_officer.current_position_title or f"ប្រធាន{d.name_kh}"
            if is_acting and 'ប្រធាន' not in pos_display:
                if 'អនុប្រធាន' in pos_display:
                    pos_display = f"{pos_display} (ទទួលបន្ទុករួម)"
                else:
                    pos_display = f"{pos_display} (តំណាង {d.name_kh})"
            elif is_acting and 'អនុប្រធាន' in pos_display and 'ទទួលបន្ទុករួម' not in pos_display:
                pos_display = f"{pos_display} (ទទួលបន្ទុករួម)"

            section_office.append({
                'section': 'III. ប្រធានការិយាល័យ',
                'section_code': 'OFFICE',
                'section_num': 3,
                'index': office_idx,
                'officer': head_officer,
                'person_id': head_officer.id,
                'person_type': 'CIVIL_SERVANT',
                'name': head_officer.full_name_kh,
                'gender': 'ប' if head_officer.gender == 'MALE' else 'ស',
                'position': pos_display,
                'department': d,
                'is_acting': is_acting,
            })
            office_idx += 1

    return section_lead, section_canton, section_office


def _get_all_leadership_officer_ids():
    """
    Returns set of all CivilServantProfile IDs who are classified as Leaders / Heads
    (Department Leadership, Canton Heads, Office Heads, or acting/sole officers representing the unit).
    These officers appear in Attendance Model B (Leadership Attendance) and must NOT
    appear in Attendance Model A (Department Staff Attendance).
    """
    sec_lead, sec_canton, sec_office = _get_leadership_roster()
    return set([l['person_id'] for l in sec_lead + sec_canton + sec_office])


def _is_chief_position_to_exclude(title):
    """
    Excludes Canton Chief (នាយខណ្ឌ) and Office Chief (ប្រធានការិយាល័យ) as well as
    Department Directors from the department/canton staff attendance rosters.
    """
    if not title:
        return False
    t = normalize_khmer_role(title)
    # Exclude Department Director & Deputy Director & Executives
    if 'ប្រធានមន្ទីរ' in t or 'អនុប្រធានមន្ទីរ' in t or 'នាយក' in t or 'ប្រធានអង្គភាព' in t:
        return True
    # Exclude Canton Chief (នាយខណ្ឌ, នាយខណ្ឌរដ្ឋបាល - keep នាយរងខណ្ឌ)
    if 'នាយខណ្ឌ' in t and 'នាយរង' not in t and 'អនុ' not in t:
        return True
    # Exclude Office Chief (ប្រធានការិយាល័យ, ប្រធានការិយាល័យ... - keep អនុប្រធានការិយាល័យ)
    if 'ប្រធានការិយាល័យ' in t and 'អនុ' not in t:
        return True
    return False


def _get_department_roster(dept):
    """
    Returns sorted list of civil servants & contract officers for department staff attendance roster.
    Strictly excludes:
    1. Department Leadership (ថ្នាក់ដឹកនាំមន្ទីរ)
    2. Canton Heads & Office Chiefs (នាយខណ្ឌ, ប្រធានការិយាល័យ)
    3. Any Acting Deputy Chief or sole officer representing the unit who is recorded in Leadership Attendance.
    """
    roster = []
    if not dept:
        return roster

    # Exclude all leaders present in the leadership roster
    leadership_officer_ids = _get_all_leadership_officer_ids()

    # 1. Civil Servants (Excluding all leaders and chiefs)
    officers = CivilServantProfile.objects.filter(
        department=dept
    ).exclude(
        officer_status__in=['DISMISSED', 'RETIRED', 'TRANSFERRED_OUT']
    ).exclude(
        id__in=leadership_officer_ids
    ).select_related('department')

    for o in officers:
        pos = o.current_position_title or "មន្ត្រី"
        if _is_chief_position_to_exclude(pos):
            continue  # Safety fallback

        sort_w = get_attendance_position_sort_weight(pos, is_contract=False)
        rank_w = get_rank_step_sort_weight(o.current_rank_and_step)
        name = o.full_name_kh
        gender = 'ប' if o.gender == 'MALE' else 'ស'

        roster.append({
            'person_type': 'CIVIL_SERVANT',
            'person_id': o.id,
            'officer': o,
            'contract_officer': None,
            'name': name,
            'gender': gender,
            'position': pos,
            'officer_id_number': o.officer_id_number or '',
            'sort_weight': sort_w,
            'rank_weight': rank_w,
            'is_contract': False,
        })

    # 2. Contract Officers
    contract_officers = ContractOfficer.objects.filter(
        department=dept,
        is_active=True
    ).exclude(
        contract_status__in=['EXPIRED', 'TERMINATED']
    ).select_related('department')

    for c in contract_officers:
        pos = c.position_title or "មន្ត្រីជាប់កិច្ចសន្យា"
        sort_w = get_attendance_position_sort_weight(pos, is_contract=True)
        name = c.full_name_kh
        gender = 'ប' if c.gender == 'MALE' else 'ស'

        roster.append({
            'person_type': 'CONTRACT_OFFICER',
            'person_id': c.id,
            'officer': None,
            'contract_officer': c,
            'name': name,
            'gender': gender,
            'position': pos,
            'officer_id_number': c.id_number or '',
            'sort_weight': sort_w,
            'rank_weight': (99, 99, 99),
            'is_contract': True,
        })

    roster.sort(key=lambda item: (item['sort_weight'], item['rank_weight'], item['name']))
    return roster


@login_required
def attendance_daily_entry(request):
    """
    Daily attendance entry interface for Officers and Contract Staff.
    Strict Department Isolation: Only user ADMIN can switch departments;
    all other users are locked to their own department/canton.
    """
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    is_admin = _is_admin_user(request.user)

    # Date handling
    date_str = request.GET.get('date', '').strip()
    if date_str:
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            target_date = timezone.now().date()
    else:
        target_date = timezone.now().date()

    # Department Isolation
    if is_admin:
        dept_id = request.GET.get('department', '').strip()
        if dept_id:
            dept = Department.objects.filter(id=dept_id, is_active=True).first()
        else:
            dept = user_dept or Department.objects.filter(is_active=True).order_by('order_index').first()
        departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')
    else:
        dept = user_dept
        departments = Department.objects.filter(id=dept.id) if dept else []

    if not dept:
        messages.warning(request, '⚠️ គណនីរបស់អ្នកពុំទាន់ត្រូវបានភ្ជាប់ទៅកាន់ការិយាល័យ/ខណ្ឌណាមួយឡើយ! សូមទាក់ទង ADMIN។')

    # POST: Save / Update Daily Attendance
    if request.method == 'POST':
        if not is_admin:
            dept = user_dept
            if not dept:
                messages.error(request, '⚠️ លោកអ្នកគ្មានការិយាល័យ/ខណ្ឌសម្រាប់កត់ត្រាវត្តមានឡើយ!')
                return redirect('dashboard')
        else:
            posted_dept_id = request.POST.get('department_id', '').strip()
            if posted_dept_id:
                dept = Department.objects.filter(id=posted_dept_id).first() or dept

        posted_date_str = request.POST.get('target_date', '').strip()
        if posted_date_str:
            try:
                target_date = datetime.strptime(posted_date_str, '%Y-%m-%d').date()
            except Exception:
                pass

        person_keys = request.POST.getlist('person_keys')
        saved_count = 0

        for pkey in person_keys:
            parts = pkey.split('_')
            if len(parts) < 3:
                continue
            ptype = f"{parts[0]}_{parts[1]}"
            pid = int(parts[2])

            status = request.POST.get(f'status_{pkey}', 'PRESENT')
            morning_in = request.POST.get(f'morning_in_{pkey}') == '1'
            morning_out = request.POST.get(f'morning_out_{pkey}') == '1'
            afternoon_in = request.POST.get(f'afternoon_in_{pkey}') == '1'
            afternoon_out = request.POST.get(f'afternoon_out_{pkey}') == '1'
            is_late = request.POST.get(f'is_late_{pkey}') == '1'
            is_early_out = request.POST.get(f'is_early_out_{pkey}') == '1'
            leave_type = request.POST.get(f'leave_type_{pkey}', '').strip()
            reference_doc = request.POST.get(f'reference_doc_{pkey}', '').strip()
            disciplinary_measure = request.POST.get(f'disciplinary_measure_{pkey}', '').strip()
            remarks = request.POST.get(f'remarks_{pkey}', '').strip()
            pos_title = request.POST.get(f'pos_{pkey}', '').strip()

            if status not in ['LEAVE_PERMISSION', 'UNPAID_LEAVE']:
                leave_type = ''

            if ptype == 'CIVIL_SERVANT':
                AttendanceRecord.objects.update_or_create(
                    department=dept,
                    date=target_date,
                    person_type='CIVIL_SERVANT',
                    officer_id=pid,
                    defaults={
                        'contract_officer': None,
                        'position_title': pos_title,
                        'morning_in': morning_in,
                        'morning_out': morning_out,
                        'afternoon_in': afternoon_in,
                        'afternoon_out': afternoon_out,
                        'status': status,
                        'is_late': is_late,
                        'is_early_out': is_early_out,
                        'leave_type': leave_type,
                        'reference_doc': reference_doc,
                        'disciplinary_measure': disciplinary_measure,
                        'remarks': remarks,
                        'recorded_by': request.user,
                    }
                )
                saved_count += 1
            elif ptype == 'CONTRACT_OFFICER':
                AttendanceRecord.objects.update_or_create(
                    department=dept,
                    date=target_date,
                    person_type='CONTRACT_OFFICER',
                    contract_officer_id=pid,
                    defaults={
                        'officer': None,
                        'position_title': pos_title,
                        'morning_in': morning_in,
                        'morning_out': morning_out,
                        'afternoon_in': afternoon_in,
                        'afternoon_out': afternoon_out,
                        'status': status,
                        'is_late': is_late,
                        'is_early_out': is_early_out,
                        'leave_type': leave_type,
                        'reference_doc': reference_doc,
                        'disciplinary_measure': disciplinary_measure,
                        'remarks': remarks,
                        'recorded_by': request.user,
                    }
                )
                saved_count += 1

        dept_name = dept.name_kh if dept else "អង្គភាព"
        messages.success(request, f"✅ បានរក្សាទុកវត្តមានមន្ត្រីចំនួន {saved_count} នាក់ សម្រាប់ «{dept_name}» កាលបរិច្ឆេទ {target_date.strftime('%d/%m/%Y')} ដោយជោគជ័យ!")
        redirect_url = f"{reverse('attendance_daily_entry')}?date={target_date.strftime('%Y-%m-%d')}"
        if is_admin and dept:
            redirect_url += f"&department={dept.id}"
        return redirect(redirect_url)

    # Build Roster & Existing Records
    roster = _get_department_roster(dept) if dept else []
    existing_records = AttendanceRecord.objects.filter(
        department=dept,
        date=target_date
    ) if dept else []

    record_map = {}
    for r in existing_records:
        if r.person_type == 'CIVIL_SERVANT' and r.officer_id:
            record_map[f"CIVIL_SERVANT_{r.officer_id}"] = r
        elif r.person_type == 'CONTRACT_OFFICER' and r.contract_officer_id:
            record_map[f"CONTRACT_OFFICER_{r.contract_officer_id}"] = r

    roster_with_records = []
    civil_count = 0
    contract_count = 0
    female_count = 0
    present_count = 0
    mission_count = 0
    leave_count = 0
    absent_count = 0

    for item in roster:
        pkey = f"{item['person_type']}_{item['person_id']}"
        rec = record_map.get(pkey)

        if item['is_contract']:
            contract_count += 1
        else:
            civil_count += 1

        if item['gender'] == 'ស':
            female_count += 1

        if rec:
            status = rec.status
            morning_in = rec.morning_in
            morning_out = rec.morning_out
            afternoon_in = rec.afternoon_in
            afternoon_out = rec.afternoon_out
            is_late = rec.is_late
            is_early_out = rec.is_early_out
            leave_type = rec.leave_type
            reference_doc = rec.reference_doc
            disciplinary_measure = rec.disciplinary_measure
            remarks = rec.remarks
            is_recorded = True
        else:
            status = 'PRESENT'
            morning_in = True
            morning_out = True
            afternoon_in = True
            afternoon_out = True
            is_late = False
            is_early_out = False
            leave_type = ''
            reference_doc = ''
            disciplinary_measure = ''
            remarks = ''
            is_recorded = False

        if status == 'PRESENT':
            present_count += 1
        elif status == 'MISSION':
            mission_count += 1
        elif status in ['LEAVE_PERMISSION', 'UNPAID_LEAVE']:
            leave_count += 1
        elif status == 'ABSENT_NO_LEAVE':
            absent_count += 1

        roster_with_records.append({
            'person': item,
            'pkey': pkey,
            'status': status,
            'morning_in': morning_in,
            'morning_out': morning_out,
            'afternoon_in': afternoon_in,
            'afternoon_out': afternoon_out,
            'is_late': is_late,
            'is_early_out': is_early_out,
            'leave_type': leave_type,
            'reference_doc': reference_doc,
            'disciplinary_measure': disciplinary_measure,
            'remarks': remarks,
            'is_recorded': is_recorded,
        })

    prev_date = target_date - timedelta(days=1)
    next_date = target_date + timedelta(days=1)
    is_today = (target_date == timezone.now().date())

    context = {
        'dept': dept,
        'departments': departments,
        'target_date': target_date,
        'target_date_str': target_date.strftime('%Y-%m-%d'),
        'prev_date_str': prev_date.strftime('%Y-%m-%d'),
        'next_date_str': next_date.strftime('%Y-%m-%d'),
        'is_today': is_today,
        'roster_with_records': roster_with_records,
        'total_staff': len(roster),
        'civil_count': civil_count,
        'contract_count': contract_count,
        'female_count': female_count,
        'present_count': present_count,
        'mission_count': mission_count,
        'leave_count': leave_count,
        'absent_count': absent_count,
        'is_admin': is_admin,
        'status_choices': AttendanceRecord.STATUS_CHOICES,
        'leave_choices': AttendanceRecord.LEAVE_TYPE_CHOICES,
    }
    return render(request, 'dms/attendance_daily_entry.html', context)


def _build_attendance_monthly_context(request):
    """
    Common helper to build the monthly attendance matrix data with strict department isolation.
    """
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    is_admin = _is_admin_user(request.user)

    now = timezone.now()
    year_str = request.GET.get('year', str(now.year)).strip()
    month_str = request.GET.get('month', str(now.month)).strip()

    try:
        year = int(to_arabic_digits(year_str))
    except Exception:
        year = now.year

    try:
        month = int(to_arabic_digits(month_str))
        if month < 1 or month > 12:
            month = now.month
    except Exception:
        month = now.month

    # Strict Department Isolation: Only ADMIN can view other departments
    if is_admin:
        dept_id = request.GET.get('department', '').strip()
        if dept_id:
            dept = Department.objects.filter(id=dept_id, is_active=True).first()
        else:
            dept = user_dept or Department.objects.filter(is_active=True).order_by('order_index').first()
        departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')
    else:
        dept = user_dept
        departments = Department.objects.filter(id=dept.id) if dept else []

    num_days = calendar.monthrange(year, month)[1]
    days_meta = []
    khmer_weekday_names = {
        0: 'ចន្ទ', 1: 'អង្គារ', 2: 'ពុធ', 3: 'ព្រហ', 4: 'សុក្រ', 5: 'សៅរ៍', 6: 'អាទិត្យ'
    }

    for d in range(1, num_days + 1):
        cur_date = date(year, month, d)
        wday = cur_date.weekday()
        days_meta.append({
            'day': d,
            'date': cur_date,
            'date_str': cur_date.strftime('%Y-%m-%d'),
            'weekday_num': wday,
            'weekday_name': khmer_weekday_names.get(wday, ''),
            'is_weekend': (wday in [5, 6]),
            'is_today': (cur_date == timezone.now().date()),
        })

    roster = _get_department_roster(dept) if dept else []

    records = AttendanceRecord.objects.filter(
        department=dept,
        date__year=year,
        date__month=month
    ) if dept else []

    rec_by_person = {}
    for r in records:
        if r.person_type == 'CIVIL_SERVANT' and r.officer_id:
            pkey = f"CIVIL_SERVANT_{r.officer_id}"
        elif r.person_type == 'CONTRACT_OFFICER' and r.contract_officer_id:
            pkey = f"CONTRACT_OFFICER_{r.contract_officer_id}"
        else:
            continue
        rec_by_person.setdefault(pkey, []).append(r)

    rows_civil = []
    rows_contract = []
    civil_idx = 1
    contract_idx = 1

    for person in roster:
        pkey = f"{person['person_type']}_{person['person_id']}"
        recs = rec_by_person.get(pkey, [])

        present = sum(1 for r in recs if r.status == 'PRESENT')
        mission = sum(1 for r in recs if r.status == 'MISSION')
        leave = sum(1 for r in recs if r.status in ['LEAVE_PERMISSION', 'UNPAID_LEAVE'])
        absent = sum(1 for r in recs if r.status == 'ABSENT_NO_LEAVE')
        late_or_early = sum(1 for r in recs if r.is_late or r.is_early_out)

        leave_types = list(set([r.get_leave_type_display() for r in recs if r.leave_type]))
        leave_type_str = ', '.join(leave_types) if leave_types else ''

        remarks_list = list(set([r.remarks for r in recs if r.remarks]))
        remarks_str = ', '.join(remarks_list) if remarks_list else ''

        disciplinary_list = list(set([r.disciplinary_measure for r in recs if r.disciplinary_measure]))
        disciplinary_str = ', '.join(disciplinary_list) if disciplinary_list else ''

        if person['person_type'] == 'CIVIL_SERVANT':
            person_copy = dict(person)
            person_copy['index'] = civil_idx
            rows_civil.append({
                'person': person_copy,
                'pkey': pkey,
                'present': present,
                'mission': mission,
                'leave': leave,
                'absent': absent,
                'late_or_early': late_or_early,
                'disciplinary': disciplinary_str,
                'leave_type': leave_type_str,
                'remarks': remarks_str,
            })
            civil_idx += 1
        else:
            person_copy = dict(person)
            person_copy['index'] = contract_idx
            rows_contract.append({
                'person': person_copy,
                'pkey': pkey,
                'present': present,
                'mission': mission,
                'leave': leave,
                'absent': absent,
                'late_or_early': late_or_early,
                'disciplinary': disciplinary_str,
                'leave_type': leave_type_str,
                'remarks': remarks_str,
            })
            contract_idx += 1

    all_rows = rows_civil + rows_contract
    total_staff = len(all_rows)
    tot_present = sum(r['present'] for r in all_rows)
    tot_mission = sum(r['mission'] for r in all_rows)
    tot_leave = sum(r['leave'] for r in all_rows)
    tot_absent = sum(r['absent'] for r in all_rows)
    tot_late_early = sum(r['late_or_early'] for r in all_rows)

    khmer_month_names = {
        1: 'មករា', 2: 'កុម្ភៈ', 3: 'មីនា', 4: 'មេសា',
        5: 'ឧសភា', 6: 'មិថុនា', 7: 'កក្កដា', 8: 'សីហា',
        9: 'កញ្ញា', 10: 'តុលា', 11: 'វិច្ឆិកា', 12: 'ធ្នូ'
    }

    year_options = [2025, 2026, 2027, 2028, 2029, 2030]
    month_options = [(i, khmer_month_names[i]) for i in range(1, 13)]

    return {
        'dept': dept,
        'departments': departments,
        'year': year,
        'month': month,
        'month_name_kh': khmer_month_names.get(month, ''),
        'num_days': num_days,
        'days_meta': days_meta,
        'rows_civil': rows_civil,
        'rows_contract': rows_contract,
        'all_rows': all_rows,
        'total_staff': total_staff,
        'tot_present': tot_present,
        'tot_mission': tot_mission,
        'tot_leave': tot_leave,
        'tot_absent': tot_absent,
        'tot_late_early': tot_late_early,
        'year_options': year_options,
        'month_options': month_options,
        'is_admin': is_admin,
    }


def _build_attendance_weekly_context(request):
    """
    Builds weekly attendance roster matrix data for an office/canton.
    Calculates Monday to Friday (standard 5 working days) for the selected week.
    Strict Department Isolation: Non-admin is restricted to user_dept only.
    """
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    is_admin = _is_admin_user(request.user)

    now = timezone.now()
    date_str = request.GET.get('date', '').strip()
    if date_str:
        try:
            base_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            base_date = now.date()
    else:
        base_date = now.date()

    # Calculate Monday of this week
    monday = base_date - timedelta(days=base_date.weekday())
    friday = monday + timedelta(days=4)
    prev_week_monday = monday - timedelta(days=7)
    next_week_monday = monday + timedelta(days=7)
    is_current_week = (monday <= now.date() <= monday + timedelta(days=6))

    # Department Isolation: Only ADMIN can view other departments
    if is_admin:
        dept_id = request.GET.get('department', '').strip()
        if dept_id:
            dept = Department.objects.filter(id=dept_id, is_active=True).first()
        else:
            dept = user_dept or Department.objects.filter(is_active=True).order_by('order_index').first()
        departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')
    else:
        dept = user_dept
        departments = Department.objects.filter(id=dept.id) if dept else []

    roster = _get_department_roster(dept) if dept else []

    records = AttendanceRecord.objects.filter(
        department=dept,
        date__gte=monday,
        date__lte=friday
    ) if dept else []

    rec_by_person = {}
    for r in records:
        if r.person_type == 'CIVIL_SERVANT' and r.officer_id:
            pkey = f"CIVIL_SERVANT_{r.officer_id}"
        elif r.person_type == 'CONTRACT_OFFICER' and r.contract_officer_id:
            pkey = f"CONTRACT_OFFICER_{r.contract_officer_id}"
        else:
            continue
        rec_by_person.setdefault(pkey, []).append(r)

    rows_civil = []
    rows_contract = []
    civil_idx = 1
    contract_idx = 1

    for person in roster:
        pkey = f"{person['person_type']}_{person['person_id']}"
        recs = rec_by_person.get(pkey, [])

        present = sum(1 for r in recs if r.status == 'PRESENT')
        mission = sum(1 for r in recs if r.status == 'MISSION')
        leave = sum(1 for r in recs if r.status in ['LEAVE_PERMISSION', 'UNPAID_LEAVE'])
        absent = sum(1 for r in recs if r.status == 'ABSENT_NO_LEAVE')
        late_or_early = sum(1 for r in recs if r.is_late or r.is_early_out)

        leave_types = list(set([r.get_leave_type_display() for r in recs if r.leave_type]))
        leave_type_str = ', '.join(leave_types) if leave_types else ''

        remarks_list = list(set([r.remarks for r in recs if r.remarks]))
        remarks_str = ', '.join(remarks_list) if remarks_list else ''

        disciplinary_list = list(set([r.disciplinary_measure for r in recs if r.disciplinary_measure]))
        disciplinary_str = ', '.join(disciplinary_list) if disciplinary_list else ''

        if person['person_type'] == 'CIVIL_SERVANT':
            person_copy = dict(person)
            person_copy['index'] = civil_idx
            rows_civil.append({
                'person': person_copy,
                'pkey': pkey,
                'present': present,
                'mission': mission,
                'leave': leave,
                'absent': absent,
                'late_or_early': late_or_early,
                'disciplinary': disciplinary_str,
                'leave_type': leave_type_str,
                'remarks': remarks_str,
            })
            civil_idx += 1
        else:
            person_copy = dict(person)
            person_copy['index'] = contract_idx
            rows_contract.append({
                'person': person_copy,
                'pkey': pkey,
                'present': present,
                'mission': mission,
                'leave': leave,
                'absent': absent,
                'late_or_early': late_or_early,
                'disciplinary': disciplinary_str,
                'leave_type': leave_type_str,
                'remarks': remarks_str,
            })
            contract_idx += 1

    all_rows = rows_civil + rows_contract
    total_staff = len(all_rows)
    tot_present = sum(r['present'] for r in all_rows)
    tot_mission = sum(r['mission'] for r in all_rows)
    tot_leave = sum(r['leave'] for r in all_rows)
    tot_absent = sum(r['absent'] for r in all_rows)
    tot_late_early = sum(r['late_or_early'] for r in all_rows)

    # Week number
    iso_year, iso_week, _ = monday.isocalendar()

    return {
        'dept': dept,
        'departments': departments,
        'base_date': base_date,
        'base_date_str': base_date.strftime('%Y-%m-%d'),
        'monday': monday,
        'friday': friday,
        'monday_str': monday.strftime('%Y-%m-%d'),
        'friday_str': friday.strftime('%Y-%m-%d'),
        'prev_week_monday_str': prev_week_monday.strftime('%Y-%m-%d'),
        'next_week_monday_str': next_week_monday.strftime('%Y-%m-%d'),
        'is_current_week': is_current_week,
        'iso_year': iso_year,
        'iso_week': iso_week,
        'rows_civil': rows_civil,
        'rows_contract': rows_contract,
        'all_rows': all_rows,
        'total_staff': total_staff,
        'tot_present': tot_present,
        'tot_mission': tot_mission,
        'tot_leave': tot_leave,
        'tot_absent': tot_absent,
        'tot_late_early': tot_late_early,
        'is_admin': is_admin,
    }


@login_required
def attendance_weekly_sheet(request):
    """
    Weekly Attendance Report (របាយការណ៍វត្តមានប្រចាំសប្តាហ៍) for Offices and Cantons.
    Displays 5 working days (Monday-Friday) with status, checkins, leaves, and totals.
    """
    context = _build_attendance_weekly_context(request)
    return render(request, 'dms/attendance_weekly_sheet.html', context)


@login_required
def attendance_weekly_print(request):
    """
    Official A4 Landscape / Portrait print format for Weekly Attendance Report.
    """
    context = _build_attendance_weekly_context(request)
    return render(request, 'dms/attendance_weekly_print.html', context)


@login_required
def attendance_weekly_export_excel(request):
    """
    Exports the Weekly Attendance Report for the selected office/canton to Excel (.xlsx).
    """
    context = _build_attendance_weekly_context(request)
    dept = context.get('dept')
    monday = context.get('monday')
    friday = context.get('friday')
    iso_week = context.get('iso_week')
    iso_year = context.get('iso_year')
    days_meta = context.get('days_meta', [])
    matrix_rows = context.get('matrix_rows', [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"វត្តមាន_សប្តាហ៍_{iso_week}"

    # Styles
    font_title = Font(name="Khmer OS Muol Light", size=13, bold=True, color="1E3A8A")
    font_sub = Font(name="Khmer OS Battambang", size=10, bold=True)
    font_header = Font(name="Khmer OS Battambang", size=9, bold=True, color="FFFFFF")
    font_cell = Font(name="Khmer OS Battambang", size=9)
    font_cell_bold = Font(name="Khmer OS Battambang", size=9, bold=True)
    font_code = Font(name="Arial", size=9, bold=True)

    fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_summary = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    dept_title = dept.name_kh if dept else "គ្រប់អង្គភាព"
    last_col_letter = openpyxl.utils.get_column_letter(4 + 5 + 7)

    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = f"របាយការណ៍វត្តមានមន្ត្រីរាជការ និងមន្ត្រីជាប់កិច្ចសន្យាប្រចាំសប្តាហ៍ (សប្តាហ៍ទី {iso_week} ឆ្នាំ {iso_year})"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center

    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f"{dept_title} - ចាប់ពីថ្ងៃទី {monday.strftime('%d/%m/%Y')} ដល់ {friday.strftime('%d/%m/%Y')}"
    ws['A2'].font = font_sub
    ws['A2'].alignment = align_center

    # Header Rows (Row 4 & 5)
    ws.merge_cells('A4:A5')
    ws['A4'] = "ល.រ"
    ws.merge_cells('B4:B5')
    ws['B4'] = "គោត្តនាម និងនាម"
    ws.merge_cells('C4:C5')
    ws['C4'] = "ភេទ"
    ws.merge_cells('D4:D5')
    ws['D4'] = "មុខតំណែង"

    for col in ['A4', 'B4', 'C4', 'D4', 'A5', 'B5', 'C5', 'D5']:
        ws[col].font = font_header
        ws[col].fill = fill_header
        ws[col].alignment = align_center
        ws[col].border = thin_border

    # Days Columns (Row 4 = Day Number & Date, Row 5 = Weekday)
    for idx, d_meta in enumerate(days_meta, 1):
        col_num = 4 + idx
        col_let = openpyxl.utils.get_column_letter(col_num)

        c4 = ws[f'{col_let}4']
        c4.value = d_meta['date'].strftime('%d/%m')
        c4.font = font_header
        c4.fill = fill_header
        c4.alignment = align_center
        c4.border = thin_border

        c5 = ws[f'{col_let}5']
        c5.value = d_meta['weekday_short']
        c5.font = font_header
        c5.fill = fill_header
        c5.alignment = align_center
        c5.border = thin_border

    # Summary Headers
    summary_headers = [
        ("វត្តមាន (✓)", 10),
        ("បេសកកម្ម (M)", 11),
        ("មានច្បាប់ (P)", 12),
        ("អត់ច្បាប់ (A)", 13),
        ("គ្មានបៀវត្ស (U)", 14),
        ("មកយឺត", 15),
        ("ចេញមុន", 16),
    ]

    for title, col_idx in summary_headers:
        col_let = openpyxl.utils.get_column_letter(col_idx)
        ws.merge_cells(f'{col_let}4:{col_let}5')
        c = ws[f'{col_let}4']
        c.value = title
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = thin_border
        ws[f'{col_let}5'].border = thin_border

    # Data Rows
    row_num = 6
    for item in matrix_rows:
        ws.row_dimensions[row_num].height = 22
        is_even = (item['index'] % 2 == 0)
        p = item['person']

        ws[f'A{row_num}'] = item['index']
        ws[f'A{row_num}'].alignment = align_center
        ws[f'A{row_num}'].font = font_cell
        ws[f'A{row_num}'].border = thin_border

        ws[f'B{row_num}'] = p['name']
        ws[f'B{row_num}'].alignment = align_left
        ws[f'B{row_num}'].font = font_cell_bold
        ws[f'B{row_num}'].border = thin_border

        ws[f'C{row_num}'] = p['gender']
        ws[f'C{row_num}'].alignment = align_center
        ws[f'C{row_num}'].font = font_cell
        ws[f'C{row_num}'].border = thin_border

        ws[f'D{row_num}'] = p['position']
        ws[f'D{row_num}'].alignment = align_left
        ws[f'D{row_num}'].font = font_cell
        ws[f'D{row_num}'].border = thin_border

        for d_idx, d_data in enumerate(item['days'], 1):
            col_let = openpyxl.utils.get_column_letter(4 + d_idx)
            c = ws[f'{col_let}{row_num}']
            c.value = d_data['status_code'] if d_data['has_rec'] else ""
            c.alignment = align_center
            c.font = font_code
            c.border = thin_border

            if is_even:
                c.fill = fill_zebra

        stats = item['stats']
        stats_vals = [
            stats['present'], stats['mission'], stats['leave_permission'],
            stats['absent_no_leave'], stats['unpaid_leave'], stats['late'], stats['early_out']
        ]

        for s_idx, s_val in enumerate(stats_vals, 1):
            col_let = openpyxl.utils.get_column_letter(9 + s_idx)
            c = ws[f'{col_let}{row_num}']
            c.value = s_val
            c.alignment = align_center
            c.font = font_cell_bold
            c.border = thin_border
            c.fill = fill_summary

        row_num += 1

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 24
    for i in range(1, 6):
        col_let = openpyxl.utils.get_column_letter(4 + i)
        ws.column_dimensions[col_let].width = 8
    for i in range(1, 8):
        col_let = openpyxl.utils.get_column_letter(9 + i)
        ws.column_dimensions[col_let].width = 10

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_weekly_{dept.code if dept else "all"}_{iso_year}_W{iso_week:02d}.xlsx"'
    wb.save(response)
    return response


@login_required
def attendance_monthly_sheet(request):
    """
    Monthly Attendance Matrix Sheet (តារាងវត្តមានមន្ត្រីប្រចាំខែ) matching AttendendA.xlsx layout.
    Displays days 1..31 with full status markers (✓, M, P, A, U), weekend highlighting,
    and automatic total calculations.
    """
    context = _build_attendance_monthly_context(request)
    return render(request, 'dms/attendance_monthly_sheet.html', context)


@login_required
def attendance_monthly_print(request):
    """
    Official A4 Landscape print format matching Cambodian Government / Ministry of Agriculture standard.
    """
    context = _build_attendance_monthly_context(request)
    return render(request, 'dms/attendance_monthly_print.html', context)


@login_required
def attendance_monthly_export_excel(request):
    """
    Exports the Department Attendance Roster for the selected month to Excel (.xlsx)
    matching the layout of AttendendA.xlsx.
    """
    context = _build_attendance_monthly_context(request)
    dept = context.get('dept')
    year = context.get('year')
    month = context.get('month')
    month_name_kh = context.get('month_name_kh')
    days_meta = context.get('days_meta', [])
    matrix_rows = context.get('matrix_rows', [])
    num_days = context.get('num_days', 31)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"វត្តមាន_{month_name_kh}_{year}"

    # Styles
    font_title = Font(name="Khmer OS Muol Light", size=13, bold=True, color="1E3A8A")
    font_sub = Font(name="Khmer OS Battambang", size=10, bold=True)
    font_header = Font(name="Khmer OS Battambang", size=9, bold=True, color="FFFFFF")
    font_cell = Font(name="Khmer OS Battambang", size=9)
    font_cell_bold = Font(name="Khmer OS Battambang", size=9, bold=True)
    font_code = Font(name="Arial", size=9, bold=True)

    fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fill_weekend = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_summary = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Header Row 1: Title
    dept_title = dept.name_kh if dept else "គ្រប់អង្គភាព"
    last_col_letter = openpyxl.utils.get_column_letter(4 + num_days + 7)

    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = f"បញ្ជីវត្តមានមន្ត្រីរាជការ និងមន្ត្រីជាប់កិច្ចសន្យា {dept_title}"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center

    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f"ប្រចាំខែ {month_name_kh} ឆ្នាំ {year}"
    ws['A2'].font = font_sub
    ws['A2'].alignment = align_center

    # Table Header Row (Row 4 & 5)
    ws.merge_cells('A4:A5')
    ws['A4'] = "ល.រ"
    ws.merge_cells('B4:B5')
    ws['B4'] = "គោត្តនាម និងនាម"
    ws.merge_cells('C4:C5')
    ws['C4'] = "ភេទ"
    ws.merge_cells('D4:D5')
    ws['D4'] = "មុខតំណែង"

    for col in ['A4', 'B4', 'C4', 'D4', 'A5', 'B5', 'C5', 'D5']:
        ws[col].font = font_header
        ws[col].fill = fill_header
        ws[col].alignment = align_center
        ws[col].border = thin_border

    # Days Columns (Row 4 = Day Number, Row 5 = Weekday)
    for idx, d_meta in enumerate(days_meta, 1):
        col_num = 4 + idx
        col_let = openpyxl.utils.get_column_letter(col_num)

        c4 = ws[f'{col_let}4']
        c4.value = d_meta['day']
        c4.font = font_header
        c4.fill = fill_header
        c4.alignment = align_center
        c4.border = thin_border

        c5 = ws[f'{col_let}5']
        c5.value = d_meta['weekday_name'][:3]
        c5.font = font_header
        c5.fill = fill_header
        c5.alignment = align_center
        c5.border = thin_border

    # Summary Headers
    summary_headers = [
        ("វត្តមាន (✓)", 5 + num_days),
        ("បេសកកម្ម (M)", 6 + num_days),
        ("មានច្បាប់ (P)", 7 + num_days),
        ("អត់ច្បាប់ (A)", 8 + num_days),
        ("គ្មានបៀវត្ស (U)", 9 + num_days),
        ("មកយឺត", 10 + num_days),
        ("ចេញមុន", 11 + num_days),
    ]

    for title, col_idx in summary_headers:
        col_let = openpyxl.utils.get_column_letter(col_idx)
        ws.merge_cells(f'{col_let}4:{col_let}5')
        c = ws[f'{col_let}4']
        c.value = title
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = thin_border
        ws[f'{col_let}5'].border = thin_border

    # Data Rows
    row_num = 6
    for item in matrix_rows:
        ws.row_dimensions[row_num].height = 22
        is_even = (item['index'] % 2 == 0)
        p = item['person']

        ws[f'A{row_num}'] = item['index']
        ws[f'A{row_num}'].alignment = align_center
        ws[f'A{row_num}'].font = font_cell
        ws[f'A{row_num}'].border = thin_border

        ws[f'B{row_num}'] = p['name']
        ws[f'B{row_num}'].alignment = align_left
        ws[f'B{row_num}'].font = font_cell_bold
        ws[f'B{row_num}'].border = thin_border

        ws[f'C{row_num}'] = p['gender']
        ws[f'C{row_num}'].alignment = align_center
        ws[f'C{row_num}'].font = font_cell
        ws[f'C{row_num}'].border = thin_border

        ws[f'D{row_num}'] = p['position']
        ws[f'D{row_num}'].alignment = align_left
        ws[f'D{row_num}'].font = font_cell
        ws[f'D{row_num}'].border = thin_border

        for d_idx, d_data in enumerate(item['days'], 1):
            col_let = openpyxl.utils.get_column_letter(4 + d_idx)
            c = ws[f'{col_let}{row_num}']
            c.value = d_data['status_code'] if d_data['has_rec'] else ""
            c.alignment = align_center
            c.font = font_code
            c.border = thin_border

            if d_data['is_weekend']:
                c.fill = fill_weekend
            elif is_even:
                c.fill = fill_zebra

        stats = item['stats']
        stats_vals = [
            stats['present'], stats['mission'], stats['leave_permission'],
            stats['absent_no_leave'], stats['unpaid_leave'], stats['late'], stats['early_out']
        ]

        for s_idx, s_val in enumerate(stats_vals, 1):
            col_let = openpyxl.utils.get_column_letter(4 + num_days + s_idx)
            c = ws[f'{col_let}{row_num}']
            c.value = s_val
            c.alignment = align_center
            c.font = font_cell_bold
            c.border = thin_border
            c.fill = fill_summary

        row_num += 1

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 24
    for i in range(1, num_days + 1):
        col_let = openpyxl.utils.get_column_letter(4 + i)
        ws.column_dimensions[col_let].width = 4.5
    for i in range(1, 8):
        col_let = openpyxl.utils.get_column_letter(4 + num_days + i)
        ws.column_dimensions[col_let].width = 10

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_{dept.code if dept else "all"}_{year}_{month:02d}.xlsx"'
    wb.save(response)
    return response


@login_required
def attendance_summary_report(request):
    """
    Overview Attendance Statistics across all departments and cantons for ADMIN ONLY.
    """
    is_admin = _is_admin_user(request.user)
    if not is_admin:
        messages.warning(request, '⚠️ ផ្ទាំងរបាយការណ៍សង្ខេបវត្តមានទូទាំងអង្គភាព ត្រូវបានអនុញ្ញាតសម្រាប់តែ User: ADMIN ប៉ុណ្ណោះ!')
        return redirect('attendance_monthly_sheet')

    now = timezone.now()
    year_str = request.GET.get('year', str(now.year)).strip()
    month_str = request.GET.get('month', str(now.month)).strip()

    try:
        year = int(to_arabic_digits(year_str))
    except Exception:
        year = now.year

    try:
        month = int(to_arabic_digits(month_str))
        if month < 1 or month > 12:
            month = now.month
    except Exception:
        month = now.month

    khmer_month_names = {
        1: 'មករា', 2: 'កុម្ភៈ', 3: 'មីនា', 4: 'មេសា',
        5: 'ឧសភា', 6: 'មិថុនា', 7: 'កក្កដា', 8: 'សីហា',
        9: 'កញ្ញា', 10: 'តុលា', 11: 'វិច្ឆិកា', 12: 'ធ្នូ'
    }

    departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')
    dept_summaries = []

    grand_total_staff = 0
    grand_present = 0
    grand_mission = 0
    grand_leave = 0
    grand_absent = 0

    for d in departments:
        roster = _get_department_roster(d)
        total_staff = len(roster)
        if total_staff == 0:
            continue

        civil_count = sum(1 for p in roster if not p['is_contract'])
        contract_count = sum(1 for p in roster if p['is_contract'])

        recs = AttendanceRecord.objects.filter(
            department=d,
            date__year=year,
            date__month=month
        )

        total_records = recs.count()
        present = recs.filter(status='PRESENT').count()
        mission = recs.filter(status='MISSION').count()
        leave = recs.filter(status__in=['LEAVE_PERMISSION', 'UNPAID_LEAVE']).count()
        absent = recs.filter(status='ABSENT_NO_LEAVE').count()

        rate = round((present / total_records * 100), 1) if total_records > 0 else 0

        grand_total_staff += total_staff
        grand_present += present
        grand_mission += mission
        grand_leave += leave
        grand_absent += absent

        dept_summaries.append({
            'department': d,
            'total_staff': total_staff,
            'civil_count': civil_count,
            'contract_count': contract_count,
            'total_records': total_records,
            'present': present,
            'mission': mission,
            'leave': leave,
            'absent': absent,
            'rate': rate,
        })

    year_options = [2025, 2026, 2027, 2028, 2029, 2030]
    month_options = [(i, khmer_month_names[i]) for i in range(1, 13)]

    context = {
        'year': year,
        'month': month,
        'month_name_kh': khmer_month_names.get(month, ''),
        'dept_summaries': dept_summaries,
        'grand_total_staff': grand_total_staff,
        'grand_present': grand_present,
        'grand_mission': grand_mission,
        'grand_leave': grand_leave,
        'grand_absent': grand_absent,
        'year_options': year_options,
        'month_options': month_options,
        'is_admin': is_admin,
    }
    return render(request, 'dms/attendance_reports.html', context)


@login_required
def api_attendance_quick_toggle(request):
    """
    AJAX endpoint for toggling or updating single attendance fields instantly.
    Strict Department Isolation: Non-admin can only modify records for their own department.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)

    profile = getattr(request.user, 'profile', None)
    is_admin = _is_admin_user(request.user)

    dept_id = request.POST.get('department_id')
    date_str = request.POST.get('date')
    person_type = request.POST.get('person_type')
    person_id = request.POST.get('person_id')
    field = request.POST.get('field')
    value = request.POST.get('value')

    dept = Department.objects.filter(id=dept_id).first()
    if not dept:
        return JsonResponse({'success': False, 'error': 'Department not found.'})

    if not is_admin and profile and profile.department != dept:
        return JsonResponse({'success': False, 'error': 'គ្មានសិទ្ធិកែប្រែវត្តមាននៃការិយាល័យផ្សេងឡើយ!'}, status=403)

    try:
        rec_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid date.'})

    filter_kwargs = {
        'department': dept,
        'date': rec_date,
        'person_type': person_type,
    }
    if person_type == 'CIVIL_SERVANT':
        filter_kwargs['officer_id'] = person_id
    else:
        filter_kwargs['contract_officer_id'] = person_id

    rec, created = AttendanceRecord.objects.get_or_create(
        defaults={'recorded_by': request.user},
        **filter_kwargs
    )

    if field == 'status':
        rec.status = value
    elif field in ['morning_in', 'morning_out', 'afternoon_in', 'afternoon_out', 'is_late', 'is_early_out']:
        setattr(rec, field, value in ['1', 'true', True])
    elif field in ['leave_type', 'reference_doc', 'remarks', 'disciplinary_measure']:
        setattr(rec, field, value)

    rec.recorded_by = request.user
    rec.save()

    return JsonResponse({
        'success': True,
        'status': rec.status,
        'status_code': rec.status_code,
        'badge_class': rec.status_badge_class,
    })


# ==============================================================================
# 📑 ANNUAL WORK RECORD / LEAVE BULLETIN MODULE (ព្រឹត្តិបត្រការងារប្រចាំឆ្នាំរបស់មន្ត្រីរាជការស៊ីវិល)
# ឧបសម្ព័ន្ធទី៤នៃអនុក្រឹត្យលេខ ៥៨ អនក្រ.បក ចុះថ្ងៃទី ០១ ខែ កញ្ញា ឆ្នាំ២០១៦
# ==============================================================================

def _build_annual_bulletin_data(officer, year):
    """
    Builds leave bulletin data according to Sub-Decree 58 Annex 4 (ឧបសម្ព័ន្ធទី៤ នៃអនុក្រឹត្យលេខ ៥៨)
    """
    records = AttendanceRecord.objects.filter(
        officer=officer,
        date__year=year
    )

    # 1. Annual Leave (១. ច្បាប់ឈប់ប្រចាំឆ្នាំ - ១៥ថ្ងៃនៃថ្ងៃធ្វើការ/១ឆ្នាំ)
    annual_used = records.filter(leave_type='ANNUAL').count()
    annual_prev_unused = 0
    prev_records = AttendanceRecord.objects.filter(officer=officer, date__year=year - 1)
    if prev_records.exists():
        prev_annual_used = prev_records.filter(leave_type='ANNUAL').count()
        annual_prev_unused = max(0, 15 - prev_annual_used)

    annual_transfer_direct = 0
    annual_transfer_other = 0
    annual_additional = 0
    annual_total_entitled = 15 + annual_prev_unused + annual_additional
    annual_remaining = max(0, annual_total_entitled - annual_used)

    # 2. Short-term Leave (២. ច្បាប់ឈប់រយៈពេលខ្លី - ១៥ថ្ងៃនៃថ្ងៃធ្វើការ/១ឆ្នាំ)
    short_used = records.filter(leave_type='SHORT').count()
    short_remaining = max(0, 15 - short_used)

    # 3. Maternity Leave (៣. ច្បាប់ឈប់សម្រាកលំហែមាតុភាព - ៣ខែ)
    maternity_used = records.filter(leave_type='MATERNITY').count()
    maternity_display = f"{maternity_used} ថ្ងៃ" if maternity_used > 0 else "០"

    # 4. Sick Leave (៤. ច្បាប់ឈប់សម្រាកព្យាបាលជំងឺ - ១២ខែ សម្រាប់មួយជីវិតជាមន្ត្រី)
    lifetime_sick_records = AttendanceRecord.objects.filter(officer=officer, leave_type='SICK').count()
    sick_used_months = round(lifetime_sick_records / 30, 1) if lifetime_sick_records > 0 else 0
    sick_remaining_months = max(0, round(12 - sick_used_months, 1))

    # 5. Personal Business Leave (៥. ច្បាប់ឈប់សម្រាកដោយមានកិច្ចផ្ទាល់ខ្លួន - ៣ខែ សម្រាប់មួយជីវិតជាមន្ត្រី)
    lifetime_personal_records = AttendanceRecord.objects.filter(officer=officer, leave_type='PERSONAL').count()
    personal_used_months = round(lifetime_personal_records / 30, 1) if lifetime_personal_records > 0 else 0
    personal_remaining_months = max(0, round(3 - personal_used_months, 1))

    # 6. Disciplinary & Sanctions (ទណ្ឌកម្មវិន័យ និងវិធានការ)
    disc_list = list(set([r.disciplinary_measure for r in records if r.disciplinary_measure]))
    remarks_list = list(set([r.remarks for r in records if r.remarks and 'វិន័យ' in r.remarks]))
    all_disc = disc_list + remarks_list
    disciplinary_str = ', '.join(all_disc) if all_disc else '-'

    # Absences & other summary stats
    authorized_leave_days = records.filter(status='LEAVE_PERMISSION').count()
    unauthorized_absent_days = records.filter(status='ABSENT_NO_LEAVE').count()
    unpaid_leave_days = records.filter(status='UNPAID_LEAVE').count()
    other_leave_days = records.filter(leave_type='OTHER').count()

    g_kh = 'ប្រុស' if officer.gender == 'MALE' or str(officer.gender).upper() in ['M', 'ប្រុស', 'ប'] else 'ស្រី'

    return {
        'officer': officer,
        'year': year,
        'year_kh': _to_khmer_digits(str(year)),
        'gender_kh': g_kh,
        'department_name': officer.department.name_kh if officer.department else 'មន្ទីរកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទខេត្តប៉ៃលិន',

        # Category 1: Annual Leave
        'annual_prev_unused': _to_khmer_digits(str(annual_prev_unused)) if annual_prev_unused else '០',
        'annual_transfer_direct': _to_khmer_digits(str(annual_transfer_direct)) if annual_transfer_direct else '០',
        'annual_transfer_other': _to_khmer_digits(str(annual_transfer_other)) if annual_transfer_other else '០',
        'annual_additional': _to_khmer_digits(str(annual_additional)) if annual_additional else '០',
        'annual_used': _to_khmer_digits(str(annual_used)),
        'annual_remaining': _to_khmer_digits(str(annual_remaining)),

        # Category 2: Short-term Leave
        'short_used': _to_khmer_digits(str(short_used)),
        'short_remaining': _to_khmer_digits(str(short_remaining)),

        # Category 3: Maternity Leave
        'maternity_display': _to_khmer_digits(maternity_display),

        # Category 4: Sick Leave
        'sick_used_months': _to_khmer_digits(str(sick_used_months)) if sick_used_months > 0 else '០',
        'sick_remaining_months': _to_khmer_digits(str(sick_remaining_months)),

        # Category 5: Personal Leave
        'personal_used_months': _to_khmer_digits(str(personal_used_months)) if personal_used_months > 0 else '០',
        'personal_remaining_months': _to_khmer_digits(str(personal_remaining_months)),

        # Category 6: Discipline
        'disciplinary_str': disciplinary_str,

        # Absence stats
        'authorized_leave_days': _to_khmer_digits(str(authorized_leave_days)),
        'unauthorized_absent_days': _to_khmer_digits(str(unauthorized_absent_days)),
        'unpaid_leave_days': _to_khmer_digits(str(unpaid_leave_days)),
        'other_leave_days': _to_khmer_digits(str(other_leave_days)) if other_leave_days else '០',

        'lunar_year_text': _get_khmer_lunar_year_info(year),
    }


def _build_annual_bulletin_workbook(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ព្រឹត្តិបត្រការងារ"

    # Page Setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3

    # Column widths (14 columns A to N)
    col_widths = {
        'A': 8.5, 'B': 8.0, 'C': 9.5, 'D': 8.5, 'E': 8.0, 'F': 8.0,
        'G': 8.0, 'H': 8.0, 'I': 10.0, 'J': 8.5, 'K': 8.5, 'L': 8.5,
        'M': 8.5, 'N': 16.0
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    font_org = Font(name='Khmer OS Muol Light', size=10, bold=False)
    font_annex = Font(name='Khmer OS Battambang', size=9, bold=False)
    font_title = Font(name='Khmer OS Muol Light', size=13, bold=False)
    font_sub_info = Font(name='Khmer OS Battambang', size=10, bold=True)
    font_th_large = Font(name='Khmer OS Muol Light', size=9.5, bold=False)
    font_th_bold = Font(name='Khmer OS Battambang', size=8.5, bold=True)
    font_th_normal = Font(name='Khmer OS Battambang', size=8, bold=False)
    font_data = Font(name='Khmer OS Battambang', size=10, bold=False)
    font_notes = Font(name='Khmer OS Battambang', size=9, bold=False)

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_data_center = Alignment(horizontal='center', vertical='center')

    # Top Header
    # Row 2
    ws['A2'] = "ក្រសួង ស្ថាប័ន៖ ក្រសួងកសិកម្ម រុក្ខាប្រមាញ់ និងនេសាទ"
    ws['A2'].font = font_org
    ws.merge_cells('D2:J2')
    ws['D2'] = "ឧបសម្ព័ន្ធទី៤នៃអនុក្រឹត្យលេខ ៥៨ អនក្រ.បក ចុះថ្ងៃទី ០១ ខែ កញ្ញា ឆ្នាំ២០១៦"
    ws['D2'].font = font_annex
    ws['D2'].alignment = align_center
    ws.merge_cells('L2:N2')
    ws['L2'] = "ព្រះរាជាណាចក្រកម្ពុជា"
    ws['L2'].font = font_org
    ws['L2'].alignment = align_center

    # Row 3
    ws['A3'] = f"អង្គភាព៖ {data['department_name']}"
    ws['A3'].font = font_org
    ws.merge_cells('L3:N3')
    ws['L3'] = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
    ws['L3'].font = font_org
    ws['L3'].alignment = align_center

    # Row 5: Title
    ws.merge_cells('A5:N5')
    ws['A5'] = "ព្រឹត្តិបត្រការងារប្រចាំឆ្នាំរបស់មន្ត្រីរាជការស៊ីវិល"
    ws['A5'].font = font_title
    ws['A5'].alignment = align_center
    ws.row_dimensions[5].height = 28

    # Row 7: Officer Info
    o = data['officer']
    g_kh = data['gender_kh']
    ws.merge_cells('A7:N7')
    ws['A7'] = f"ឈ្មោះ៖  {o.full_name_kh}        ភេទ៖  {g_kh}        ថ្ងៃខែឆ្នាំកំណើត៖  {_format_d1_dob(o.dob)}        អត្តលេខ៖  {o.officer_id_number or '-'}"
    ws['A7'].font = font_sub_info
    ws['A7'].alignment = align_left
    ws.row_dimensions[7].height = 22

    # Table Headers: Rows 9, 10, 11, 12
    ws.merge_cells('A9:M9')
    ws['A9'] = "ប្រភេទនៃច្បាប់ឈប់សម្រាករបស់មន្ត្រីរាជការស៊ីវិល"
    ws['A9'].font = font_th_large
    ws['A9'].alignment = align_center

    ws.merge_cells('N9:N12')
    ws['N9'] = "ទណ្ឌកម្មវិន័យ\nនិងវិធានការ"
    ws['N9'].font = font_th_bold
    ws['N9'].alignment = align_center

    # Row 10: The 5 Categories
    ws.merge_cells('A10:F10')
    ws['A10'] = "១. ច្បាប់ឈប់ប្រចាំឆ្នាំ\nមានចំនួន ១៥ថ្ងៃនៃថ្ងៃធ្វើការ/១ឆ្នាំ"
    ws['A10'].font = font_th_bold
    ws['A10'].alignment = align_center

    ws.merge_cells('G10:H10')
    ws['G10'] = "២. ច្បាប់ឈប់រយៈពេលខ្លី\nមានចំនួន ១៥ថ្ងៃ\nនៃថ្ងៃធ្វើការ/១ឆ្នាំ"
    ws['G10'].font = font_th_bold
    ws['G10'].alignment = align_center

    ws.merge_cells('I10:I12')
    ws['I10'] = "៣. ច្បាប់\nឈប់សម្រាក\nលំហែ\nមាតុភាព\nមានចំនួន\n៣ខែ"
    ws['I10'].font = font_th_bold
    ws['I10'].alignment = align_center

    ws.merge_cells('J10:K10')
    ws['J10'] = "៤. ច្បាប់ឈប់សម្រាកព្យាបាលជំងឺ\nមានចំនួន ១២ខែ\n(សម្រាប់មួយជីវិតជាមន្ត្រី)"
    ws['J10'].font = font_th_bold
    ws['J10'].alignment = align_center

    ws.merge_cells('L10:M10')
    ws['L10'] = "៥. ច្បាប់ឈប់សម្រាកដោយមាន\nកិច្ចផ្ទាល់ខ្លួន (៣ខែ)\n(សម្រាប់មួយជីវិតជាមន្ត្រី)"
    ws['L10'].font = font_th_bold
    ws['L10'].alignment = align_center

    # Row 11: Sub-headers
    ws.merge_cells('A11:A12')
    ws['A11'] = "ចំនួនថ្ងៃដែល\nមិនទាន់បាន\nអនុវត្តក្នុង\nឆ្នាំចាស់"
    ws['A11'].font = font_th_normal
    ws['A11'].alignment = align_center

    ws.merge_cells('B11:C11')
    ws['B11'] = "ចំនួនថ្ងៃដែលត្រូវផ្លាស់ប្តូរ"
    ws['B11'].font = font_th_normal
    ws['B11'].alignment = align_center

    ws.merge_cells('D11:D12')
    ws['D11'] = "ចំនួនថ្ងៃឈប់\nសម្រាកប្រចាំ\nឆ្នាំបន្ថែម"
    ws['D11'].font = font_th_normal
    ws['D11'].alignment = align_center

    ws.merge_cells('E11:E12')
    ws['E11'] = "ចំនួនថ្ងៃដែល\nបានអនុវត្ត"
    ws['E11'].font = font_th_normal
    ws['E11'].alignment = align_center

    ws.merge_cells('F11:F12')
    ws['F11'] = "ចំនួនថ្ងៃដែល\nមិនទាន់បាន\nអនុវត្ត"
    ws['F11'].font = font_th_normal
    ws['F11'].alignment = align_center

    ws.merge_cells('G11:G12')
    ws['G11'] = "ចំនួនថ្ងៃដែល\nបានអនុវត្ត"
    ws['G11'].font = font_th_normal
    ws['G11'].alignment = align_center

    ws.merge_cells('H11:H12')
    ws['H11'] = "ចំនួនថ្ងៃដែល\nមិនទាន់បាន\nអនុវត្ត"
    ws['H11'].font = font_th_normal
    ws['H11'].alignment = align_center

    ws.merge_cells('J11:J12')
    ws['J11'] = "ចំនួនខែដែល\nបានអនុវត្ត"
    ws['J11'].font = font_th_normal
    ws['J11'].alignment = align_center

    ws.merge_cells('K11:K12')
    ws['K11'] = "ចំនួនខែដែល\nមិនទាន់បាន\nអនុវត្ត"
    ws['K11'].font = font_th_normal
    ws['K11'].alignment = align_center

    ws.merge_cells('L11:L12')
    ws['L11'] = "ចំនួនខែដែល\nបានអនុវត្ត"
    ws['L11'].font = font_th_normal
    ws['L11'].alignment = align_center

    ws.merge_cells('M11:M12')
    ws['M11'] = "ចំនួនខែដែល\nមិនទាន់បាន\nអនុវត្ត"
    ws['M11'].font = font_th_normal
    ws['M11'].alignment = align_center

    # Row 12: B12 & C12
    ws['B12'] = "ផ្ទាល់"
    ws['B12'].font = font_th_normal
    ws['B12'].alignment = align_center

    ws['C12'] = "ឈប់សម្រាក\nផ្សេងៗ"
    ws['C12'].font = font_th_normal
    ws['C12'].alignment = align_center

    # Apply thin border to all cells in rows 9-12
    for r in range(9, 13):
        for c in range(1, 15):
            ws.cell(r, c).border = thin_border

    # Row 13: Data Row
    data_row = [
        data['annual_prev_unused'],
        data['annual_transfer_direct'],
        data['annual_transfer_other'],
        data['annual_additional'],
        data['annual_used'],
        data['annual_remaining'],
        data['short_used'],
        data['short_remaining'],
        data['maternity_display'],
        data['sick_used_months'],
        data['sick_remaining_months'],
        data['personal_used_months'],
        data['personal_remaining_months'],
        data['disciplinary_str'],
    ]
    ws.row_dimensions[13].height = 26
    for c_idx, val in enumerate(data_row, 1):
        cell = ws.cell(13, c_idx)
        cell.value = val
        cell.font = font_data
        cell.alignment = align_data_center
        cell.border = thin_border

    # Rows 14 to 16: Extra lined rows matching official printed ledger
    for r_idx in range(14, 17):
        ws.row_dimensions[r_idx].height = 24
        for c_idx in range(1, 15):
            cell = ws.cell(r_idx, c_idx)
            cell.value = ""
            cell.border = thin_border

    # Footer Notes & Signatures
    ws['A18'] = f"- អវត្តមានមានច្បាប់អនុញ្ញាតប្រចាំឆ្នាំ ៖  {data['authorized_leave_days']} ថ្ងៃ"
    ws['A18'].font = font_notes
    ws['A19'] = f"- អវត្តមានគ្មានច្បាប់អនុញ្ញាតប្រចាំឆ្នាំ ៖  {data['unauthorized_absent_days']} ថ្ងៃ"
    ws['A19'].font = font_notes
    ws['A20'] = f"- ទំនេរគ្មានបៀវត្ស ៖  {data['unpaid_leave_days']} ថ្ងៃ"
    ws['A20'].font = font_notes
    ws['A21'] = f"- ផ្សេងៗ ៖  {data['other_leave_days']} ថ្ងៃ"
    ws['A21'].font = font_notes

    disc_text = (
        "សម្គាល់ ៖ ទណ្ឌកម្មវិន័យអនុវត្តចំពោះអវត្តមានគ្មានច្បាប់អនុញ្ញាត៖\n"
        "១- មន្ត្រីរាជការស៊ីវិល\n"
        "  - ការស្តីបន្ទោស\n"
        "  - ការស្តីបន្ទោសដោយមានចំណារក្នុងសំណុំលិខិតផ្ទាល់ខ្លួន\n"
        "  - ការផ្លាស់ដោយបង្ខំតាមវិធានការវិន័យឬការលុបឈ្មោះចេញពីតារាងដំឡើងថ្នាក់ឬឋានន្តរស័ក្តិ\n"
        "  - ការលុបឈ្មោះចេញពីក្របខណ្ឌ។\n"
        "២- មន្ត្រីជាប់កិច្ចសន្យា\n"
        "  - ណែនាំលើកទី១\n"
        "  - ណែនាំចុងក្រោយ\n"
        "  - លុបឈ្មោះពីអង្គភាពសាមី។"
    )
    ws.merge_cells('A23:G29')
    ws['A23'] = disc_text
    ws['A23'].font = Font(name='Khmer OS Battambang', size=8.5)
    ws['A23'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws.merge_cells('I18:N18')
    ws['I18'] = f"ធ្វើនៅប៉ៃលិន, ថ្ងៃទី....... ខែ............... ឆ្នាំ{data['year_kh']}"
    ws['I18'].font = font_notes
    ws['I18'].alignment = align_center

    ws.merge_cells('I19:N19')
    ws['I19'] = "ប្រធានអង្គភាព"
    ws['I19'].font = Font(name='Khmer OS Muol Light', size=10)
    ws['I19'].alignment = align_center

    leave_desc = (
        "ប្រភេទច្បាប់ឈប់សម្រាករបស់មន្ត្រីរាជការស៊ីវិលរួមមាន៖\n"
        "១- ច្បាប់ឈប់ប្រចាំឆ្នាំ មានរយៈពេល១៥ថ្ងៃនៃថ្ងៃធ្វើការ/១ឆ្នាំ\n"
        "២- ច្បាប់ឈប់រយៈពេលខ្លី មានរយៈពេល១៥ថ្ងៃនៃថ្ងៃធ្វើការ/១ឆ្នាំ\n"
        "៣- ច្បាប់ឈប់សម្រាកលំហែមាតុភាព មានរយៈពេល៣ខែ\n"
        "៤- ច្បាប់ឈប់សម្រាកព្យាបាលជំងឺ មានរយៈពេល១២ខែក្នុងអំឡុងពេលបម្រើការងារជាមន្ត្រី\n"
        "៥- ច្បាប់ឈប់សម្រាកដោយមានកិច្ចការផ្ទាល់ខ្លួន មានរយៈពេល៣ខែក្នុងអំឡុងពេលបម្រើការងារជាមន្ត្រី"
    )
    ws.merge_cells('I23:N29')
    ws['I23'] = leave_desc
    ws['I23'].font = Font(name='Khmer OS Battambang', size=8.5)
    ws['I23'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    return wb


@login_required
def attendance_annual_bulletin(request):
    """
    Annual Civil Servant Work / Leave Bulletin (ព្រឹត្តិបត្រការងារប្រចាំឆ្នាំរបស់មន្ត្រីរាជការស៊ីវិល)
    Annex 4 of Sub-Decree 58 ANKr.BK dated 01 September 2016.
    """
    profile = getattr(request.user, 'profile', None)
    user_dept = profile.department if profile else None
    is_admin = _is_admin_user(request.user)

    now = timezone.now()
    year_str = request.GET.get('year', str(now.year)).strip()
    try:
        year = int(to_arabic_digits(year_str))
    except Exception:
        year = now.year

    # Department handling with strict isolation
    if is_admin:
        dept_id = request.GET.get('department', '').strip()
        if dept_id:
            dept = Department.objects.filter(id=dept_id, is_active=True).first()
        else:
            dept = user_dept or Department.objects.filter(is_active=True).order_by('order_index').first()
        departments = Department.objects.filter(is_active=True).order_by('order_index', 'name_kh')
    else:
        dept = user_dept
        dept_id = str(dept.id) if dept else ''
        departments = Department.objects.filter(id=dept.id) if dept else []

    # Get civil servants for selector
    if is_admin and not dept_id:
        officers_qs = CivilServantProfile.objects.exclude(
            officer_status__in=['DISMISSED', 'RETIRED', 'TRANSFERRED_OUT']
        ).select_related('department').order_by('department__order_index', 'khmer_last_name', 'khmer_first_name')
    elif dept:
        officers_qs = CivilServantProfile.objects.filter(
            department=dept
        ).exclude(
            officer_status__in=['DISMISSED', 'RETIRED', 'TRANSFERRED_OUT']
        ).order_by('khmer_last_name', 'khmer_first_name')
    else:
        officers_qs = CivilServantProfile.objects.none()

    officer_id = request.GET.get('officer_id', '').strip()
    selected_officer = None
    if officer_id:
        selected_officer = officers_qs.filter(id=officer_id).first()
    if not selected_officer and officers_qs.exists():
        selected_officer = officers_qs.first()

    bulletin_data = None
    if selected_officer:
        bulletin_data = _build_annual_bulletin_data(selected_officer, year)

    years = list(range(now.year + 1, 2019, -1))

    context = {
        'dept': dept,
        'departments': departments,
        'officers': officers_qs,
        'selected_officer': selected_officer,
        'selected_officer_id': selected_officer.id if selected_officer else None,
        'year': year,
        'years': years,
        'bulletin': bulletin_data,
        'is_admin': is_admin,
    }
    return render(request, 'dms/attendance_annual_bulletin.html', context)


@login_required
def attendance_annual_bulletin_print(request):
    """
    Official A4 Landscape Printable / PDF Preview format for Annual Work Bulletin
    (ព្រឹត្តិបត្រការងារប្រចាំឆ្នាំរបស់មន្ត្រីរាជការស៊ីវិល - ឧបសម្ព័ន្ធទី៤ នៃអនុក្រឹត្យលេខ ៥៨).
    """
    officer_id = request.GET.get('officer_id', '').strip()
    if not officer_id:
        messages.warning(request, '⚠️ សូមជ្រើសរើសមន្ត្រីរាជការដើម្បីបោះពុម្ពព្រឹត្តិបត្រការងារ!')
        return redirect('attendance_annual_bulletin')

    officer = get_object_or_404(CivilServantProfile, pk=officer_id)

    profile = getattr(request.user, 'profile', None)
    is_admin = _is_admin_user(request.user)
    if not is_admin and profile and profile.department:
        if officer.department_id != profile.department_id:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិមើលព្រឹត្តិបត្រការងាររបស់មន្ត្រីនៃអង្គភាពផ្សេងឡើយ!')
            return redirect('attendance_annual_bulletin')

    now = timezone.now()
    year_str = request.GET.get('year', str(now.year)).strip()
    try:
        year = int(to_arabic_digits(year_str))
    except Exception:
        year = now.year

    bulletin_data = _build_annual_bulletin_data(officer, year)

    context = {
        'bulletin': bulletin_data,
        'officer': officer,
        'year': year,
        'query_params': request.GET.urlencode(),
    }
    return render(request, 'dms/attendance_annual_bulletin_print.html', context)


@login_required
def attendance_annual_bulletin_export_excel(request):
    """
    Exports the official Annual Civil Servant Work Bulletin to Excel (.xlsx)
    matching Sub-Decree 58 Annex 4 layout.
    """
    officer_id = request.GET.get('officer_id', '').strip()
    if not officer_id:
        messages.warning(request, '⚠️ សូមជ្រើសរើសមន្ត្រីរាជការដើម្បីទាញយកព្រឹត្តិបត្រការងារ!')
        return redirect('attendance_annual_bulletin')

    officer = get_object_or_404(CivilServantProfile, pk=officer_id)

    profile = getattr(request.user, 'profile', None)
    is_admin = _is_admin_user(request.user)
    if not is_admin and profile and profile.department:
        if officer.department_id != profile.department_id:
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិទាញយកព្រឹត្តិបត្រការងាររបស់មន្ត្រីនៃអង្គភាពផ្សេងឡើយ!')
            return redirect('attendance_annual_bulletin')

    now = timezone.now()
    year_str = request.GET.get('year', str(now.year)).strip()
    try:
        year = int(to_arabic_digits(year_str))
    except Exception:
        year = now.year

    bulletin_data = _build_annual_bulletin_data(officer, year)
    wb = _build_annual_bulletin_workbook(bulletin_data)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = officer.full_name_latin or officer.officer_id_number or str(officer.id)
    response['Content-Disposition'] = f'attachment; filename="annual_bulletin_{safe_name}_{year}.xlsx"'
    wb.save(response)
    return response


# ==============================================================================
# 👑 LEADERSHIP & HEADS ATTENDANCE MODULE (គំរូ Attendend B - ថ្នាក់ដឹកនាំ ខណ្ឌ និងប្រធានការិយាល័យ)
# ==============================================================================

def _can_manage_leadership_attendance(user):
    """
    Leadership Attendance (Attendend B) is strictly restricted to:
    1. User 'ADMIN' or superuser
    2. Users belonging to the Administration & Personnel Office (ការិយាល័យរដ្ឋបាល-បុគ្គលិក)
    """
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.username.upper() == 'ADMIN':
        return True
    profile = getattr(user, 'profile', None)
    if profile and profile.department:
        dept_code = (profile.department.code or '').strip().upper()
        dept_name = (profile.department.name_kh or '').strip()
        if dept_code in ['ADMIN_PERS', 'ADMIN_PERSONNEL'] or ('រដ្ឋបាល' in dept_name and 'បុគ្គលិក' in dept_name):
            return True
    return False


# Note: _get_leadership_roster() is canonically defined in the attendance header above.


@login_required
def attendance_leadership_daily(request):
    """
    Daily Attendance for Department Leadership, Canton Heads, and Office Chiefs (Attendend B).
    Strictly restricted to User ADMIN and Administration & Personnel Office (ការិយាល័យរដ្ឋបាល-បុគ្គលិក).
    """
    if not _can_manage_leadership_attendance(request.user):
        messages.error(request, '⚠️ មុខងារគ្រប់គ្រងវត្តមានថ្នាក់ដឹកនាំ (Attendend B) ត្រូវបានអនុញ្ញាតសម្រាប់តែ User: ADMIN និងមន្ត្រីនៃការិយាល័យរដ្ឋបាល-បុគ្គលិក ប៉ុណ្ណោះ!')
        return redirect('attendance_monthly_sheet')

    profile = getattr(request.user, 'profile', None)
    is_admin = _is_admin_user(request.user)

    date_str = request.GET.get('date', '').strip()
    if date_str:
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            target_date = timezone.now().date()
    else:
        target_date = timezone.now().date()

    sec_lead, sec_canton, sec_office = _get_leadership_roster()
    all_leaders = sec_lead + sec_canton + sec_office

    # POST: Save Leadership Daily Attendance
    if request.method == 'POST':
        if not (is_admin or is_lead):
            messages.error(request, '⚠️ លោកអ្នកគ្មានសិទ្ធិកត់ត្រាវត្តមានសម្រាប់ថ្នាក់ដឹកនាំឡើយ!')
            return redirect('attendance_leadership_daily')

        posted_date_str = request.POST.get('target_date', '').strip()
        if posted_date_str:
            try:
                target_date = datetime.strptime(posted_date_str, '%Y-%m-%d').date()
            except Exception:
                pass

        person_keys = request.POST.getlist('person_keys')
        saved_count = 0

        for pkey in person_keys:
            parts = pkey.split('_')
            if len(parts) < 3:
                continue
            pid = int(parts[2])

            # Find leader info
            leader_match = next((l for l in all_leaders if l['person_id'] == pid), None)
            if not leader_match:
                continue

            status = request.POST.get(f'status_{pkey}', 'PRESENT')
            morning_in = request.POST.get(f'morning_in_{pkey}') == '1'
            morning_out = request.POST.get(f'morning_out_{pkey}') == '1'
            afternoon_in = request.POST.get(f'afternoon_in_{pkey}') == '1'
            afternoon_out = request.POST.get(f'afternoon_out_{pkey}') == '1'
            is_late = request.POST.get(f'is_late_{pkey}') == '1'
            is_early_out = request.POST.get(f'is_early_out_{pkey}') == '1'
            leave_type = request.POST.get(f'leave_type_{pkey}', '').strip()
            reference_doc = request.POST.get(f'reference_doc_{pkey}', '').strip()
            disciplinary_measure = request.POST.get(f'disciplinary_measure_{pkey}', '').strip()
            remarks = request.POST.get(f'remarks_{pkey}', '').strip()
            pos_title = leader_match['position']

            if status not in ['LEAVE_PERMISSION', 'UNPAID_LEAVE']:
                leave_type = ''

            AttendanceRecord.objects.update_or_create(
                department=leader_match['department'],
                date=target_date,
                person_type='CIVIL_SERVANT',
                officer_id=pid,
                defaults={
                    'contract_officer': None,
                    'position_title': pos_title,
                    'morning_in': morning_in,
                    'morning_out': morning_out,
                    'afternoon_in': afternoon_in,
                    'afternoon_out': afternoon_out,
                    'status': status,
                    'is_late': is_late,
                    'is_early_out': is_early_out,
                    'leave_type': leave_type,
                    'reference_doc': reference_doc,
                    'disciplinary_measure': disciplinary_measure,
                    'remarks': remarks,
                    'recorded_by': request.user,
                }
            )
            saved_count += 1

        messages.success(request, f"✅ បានរក្សាទុកវត្តមានថ្នាក់ដឹកនាំ និងប្រធានអង្គភាពចំនួន {saved_count} រូប សម្រាប់កាលបរិច្ឆេទ {target_date.strftime('%d/%m/%Y')} ដោយជោគជ័យ!")
        return redirect(f"{reverse('attendance_leadership_daily')}?date={target_date.strftime('%Y-%m-%d')}")

    # Build Existing Records
    officer_ids = [l['person_id'] for l in all_leaders]
    existing_records = AttendanceRecord.objects.filter(
        date=target_date,
        person_type='CIVIL_SERVANT',
        officer_id__in=officer_ids
    )

    record_map = {f"CIVIL_SERVANT_{r.officer_id}": r for r in existing_records}

    def _attach_records(sec_list):
        result = []
        for item in sec_list:
            pkey = f"CIVIL_SERVANT_{item['person_id']}"
            rec = record_map.get(pkey)
            if rec:
                status = rec.status
                morning_in = rec.morning_in
                morning_out = rec.morning_out
                afternoon_in = rec.afternoon_in
                afternoon_out = rec.afternoon_out
                is_late = rec.is_late
                is_early_out = rec.is_early_out
                leave_type = rec.leave_type
                reference_doc = rec.reference_doc
                remarks = rec.remarks
                is_recorded = True
            else:
                status = 'PRESENT'
                morning_in = True
                morning_out = True
                afternoon_in = True
                afternoon_out = True
                is_late = False
                is_early_out = False
                leave_type = ''
                reference_doc = ''
                remarks = ''
                is_recorded = False

            result.append({
                'person': item,
                'pkey': pkey,
                'status': status,
                'morning_in': morning_in,
                'morning_out': morning_out,
                'afternoon_in': afternoon_in,
                'afternoon_out': afternoon_out,
                'is_late': is_late,
                'is_early_out': is_early_out,
                'leave_type': leave_type,
                'reference_doc': reference_doc,
                'remarks': remarks,
                'is_recorded': is_recorded,
            })
        return result

    rows_lead = _attach_records(sec_lead)
    rows_canton = _attach_records(sec_canton)
    rows_office = _attach_records(sec_office)

    all_rows = rows_lead + rows_canton + rows_office
    total_leaders = len(all_rows)
    present_count = sum(1 for r in all_rows if r['status'] == 'PRESENT')
    mission_count = sum(1 for r in all_rows if r['status'] == 'MISSION')
    leave_count = sum(1 for r in all_rows if r['status'] in ['LEAVE_PERMISSION', 'UNPAID_LEAVE'])
    absent_count = sum(1 for r in all_rows if r['status'] == 'ABSENT_NO_LEAVE')

    prev_date = target_date - timedelta(days=1)
    next_date = target_date + timedelta(days=1)
    is_today = (target_date == timezone.now().date())

    context = {
        'target_date': target_date,
        'target_date_str': target_date.strftime('%Y-%m-%d'),
        'prev_date_str': prev_date.strftime('%Y-%m-%d'),
        'next_date_str': next_date.strftime('%Y-%m-%d'),
        'is_today': is_today,
        'rows_lead': rows_lead,
        'rows_canton': rows_canton,
        'rows_office': rows_office,
        'total_leaders': total_leaders,
        'present_count': present_count,
        'mission_count': mission_count,
        'leave_count': leave_count,
        'absent_count': absent_count,
        'is_admin': is_admin or is_lead,
        'status_choices': AttendanceRecord.STATUS_CHOICES,
        'leave_choices': AttendanceRecord.LEAVE_TYPE_CHOICES,
    }
    return render(request, 'dms/attendance_leadership_daily.html', context)


def _build_leadership_weekly_context(request):
    """
    Builds weekly leadership attendance context matching Attendend B.xlsx (ថ្នាក់ដឹកនាំប្រចាំសប្តាហ).
    """
    now = timezone.now()
    date_str = request.GET.get('date', '').strip()
    if date_str:
        try:
            base_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except Exception:
            base_date = now.date()
    else:
        base_date = now.date()

    monday = base_date - timedelta(days=base_date.weekday())
    friday = monday + timedelta(days=4)
    prev_week_monday = monday - timedelta(days=7)
    next_week_monday = monday + timedelta(days=7)
    is_current_week = (monday <= now.date() <= monday + timedelta(days=6))

    sec_lead, sec_canton, sec_office = _get_leadership_roster()
    all_leaders = sec_lead + sec_canton + sec_office
    officer_ids = [l['person_id'] for l in all_leaders]

    records = AttendanceRecord.objects.filter(
        date__gte=monday,
        date__lte=friday,
        person_type='CIVIL_SERVANT',
        officer_id__in=officer_ids
    )

    rec_by_officer = {}
    for r in records:
        rec_by_officer.setdefault(r.officer_id, []).append(r)

    def _process_section(sec_list):
        result = []
        for item in sec_list:
            recs = rec_by_officer.get(item['person_id'], [])
            present = sum(1 for r in recs if r.status == 'PRESENT')
            mission = sum(1 for r in recs if r.status == 'MISSION')
            leave = sum(1 for r in recs if r.status in ['LEAVE_PERMISSION', 'UNPAID_LEAVE'])
            absent = sum(1 for r in recs if r.status == 'ABSENT_NO_LEAVE')
            late_or_early = sum(1 for r in recs if r.is_late or r.is_early_out)

            leave_types = list(set([r.get_leave_type_display() for r in recs if r.leave_type]))
            leave_type_str = ', '.join(leave_types) if leave_types else ''

            remarks_list = list(set([r.remarks for r in recs if r.remarks]))
            remarks_str = ', '.join(remarks_list) if remarks_list else ''

            disciplinary_list = list(set([r.disciplinary_measure for r in recs if r.disciplinary_measure]))
            disciplinary_str = ', '.join(disciplinary_list) if disciplinary_list else ''

            result.append({
                'person': item,
                'present': present,
                'mission': mission,
                'leave': leave,
                'absent': absent,
                'late_or_early': late_or_early,
                'leave_type': leave_type_str,
                'disciplinary': disciplinary_str,
                'remarks': remarks_str,
            })
        return result

    rows_lead = _process_section(sec_lead)
    rows_canton = _process_section(sec_canton)
    rows_office = _process_section(sec_office)

    all_rows = rows_lead + rows_canton + rows_office
    total_leaders = len(all_rows)
    tot_present = sum(r['present'] for r in all_rows)
    tot_mission = sum(r['mission'] for r in all_rows)
    tot_leave = sum(r['leave'] for r in all_rows)
    tot_absent = sum(r['absent'] for r in all_rows)
    tot_late_early = sum(r['late_or_early'] for r in all_rows)

    iso_year, iso_week, _ = monday.isocalendar()

    return {
        'base_date': base_date,
        'base_date_str': base_date.strftime('%Y-%m-%d'),
        'monday': monday,
        'friday': friday,
        'monday_str': monday.strftime('%Y-%m-%d'),
        'friday_str': friday.strftime('%Y-%m-%d'),
        'prev_week_monday_str': prev_week_monday.strftime('%Y-%m-%d'),
        'next_week_monday_str': next_week_monday.strftime('%Y-%m-%d'),
        'is_current_week': is_current_week,
        'iso_year': iso_year,
        'iso_week': iso_week,
        'rows_lead': rows_lead,
        'rows_canton': rows_canton,
        'rows_office': rows_office,
        'total_leaders': total_leaders,
        'tot_present': tot_present,
        'tot_mission': tot_mission,
        'tot_leave': tot_leave,
        'tot_absent': tot_absent,
        'tot_late_early': tot_late_early,
    }


@login_required
def attendance_leadership_weekly(request):
    """
    Weekly Attendance Report for Leadership, Canton & Office Heads (Attendend B).
    Strictly restricted to User ADMIN and Administration & Personnel Office.
    """
    if not _can_manage_leadership_attendance(request.user):
        messages.error(request, '⚠️ មុខងារគ្រប់គ្រងវត្តមានថ្នាក់ដឹកនាំ (Attendend B) ត្រូវបានអនុញ្ញាតសម្រាប់តែ User: ADMIN និងមន្ត្រីនៃការិយាល័យរដ្ឋបាល-បុគ្គលិក ប៉ុណ្ណោះ!')
        return redirect('attendance_monthly_sheet')
    context = _build_leadership_weekly_context(request)
    return render(request, 'dms/attendance_leadership_weekly.html', context)


@login_required
def attendance_leadership_weekly_print(request):
    """
    Print format matching Attendend B.xlsx sheet ថ្នាក់ដឹកនាំប្រចាំសប្តាហ.
    """
    if not _can_manage_leadership_attendance(request.user):
        messages.error(request, '⚠️ មុខងារគ្រប់គ្រងវត្តមានថ្នាក់ដឹកនាំ (Attendend B) ត្រូវបានអនុញ្ញាតសម្រាប់តែ User: ADMIN និងមន្ត្រីនៃការិយាល័យរដ្ឋបាល-បុគ្គលិក ប៉ុណ្ណោះ!')
        return redirect('attendance_monthly_sheet')
    context = _build_leadership_weekly_context(request)
    return render(request, 'dms/attendance_leadership_weekly_print.html', context)


@login_required
def attendance_leadership_weekly_export_excel(request):
    """
    Exports Weekly Leadership Attendance Report to Excel matching Attendend B.xlsx.
    """
    if not _can_manage_leadership_attendance(request.user):
        messages.error(request, '⚠️ មុខងារគ្រប់គ្រងវត្តមានថ្នាក់ដឹកនាំ (Attendend B) ត្រូវបានអនុញ្ញាតសម្រាប់តែ User: ADMIN និងមន្ត្រីនៃការិយាល័យរដ្ឋបាល-បុគ្គលិក ប៉ុណ្ណោះ!')
        return redirect('attendance_monthly_sheet')
    context = _build_leadership_weekly_context(request)
    monday = context['monday']
    friday = context['friday']
    iso_week = context['iso_week']
    iso_year = context['iso_year']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"ថ្នាក់ដឹកនាំ_សប្តាហ៍_{iso_week}"

    font_title = Font(name="Khmer OS Muol Light", size=12, bold=True, color="1E3A8A")
    font_sub = Font(name="Khmer OS Battambang", size=10, bold=True)
    font_sec = Font(name="Khmer OS Battambang", size=10, bold=True, color="1E40AF")
    font_header = Font(name="Khmer OS Battambang", size=9, bold=True, color="FFFFFF")
    font_cell = Font(name="Khmer OS Battambang", size=9)
    font_cell_bold = Font(name="Khmer OS Battambang", size=9, bold=True)

    fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fill_sec = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_summary = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.merge_cells('A1:I1')
    ws['A1'] = f"របាយការណ៍វត្តមានរបស់ថ្នាក់ដឹកនាំមន្ទីរ ខណ្ឌរដ្ឋបាល និងប្រធានការិយាល័យ"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center

    ws.merge_cells('A2:I2')
    ws['A2'] = f"ប្រចាំសប្តាហ៍ទី {iso_week} ឆ្នាំ {iso_year} (ពីថ្ងៃទី {monday.strftime('%d/%m/%Y')} ដល់ {friday.strftime('%d/%m/%Y')})"
    ws['A2'].font = font_sub
    ws['A2'].alignment = align_center

    # Header Row 4 & 5
    headers = [
        ('A4:A5', 'A4', 'ល.រ', 5),
        ('B4:B5', 'B4', 'គោត្តនាម-នាម', 22),
        ('C4:C5', 'C4', 'ភេទ', 6),
        ('D4:D5', 'D4', 'តួនាទី / មុខតំណែង', 28),
        ('E4:E5', 'E4', 'មានច្បាប់', 10),
        ('F4:F5', 'F4', 'អត់ច្បាប់', 10),
        ('G4:G5', 'G4', 'យឺត/មុន', 10),
        ('H4:H5', 'H4', 'ប្រភេទឈប់សម្រាក', 20),
        ('I4:I5', 'I4', 'ផ្សេងៗ', 18),
    ]

    for merge_range, cell_ref, label, width in headers:
        ws.merge_cells(merge_range)
        ws[cell_ref] = label
        ws[cell_ref].font = font_header
        ws[cell_ref].fill = fill_header
        ws[cell_ref].alignment = align_center
        ws[cell_ref].border = thin_border
        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width

    row_num = 6
    sections_data = [
        ('I. ថ្នាក់ដឹកនាំមន្ទីរ', context['rows_lead']),
        ('II. ខណ្ឌរដ្ឋបាល', context['rows_canton']),
        ('III. ប្រធានការិយាល័យ', context['rows_office']),
    ]

    for sec_title, sec_rows in sections_data:
        # Section Header Row
        ws.merge_cells(f'A{row_num}:I{row_num}')
        ws[f'A{row_num}'] = sec_title
        ws[f'A{row_num}'].font = font_sec
        ws[f'A{row_num}'].fill = fill_sec
        ws[f'A{row_num}'].alignment = align_left
        ws[f'A{row_num}'].border = thin_border
        row_num += 1

        for r in sec_rows:
            p = r['person']
            ws[f'A{row_num}'] = p['index']
            ws[f'A{row_num}'].alignment = align_center
            ws[f'A{row_num}'].font = font_cell
            ws[f'A{row_num}'].border = thin_border

            ws[f'B{row_num}'] = p['name']
            ws[f'B{row_num}'].alignment = align_left
            ws[f'B{row_num}'].font = font_cell_bold
            ws[f'B{row_num}'].border = thin_border

            ws[f'C{row_num}'] = p['gender']
            ws[f'C{row_num}'].alignment = align_center
            ws[f'C{row_num}'].font = font_cell
            ws[f'C{row_num}'].border = thin_border

            ws[f'D{row_num}'] = p['position']
            ws[f'D{row_num}'].alignment = align_left
            ws[f'D{row_num}'].font = font_cell
            ws[f'D{row_num}'].border = thin_border

            ws[f'E{row_num}'] = r['leave'] if r['leave'] > 0 else ""
            ws[f'E{row_num}'].alignment = align_center
            ws[f'E{row_num}'].font = font_cell_bold
            ws[f'E{row_num}'].border = thin_border

            ws[f'F{row_num}'] = r['absent'] if r['absent'] > 0 else ""
            ws[f'F{row_num}'].alignment = align_center
            ws[f'F{row_num}'].font = font_cell_bold
            ws[f'F{row_num}'].border = thin_border

            ws[f'G{row_num}'] = r['late_or_early'] if r['late_or_early'] > 0 else ""
            ws[f'G{row_num}'].alignment = align_center
            ws[f'G{row_num}'].font = font_cell
            ws[f'G{row_num}'].border = thin_border

            ws[f'H{row_num}'] = r['leave_type']
            ws[f'H{row_num}'].alignment = align_left
            ws[f'H{row_num}'].font = font_cell
            ws[f'H{row_num}'].border = thin_border

            ws[f'I{row_num}'] = r['remarks']
            ws[f'I{row_num}'].alignment = align_left
            ws[f'I{row_num}'].font = font_cell
            ws[f'I{row_num}'].border = thin_border

            row_num += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_leadership_weekly_{iso_year}_W{iso_week:02d}.xlsx"'
    wb.save(response)
    return response


def _build_leadership_monthly_context(request):
    """
    Builds monthly leadership attendance context matching Attendend B.xlsx (ថ្នាក់ដឹកនាំប្រចាំខែ).
    """
    now = timezone.now()
    year_str = request.GET.get('year', str(now.year)).strip()
    month_str = request.GET.get('month', str(now.month)).strip()

    try:
        year = int(to_arabic_digits(year_str))
    except Exception:
        year = now.year

    try:
        month = int(to_arabic_digits(month_str))
        if month < 1 or month > 12:
            month = now.month
    except Exception:
        month = now.month

    sec_lead, sec_canton, sec_office = _get_leadership_roster()
    all_leaders = sec_lead + sec_canton + sec_office
    officer_ids = [l['person_id'] for l in all_leaders]

    records = AttendanceRecord.objects.filter(
        date__year=year,
        date__month=month,
        person_type='CIVIL_SERVANT',
        officer_id__in=officer_ids
    )

    rec_by_officer = {}
    for r in records:
        rec_by_officer.setdefault(r.officer_id, []).append(r)

    def _process_section(sec_list):
        result = []
        for item in sec_list:
            recs = rec_by_officer.get(item['person_id'], [])
            present = sum(1 for r in recs if r.status == 'PRESENT')
            mission = sum(1 for r in recs if r.status == 'MISSION')
            leave = sum(1 for r in recs if r.status in ['LEAVE_PERMISSION', 'UNPAID_LEAVE'])
            absent = sum(1 for r in recs if r.status == 'ABSENT_NO_LEAVE')
            late_or_early = sum(1 for r in recs if r.is_late or r.is_early_out)

            leave_types = list(set([r.get_leave_type_display() for r in recs if r.leave_type]))
            leave_type_str = ', '.join(leave_types) if leave_types else ''

            remarks_list = list(set([r.remarks for r in recs if r.remarks]))
            remarks_str = ', '.join(remarks_list) if remarks_list else ''

            result.append({
                'person': item,
                'present': present,
                'mission': mission,
                'leave': leave,
                'absent': absent,
                'late_or_early': late_or_early,
                'leave_type': leave_type_str,
                'remarks': remarks_str,
            })
        return result

    rows_lead = _process_section(sec_lead)
    rows_canton = _process_section(sec_canton)
    rows_office = _process_section(sec_office)

    all_rows = rows_lead + rows_canton + rows_office
    total_leaders = len(all_rows)
    tot_present = sum(r['present'] for r in all_rows)
    tot_mission = sum(r['mission'] for r in all_rows)
    tot_leave = sum(r['leave'] for r in all_rows)
    tot_absent = sum(r['absent'] for r in all_rows)
    tot_late_early = sum(r['late_or_early'] for r in all_rows)

    khmer_month_names = {
        1: 'មករា', 2: 'កុម្ភៈ', 3: 'មីនា', 4: 'មេសា',
        5: 'ឧសភា', 6: 'មិថុនា', 7: 'កក្កដា', 8: 'សីហា',
        9: 'កញ្ញា', 10: 'តុលា', 11: 'វិច្ឆិកា', 12: 'ធ្នូ'
    }

    year_options = [2025, 2026, 2027, 2028, 2029, 2030]
    month_options = [(i, khmer_month_names[i]) for i in range(1, 13)]

    return {
        'year': year,
        'month': month,
        'month_name_kh': khmer_month_names.get(month, ''),
        'rows_lead': rows_lead,
        'rows_canton': rows_canton,
        'rows_office': rows_office,
        'total_leaders': total_leaders,
        'tot_present': tot_present,
        'tot_mission': tot_mission,
        'tot_leave': tot_leave,
        'tot_absent': tot_absent,
        'tot_late_early': tot_late_early,
        'year_options': year_options,
        'month_options': month_options,
    }


@login_required
def attendance_leadership_monthly(request):
    """
    Monthly Attendance Matrix for Department Leadership, Canton & Office Heads (Attendend B).
    Strictly restricted to User ADMIN and Administration & Personnel Office.
    """
    if not _can_manage_leadership_attendance(request.user):
        messages.error(request, '⚠️ មុខងារគ្រប់គ្រងវត្តមានថ្នាក់ដឹកនាំ (Attendend B) ត្រូវបានអនុញ្ញាតសម្រាប់តែ User: ADMIN និងមន្ត្រីនៃការិយាល័យរដ្ឋបាល-បុគ្គលិក ប៉ុណ្ណោះ!')
        return redirect('attendance_monthly_sheet')
    context = _build_leadership_monthly_context(request)
    return render(request, 'dms/attendance_leadership_monthly.html', context)


@login_required
def attendance_leadership_monthly_print(request):
    """
    Print format matching Attendend B.xlsx sheet ថ្នាក់ដឹកនាំប្រចាំខែ.
    """
    if not _can_manage_leadership_attendance(request.user):
        messages.error(request, '⚠️ មុខងារគ្រប់គ្រងវត្តមានថ្នាក់ដឹកនាំ (Attendend B) ត្រូវបានអនុញ្ញាតសម្រាប់តែ User: ADMIN និងមន្ត្រីនៃការិយាល័យរដ្ឋបាល-បុគ្គលិក ប៉ុណ្ណោះ!')
        return redirect('attendance_monthly_sheet')
    context = _build_leadership_monthly_context(request)
    return render(request, 'dms/attendance_leadership_monthly_print.html', context)


@login_required
def attendance_leadership_monthly_export_excel(request):
    """
    Exports Monthly Leadership Attendance Report to Excel matching Attendend B.xlsx.
    """
    if not _can_manage_leadership_attendance(request.user):
        messages.error(request, '⚠️ មុខងារគ្រប់គ្រងវត្តមានថ្នាក់ដឹកនាំ (Attendend B) ត្រូវបានអនុញ្ញាតសម្រាប់តែ User: ADMIN និងមន្ត្រីនៃការិយាល័យរដ្ឋបាល-បុគ្គលិក ប៉ុណ្ណោះ!')
        return redirect('attendance_monthly_sheet')
    context = _build_leadership_monthly_context(request)
    year = context['year']
    month = context['month']
    month_name_kh = context['month_name_kh']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"ថ្នាក់ដឹកនាំ_{month_name_kh}_{year}"

    font_title = Font(name="Khmer OS Muol Light", size=12, bold=True, color="1E3A8A")
    font_sub = Font(name="Khmer OS Battambang", size=10, bold=True)
    font_sec = Font(name="Khmer OS Battambang", size=10, bold=True, color="1E40AF")
    font_header = Font(name="Khmer OS Battambang", size=9, bold=True, color="FFFFFF")
    font_cell = Font(name="Khmer OS Battambang", size=9)
    font_cell_bold = Font(name="Khmer OS Battambang", size=9, bold=True)

    fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fill_sec = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.merge_cells('A1:I1')
    ws['A1'] = f"របាយការណ៍វត្តមានរបស់ថ្នាក់ដឹកនាំមន្ទីរ ខណ្ឌរដ្ឋបាល និងប្រធានការិយាល័យ"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center

    ws.merge_cells('A2:I2')
    ws['A2'] = f"ប្រចាំខែ {month_name_kh} ឆ្នាំ {year}"
    ws['A2'].font = font_sub
    ws['A2'].alignment = align_center

    # Header Row 4 & 5
    headers = [
        ('A4:A5', 'A4', 'ល.រ', 5),
        ('B4:B5', 'B4', 'គោត្តនាម-នាម', 22),
        ('C4:C5', 'C4', 'ភេទ', 6),
        ('D4:D5', 'D4', 'តួនាទី / មុខតំណែង', 28),
        ('E4:E5', 'E4', 'មានច្បាប់', 10),
        ('F4:F5', 'F4', 'អត់ច្បាប់', 10),
        ('G4:G5', 'G4', 'យឺត/មុន', 10),
        ('H4:H5', 'H4', 'ប្រភេទឈប់សម្រាក', 20),
        ('I4:I5', 'I4', 'ផ្សេងៗ', 18),
    ]

    for merge_range, cell_ref, label, width in headers:
        ws.merge_cells(merge_range)
        ws[cell_ref] = label
        ws[cell_ref].font = font_header
        ws[cell_ref].fill = fill_header
        ws[cell_ref].alignment = align_center
        ws[cell_ref].border = thin_border
        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = width

    row_num = 6
    sections_data = [
        ('I. ថ្នាក់ដឹកនាំមន្ទីរ', context['rows_lead']),
        ('II. ខណ្ឌរដ្ឋបាល', context['rows_canton']),
        ('III. ប្រធានការិយាល័យ', context['rows_office']),
    ]

    for sec_title, sec_rows in sections_data:
        ws.merge_cells(f'A{row_num}:I{row_num}')
        ws[f'A{row_num}'] = sec_title
        ws[f'A{row_num}'].font = font_sec
        ws[f'A{row_num}'].fill = fill_sec
        ws[f'A{row_num}'].alignment = align_left
        ws[f'A{row_num}'].border = thin_border
        row_num += 1

        for r in sec_rows:
            p = r['person']
            ws[f'A{row_num}'] = p['index']
            ws[f'A{row_num}'].alignment = align_center
            ws[f'A{row_num}'].font = font_cell
            ws[f'A{row_num}'].border = thin_border

            ws[f'B{row_num}'] = p['name']
            ws[f'B{row_num}'].alignment = align_left
            ws[f'B{row_num}'].font = font_cell_bold
            ws[f'B{row_num}'].border = thin_border

            ws[f'C{row_num}'] = p['gender']
            ws[f'C{row_num}'].alignment = align_center
            ws[f'C{row_num}'].font = font_cell
            ws[f'C{row_num}'].border = thin_border

            ws[f'D{row_num}'] = p['position']
            ws[f'D{row_num}'].alignment = align_left
            ws[f'D{row_num}'].font = font_cell
            ws[f'D{row_num}'].border = thin_border

            ws[f'E{row_num}'] = r['leave'] if r['leave'] > 0 else ""
            ws[f'E{row_num}'].alignment = align_center
            ws[f'E{row_num}'].font = font_cell_bold
            ws[f'E{row_num}'].border = thin_border

            ws[f'F{row_num}'] = r['absent'] if r['absent'] > 0 else ""
            ws[f'F{row_num}'].alignment = align_center
            ws[f'F{row_num}'].font = font_cell_bold
            ws[f'F{row_num}'].border = thin_border

            ws[f'G{row_num}'] = r['late_or_early'] if r['late_or_early'] > 0 else ""
            ws[f'G{row_num}'].alignment = align_center
            ws[f'G{row_num}'].font = font_cell
            ws[f'G{row_num}'].border = thin_border

            ws[f'H{row_num}'] = r['leave_type']
            ws[f'H{row_num}'].alignment = align_left
            ws[f'H{row_num}'].font = font_cell
            ws[f'H{row_num}'].border = thin_border

            ws[f'I{row_num}'] = r['remarks']
            ws[f'I{row_num}'].alignment = align_left
            ws[f'I{row_num}'].font = font_cell
            ws[f'I{row_num}'].border = thin_border

            row_num += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_leadership_monthly_{year}_{month:02d}.xlsx"'
    wb.save(response)
    return response




